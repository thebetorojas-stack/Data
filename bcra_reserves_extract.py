#!/usr/bin/env python3
"""
BCRA Reserves Extractor & Historical Accumulator
=================================================

Reads a BCRA-downloaded ``*_series.xlsm`` (or .xlsx) file, extracts the
four key reserve series, and *merges* them into a persistent historical
workbook so old data is never lost.

Every run is idempotent:
    - new dates       -> appended
    - existing dates  -> values refreshed (BCRA does revise prior days)
    - dropped dates   -> kept in history (history never loses rows)

Series extracted (only these four; daily, Tipo de serie = 'D')
--------------------------------------------------------------
    1. Reservas Internacionales        (col C)  -- stock, USD MM
    2. Compra de Divisas               (col H)  -- daily BCRA FX intervention
    3. Pagos a Organismos Internac.    (col I)  -- net flows w/ IMF & multilaterals
    4. Tipo de Cambio USD/ARS          (col P)  -- BCRA reference rate

Workbook tabs
-------------
    Snapshot         -- one-page latest-print summary
    Monthly          -- the main analytical view: EoM stocks + monthly flows
                        + YTD Compra de Divisas + MoM/YoY reserves
    Daily            -- full daily history, sorted ascending
    Metadata         -- last update timestamp, source file, row counts

Refresh-in-place
----------------
The script *opens* the existing workbook on every run and replaces only the
four managed tabs above. ANY sheet you add yourself (your charts, your
custom analysis, your formulas referencing Daily/Monthly) is preserved
across runs. Build your work on NEW tabs — not on the four managed tabs,
because those get rewritten each run.

Usage
-----
    python3 scripts/bcra_reserves_extract.py [path_to_new_bcra_file.xlsm]

With no argument the most recently modified .xlsm/.xlsx in inputs/ is used.
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUTS_DIR   = PROJECT_ROOT / "inputs"
OUTPUTS_DIR  = PROJECT_ROOT / "outputs"
ARCHIVE_DIR  = INPUTS_DIR / "bcra_archive"
HIST_PATH    = OUTPUTS_DIR / "argentina_reserves_history.xlsx"

# -----------------------------------------------------------------------------
# BCRA RESERVAS sheet layout
# -----------------------------------------------------------------------------
SHEET           = "RESERVAS"
DATA_START_ROW  = 10
TIPO_SERIE_COL  = 17  # Q
DATE_COL        = 1   # A
COLS = [
    (3,  "Reservas"),
    (8,  "Compra_Divisas"),
    (9,  "Pagos_Organismos"),
    (16, "Tipo_Cambio"),
]
OUT_COLS = ["Fecha"] + [n for _, n in COLS]

# -----------------------------------------------------------------------------
# Styling — soft, scannable palette
# -----------------------------------------------------------------------------
INK        = "1F2A44"   # near-black navy, headline text
SUBTLE_INK = "5A6478"   # muted slate, secondary text
ACCENT     = "9D8CC4"   # soft violet accent
ACCENT_DK  = "5E4F8C"
BAND_LIGHT = "F7F7FA"   # very light grey, banded rows
BAND_NONE  = "FFFFFFFF" # transparent / white
HEADER_BG  = "EFEAF6"   # lavender wash for headers
SUBHEADER  = "E8EEF7"   # light blue wash for sub-headers
DIVIDER    = "D9D2EB"   # accent line between header & body
TOTAL_BG   = "F0EDE0"   # warm cream for total/highlight rows

FMT_USD     = "#,##0;[Red](#,##0)"
FMT_USD_1D  = "#,##0.0;[Red](#,##0.0)"
FMT_FX      = "#,##0.00"
FMT_PCT     = "0.0%;[Red](0.0%)"
FMT_DATE_D  = "yyyy-mm-dd"
FMT_DATE_M  = "mmm-yy"

THIN_DIV = Side(style="thin", color=DIVIDER)
NO_BORDER = Border()


def font(bold=False, size=10, color=INK, italic=False):
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)


def fill(rgb):
    return PatternFill("solid", fgColor=rgb)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def right():
    return Alignment(horizontal="right", vertical="center")


def left():
    return Alignment(horizontal="left", vertical="center", indent=1)


# -----------------------------------------------------------------------------
# Extraction
# -----------------------------------------------------------------------------
def extract_daily(src: Path) -> pd.DataFrame:
    print(f"[extract] reading {src.name} ...")
    wb = openpyxl.load_workbook(src, data_only=True, keep_vba=False, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' not found in {src.name}.")
    ws = wb[SHEET]

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if len(row) < TIPO_SERIE_COL or row[TIPO_SERIE_COL - 1] != "D":
            continue
        d = row[DATE_COL - 1]
        if not isinstance(d, (dt.datetime, dt.date)):
            continue
        rec = {"Fecha": pd.Timestamp(d).normalize()}
        for idx, name in COLS:
            v = row[idx - 1] if len(row) >= idx else None
            rec[name] = v
        rows.append(rec)
    wb.close()

    df = pd.DataFrame.from_records(rows, columns=OUT_COLS)
    df = df.sort_values("Fecha").drop_duplicates(subset=["Fecha"], keep="last")
    for _, n in COLS:
        df[n] = pd.to_numeric(df[n], errors="coerce")
    print(f"[extract] {len(df):,} daily rows "
          f"({df['Fecha'].min().date()} → {df['Fecha'].max().date()})")
    return df.reset_index(drop=True)


# -----------------------------------------------------------------------------
# Merge with existing history
# -----------------------------------------------------------------------------
def load_existing() -> pd.DataFrame | None:
    """Read the historical Daily tab. The workbook has decorative title rows
    above the actual header (header is on row 5, i.e. zero-index 4), so we
    pass header=4 to pandas. We also remap display column names back to the
    internal short names used by the merge logic."""
    if not HIST_PATH.exists():
        return None

    rename_map = {
        "Fecha":                       "Fecha",
        "Reservas (USD MM)":           "Reservas",
        "Compra de Divisas (USD MM)":  "Compra_Divisas",
        "Pagos a Organismos (USD MM)": "Pagos_Organismos",
        "Tipo de Cambio (USD/ARS)":    "Tipo_Cambio",
        # Backward compatibility with older raw schema
        "Reservas":                    "Reservas",
        "Compra_Divisas":              "Compra_Divisas",
        "Pagos_Organismos":            "Pagos_Organismos",
        "Tipo_Cambio":                 "Tipo_Cambio",
    }

    for header_row in (4, 0):   # try the styled layout first, then a flat one
        try:
            df = pd.read_excel(HIST_PATH, sheet_name="Daily", header=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            df = df.rename(columns=rename_map)
            if "Fecha" not in df.columns:
                continue
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.normalize()
            df = df.dropna(subset=["Fecha"])
            for c in [n for _, n in COLS]:
                if c not in df.columns:
                    df[c] = pd.NA
            df = df[OUT_COLS].sort_values("Fecha").reset_index(drop=True)
            print(f"[merge] loaded existing history ({len(df):,} rows)")
            return df
        except Exception as e:
            last_err = e
            continue

    print(f"[merge] could not read existing history ({last_err}); starting fresh.")
    return None


def merge(new: pd.DataFrame, old: pd.DataFrame | None) -> tuple[pd.DataFrame, dict]:
    if old is None or old.empty:
        return new.copy(), dict(rows_before=0, rows_after=len(new),
                                rows_added=len(new), rows_updated=0)
    old_dates = set(old["Fecha"]); new_dates = set(new["Fecha"])
    added = new_dates - old_dates
    overlap = old_dates & new_dates

    combined = pd.concat([
        old[~old["Fecha"].isin(new_dates)],
        new,
    ], ignore_index=True).sort_values("Fecha").reset_index(drop=True)

    rows_updated = 0
    if overlap:
        oi = old.set_index("Fecha"); ni = new.set_index("Fecha")
        for d in overlap:
            if not oi.loc[d].equals(ni.loc[d]):
                rows_updated += 1

    return combined, dict(rows_before=len(old), rows_after=len(combined),
                          rows_added=len(added), rows_updated=rows_updated)


# -----------------------------------------------------------------------------
# Monthly aggregation
# -----------------------------------------------------------------------------
def build_monthly(daily: pd.DataFrame) -> pd.DataFrame:
    df = daily.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Year"]  = df["Fecha"].dt.year
    df["Month"] = df["Fecha"].dt.to_period("M")

    # End-of-month values (stocks & FX = last observation in the month)
    eom = (df.sort_values("Fecha")
             .groupby("Month")[["Reservas", "Tipo_Cambio"]]
             .last())

    # Sum of daily flows during the month
    flows = df.groupby("Month")[["Compra_Divisas", "Pagos_Organismos"]].sum()

    m = eom.join(flows).reset_index()
    m["Month_End"] = m["Month"].dt.to_timestamp("M")
    m["Year"]      = m["Month_End"].dt.year

    # Reserves dynamics
    m["MoM_Reservas_USD"] = m["Reservas"].diff()
    m["MoM_Reservas_pct"] = m["Reservas"].pct_change()
    m["YoY_Reservas_pct"] = m["Reservas"].pct_change(12)

    # YTD compra & pagos (cumulative within each calendar year)
    m["YTD_Compra_Divisas"]   = m.groupby("Year")["Compra_Divisas"].cumsum()
    m["YTD_Pagos_Organismos"] = m.groupby("Year")["Pagos_Organismos"].cumsum()

    # FX devaluation %
    m["MoM_TC_pct"] = m["Tipo_Cambio"].pct_change()

    cols = [
        "Month_End", "Reservas", "MoM_Reservas_USD", "MoM_Reservas_pct",
        "YoY_Reservas_pct", "Compra_Divisas", "YTD_Compra_Divisas",
        "Pagos_Organismos", "YTD_Pagos_Organismos",
        "Tipo_Cambio", "MoM_TC_pct",
    ]
    return m[cols]


# -----------------------------------------------------------------------------
# Workbook writers
# -----------------------------------------------------------------------------
PRETTY_DAILY = [
    ("Fecha",            "Fecha",                          FMT_DATE_D, "date"),
    ("Reservas",         "Reservas (USD MM)",              FMT_USD,    "num"),
    ("Compra_Divisas",   "Compra de Divisas (USD MM)",     FMT_USD_1D, "num"),
    ("Pagos_Organismos", "Pagos a Organismos (USD MM)",    FMT_USD_1D, "num"),
    ("Tipo_Cambio",      "Tipo de Cambio (USD/ARS)",       FMT_FX,     "num"),
]

# Monthly tab uses a two-row banner: group headers (row 1) + column labels (row 2)
MONTHLY_GROUPS = [
    # (group_label, [(internal_name, display_name, number_format), ...])
    ("Período",           [
        ("Month_End",          "Mes",                            FMT_DATE_M),
    ]),
    ("Reservas Internacionales",       [
        ("Reservas",            "Stock (USD MM)",                 FMT_USD),
        ("MoM_Reservas_USD",    "Δ m/m (USD MM)",                 FMT_USD),
        ("MoM_Reservas_pct",    "Δ m/m %",                        FMT_PCT),
        ("YoY_Reservas_pct",    "Δ a/a %",                        FMT_PCT),
    ]),
    ("Compra de Divisas (BCRA)",       [
        ("Compra_Divisas",      "Mes (USD MM)",                   FMT_USD_1D),
        ("YTD_Compra_Divisas",  "YTD (USD MM)",                   FMT_USD),
    ]),
    ("Pagos a Organismos Internac.",   [
        ("Pagos_Organismos",    "Mes (USD MM)",                   FMT_USD_1D),
        ("YTD_Pagos_Organismos","YTD (USD MM)",                   FMT_USD),
    ]),
    ("Tipo de Cambio",                 [
        ("Tipo_Cambio",         "USD/ARS (fin mes)",              FMT_FX),
        ("MoM_TC_pct",          "Δ m/m %",                        FMT_PCT),
    ]),
]


def _set_widths(ws, widths: dict[int, float]):
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def _landscape(ws, fit_width=True):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    if fit_width:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def style_daily_tab(ws, daily: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT_DK
    _landscape(ws)

    # Title
    ws["B2"] = "Argentina · Reservas Internacionales · Diario"
    ws["B2"].font = font(bold=True, size=16, color=INK)
    ws.row_dimensions[2].height = 24

    # Subtitle
    ws["B3"] = f"Fuente: BCRA · {len(daily):,} obs · " \
               f"{daily['Fecha'].min().date()} → {daily['Fecha'].max().date()}"
    ws["B3"].font = font(size=9, italic=True, color=SUBTLE_INK)

    HEAD_ROW = 5
    BODY_START = 6

    # Header
    for j, (_, display, _, _) in enumerate(PRETTY_DAILY, start=2):
        c = ws.cell(row=HEAD_ROW, column=j, value=display)
        c.font = font(bold=True, size=10, color=INK)
        c.fill = fill(HEADER_BG)
        c.alignment = center()
        c.border = Border(bottom=Side(style="medium", color=ACCENT_DK))
    ws.row_dimensions[HEAD_ROW].height = 30
    ws.freeze_panes = ws.cell(row=BODY_START, column=3)

    # Body
    df = daily.sort_values("Fecha").reset_index(drop=True)
    for i, row in enumerate(df.itertuples(index=False), start=BODY_START):
        band = fill(BAND_LIGHT) if (i - BODY_START) % 2 == 1 else None
        for j, (src, _, num_fmt, kind) in enumerate(PRETTY_DAILY, start=2):
            val = getattr(row, src)
            c = ws.cell(row=i, column=j,
                        value=(None if pd.isna(val) else val))
            c.font = font(size=10, color=INK)
            c.number_format = num_fmt
            if kind == "date":
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                c.alignment = right()
            if band:
                c.fill = band

    _set_widths(ws, {1: 2, 2: 14, 3: 22, 4: 28, 5: 28, 6: 22})


def style_monthly_tab(ws, monthly: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT
    _landscape(ws)

    # Title
    ws["B2"] = "Argentina · Reservas Internacionales · Mensual"
    ws["B2"].font = font(bold=True, size=16, color=INK)
    ws.row_dimensions[2].height = 24
    ws["B3"] = f"Fuente: BCRA · {len(monthly)} meses · YTD recalculado por año calendario"
    ws["B3"].font = font(size=9, italic=True, color=SUBTLE_INK)

    GROUP_ROW = 5
    HEAD_ROW  = 6
    BODY_START = 7

    # Build flat column list with group spans
    col_idx = 2  # leave A blank
    flat_cols: list[tuple[str, str, str]] = []  # (src, display, fmt)
    group_spans: list[tuple[str, int, int]] = []  # (label, start_col, end_col)

    for label, cols in MONTHLY_GROUPS:
        start = col_idx
        for src, disp, f in cols:
            flat_cols.append((src, disp, f))
            col_idx += 1
        end = col_idx - 1
        group_spans.append((label, start, end))

    # Group header row
    for label, start, end in group_spans:
        ws.cell(row=GROUP_ROW, column=start, value=label)
        if end > start:
            ws.merge_cells(start_row=GROUP_ROW, end_row=GROUP_ROW,
                           start_column=start, end_column=end)
        c = ws.cell(row=GROUP_ROW, column=start)
        c.font = font(bold=True, size=10, color="FFFFFF")
        c.fill = fill(ACCENT_DK)
        c.alignment = center()
        c.border = Border(bottom=Side(style="thin", color="FFFFFF"))
    ws.row_dimensions[GROUP_ROW].height = 22

    # Column header row
    for j, (_, disp, _) in enumerate(flat_cols, start=2):
        c = ws.cell(row=HEAD_ROW, column=j, value=disp)
        c.font = font(bold=True, size=10, color=INK)
        c.fill = fill(HEADER_BG)
        c.alignment = center()
        c.border = Border(bottom=Side(style="medium", color=ACCENT_DK))
    ws.row_dimensions[HEAD_ROW].height = 36

    ws.freeze_panes = ws.cell(row=BODY_START, column=3)

    # Body
    df = monthly.copy().reset_index(drop=True)
    for i, row in enumerate(df.itertuples(index=False), start=BODY_START):
        # Year-boundary visual marker: thicker top border on January rows
        is_jan = (getattr(row, "Month_End").month == 1)
        band = fill(BAND_LIGHT) if (i - BODY_START) % 2 == 1 else None
        for j, (src, _, num_fmt) in enumerate(flat_cols, start=2):
            val = getattr(row, src)
            c = ws.cell(row=i, column=j, value=(None if pd.isna(val) else val))
            c.font = font(size=10, color=INK)
            c.number_format = num_fmt
            c.alignment = right() if src != "Month_End" else \
                Alignment(horizontal="left", vertical="center", indent=1)
            if band:
                c.fill = band
            if is_jan and i != BODY_START:
                c.border = Border(top=Side(style="thin", color=ACCENT))

    # Column widths
    widths = {1: 2}
    col = 2
    for src, _, _ in flat_cols:
        if src == "Month_End":
            widths[col] = 11
        elif "pct" in src:
            widths[col] = 11
        elif src == "Tipo_Cambio":
            widths[col] = 13
        else:
            widths[col] = 15
        col += 1
    _set_widths(ws, widths)


def style_snapshot_tab(ws, daily: pd.DataFrame, monthly: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4D3D7D"
    _landscape(ws)

    latest = daily.sort_values("Fecha").iloc[-1]
    fecha = latest["Fecha"]

    # Year-to-date totals across the calendar year of the latest observation
    yr = fecha.year
    ytd_mask = (daily["Fecha"].dt.year == yr)
    ytd_compra = daily.loc[ytd_mask, "Compra_Divisas"].sum()
    ytd_pagos  = daily.loc[ytd_mask, "Pagos_Organismos"].sum()

    # Month-over-month from monthly tab
    last_m = monthly.iloc[-1]
    prev_m = monthly.iloc[-2] if len(monthly) > 1 else None

    ws["B2"] = "Argentina · Snapshot de Reservas"
    ws["B2"].font = font(bold=True, size=20, color=INK)
    ws.row_dimensions[2].height = 30
    ws["B3"] = f"Último dato: {fecha.date()}  ·  Fuente: BCRA"
    ws["B3"].font = font(size=10, italic=True, color=SUBTLE_INK)

    # KPI cards laid out as a 2x2 grid
    def kpi(row, col, label, value, fmt, sublabel=""):
        # Card spans two columns wide x 3 rows tall
        ws.merge_cells(start_row=row,   end_row=row,   start_column=col, end_column=col+2)
        ws.merge_cells(start_row=row+1, end_row=row+1, start_column=col, end_column=col+2)
        ws.merge_cells(start_row=row+2, end_row=row+2, start_column=col, end_column=col+2)

        lbl = ws.cell(row=row, column=col, value=label)
        lbl.font = font(bold=True, size=9, color=SUBTLE_INK)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.fill = fill(HEADER_BG)
        lbl.border = Border(top=Side(style="medium", color=ACCENT_DK),
                            left=Side(style="thin", color=DIVIDER),
                            right=Side(style="thin", color=DIVIDER))

        val = ws.cell(row=row+1, column=col, value=value)
        val.font = font(bold=True, size=20, color=INK)
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val.number_format = fmt
        val.fill = fill("FFFFFF")
        val.border = Border(left=Side(style="thin", color=DIVIDER),
                            right=Side(style="thin", color=DIVIDER))

        sub = ws.cell(row=row+2, column=col, value=sublabel)
        sub.font = font(size=9, color=SUBTLE_INK, italic=True)
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sub.fill = fill("FFFFFF")
        sub.border = Border(bottom=Side(style="thin", color=DIVIDER),
                            left=Side(style="thin", color=DIVIDER),
                            right=Side(style="thin", color=DIVIDER))

        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row+1].height = 30
        ws.row_dimensions[row+2].height = 15

    def signed_label(v, fmt="USD MM"):
        if pd.isna(v):
            return ""
        sign = "+" if v >= 0 else "−"
        return f"{sign}{abs(v):,.0f} {fmt}"

    delta_res = last_m["MoM_Reservas_USD"]
    delta_tc  = last_m["MoM_TC_pct"]

    # Row 5: Reservas | Compra YTD
    kpi(5, 2, "RESERVAS INTERNACIONALES",
        latest["Reservas"], FMT_USD,
        sublabel=f"Δ m/m {signed_label(delta_res)}")
    kpi(5, 6, f"COMPRA DE DIVISAS · YTD {yr}",
        ytd_compra, FMT_USD,
        sublabel=f"Mes en curso: {signed_label(last_m['Compra_Divisas'])}")

    # Row 9: Pagos YTD | TC
    kpi(9, 2, f"PAGOS A ORGANISMOS · YTD {yr}",
        ytd_pagos, FMT_USD,
        sublabel=f"Mes en curso: {signed_label(last_m['Pagos_Organismos'])}")
    kpi(9, 6, "TIPO DE CAMBIO (USD/ARS)",
        latest["Tipo_Cambio"], FMT_FX,
        sublabel=f"Δ m/m: {'+' if delta_tc>=0 else '−'}{abs(delta_tc):.1%}" if not pd.isna(delta_tc) else "")

    # Recent months strip (last 6 months)
    strip_head = 14
    ws.cell(row=strip_head, column=2,
            value="Últimos 6 meses").font = font(bold=True, size=11, color=INK)
    headers = ["Mes", "Reservas", "Δ m/m", "Compra Divisas",
               "YTD Compra", "Pagos Org.", "TC"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=strip_head + 1, column=j, value=h)
        c.font = font(bold=True, size=10, color=INK)
        c.fill = fill(HEADER_BG)
        c.alignment = center()
        c.border = Border(bottom=Side(style="medium", color=ACCENT_DK))
    ws.row_dimensions[strip_head + 1].height = 24

    recent = monthly.tail(6)
    for i, row in enumerate(recent.itertuples(index=False), start=strip_head + 2):
        band = fill(BAND_LIGHT) if (i - strip_head - 2) % 2 == 1 else None
        vals = [
            (row.Month_End,         FMT_DATE_M, "left"),
            (row.Reservas,          FMT_USD,    "right"),
            (row.MoM_Reservas_USD,  FMT_USD,    "right"),
            (row.Compra_Divisas,    FMT_USD_1D, "right"),
            (row.YTD_Compra_Divisas,FMT_USD,    "right"),
            (row.Pagos_Organismos,  FMT_USD_1D, "right"),
            (row.Tipo_Cambio,       FMT_FX,     "right"),
        ]
        for j, (val, num_fmt, align) in enumerate(vals, start=2):
            c = ws.cell(row=i, column=j,
                        value=(None if pd.isna(val) else val))
            c.font = font(size=10, color=INK)
            c.number_format = num_fmt
            c.alignment = (Alignment(horizontal="left", vertical="center", indent=1)
                           if align == "left" else right())
            if band:
                c.fill = band

    _set_widths(ws, {1: 2, 2: 13, 3: 16, 4: 14, 5: 18, 6: 16, 7: 16, 8: 14})


def style_metadata_tab(ws, daily: pd.DataFrame, source: Path, stats: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = SUBTLE_INK

    ws["B2"] = "Metadata"
    ws["B2"].font = font(bold=True, size=16, color=INK)
    ws.row_dimensions[2].height = 24

    rows = [
        ("Generado",                dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Archivo fuente",          source.name),
        ("Primera fecha",           str(daily["Fecha"].min().date())),
        ("Última fecha",            str(daily["Fecha"].max().date())),
        ("Observaciones diarias",   f"{len(daily):,}"),
        ("Filas antes del merge",   f"{stats['rows_before']:,}"),
        ("Filas después del merge", f"{stats['rows_after']:,}"),
        ("Filas nuevas",            f"{stats['rows_added']:,}"),
        ("Filas revisadas",         f"{stats['rows_updated']:,}"),
        ("", ""),
        ("Series", ""),
        ("  Reservas",          "Stock de Reservas Internacionales (USD MM)"),
        ("  Compra de Divisas", "Intervención cambiaria diaria del BCRA (USD MM)"),
        ("  Pagos a Organismos","Flujos netos con organismos internacionales (USD MM)"),
        ("  Tipo de Cambio",    "USD/ARS de referencia BCRA"),
    ]
    for r, (k, v) in enumerate(rows, start=4):
        a = ws.cell(row=r, column=2, value=k)
        b = ws.cell(row=r, column=3, value=v)
        if k.startswith("  "):
            a.font = font(size=10, color=SUBTLE_INK, italic=True)
        elif k == "" or k == "Series":
            a.font = font(bold=True, size=11, color=INK)
        else:
            a.font = font(bold=True, size=10, color=INK)
        b.font = font(size=10, color=INK)
    _set_widths(ws, {1: 2, 2: 28, 3: 55})


# -----------------------------------------------------------------------------
# Write workbook (refresh-in-place: preserves user-added sheets/charts)
# -----------------------------------------------------------------------------
MANAGED_TABS = ["Snapshot", "Monthly", "Daily", "Metadata"]


def _open_or_create_workbook() -> openpyxl.Workbook:
    """Open the historical workbook if it exists, else create an empty one.
    Opening preserves any user-added tabs, charts, named ranges, etc."""
    if HIST_PATH.exists() and HIST_PATH.stat().st_size > 0:
        try:
            wb = openpyxl.load_workbook(HIST_PATH)
            preserved = [s for s in wb.sheetnames if s not in MANAGED_TABS]
            if preserved:
                print(f"[write] preserving user sheets: {preserved}")
            return wb
        except Exception as e:
            print(f"[write] could not open existing workbook ({e}); creating new.")
    wb = openpyxl.Workbook()
    # Remove default 'Sheet'
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb


def _reset_managed_sheet(wb: openpyxl.Workbook, name: str, index: int) -> "openpyxl.worksheet.worksheet.Worksheet":
    """Remove the existing managed sheet (if any) and create a fresh one
    in a stable position so the user's references to Daily!Cx etc. stay
    valid run-to-run."""
    if name in wb.sheetnames:
        del wb[name]
    # Clamp index to current size
    idx = min(index, len(wb.sheetnames))
    return wb.create_sheet(name, idx)


def write_workbook(daily: pd.DataFrame, monthly: pd.DataFrame,
                   source: Path, stats: dict) -> None:
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = _open_or_create_workbook()

    style_snapshot_tab(_reset_managed_sheet(wb, "Snapshot", 0), daily, monthly)
    style_monthly_tab (_reset_managed_sheet(wb, "Monthly",  1), monthly)
    style_daily_tab   (_reset_managed_sheet(wb, "Daily",    2), daily)
    style_metadata_tab(_reset_managed_sheet(wb, "Metadata", 3), daily, source, stats)

    wb.save(HIST_PATH)
    user_tabs = [s for s in wb.sheetnames if s not in MANAGED_TABS]
    extra = f"  (kept {len(user_tabs)} user sheet{'s' if len(user_tabs)!=1 else ''})" if user_tabs else ""
    print(f"[write] saved {HIST_PATH}{extra}")


# -----------------------------------------------------------------------------
# Archive raw input
# -----------------------------------------------------------------------------
def archive_input(src: Path) -> None:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = ARCHIVE_DIR / f"{src.stem}__{stamp}{src.suffix}"
    try:
        shutil.copy2(src, dest)
        print(f"[archive] copied raw input -> {dest.name}")
    except Exception as e:
        print(f"[archive] WARNING: could not archive raw input: {e}")


def find_latest_input() -> Path:
    cands = sorted(
        list(INPUTS_DIR.glob("*.xlsm")) + list(INPUTS_DIR.glob("*.xlsx")),
        key=lambda p: p.stat().st_mtime, reverse=True,
    )
    cands = [c for c in cands if c.resolve() != HIST_PATH.resolve()]
    if not cands:
        raise SystemExit(f"No .xlsm/.xlsx files found in {INPUTS_DIR}")
    return cands[0]


# -----------------------------------------------------------------------------
# Driver
# -----------------------------------------------------------------------------
def main(argv=None):
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("source", nargs="?",
                   help="BCRA series file. If omitted, the most recent file "
                        "in inputs/ is used.")
    p.add_argument("--no-archive", action="store_true")
    args = p.parse_args(argv)

    src = Path(args.source) if args.source else find_latest_input()
    if not src.exists():
        raise SystemExit(f"Source file not found: {src}")

    new_df = extract_daily(src)
    old_df = load_existing()
    combined, stats = merge(new_df, old_df)
    monthly = build_monthly(combined)
    write_workbook(combined, monthly, src, stats)

    if not args.no_archive:
        archive_input(src)

    print(
        "[done] rows before={rows_before:,}  after={rows_after:,}  "
        "added={rows_added:,}  revised={rows_updated:,}".format(**stats)
    )


if __name__ == "__main__":
    main()
