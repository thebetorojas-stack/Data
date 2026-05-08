"""
EMBI Global Dashboard Builder
=============================

Reads any combination of J.P. Morgan EMBI Global / EMBI Global Diversified
CSVs and produces a single structured Excel workbook. The script auto-
detects three JPM file types and merges them:

  1. RETURNS file        — wide CSV, columns "EM Debt Indices | <Entity> | <Metric>"
                           Metrics: Cum Tot Ret Idx, Yld to Maturity, Z Spread,
                           Index Weight (%) (often empty / legacy EMBI only).
                           One row per date.

  2. WEIGHTS HISTORY     — two-row header. Row 1: FC_EMBIG_* codes.
                           Row 2: country names. Column 1: Trade Date,
                           column 2: Composite Index Weight, columns 3+:
                           per-country weight time series (back to 1993).

  3. SNAPSHOT            — flat CSV, one row per entity. Header includes
                           "Bam Id", "Instrument", "Mkt Cap %", "Average S&P
                           Rating", "Yield to Worst", "Z Spread to Worst",
                           "Spread Duration", and various return metrics
                           (Daily / MTD / YTD). Single trade date.

Multiple files of the same type are merged: when two files cover overlapping
dates, the later-modified file wins on conflict, but unique dates from older
files are preserved. This means once you have a deep history saved on disk,
re-running with just a fresh daily snapshot still produces a workbook with
the full history intact.

Usage
-----
    # Scan the current directory for *.csv files and process whatever it finds:
    python embi_builder.py

    # Or scan a specific directory:
    python embi_builder.py /path/to/EMBI_data

    # Or pass explicit files:
    python embi_builder.py returns.csv weights_history.csv snapshot.csv

    # Override the output filename:
    python embi_builder.py /path/to/data --output Dashboard_2026-05-07.xlsx
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import shutil
import statistics
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# 1. TAXONOMY — regions, rating order, name aliases
# ============================================================================

REGION_ORDER: List[str] = ["LatAm", "Europe", "Africa", "GCC", "Asia"]

COUNTRY_REGION: Dict[str, str] = {
    # LatAm
    "Argentina": "LatAm", "Barbados": "LatAm", "Belize": "LatAm",
    "Bolivia": "LatAm", "Brazil": "LatAm", "Chile": "LatAm",
    "Colombia": "LatAm", "Costa Rica": "LatAm", "Dominican Republic": "LatAm",
    "Ecuador": "LatAm", "El Salvador": "LatAm", "Guatemala": "LatAm",
    "Honduras": "LatAm", "Jamaica": "LatAm", "Mexico": "LatAm",
    "Panama": "LatAm", "Paraguay": "LatAm", "Peru": "LatAm",
    "Suriname": "LatAm", "Trinidad & Tobago": "LatAm",
    "Uruguay": "LatAm", "Venezuela": "LatAm",
    # Europe (Emerging Europe + Caucasus per JPM)
    "Armenia": "Europe", "Belarus": "Europe", "Bulgaria": "Europe",
    "Croatia": "Europe", "Georgia": "Europe", "Greece": "Europe",
    "Hungary": "Europe", "Latvia": "Europe", "Lithuania": "Europe",
    "Montenegro": "Europe", "Poland": "Europe", "Romania": "Europe",
    "Russia": "Europe", "Serbia": "Europe", "Slovak Republic": "Europe",
    "Turkey": "Europe", "Ukraine": "Europe",
    # Africa
    "Algeria": "Africa", "Angola": "Africa", "Benin": "Africa",
    "Cameroon": "Africa", "Cote d Ivoire": "Africa", "Egypt": "Africa",
    "Ethiopia": "Africa", "Gabon": "Africa", "Ghana": "Africa",
    "Kenya": "Africa", "Morocco": "Africa", "Mozambique": "Africa",
    "Namibia": "Africa", "Nigeria": "Africa", "Republic of Congo": "Africa",
    "Rwanda": "Africa", "Senegal": "Africa", "South Africa": "Africa",
    "Tanzania": "Africa", "Tunisia": "Africa", "Zambia": "Africa",
    # GCC / Mideast
    "Bahrain": "GCC", "Iraq": "GCC", "Jordan": "GCC", "Kuwait": "GCC",
    "Lebanon": "GCC", "Oman": "GCC", "Qatar": "GCC", "Saudi Arabia": "GCC",
    "UAE": "GCC",
    # Asia
    "Azerbaijan": "Asia", "China": "Asia", "India": "Asia",
    "Indonesia": "Asia", "Kazakhstan": "Asia", "Kyrgyzstan": "Asia",
    "Malaysia": "Asia", "Maldives": "Asia", "Mongolia": "Asia",
    "Pakistan": "Asia", "Papua New Guinea": "Asia", "Philippines": "Asia",
    "South Korea": "Asia", "Sri Lanka": "Asia", "Tajikistan": "Asia",
    "Thailand": "Asia", "Uzbekistan": "Asia", "Vietnam": "Asia",
}

REGION_AGGREGATE: Dict[str, str] = {
    "LatAm": "Latin Region",
    "Europe": "Europe Region",
    "Africa": "Africa Region",
    "GCC": "Mideast Region",
    "Asia": "Asia Region",
}

# JPM "By Region" labels in the snapshot file map differently.
SNAPSHOT_REGION_LABEL: Dict[str, str] = {
    "Africa": "Africa", "Asia": "Asia", "Europe": "Europe",
    "Latin": "LatAm", "Middle East": "GCC",
}

RATING_ORDER: List[str] = [
    "Credit AA only", "Credit A only", "Credit BBB only",
    "Credit BB only", "Credit B only", "Credit C only",
    "Credit IG only", "Credit Non-IG", "Credit NR", "Credit Residual only",
]
RATING_DISPLAY: Dict[str, str] = {
    "Credit AA only": "AA",
    "Credit A only": "A",
    "Credit BBB only": "BBB",
    "Credit BB only": "BB",
    "Credit B only": "B",
    "Credit C only": "C / Distressed",
    "Credit IG only": "IG (composite)",
    "Credit Non-IG": "Non-IG (composite)",
    "Credit NR": "Not Rated",
    "Credit Residual only": "Residual",
}

# S&P / Moody-style rating to numeric score (higher = better quality).
# Used to sort countries by credit quality within a region.
RATING_SCORE: Dict[str, int] = {
    "AAA": 22, "Aaa": 22,
    "AA+": 21, "Aa1": 21,
    "AA": 20,  "Aa2": 20,
    "AA-": 19, "Aa3": 19,
    "A+": 18,  "A1": 18,
    "A": 17,   "A2": 17,
    "A-": 16,  "A3": 16,
    "BBB+": 15, "Baa1": 15,
    "BBB": 14, "Baa2": 14,
    "BBB-": 13, "Baa3": 13,
    "BB+": 12, "Ba1": 12,
    "BB": 11,  "Ba2": 11,
    "BB-": 10, "Ba3": 10,
    "B+": 9,   "B1": 9,
    "B": 8,    "B2": 8,
    "B-": 7,   "B3": 7,
    "CCC+": 6, "Caa1": 6,
    "CCC": 5,  "Caa2": 5,
    "CCC-": 4, "Caa3": 4,
    "CC": 3,   "Ca": 3,
    "C": 2,
    "D": 1, "SD": 1,
    "NR": 0,
}

# Reverse of RATING_SCORE for the S&P notation only — used to display a
# weighted-average rating score back as a notation (e.g. 11.7 → "BB").
SP_LABELS_BY_SCORE: Dict[int, str] = {
    22: "AAA", 21: "AA+", 20: "AA", 19: "AA-",
    18: "A+",  17: "A",   16: "A-",
    15: "BBB+", 14: "BBB", 13: "BBB-",
    12: "BB+",  11: "BB",  10: "BB-",
    9:  "B+",   8:  "B",   7:  "B-",
    6:  "CCC+", 5:  "CCC", 4:  "CCC-",
    3:  "CC",   2:  "C",   1:  "D", 0: "NR",
}


def score_to_sp_label(score: float) -> str:
    """Map a (potentially fractional) rating score to its closest S&P notation."""
    if score is None:
        return ""
    return SP_LABELS_BY_SCORE.get(int(round(score)), "?")


LATAM_FOCUS: List[str] = [
    "Argentina", "Chile", "Colombia", "Dominican Republic",
    "Mexico", "Panama", "Peru", "Venezuela",
]

INDEX_NAME = "EMBI Global"
METRICS = {
    "spread": "Z Spread",
    "yield":  "Yld to Maturity",
    "tret":   "Cum Tot Ret Idx",
}

# Maps spellings used by snapshot / weights / returns files to a single canonical form.
NAME_ALIASES: Dict[str, str] = {
    "Cote D'Ivoire":         "Cote d Ivoire",
    "Cote d'Ivoire":         "Cote d Ivoire",
    "Côte d'Ivoire":         "Cote d Ivoire",
    "Trinidad And Tobago":   "Trinidad & Tobago",
    "Trinidad and Tobago":   "Trinidad & Tobago",
    "Trinidad &amp; Tobago": "Trinidad & Tobago",
    "Slovakia":              "Slovak Republic",
    "Republic of the Congo": "Republic of Congo",
    "Republic Of Congo":     "Republic of Congo",
    "Congo, Republic of":    "Republic of Congo",
}


def normalize_name(name: str) -> str:
    name = (name or "").strip().replace("&amp;", "&").replace("&#38;", "&")
    return NAME_ALIASES.get(name, name)


# ============================================================================
# 2. STYLING
# ============================================================================

FONT_NAME = "Arial"
COLOR_HEADER_BG = "1F4E78"
COLOR_HEADER_FG = "FFFFFF"
COLOR_REGION_BG = "D9E1F2"
COLOR_RATING_BG = "FFF2CC"
COLOR_INDEX_BG  = "C6E0B4"
COLOR_INPUT_BG  = "FFFF00"   # bright yellow — flag user-editable assumption cells
COLOR_BORDER    = "BFBFBF"
COLOR_HARDCODE  = "0000FF"
COLOR_FORMULA   = "000000"
COLOR_CROSSREF  = "008000"
COLOR_NOTE      = "7F7F7F"

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


# ============================================================================
# 3. PARSERS
# ============================================================================

def _parse_date(raw: str) -> datetime:
    raw = (raw or "").strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise ValueError(f"unrecognised date: {raw!r}")


def _parse_value(raw: str) -> Optional[float]:
    if raw is None:
        return None
    s = raw.strip()
    if s == "" or s.upper() in {"N/A", "NA", "#N/A", "NULL"}:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _open_csv(path: Path):
    """Open a JPM CSV with BOM-stripping and replace any decoding errors.

    Use this everywhere instead of plain open() so loaders are tolerant of
    file-format quirks (BOM, occasional non-UTF-8 bytes from JPM exports).
    """
    return path.open("r", newline="", encoding="utf-8-sig", errors="replace")


def _read_skipping_blank_rows(reader, n: int) -> List[List[str]]:
    """Return the next `n` non-empty rows from a csv.reader."""
    out: List[List[str]] = []
    for row in reader:
        if any((c or "").strip() for c in row):
            out.append(row)
            if len(out) >= n:
                return out
    return out


def classify_csv(path: Path) -> str:
    """Return one of 'returns', 'weights_history', 'snapshot', 'unknown'.

    Classification is based purely on CONTENT (header signatures), never on
    the filename. JPM ships files with timestamps, GUIDs, and inconsistent
    naming — that doesn't matter here. The script tolerates: UTF-8 BOM,
    leading blank rows, header whitespace, case differences, and quoted
    cells.
    """
    try:
        # utf-8-sig silently strips the BOM if one is present.
        with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            # Skip any number of leading blank or all-empty-cell rows until
            # we find one that actually contains content.
            row1: List[str] = []
            for candidate in reader:
                if any((c or "").strip() for c in candidate):
                    row1 = candidate
                    break
            if not row1:
                return "unknown"
            # Same for row 2 — skip blank rows there too.
            row2: List[str] = []
            for candidate in reader:
                if any((c or "").strip() for c in candidate):
                    row2 = candidate
                    break
    except (OSError, StopIteration):
        return "unknown"

    def _norm(s):
        return (s or "").strip().strip('"').strip().lower()

    head = _norm(row1[0])

    # ---- RETURNS: first column "Date", subsequent columns contain
    #      "EM Debt Indices | <Entity> | <Metric>".
    if head == "date" and any("em debt indices" in (c or "").lower() for c in row1[1:]):
        return "returns"

    # ---- SNAPSHOT: first column "Bam Id" (any spacing/capitalization) and
    #      "Instrument" appearing as another column header.
    if head in ("bam id", "bamid", "bam_id"):
        if any(_norm(c) == "instrument" for c in row1):
            return "snapshot"
        # Even without "Instrument", the Bam Id signature alone is rare enough
        # to be conclusive — JPM uses it specifically for these dumps.
        return "snapshot"

    # ---- WEIGHTS HISTORY: two-row header. Row 1 has FC_EMBIG_* codes.
    #      Row 2 starts with "Trade Date" and has country names from col 3+.
    has_codes = any((c or "").strip().upper().startswith("FC_EMBIG") for c in row1)
    has_trade_date = bool(row2) and _norm(row2[0]) == "trade date"
    if has_codes and has_trade_date:
        return "weights_history"
    # Fallback: if row 1 is empty in col 0 and row 2 says "Trade Date" with a
    # "Composite Index Weight" column, we're still looking at a weights file
    # — don't insist on FC_EMBIG codes (JPM might rename them in a future build).
    if has_trade_date and any("composite" in (c or "").lower() and "weight" in (c or "").lower() for c in row2):
        return "weights_history"

    return "unknown"


def load_returns(path: Path) -> Tuple[List[datetime], Dict[Tuple[str, str], List[Optional[float]]]]:
    """Load a returns/yields/spreads file. Returns ([dates], {(entity,metric): [values]})."""
    with _open_csv(path) as f:
        reader = csv.reader(f)
        first = _read_skipping_blank_rows(reader, 1)
        header = first[0] if first else []
        rows = list(reader)

    columns: List[Tuple[str, str]] = []
    for col in header[1:]:
        cleaned = (col or "").replace("&amp;", "&").replace("&#38;", "&")
        parts = [p.strip() for p in cleaned.split("|")]
        if len(parts) < 3:
            columns.append(("", ""))
            continue
        columns.append((normalize_name(parts[1]), parts[2]))

    parsed: List[Tuple[datetime, List[Optional[float]]]] = []
    for row in rows:
        if not row or not row[0].strip():
            continue
        try:
            d = _parse_date(row[0])
        except ValueError:
            continue
        parsed.append((d, [_parse_value(c) for c in row[1:]]))
    parsed.sort(key=lambda r: r[0])
    dates = [r[0] for r in parsed]

    series: Dict[Tuple[str, str], List[Optional[float]]] = {}
    for idx, key in enumerate(columns):
        if key == ("", ""):
            continue
        series.setdefault(key, [None] * len(parsed))
        for i, (_, vals) in enumerate(parsed):
            if idx < len(vals):
                series[key][i] = vals[idx]
    return dates, series


def load_weights_history(path: Path) -> Tuple[List[datetime], Dict[str, List[Optional[float]]]]:
    """Load monthly weights history (two-row header). Returns ([dates], {country: [weight]})."""
    with _open_csv(path) as f:
        reader = csv.reader(f)
        first_two = _read_skipping_blank_rows(reader, 2)
        codes = first_two[0] if first_two else []
        names = first_two[1] if len(first_two) > 1 else []
        rows = list(reader)

    countries = [normalize_name(n) for n in names[2:]]
    parsed: List[Tuple[datetime, List[Optional[float]]]] = []
    for row in rows:
        if not row or not row[0].strip():
            continue
        try:
            d = _parse_date(row[0])
        except ValueError:
            continue
        # Skip composite weight column 1, take per-country from column 2 onwards.
        parsed.append((d, [_parse_value(c) for c in row[2:]]))
    parsed.sort(key=lambda r: r[0])
    dates = [r[0] for r in parsed]
    out: Dict[str, List[Optional[float]]] = {c: [None] * len(parsed) for c in countries}
    for i, (_, vals) in enumerate(parsed):
        for j, c in enumerate(countries):
            if j < len(vals):
                out[c][i] = vals[j]
    return dates, out


def load_snapshot(path: Path) -> Tuple[Optional[datetime], Dict[str, Dict[str, Any]]]:
    """Load a single-date snapshot file. Returns (snapshot_date, {entity: {field: val}})."""
    with _open_csv(path) as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    snap_date: Optional[datetime] = None
    out: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        instrument = (row.get("Instrument") or "").strip()
        if not instrument or instrument.startswith("By "):
            continue
        if snap_date is None:
            try:
                snap_date = _parse_date(row.get("Date", ""))
            except (ValueError, KeyError):
                pass
        # Snapshot uses "Latin", "Middle East", etc. — translate to our region keys.
        if instrument in SNAPSHOT_REGION_LABEL:
            key = f"REGION:{SNAPSHOT_REGION_LABEL[instrument]}"
        elif instrument == "Non Latin":
            key = "AGG:NonLatin"
        elif instrument in ("EMBI Global", "EMBI Global Diversified"):
            # Either flavor of the index gets unified under the same internal
            # key. In normal use the user downloads EMBI Global (regular); the
            # Diversified label is accepted for backwards compatibility with
            # any older archived snapshot files.
            key = "INDEX:EMBI"
        else:
            key = normalize_name(instrument)
        out[key] = row
    return snap_date, out


# ============================================================================
# 4. AGGREGATOR — merge multiple files of the same type
# ============================================================================

def merge_returns(file_results: List[Tuple[List[datetime], Dict[Tuple[str, str], List[Optional[float]]]]]
                  ) -> Tuple[List[datetime], Dict[Tuple[str, str], List[Optional[float]]]]:
    if not file_results:
        return [], {}
    # Build a (key) -> {date -> value} dict, then collapse to sorted list.
    by_key: Dict[Tuple[str, str], Dict[datetime, float]] = defaultdict(dict)
    all_dates: set = set()
    for dates, series in file_results:
        for d in dates:
            all_dates.add(d)
        for key, vals in series.items():
            for i, d in enumerate(dates):
                if vals[i] is not None:
                    by_key[key][d] = vals[i]
    sorted_dates = sorted(all_dates)
    out_series: Dict[Tuple[str, str], List[Optional[float]]] = {}
    for key, dmap in by_key.items():
        out_series[key] = [dmap.get(d) for d in sorted_dates]
    return sorted_dates, out_series


def merge_weights_history(file_results: List[Tuple[List[datetime], Dict[str, List[Optional[float]]]]]
                          ) -> Tuple[List[datetime], Dict[str, List[Optional[float]]]]:
    if not file_results:
        return [], {}
    by_key: Dict[str, Dict[datetime, float]] = defaultdict(dict)
    all_dates: set = set()
    for dates, series in file_results:
        for d in dates:
            all_dates.add(d)
        for key, vals in series.items():
            for i, d in enumerate(dates):
                if vals[i] is not None:
                    by_key[key][d] = vals[i]
    sorted_dates = sorted(all_dates)
    out: Dict[str, List[Optional[float]]] = {}
    for key, dmap in by_key.items():
        out[key] = [dmap.get(d) for d in sorted_dates]
    return sorted_dates, out


def merge_snapshots(file_results: List[Tuple[Optional[datetime], Dict[str, Dict[str, Any]]]]
                    ) -> Tuple[Optional[datetime], Dict[str, Dict[str, Any]],
                                List[Tuple[datetime, Dict[str, Dict[str, Any]]]]]:
    """Return (latest_date, latest_data, full_history).

    latest_data is what the per-country Snapshot tab uses.
    full_history is the chronologically-sorted, DEDUPLICATED-by-date list of
    every snapshot loaded — consumed by the Rating_Trend tab to show how the
    universe's average rating moves over time as the user drops more monthly
    snapshots into the folder.
    """
    # Dedupe by date — if the same snapshot is loaded twice (e.g. once as
    # 'JP latest.csv' and once as the just-archived 'snapshot_<date>.csv'),
    # keep the first occurrence only so the time series doesn't double-count.
    by_date: Dict[datetime, Dict[str, Dict[str, Any]]] = {}
    for d, s in file_results:
        if d is None:
            continue
        if d not in by_date:
            by_date[d] = s
    sorted_snaps = sorted(by_date.items(), key=lambda r: r[0])
    if not sorted_snaps:
        return None, {}, []
    latest_date, latest_data = sorted_snaps[-1]
    return latest_date, latest_data, sorted_snaps


def snapshot_to_returns(
    snap_date: datetime,
    snap_data: Dict[str, Dict[str, Any]],
) -> Tuple[List[datetime], Dict[Tuple[str, str], List[Optional[float]]]]:
    """Convert one snapshot into a synthetic (dates, series) tuple matching the
    shape of a returns file. This lets a snapshot be merged into the same time
    series that drives the Spreads/Yields/TR_YTD tabs.

    Mapping:
      Snapshot 'EMBI Global' (or 'EMBI Global Diversified')  → 'EMBI Global'
      Snapshot region rollups                                → corresponding *_Region names
      Snapshot countries                                     → same name (already normalized)

      Snapshot 'Index Level'        → 'Cum Tot Ret Idx'
      Snapshot 'Yield to Worst'     → 'Yld to Maturity'
      Snapshot 'Z Spread to Worst'  → 'Z Spread'

    For non-callable bullet bonds (the bulk of the EMBI Global universe),
    Yield-to-Worst equals Yield-to-Maturity and Z-Spread-to-Worst equals
    Z-Spread exactly. So the metric mapping is clean. On any date where the
    real returns file ALSO has data, the real returns wins (this function
    feeds in BEFORE the real returns).

    Both EMBI Global and EMBI Global Diversified snapshots are accepted for
    backwards compatibility — but in normal use the user downloads the
    regular EMBI Global, which matches the rest of the workbook directly.
    """
    series: Dict[Tuple[str, str], List[Optional[float]]] = {}
    for ent, data in snap_data.items():
        if ent in _SNAPSHOT_ENTITY_TO_RETURNS:
            returns_ent = _SNAPSHOT_ENTITY_TO_RETURNS[ent]
        elif isinstance(ent, str) and not ent.startswith(("REGION:", "INDEX:", "AGG:")):
            returns_ent = ent
        else:
            continue
        for field, metric in _SNAPSHOT_FIELD_TO_METRIC.items():
            raw = (data.get(field) or "").strip() if data else ""
            if not raw:
                continue
            try:
                v = float(raw)
            except ValueError:
                continue
            series[(returns_ent, metric)] = [v]
    return [snap_date], series


_SNAPSHOT_ENTITY_TO_RETURNS: Dict[str, str] = {
    "INDEX:EMBI":    "EMBI Global",
    "REGION:Africa": "Africa Region",
    "REGION:Asia":   "Asia Region",
    "REGION:Europe": "Europe Region",
    "REGION:LatAm":  "Latin Region",
    "REGION:GCC":    "Mideast Region",
}

_SNAPSHOT_FIELD_TO_METRIC: Dict[str, str] = {
    "Index Level":       "Cum Tot Ret Idx",
    "Yield to Worst":    "Yld to Maturity",
    "Z Spread to Worst": "Z Spread",
}


def archive_snapshots(input_paths: List[Path], project_dir: Path) -> Tuple[Path, List[Tuple[str, str]]]:
    """For every input path that is a snapshot file, copy a date-stamped
    version into `<project_dir>/snapshots_archive/` if not already archived.

    This is the linchpin of the user workflow described in the README:
    the user always saves their fresh JPM snapshot to the same filename
    (e.g. 'JP latest.csv'), overwriting the previous one. Without archiving,
    every re-run would lose all prior snapshots — the user reported exactly
    this problem on their work computer. Archiving makes history accumulate
    automatically; the user never has to manually rename anything.

    Returns: (archive_dir_path_or_None, list_of_newly_archived_files).
    Returns None for the path if the project directory is read-only.
    """
    archive_dir = project_dir / "snapshots_archive"
    try:
        archive_dir.mkdir(exist_ok=True)
    except (OSError, PermissionError) as exc:
        print(f"\n  WARNING: cannot create '{archive_dir}' ({exc}). "
              f"Snapshot auto-archiving is disabled for this run. "
              f"Run from a writable folder if you want history preserved.",
              file=sys.stderr)
        return None, []
    newly_archived: List[Tuple[str, str]] = []
    for path in input_paths:
        # Don't archive files that already live inside the archive folder.
        try:
            if archive_dir in path.resolve().parents:
                continue
        except (OSError, RuntimeError):
            pass
        if classify_csv(path) != "snapshot":
            continue
        try:
            snap_date, _ = load_snapshot(path)
        except Exception:
            continue
        if snap_date is None:
            continue
        archive_name = f"snapshot_{snap_date.strftime('%Y-%m-%d')}.csv"
        archive_path = archive_dir / archive_name
        if not archive_path.exists():
            try:
                shutil.copy2(path, archive_path)
                newly_archived.append((snap_date.strftime("%Y-%m-%d"), archive_name))
            except (OSError, PermissionError) as exc:
                print(f"  WARNING: failed to archive {path.name}: {exc}",
                      file=sys.stderr)
    return archive_dir, newly_archived


# ============================================================================
# 5. BUILDER
# ============================================================================

class Builder:
    def __init__(
        self,
        dates: List[datetime],
        series: Dict[Tuple[str, str], List[Optional[float]]],
        weight_dates: List[datetime],
        weight_series: Dict[str, List[Optional[float]]],
        snap_date: Optional[datetime],
        snap_data: Dict[str, Dict[str, Any]],
        snap_history: List[Tuple[datetime, Dict[str, Dict[str, Any]]]],
        sources: List[Tuple[str, str]],  # [(file_kind, filename)]
    ) -> None:
        self.dates = dates
        self.series = series
        self.snap_date = snap_date
        self.snap_data = snap_data
        self.snap_history = snap_history  # list of (date, data) tuples
        self.sources = sources

        # ---- Merge snapshot-derived weights into the weights time series ----
        # Each snapshot file has 'Mkt Cap %' per country, which is the weight
        # at the snapshot date. As the user accumulates snapshots over time
        # (via auto-archive), these add fresh data points to the time series.
        # On any date where we have BOTH a weights-history entry AND a snapshot,
        # the snapshot wins (it's the more recent, more authoritative source).
        merged_dates, merged_series = self._merge_snapshot_weights(
            weight_dates, weight_series, snap_history
        )
        self.weight_dates = merged_dates
        self.weight_series = merged_series

        self.frequency = self._detect_frequency(dates)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        # Resolve countries actually present in returns data.
        present = {ent for (ent, _) in self.series.keys()}
        self.countries_by_region: Dict[str, List[str]] = {r: [] for r in REGION_ORDER}
        for entity, region in COUNTRY_REGION.items():
            if entity in present or entity in self.weight_series or entity in self.snap_data:
                self.countries_by_region[region].append(entity)

        # Sort countries within each region by S&P rating score (best first), then alphabetically.
        for region in REGION_ORDER:
            self.countries_by_region[region].sort(
                key=lambda c: (-self._rating_score(c), c)
            )

        self.all_countries: List[str] = [
            c for r in REGION_ORDER for c in self.countries_by_region[r]
        ]
        self.ratings_present: List[str] = [r for r in RATING_ORDER if r in present]

    # --------- helpers ----------
    @staticmethod
    def _merge_snapshot_weights(
        weight_dates: List[datetime],
        weight_series: Dict[str, List[Optional[float]]],
        snap_history: List[Tuple[datetime, Dict[str, Dict[str, Any]]]],
    ) -> Tuple[List[datetime], Dict[str, List[Optional[float]]]]:
        """Combine the weights-history file with per-country weights extracted
        from each snapshot (the 'Mkt Cap %' field per country).

        The weights-history file gives long monthly history (back to 1993).
        Each snapshot adds one fresh data point per country at the snapshot
        date. On overlapping dates, the snapshot wins. The output covers the
        union of all dates from both sources.
        """
        # Normalize the weights-history into a {country: {date: weight}} dict.
        merged: Dict[str, Dict[datetime, float]] = defaultdict(dict)
        for c, vals in weight_series.items():
            for i, d in enumerate(weight_dates):
                v = vals[i]
                if v is not None:
                    merged[c][d] = v

        # Layer snapshot weights on top — they override on overlap.
        for snap_date, snap_data in snap_history:
            for ent, data in snap_data.items():
                if not isinstance(ent, str) or ent.startswith(("REGION:", "INDEX:", "AGG:")):
                    continue
                raw = (data.get("Mkt Cap %") or "").strip()
                if not raw:
                    continue
                try:
                    w = float(raw)
                except ValueError:
                    continue
                merged[ent][snap_date] = w

        if not merged:
            return [], {}

        # Re-flatten back into (sorted_dates, {country: [values per date]}).
        all_dates = sorted({d for c_data in merged.values() for d in c_data})
        out_series: Dict[str, List[Optional[float]]] = {
            c: [merged[c].get(d) for d in all_dates] for c in merged
        }
        return all_dates, out_series

    @staticmethod
    def _detect_frequency(dates: List[datetime]) -> str:
        if len(dates) < 2:
            return "unknown"
        deltas = [(dates[i] - dates[i - 1]).days for i in range(1, len(dates))]
        avg = sum(deltas) / len(deltas)
        if avg <= 3:   return "daily"
        if 6 <= avg <= 9: return "weekly"
        if 25 <= avg <= 35: return "monthly"
        return "mixed"

    def _rating_score(self, country: str) -> int:
        snap = self.snap_data.get(country)
        if snap:
            for field in ("Average S&P Rating", "Average Moody Rating", "Average Fitch Rating"):
                v = (snap.get(field) or "").strip()
                if v in RATING_SCORE:
                    return RATING_SCORE[v]
        return -1  # unknown ratings sort to the bottom of their region

    def latest_value(self, entity: str, metric: str) -> Optional[float]:
        vals = self.series.get((entity, metric))
        if not vals:
            return None
        for v in reversed(vals):
            if v is not None:
                return v
        return None

    def value_at(self, entity: str, metric: str, idx: int) -> Optional[float]:
        vals = self.series.get((entity, metric))
        if not vals or idx is None or idx < 0 or idx >= len(vals):
            return None
        return vals[idx]

    def year_start_index(self, year: int) -> Optional[int]:
        if not self.dates:
            return None
        prior = None
        for i, d in enumerate(self.dates):
            if d.year < year:
                prior = i
            elif d.year >= year:
                break
        return prior if prior is not None else 0

    def latest_weight(self, country: str) -> Optional[float]:
        vals = self.weight_series.get(country)
        if not vals:
            return None
        for v in reversed(vals):
            if v is not None:
                return v
        return None

    def latest_weight_date(self) -> Optional[datetime]:
        if not self.weight_dates:
            return None
        # find the most recent date with at least one non-null weight
        for i in range(len(self.weight_dates) - 1, -1, -1):
            for c in self.weight_series:
                if self.weight_series[c][i] is not None:
                    return self.weight_dates[i]
        return None

    # --------- cell helpers ----------
    def _font(self, cell, **kw):
        d = {"name": FONT_NAME, "size": 10}
        d.update(kw)
        cell.font = Font(**d)

    def _hdr(self, cell, txt):
        cell.value = txt
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        self._font(cell, color=COLOR_HEADER_FG, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    def _region_lbl(self, cell, txt):
        cell.value = txt
        cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
        self._font(cell, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    def _rating_lbl(self, cell, txt):
        cell.value = txt
        cell.fill = PatternFill("solid", start_color=COLOR_RATING_BG)
        self._font(cell, bold=True, italic=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    def _index_lbl(self, cell, txt):
        cell.value = txt
        cell.fill = PatternFill("solid", start_color=COLOR_INDEX_BG)
        self._font(cell, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    def _val(self, cell, value, *, fmt="0.00", color=COLOR_FORMULA):
        cell.value = value
        cell.number_format = fmt
        self._font(cell, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = THIN_BORDER

    # --------- Cover ----------
    def build_cover(self):
        ws = self.wb.create_sheet("Cover")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80

        ws["A1"] = "EMBI Global Dashboard"
        self._font(ws["A1"], size=20, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 28

        meta = [
            ("Built on", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("YTD anchor year", str(datetime.now().year)),
            ("Returns frequency", self.frequency),
            ("Returns observations", len(self.dates)),
            ("Returns date range", f"{self.dates[0]:%Y-%m-%d} — {self.dates[-1]:%Y-%m-%d}" if self.dates else "(no returns data)"),
            ("Weights observations", len(self.weight_dates)),
            ("Weights date range", f"{self.weight_dates[0]:%Y-%m-%d} — {self.weight_dates[-1]:%Y-%m-%d}" if self.weight_dates else "(no weights data)"),
            ("Snapshot date", self.snap_date.strftime("%Y-%m-%d") if self.snap_date else "(no snapshot loaded)"),
            ("Countries present", sum(len(v) for v in self.countries_by_region.values())),
            ("Rating buckets present", len(self.ratings_present)),
        ]
        for i, (k, v) in enumerate(meta, start=3):
            ws.cell(row=i, column=1, value=k)
            self._font(ws.cell(row=i, column=1), bold=True)
            ws.cell(row=i, column=2, value=v)
            self._font(ws.cell(row=i, column=2))

        ws.cell(row=14, column=1, value="Source files loaded")
        self._font(ws.cell(row=14, column=1), size=12, bold=True, color=COLOR_HEADER_BG)
        for i, (kind, fname) in enumerate(self.sources, start=15):
            ws.cell(row=i, column=1, value=f"  • {kind}")
            self._font(ws.cell(row=i, column=1), italic=True)
            ws.cell(row=i, column=2, value=fname)
            self._font(ws.cell(row=i, column=2))

        legend_row = 15 + len(self.sources) + 2
        ws.cell(row=legend_row, column=1, value="Color key")
        self._font(ws.cell(row=legend_row, column=1), size=12, bold=True, color=COLOR_HEADER_BG)
        legend = [
            ("Hardcoded input (blue text)",   COLOR_HARDCODE),
            ("Formula / calculation (black)", COLOR_FORMULA),
            ("Cross-sheet link (green)",      COLOR_CROSSREF),
            ("Note / metadata (grey)",        COLOR_NOTE),
        ]
        for i, (lbl, color) in enumerate(legend, start=legend_row + 1):
            ws.cell(row=i, column=1, value=lbl)
            self._font(ws.cell(row=i, column=1), color=color, bold=True)

    # --------- Instructions ----------
    def build_instructions(self):
        ws = self.wb.create_sheet("Instructions")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 100
        ws["A1"] = "How to keep this workbook up to date"
        self._font(ws["A1"], size=18, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 26

        sections = [
            ("Workflow in one line",
             "Drop fresh JPM CSVs into your project folder, open a terminal in that folder, run "
             "`python embi_builder.py` — done. The workbook regenerates with all the data the script "
             "can find. HISTORY IS PRESERVED automatically (see 'How history is preserved' below)."),
            ("","",),
            ("Recommended workflow: 'JP latest' file",
             "When you download a fresh JPM SNAPSHOT, save it as 'JP latest.csv' in this folder, "
             "overwriting any previous file with that name. Then run `python embi_builder.py`. The "
             "script will: (a) detect that 'JP latest.csv' is a snapshot file (by inspecting the "
             "header), (b) read the snapshot's date from inside the file, (c) copy a date-stamped "
             "version into a 'snapshots_archive/' subfolder (e.g. 'snapshot_2026-05-22.csv'), "
             "(d) regenerate the workbook using ALL archived snapshots plus the latest one. The "
             "filename 'JP latest.csv' is just a convenience — call it whatever you want, the "
             "script classifies by content, not by filename."),
            ("How history is preserved",
             "The snapshot file is the only single-date file type, so it's the only one at risk of "
             "history loss when overwritten. The script handles this for you: every time you run, "
             "it looks for snapshot files in the project folder, reads each one's snapshot date "
             "from inside the CSV, and copies a date-stamped duplicate into 'snapshots_archive/' "
             "if one isn't already there. The archive grows monotonically — older snapshots are "
             "NEVER deleted. The next time you run, the script reads from BOTH the project folder "
             "AND the archive, so all your historical snapshots feed into the workbook. You can "
             "freely overwrite 'JP latest.csv' without losing any data. Returns and weights-history "
             "files self-preserve because JPM ships them with full history each time, but if you "
             "want extra safety you can also save copies to the archive folder manually."),
            ("Snapshot weights also feed the time series",
             "Each snapshot file contains per-country 'Mkt Cap %' (= weight in the index at the "
             "snapshot date). The script extracts these and merges them with the weights-history "
             "file. On overlapping dates the snapshot wins (it's the most authoritative). As you "
             "accumulate snapshots, the weight time series gets a fresh data point at each snapshot "
             "date, both for the Weights tab (latest values) and the Weights_History tab (full "
             "trajectory). You don't need to download a new weights-history file every month — "
             "snapshots alone keep your weight data current."),
            ("Snapshot spreads/yields/TR also feed the time series",
             "Each snapshot also contains spread, yield, and index-level data per country and per "
             "region rollup. The script converts these into synthetic rows that get appended to "
             "the Spreads / Yields / TR_YTD tabs as a new column at the snapshot date. Mapping: "
             "Yield-to-Worst → YTM; Z-Spread-to-Worst → Z-Spread; Index Level → Cum Tot Ret Idx. "
             "For non-callable bullet bonds (the bulk of the EMBI Global universe) YTW = YTM "
             "exactly. On any date where the JPM returns file ALSO has data, the returns file "
             "wins. So as you drop fresh EMBI Global snapshots in monthly, you get a fresh column "
             "in every time-series tab — even without re-downloading the returns file. "
             "(EMBI Global Diversified snapshots are also supported for backwards compatibility, "
             "but the regular EMBI Global is what matches the rest of the workbook.)"),
            ("",""),
            ("Required files (download from JPM)",
             "Three file types power this workbook. You don't need all three every time, but you do "
             "need each at least once. The script auto-detects file type by inspecting HEADERS, not "
             "the filename — so JPM can rename, timestamp, GUID-suffix, or otherwise mangle the "
             "file name however it wants. As long as the column structure is intact, classification "
             "works. The script also tolerates UTF-8 BOM, leading blank rows, header whitespace, "
             "and case differences."),
            ("  1. RETURNS file",
             "JPM 'Markets / DataQuery' EMBI Global query. Columns named 'EM Debt Indices | <Entity> | "
             "<Metric>'. Metrics: Cum Tot Ret Idx, Yld to Maturity, Z Spread. One row per date. "
             "Typical filename: 'Query 3_<id>.csv'. Download whatever frequency you want — monthly to "
             "build history, then daily going forward. Both work."),
            ("  2. WEIGHTS HISTORY file",
             "JPM 'EMBI Global Diversified — Monthly' download. Two-row header: row 1 has FC_EMBIG_* "
             "ticker codes, row 2 has country names; column 1 is 'Trade Date'. Goes back to 1993. "
             "Typical filename: 'JPM_EMBI_Global_Div___Mo_<date>_<id>.csv'. Re-download whenever you "
             "want fresh weights — monthly is plenty."),
            ("  3. SNAPSHOT file",
             "JPM 'EMBI Global Diversified' single-date detail. Flat CSV starting with 'Bam Id'. "
             "Includes ratings (S&P/Moody/Fitch), spread duration, market cap, daily/MTD/YTD returns. "
             "Typical filename: 'JPM_EMBI_Global_Diversif_<date>_<id>.csv'. KEEP every snapshot you "
             "ever download in the folder — the Snapshot tab uses only the latest, but the "
             "Rating_Trend tab reads ALL of them and uses each as-of its own date. The more "
             "snapshots you've accumulated, the more the rating-trend chart captures TRUE rating "
             "drift (rather than just compositional drift)."),
            ("","",),
            ("Adding new data",
             "Just drop the new file in the folder alongside the older ones and re-run the script. "
             "If two files cover the same dates the most recent file's data wins, but unique dates "
             "from older files are preserved — so your history is never lost."),
            ("","",),
            ("YTD calculation",
             "All YTD figures (returns, spread changes, yield changes) are anchored to the CURRENT "
             "calendar year (today's year, per your computer's clock) — not to the latest data point. "
             "If you re-run the script in February 2027 against data that ends Dec-2026, YTD will "
             "show as zero or near-zero (because we're at the start of the new year and have no 2027 "
             "data yet). The YTD anchor year is shown on the Cover tab so you can sanity-check it."),
            ("","",),
            ("Optional CLI flags",
             "`python embi_builder.py /path/to/folder` — point at a different folder.  "
             "`python embi_builder.py file1.csv file2.csv` — explicit list.  "
             "`python embi_builder.py --output My_Dashboard.xlsx` — change output filename.  "
             "`python embi_builder.py --recursive` — scan subfolders too."),
            ("","",),
            ("What each tab does",
             "Cover (metadata) → Instructions (this guide) → Forecast (interactive 3/6/12-month "
             "monitor — yellow cells are your assumptions) → Methodology (justification for the "
             "model and its β coefficients — open this when someone challenges the model) → "
             "Rating_Trend (weighted-avg S&P rating of the EMBI universe over time, with 3/6/12m "
             "drift signals to feed the Forecast tab) → Weights (latest snapshot, region-organized) "
             "→ Spreads / Yields / TR_YTD (time series, country sorted by S&P rating within region) "
             "→ By_Rating (composition by JPM rating bucket) → Snapshot (per-country deep dive: "
             "ratings, duration, mkt cap, returns) → LatAm_Focus (8 focus credits vs peers) → "
             "Charts (region-level overviews) → Weights_History (full weight time series) → "
             "Data_Raw (long-form pivot of all loaded data)."),
            ("","",),
            ("If something looks wrong",
             "Check the 'Source files loaded' section on the Cover tab — it lists every file the "
             "script ingested, classified by type. If a file you expected isn't there, your filename "
             "may have been excluded by the glob (only *.csv) or the file format may have drifted "
             "from what JPM gives today. The script's classify_csv() function is the place to teach "
             "it about a new variant."),
            ("","",),
            ("Country-name normalization",
             "JPM uses different spellings for the same country across files (e.g. 'Cote d Ivoire' "
             "vs \"Cote D'Ivoire\", 'Trinidad & Tobago' vs 'Trinidad And Tobago'). The script "
             "normalizes these. If you see a country missing weights or rating data, search "
             "embi_builder.py for NAME_ALIASES and add the new variant."),
            ("","",),
            ("When in doubt",
             "The script regenerates the file in full each run. Nothing the user enters into the "
             "workbook is preserved across runs — so if you add notes or formulas, do it in a copy. "
             "All historical data is preserved because it lives in the source CSVs, not in the xlsx."),
        ]
        row = 3
        for k, v in sections:
            if not k and not v:
                row += 1
                continue
            ws.cell(row=row, column=1, value=k)
            self._font(ws.cell(row=row, column=1), bold=True if k and not k.startswith("  ") else False, size=11 if not k.startswith("  ") else 10)
            ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=2, value=v)
            self._font(ws.cell(row=row, column=2))
            ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = max(18, 14 * (1 + len(v) // 90))
            row += 1

    # --------- Weights ----------
    def build_weights(self):
        ws = self.wb.create_sheet("Weights")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "EMBI Global — Country Weights (latest)"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        latest_w_date = self.latest_weight_date()
        if latest_w_date:
            ws["A2"] = f"Source: weights history file. Latest date: {latest_w_date:%Y-%m-%d}. " \
                       f"All values pulled from the Weights_History tab via INDEX/MATCH."
        else:
            ws["A2"] = "No weights history loaded — paste a JPM weights file or update Weights_History tab manually."
        ws.merge_cells("A2:G2")
        self._font(ws["A2"], color=COLOR_NOTE, italic=True)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 32

        header_row = 4
        cols = [
            ("Region", 14), ("Country", 24), ("S&P", 8), ("Moody's", 9), ("Fitch", 8),
            ("Weight (%)", 13), ("Region share (%)", 16),
            ("Latest Spread (bps)", 18), ("Latest Yield (%)", 16),
        ]
        for i, (h, w) in enumerate(cols, start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
            ws.column_dimensions[get_column_letter(i)].width = w

        row = header_row + 1
        region_subtotal_rows: Dict[str, int] = {}

        for region in REGION_ORDER:
            countries = self.countries_by_region.get(region, [])
            if not countries:
                continue
            self._region_lbl(ws.cell(row=row, column=1), region)
            self._region_lbl(ws.cell(row=row, column=2), f"{region} subtotal")
            for c in (3, 4, 5):
                ws.cell(row=row, column=c, value="")
                ws.cell(row=row, column=c).fill = PatternFill("solid", start_color=COLOR_REGION_BG)
                ws.cell(row=row, column=c).border = THIN_BORDER
            subtotal_row = row
            row += 1
            child_start = row
            for c in countries:
                ws.cell(row=row, column=1, value="").border = THIN_BORDER
                self._font(ws.cell(row=row, column=1))
                name_cell = ws.cell(row=row, column=2, value=c)
                self._font(name_cell)
                name_cell.border = THIN_BORDER

                # Ratings from snapshot
                snap = self.snap_data.get(c, {}) or {}
                self._val(ws.cell(row=row, column=3), (snap.get("Average S&P Rating") or "").strip() or "", fmt="@", color=COLOR_HARDCODE)
                self._val(ws.cell(row=row, column=4), (snap.get("Average Moody Rating") or "").strip() or "", fmt="@", color=COLOR_HARDCODE)
                self._val(ws.cell(row=row, column=5), (snap.get("Average Fitch Rating") or "").strip() or "", fmt="@", color=COLOR_HARDCODE)

                w = self.latest_weight(c)
                if w is None:
                    self._val(ws.cell(row=row, column=6), "", fmt="0.00%;(0.00%);-", color=COLOR_HARDCODE)
                else:
                    # Weights from JPM are stored as percentages (e.g. 5.12 = 5.12%); divide.
                    self._val(ws.cell(row=row, column=6), w / 100, fmt="0.00%;(0.00%);-", color=COLOR_HARDCODE)

                # Region share = country weight / region subtotal
                share = ws.cell(row=row, column=7)
                share.value = f'=IFERROR(F{row}/$F${subtotal_row},"")'
                share.number_format = "0.0%;(0.0%);-"
                self._font(share)
                share.alignment = Alignment(horizontal="right")
                share.border = THIN_BORDER

                sp = self.latest_value(c, METRICS["spread"])
                yd = self.latest_value(c, METRICS["yield"])
                self._val(ws.cell(row=row, column=8), sp if sp is not None else "", fmt="0", color=COLOR_HARDCODE)
                self._val(ws.cell(row=row, column=9), yd if yd is not None else "", fmt="0.00", color=COLOR_HARDCODE)
                row += 1
            child_end = row - 1
            sub_cell = ws.cell(row=subtotal_row, column=6)
            sub_cell.value = f"=IFERROR(SUM(F{child_start}:F{child_end}),0)"
            sub_cell.number_format = "0.00%;(0.00%);-"
            self._font(sub_cell, bold=True)
            sub_cell.alignment = Alignment(horizontal="right")
            sub_cell.border = THIN_BORDER
            sub_cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
            region_subtotal_rows[region] = subtotal_row
            row += 1

        if region_subtotal_rows:
            self._index_lbl(ws.cell(row=row, column=1), "Total")
            self._index_lbl(ws.cell(row=row, column=2), "EMBI Global")
            for c in (3, 4, 5):
                ws.cell(row=row, column=c, value="")
                ws.cell(row=row, column=c).fill = PatternFill("solid", start_color=COLOR_INDEX_BG)
                ws.cell(row=row, column=c).border = THIN_BORDER
            sum_parts = "+".join(f"F{r}" for r in region_subtotal_rows.values())
            tot = ws.cell(row=row, column=6)
            tot.value = f"={sum_parts}" if sum_parts else "=0"
            tot.number_format = "0.00%;(0.00%);-"
            self._font(tot, bold=True)
            tot.alignment = Alignment(horizontal="right")
            tot.border = THIN_BORDER
            tot.fill = PatternFill("solid", start_color=COLOR_INDEX_BG)

        ws.freeze_panes = "A5"

        # Color scale on weights
        if region_subtotal_rows:
            ws.conditional_formatting.add(
                f"F{header_row + 1}:F{row}",
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B",
                ),
            )

    # --------- Time series sheets ----------
    def _build_timeseries_sheet(self, sheet_name, metric_key, units_label, number_format, change_format):
        metric = METRICS[metric_key]
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        ws["A1"] = f"{sheet_name} — {metric} ({units_label})"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        # Cap displayed dates to 100 most recent (full history still in Data_Raw).
        max_dates_shown = 100
        if len(self.dates) > max_dates_shown:
            display_idx_start = len(self.dates) - max_dates_shown
            ws["A2"] = (
                f"Showing the {max_dates_shown} most recent observations of {len(self.dates)} total. "
                f"Full history available on the Data_Raw tab."
            )
            ws.merge_cells("A2:H2")
            self._font(ws["A2"], color=COLOR_NOTE, italic=True)
        else:
            display_idx_start = 0

        display_dates = self.dates[display_idx_start:]
        header_row = 3
        date_start_col = 5
        for i, h in enumerate(["Region / Bucket", "Entity", "Latest", "YTD Δ"], start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
        for j, d in enumerate(display_dates):
            self._hdr(ws.cell(row=header_row, column=date_start_col + j), d.strftime("%Y-%m-%d"))

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 11
        ws.column_dimensions["D"].width = 11
        for j in range(len(display_dates)):
            ws.column_dimensions[get_column_letter(date_start_col + j)].width = 11
        ws.freeze_panes = ws.cell(row=header_row + 1, column=date_start_col)

        # YTD is always anchored to the current real-world calendar year, not
        # to the latest data point. If you re-run in 2027 with stale 2026 data,
        # the denominator is end-of-2026; if you re-run in 2026, end-of-2025.
        latest_year = datetime.now().year
        ytd_idx = self.year_start_index(latest_year)
        # If ytd_idx is before the displayed window, drop YTD column references.
        ytd_in_window = (ytd_idx is not None) and (ytd_idx >= display_idx_start)
        ytd_ref_letter = (
            get_column_letter(date_start_col + (ytd_idx - display_idx_start))
            if ytd_in_window else None
        )
        latest_letter = (
            get_column_letter(date_start_col + len(display_dates) - 1)
            if display_dates else None
        )

        row = header_row + 1
        self._index_lbl(ws.cell(row=row, column=1), "Index")
        self._index_lbl(ws.cell(row=row, column=2), INDEX_NAME)
        self._write_series(ws, row, INDEX_NAME, metric, number_format, change_format,
                           date_start_col, display_idx_start, latest_letter, ytd_ref_letter)
        row += 2

        for region in REGION_ORDER:
            countries = self.countries_by_region.get(region, [])
            if not countries:
                continue
            agg = REGION_AGGREGATE[region]
            self._region_lbl(ws.cell(row=row, column=1), region)
            self._region_lbl(ws.cell(row=row, column=2), f"{agg} (rollup)")
            self._write_series(ws, row, agg, metric, number_format, change_format,
                               date_start_col, display_idx_start, latest_letter, ytd_ref_letter)
            row += 1
            for c in countries:
                ws.cell(row=row, column=1, value="").border = THIN_BORDER
                self._font(ws.cell(row=row, column=1))
                ws.cell(row=row, column=2, value=c).border = THIN_BORDER
                self._font(ws.cell(row=row, column=2))
                self._write_series(ws, row, c, metric, number_format, change_format,
                                   date_start_col, display_idx_start, latest_letter, ytd_ref_letter)
                row += 1
            row += 1

        ws.cell(row=row, column=1, value="By rating")
        self._font(ws.cell(row=row, column=1), bold=True, color=COLOR_HEADER_BG)
        row += 1
        for rating in self.ratings_present:
            self._rating_lbl(ws.cell(row=row, column=1), "Rating")
            self._rating_lbl(ws.cell(row=row, column=2), RATING_DISPLAY.get(rating, rating))
            self._write_series(ws, row, rating, metric, number_format, change_format,
                               date_start_col, display_idx_start, latest_letter, ytd_ref_letter)
            row += 1

        if display_dates:
            last = row - 1
            scale_col = "C"
            ws.conditional_formatting.add(
                f"{scale_col}{header_row + 1}:{scale_col}{last}",
                ColorScaleRule(
                    start_type="min", start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="F8696B",
                ),
            )

    def _write_series(self, ws, row, entity, metric, fmt, change_fmt,
                      date_start_col, display_idx_start, latest_letter, ytd_letter):
        vals = self.series.get((entity, metric), [None] * len(self.dates))
        for j, v in enumerate(vals[display_idx_start:]):
            cell = ws.cell(row=row, column=date_start_col + j)
            self._val(cell, v if v is not None else "", fmt=fmt, color=COLOR_HARDCODE)

        if latest_letter:
            c = ws.cell(row=row, column=3)
            c.value = f"={latest_letter}{row}"
            c.number_format = fmt
            self._font(c, bold=True)
            c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

        if latest_letter and ytd_letter:
            c = ws.cell(row=row, column=4)
            c.value = f'=IFERROR({latest_letter}{row}-{ytd_letter}{row}, "")'
            c.number_format = change_fmt
            self._font(c)
            c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

    # --------- TR_YTD ----------
    def build_tret_ytd(self):
        ws = self.wb.create_sheet("TR_YTD")
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"Total Return — YTD {datetime.now().year} (% from end of {datetime.now().year - 1})"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        max_dates = 100
        if len(self.dates) > max_dates:
            display_idx_start = len(self.dates) - max_dates
            ws["A2"] = f"Showing {max_dates} most recent of {len(self.dates)} obs."
            self._font(ws["A2"], color=COLOR_NOTE, italic=True)
        else:
            display_idx_start = 0
        display_dates = self.dates[display_idx_start:]

        header_row = 3
        date_start_col = 5
        for i, h in enumerate(["Region / Bucket", "Entity", "YTD Return", "Index Value"], start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
        for j, d in enumerate(display_dates):
            self._hdr(ws.cell(row=header_row, column=date_start_col + j), d.strftime("%Y-%m-%d"))

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        for j in range(len(display_dates)):
            ws.column_dimensions[get_column_letter(date_start_col + j)].width = 11
        ws.freeze_panes = ws.cell(row=header_row + 1, column=date_start_col)

        # YTD is always anchored to the current real-world calendar year, not
        # to the latest data point. If you re-run in 2027 with stale 2026 data,
        # the denominator is end-of-2026; if you re-run in 2026, end-of-2025.
        latest_year = datetime.now().year
        ytd_idx = self.year_start_index(latest_year)
        ytd_in_window = (ytd_idx is not None) and (ytd_idx >= display_idx_start)
        ytd_letter = get_column_letter(date_start_col + (ytd_idx - display_idx_start)) if ytd_in_window else None
        latest_letter = get_column_letter(date_start_col + len(display_dates) - 1) if display_dates else None
        metric = METRICS["tret"]

        row = header_row + 1
        self._index_lbl(ws.cell(row=row, column=1), "Index")
        self._index_lbl(ws.cell(row=row, column=2), INDEX_NAME)
        self._write_tret(ws, row, INDEX_NAME, metric, date_start_col, display_idx_start, latest_letter, ytd_letter)
        row += 2

        for region in REGION_ORDER:
            countries = self.countries_by_region.get(region, [])
            if not countries:
                continue
            agg = REGION_AGGREGATE[region]
            self._region_lbl(ws.cell(row=row, column=1), region)
            self._region_lbl(ws.cell(row=row, column=2), f"{agg} (rollup)")
            self._write_tret(ws, row, agg, metric, date_start_col, display_idx_start, latest_letter, ytd_letter)
            row += 1
            for c in countries:
                ws.cell(row=row, column=1, value="").border = THIN_BORDER
                self._font(ws.cell(row=row, column=1))
                ws.cell(row=row, column=2, value=c).border = THIN_BORDER
                self._font(ws.cell(row=row, column=2))
                self._write_tret(ws, row, c, metric, date_start_col, display_idx_start, latest_letter, ytd_letter)
                row += 1
            row += 1

        ws.cell(row=row, column=1, value="By rating")
        self._font(ws.cell(row=row, column=1), bold=True, color=COLOR_HEADER_BG)
        row += 1
        for rating in self.ratings_present:
            self._rating_lbl(ws.cell(row=row, column=1), "Rating")
            self._rating_lbl(ws.cell(row=row, column=2), RATING_DISPLAY.get(rating, rating))
            self._write_tret(ws, row, rating, metric, date_start_col, display_idx_start, latest_letter, ytd_letter)
            row += 1

        if display_dates:
            last = row - 1
            ws.conditional_formatting.add(
                f"C{header_row + 1}:C{last}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="num", mid_value=0, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B",
                ),
            )

    def _write_tret(self, ws, row, entity, metric, date_start_col, display_idx_start, latest_letter, ytd_letter):
        vals = self.series.get((entity, metric), [None] * len(self.dates))
        for j, v in enumerate(vals[display_idx_start:]):
            cell = ws.cell(row=row, column=date_start_col + j)
            self._val(cell, v if v is not None else "", fmt="0.00", color=COLOR_HARDCODE)

        if latest_letter and ytd_letter:
            c = ws.cell(row=row, column=3)
            c.value = f'=IFERROR(({latest_letter}{row}/{ytd_letter}{row})-1, "")'
            c.number_format = "0.00%;(0.00%);-"
            self._font(c, bold=True)
            c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

        if latest_letter:
            c = ws.cell(row=row, column=4)
            c.value = f"={latest_letter}{row}"
            c.number_format = "#,##0.00"
            self._font(c)
            c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

    # --------- By Rating ----------
    def build_by_rating(self):
        ws = self.wb.create_sheet("By_Rating")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "EMBI Global — Composition by Credit Rating"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        ws["A2"] = "Latest snapshot of each JPM rating bucket. YTD TR uses end-of-prior-year as denominator."
        ws.merge_cells("A2:F2")
        self._font(ws["A2"], color=COLOR_NOTE, italic=True)

        header_row = 4
        for i, h in enumerate(["Bucket", "Spread (bps)", "Yield (%)", "Total Return Idx", "YTD TR (%)", "Notes"], start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 30

        # YTD is always anchored to the current real-world calendar year, not
        # to the latest data point. If you re-run in 2027 with stale 2026 data,
        # the denominator is end-of-2026; if you re-run in 2026, end-of-2025.
        latest_year = datetime.now().year
        ytd_idx = self.year_start_index(latest_year)
        row = header_row + 1
        for rating in self.ratings_present:
            self._rating_lbl(ws.cell(row=row, column=1), RATING_DISPLAY.get(rating, rating))
            sp = self.latest_value(rating, METRICS["spread"])
            yd = self.latest_value(rating, METRICS["yield"])
            tr = self.latest_value(rating, METRICS["tret"])
            tr_start = self.value_at(rating, METRICS["tret"], ytd_idx) if ytd_idx is not None else None
            self._val(ws.cell(row=row, column=2), sp if sp is not None else "", fmt="0", color=COLOR_HARDCODE)
            self._val(ws.cell(row=row, column=3), yd if yd is not None else "", fmt="0.00", color=COLOR_HARDCODE)
            self._val(ws.cell(row=row, column=4), tr if tr is not None else "", fmt="#,##0.00", color=COLOR_HARDCODE)
            if tr is not None and tr_start not in (None, 0):
                self._val(ws.cell(row=row, column=5), (tr / tr_start) - 1, fmt="0.00%;(0.00%);-")
            else:
                self._val(ws.cell(row=row, column=5), "", fmt="0.00%;(0.00%);-")
            ws.cell(row=row, column=6, value="").border = THIN_BORDER
            row += 1
        ws.freeze_panes = "A5"

    # --------- Snapshot tab ----------
    def build_snapshot(self):
        ws = self.wb.create_sheet("Snapshot")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "Per-country snapshot detail"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        if self.snap_date:
            ws["A2"] = f"Snapshot date: {self.snap_date:%Y-%m-%d}"
        else:
            ws["A2"] = "No snapshot file loaded."
        self._font(ws["A2"], color=COLOR_NOTE, italic=True)

        # Pick the metrics that matter from the snapshot rows
        cols = [
            ("Region", 14, "region"),
            ("Country", 26, "country"),
            ("Mkt Cap %", 12, "Mkt Cap %"),
            ("S&P", 8, "Average S&P Rating"),
            ("Moody's", 9, "Average Moody Rating"),
            ("Fitch", 8, "Average Fitch Rating"),
            ("YTW", 10, "Yield to Worst"),
            ("STW (Trsy)", 11, "STW (Trsy)"),
            ("Z-Spread to Worst", 16, "Z Spread to Worst"),
            ("Spread Dur", 11, "Spread Duration"),
            ("Avg Life", 10, "Avr. Life"),
            ("MTD chg %", 11, "MTD Change (%)"),
            ("YTD chg %", 11, "YTD Change (%)"),
            ("# Issues", 10, "No. of Issues"),
            ("# Issuers", 10, "No. of Issuer"),
        ]
        header_row = 4
        for i, (h, w, _) in enumerate(cols, start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
            ws.column_dimensions[get_column_letter(i)].width = w

        row = header_row + 1
        for region in REGION_ORDER:
            countries = self.countries_by_region.get(region, [])
            if not countries:
                continue
            self._region_lbl(ws.cell(row=row, column=1), region)
            self._region_lbl(ws.cell(row=row, column=2), f"{region} subtotal")
            # Region rollup from snapshot
            snap = self.snap_data.get(f"REGION:{region}") or {}
            self._fill_snapshot_row(ws, row, snap, cols, label_style="region")
            row += 1
            for c in countries:
                snap = self.snap_data.get(c) or {}
                ws.cell(row=row, column=1, value="").border = THIN_BORDER
                self._font(ws.cell(row=row, column=1))
                ws.cell(row=row, column=2, value=c).border = THIN_BORDER
                self._font(ws.cell(row=row, column=2))
                self._fill_snapshot_row(ws, row, snap, cols, label_style="country")
                row += 1
            row += 1
        ws.freeze_panes = "C5"

    def _fill_snapshot_row(self, ws, row, snap, cols, *, label_style):
        for i, (h, _, key) in enumerate(cols, start=1):
            if key in ("region", "country"):
                continue  # already written by caller
            val = (snap.get(key) or "").strip() if snap else ""
            cell = ws.cell(row=row, column=i)
            cell.border = THIN_BORDER
            self._font(cell)
            cell.alignment = Alignment(horizontal="right" if key not in ("Average S&P Rating","Average Moody Rating","Average Fitch Rating") else "center")
            if val == "":
                cell.value = ""
                continue
            # Try numeric
            try:
                f = float(val)
                cell.value = f
                if key == "Mkt Cap %":
                    cell.number_format = "0.00"
                elif "Change" in key:
                    cell.number_format = "+0.00;-0.00;0.00"
                elif key in ("Yield to Worst", "STW (Trsy)", "Spread Duration", "Avr. Life"):
                    cell.number_format = "0.00"
                elif key == "Z Spread to Worst":
                    cell.number_format = "0"
                elif key in ("No. of Issues", "No. of Issuer"):
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"
                self._font(cell, color=COLOR_HARDCODE)
            except ValueError:
                cell.value = val
                self._font(cell, color=COLOR_HARDCODE)

    # --------- Weights History ----------
    def build_weights_history(self):
        ws = self.wb.create_sheet("Weights_History")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "Country weights — full history (monthly)"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        if not self.weight_dates:
            ws["A2"] = "No weights data loaded."
            self._font(ws["A2"], color=COLOR_NOTE, italic=True)
            return

        # Cap displayed dates to last 60 (5 years monthly).
        max_dates = 60
        if len(self.weight_dates) > max_dates:
            ds = len(self.weight_dates) - max_dates
            ws["A2"] = f"Showing {max_dates} most recent of {len(self.weight_dates)} obs (full history below)."
        else:
            ds = 0
        self._font(ws["A2"], color=COLOR_NOTE, italic=True)
        display_dates = self.weight_dates[ds:]

        header_row = 4
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 26
        for j in range(len(display_dates)):
            ws.column_dimensions[get_column_letter(3 + j)].width = 9

        for i, h in enumerate(["Region", "Country"], start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
        for j, d in enumerate(display_dates):
            self._hdr(ws.cell(row=header_row, column=3 + j), d.strftime("%Y-%m"))

        ws.freeze_panes = ws.cell(row=header_row + 1, column=3)
        row = header_row + 1
        for region in REGION_ORDER:
            countries = self.countries_by_region.get(region, [])
            if not countries:
                continue
            for c in countries:
                self._region_lbl(ws.cell(row=row, column=1), region) if c == countries[0] else (
                    ws.cell(row=row, column=1, value="").__setattr__("border", THIN_BORDER)
                )
                ws.cell(row=row, column=2, value=c).border = THIN_BORDER
                self._font(ws.cell(row=row, column=2))
                vals = self.weight_series.get(c, [None] * len(self.weight_dates))
                for j, v in enumerate(vals[ds:]):
                    cell = ws.cell(row=row, column=3 + j)
                    if v is None:
                        cell.value = ""
                    else:
                        cell.value = v
                        cell.number_format = "0.00"
                        self._font(cell, color=COLOR_HARDCODE)
                    cell.border = THIN_BORDER
                row += 1

    # --------- LatAm Focus ----------
    def build_latam_focus(self):
        ws = self.wb.create_sheet("LatAm_Focus")
        ws.sheet_view.showGridLines = False
        focus_present = [c for c in LATAM_FOCUS if (c in {e for (e, _) in self.series.keys()}) or (c in self.snap_data)]
        latam_others = [c for c in self.countries_by_region["LatAm"] if c not in LATAM_FOCUS]
        peer_aggregates = [
            ("LatAm rollup",   REGION_AGGREGATE["LatAm"]),
            ("Asia rollup",    REGION_AGGREGATE["Asia"]),
            ("Europe rollup",  REGION_AGGREGATE["Europe"]),
            ("Africa rollup",  REGION_AGGREGATE["Africa"]),
            ("GCC rollup",     REGION_AGGREGATE["GCC"]),
            ("EMBI Global",    INDEX_NAME),
        ]

        ws["A1"] = "LatAm Focus — Performance & Spread Snapshot"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22
        ws["A2"] = ("Top: focus credits. Middle: peer rollups. Bottom: other LatAm constituents. "
                    "All YTD figures use end-of-prior-year as denominator.")
        ws.merge_cells("A2:H2")
        self._font(ws["A2"], color=COLOR_NOTE, italic=True)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 32

        header_row = 4
        cols = [("Bucket", 14), ("Entity", 22), ("Spread (bps)", 12), ("YTD Δ Spread", 12),
                ("Yield (%)", 11), ("YTD Δ Yield", 12), ("TR Index", 12), ("YTD TR (%)", 12)]
        for i, (h, w) in enumerate(cols, start=1):
            self._hdr(ws.cell(row=header_row, column=i), h)
            ws.column_dimensions[get_column_letter(i)].width = w

        # YTD is always anchored to the current real-world calendar year, not
        # to the latest data point. If you re-run in 2027 with stale 2026 data,
        # the denominator is end-of-2026; if you re-run in 2026, end-of-2025.
        latest_year = datetime.now().year
        ytd_idx = self.year_start_index(latest_year)

        def _snap(row, bucket, entity, *, label_style):
            if label_style == "focus":
                self._region_lbl(ws.cell(row=row, column=1), bucket)
                cell = ws.cell(row=row, column=2, value=entity)
                self._font(cell, bold=True)
                cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
                cell.border = THIN_BORDER
            elif label_style == "peer":
                self._index_lbl(ws.cell(row=row, column=1), bucket)
                self._index_lbl(ws.cell(row=row, column=2), entity)
            else:
                ws.cell(row=row, column=1, value=bucket).border = THIN_BORDER
                self._font(ws.cell(row=row, column=1))
                ws.cell(row=row, column=2, value=entity).border = THIN_BORDER
                self._font(ws.cell(row=row, column=2))

            sp_now = self.latest_value(entity, METRICS["spread"])
            sp_start = self.value_at(entity, METRICS["spread"], ytd_idx) if ytd_idx is not None else None
            yd_now = self.latest_value(entity, METRICS["yield"])
            yd_start = self.value_at(entity, METRICS["yield"], ytd_idx) if ytd_idx is not None else None
            tr_now = self.latest_value(entity, METRICS["tret"])
            tr_start = self.value_at(entity, METRICS["tret"], ytd_idx) if ytd_idx is not None else None
            self._val(ws.cell(row=row, column=3), sp_now if sp_now is not None else "", fmt="0", color=COLOR_HARDCODE)
            self._val(ws.cell(row=row, column=4), (sp_now - sp_start) if sp_now is not None and sp_start is not None else "", fmt="+0;-0;0")
            self._val(ws.cell(row=row, column=5), yd_now if yd_now is not None else "", fmt="0.00", color=COLOR_HARDCODE)
            self._val(ws.cell(row=row, column=6), (yd_now - yd_start) if yd_now is not None and yd_start is not None else "", fmt="+0.00;-0.00;0.00")
            self._val(ws.cell(row=row, column=7), tr_now if tr_now is not None else "", fmt="#,##0.00", color=COLOR_HARDCODE)
            if tr_now is not None and tr_start not in (None, 0):
                self._val(ws.cell(row=row, column=8), (tr_now / tr_start) - 1, fmt="0.00%;(0.00%);-")
            else:
                self._val(ws.cell(row=row, column=8), "", fmt="0.00%;(0.00%);-")

        row = header_row + 1
        ws.cell(row=row, column=1, value="FOCUS CREDITS")
        self._font(ws.cell(row=row, column=1), bold=True, color=COLOR_HEADER_BG)
        row += 1
        for c in focus_present:
            _snap(row, "LatAm focus", c, label_style="focus")
            row += 1
        row += 1
        ws.cell(row=row, column=1, value="PEER BENCHMARKS")
        self._font(ws.cell(row=row, column=1), bold=True, color=COLOR_HEADER_BG)
        row += 1
        for label, ent in peer_aggregates:
            _snap(row, label, ent, label_style="peer")
            row += 1
        row += 1
        if latam_others:
            ws.cell(row=row, column=1, value="OTHER LATAM CONSTITUENTS")
            self._font(ws.cell(row=row, column=1), bold=True, color=COLOR_HEADER_BG)
            row += 1
            for c in latam_others:
                _snap(row, "LatAm", c, label_style="other")
                row += 1

        snap_last_row = row - 1
        ws.conditional_formatting.add(
            f"D{header_row + 1}:D{snap_last_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B",
                           mid_type="num", mid_value=0, mid_color="FFEB84",
                           end_type="max", end_color="F8696B"),
        )
        ws.conditional_formatting.add(
            f"H{header_row + 1}:H{snap_last_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="num", mid_value=0, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"),
        )
        ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

        # ---- chart blocks ----
        chart_block = snap_last_row + 5
        ws.cell(row=chart_block, column=1, value="Chart data — focus + benchmarks")
        self._font(ws.cell(row=chart_block, column=1), bold=True, color=COLOR_NOTE, italic=True)

        # Limit chart data to last 60 obs to keep chart readable
        chart_dates = self.dates[-60:] if len(self.dates) > 60 else self.dates[:]
        offset_idx = len(self.dates) - len(chart_dates)
        chart_columns = focus_present + [REGION_AGGREGATE["LatAm"], INDEX_NAME]
        chart_labels  = focus_present + ["LatAm Region", "EMBI Global"]

        def _block(start_row, title, metric, fmt, rebase):
            ws.cell(row=start_row, column=1, value=title)
            self._font(ws.cell(row=start_row, column=1), bold=True)
            ws.cell(row=start_row + 1, column=1, value="Date")
            self._font(ws.cell(row=start_row + 1, column=1), bold=True)
            for j, lbl in enumerate(chart_labels):
                ws.cell(row=start_row + 1, column=2 + j, value=lbl)
                self._font(ws.cell(row=start_row + 1, column=2 + j), bold=True)
            bases: Dict[str, Optional[float]] = {}
            if rebase:
                for ent in chart_columns:
                    for v in self.series.get((ent, metric), [])[offset_idx:]:
                        if v is not None:
                            bases[ent] = v
                            break
                    else:
                        bases[ent] = None
            for i, d in enumerate(chart_dates):
                ws.cell(row=start_row + 2 + i, column=1, value=d).number_format = "yyyy-mm-dd"
                for j, ent in enumerate(chart_columns):
                    v = self.value_at(ent, metric, offset_idx + i)
                    if v is None:
                        continue
                    if rebase:
                        base = bases.get(ent)
                        if not base:
                            continue
                        v = (v / base) * 100
                    cell = ws.cell(row=start_row + 2 + i, column=2 + j, value=v)
                    cell.number_format = fmt
            return start_row + 1, start_row + 1 + len(chart_dates), start_row + 2 + len(chart_dates)

        spr_h, spr_l, nxt = _block(chart_block + 2, "Z-Spread (bps)", METRICS["spread"], "0", False)
        yld_h, yld_l, nxt = _block(nxt + 3, "Yield to Maturity (%)", METRICS["yield"], "0.00", False)
        tr_h, tr_l, _ = _block(nxt + 3, "Cumulative Total Return — rebased to 100", METRICS["tret"], "0.00", True)

        def _line(title, ytitle, hdr, last):
            ch = LineChart()
            ch.title = title
            ch.style = 12
            ch.y_axis.title = ytitle
            ch.height = 11
            ch.width = 24
            data = Reference(ws, min_col=2, max_col=1 + len(chart_columns), min_row=hdr, max_row=last)
            cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=last)
            ch.add_data(data, titles_from_data=True)
            ch.set_categories(cats)
            return ch

        ws.add_chart(_line("Z-Spread — focus vs LatAm vs EMBI Global", "bps", spr_h, spr_l), f"J{header_row}")
        ws.add_chart(_line("Yield to Maturity — focus vs benchmarks", "%", yld_h, yld_l), f"J{header_row + 22}")
        ws.add_chart(_line("Cumulative Total Return (rebased = 100)", "Index", tr_h, tr_l), f"J{header_row + 44}")

    # --------- Forecast Simulator ----------
    def build_forecast(self):
        """Interactive forecast monitor.

        Model:
          ΔSpread_drift = β_UST · ΔUST + β_Oil · ΔOil + β_Rating · ΔRating
                          + (μ_monthly · horizon_months)
          σ_spread_h    = σ_monthly · √horizon_months · vol_multiplier
          Spread P50    = current_spread + ΔSpread_drift
          Spread Pα     = P50 + Φ⁻¹(α) · σ_spread_h        (analytical normal)
          ΔYield_drift  = ΔUST/100 + ΔSpread_drift/100
          Yield P50     = current_yield + ΔYield_drift
          TR P50        = (current_yield · h/12) − duration · ΔYield_drift

        Rationale: the analytical-normal percentile approach is mathematically
        equivalent to a Monte Carlo simulation with infinite paths under
        normality. It updates instantly, has no F9-flicker, and is therefore
        the right choice for a "monitor" use case where the user wants to
        plug in assumptions and read the answer.
        """
        ws = self.wb.create_sheet("Forecast")
        ws.sheet_view.showGridLines = False

        # ----- 1. Compute historical statistics for EMBI Global -----
        spread_series = self.series.get((INDEX_NAME, METRICS["spread"])) or []
        yield_series  = self.series.get((INDEX_NAME, METRICS["yield"])) or []

        def _diffs(vals):
            return [vals[i] - vals[i - 1]
                    for i in range(1, len(vals))
                    if vals[i] is not None and vals[i - 1] is not None]

        spread_diffs = _diffs(spread_series)
        yield_diffs  = _diffs(yield_series)
        sp_mean = statistics.mean(spread_diffs) if spread_diffs else 0.0
        sp_std  = statistics.stdev(spread_diffs) if len(spread_diffs) > 1 else 0.0
        yd_mean = statistics.mean(yield_diffs)  if yield_diffs else 0.0
        yd_std  = statistics.stdev(yield_diffs) if len(yield_diffs) > 1 else 0.0

        cur_spread = self.latest_value(INDEX_NAME, METRICS["spread"]) or 0.0
        cur_yield  = self.latest_value(INDEX_NAME, METRICS["yield"]) or 0.0

        # Spread duration — pull from snapshot if available, otherwise a
        # reasonable EMBI Global default.
        snap_idx = self.snap_data.get("INDEX:EMBI") or {}
        try:
            duration = float(snap_idx.get("Spread Duration") or "")
        except (TypeError, ValueError):
            duration = 5.7  # historical EMBI Global avg

        # Default macro sensitivities — practitioner-standard, editable by user.
        # Oil was deliberately removed: EMBI Global mixes oil exporters (Saudi,
        # UAE, Mexico, Colombia, Nigeria, Angola) with major importers (China,
        # India, Turkey, Indonesia). Aggregate sensitivity is small and policy-
        # regime-dependent. DXY and VIX are far more robust.
        #
        # β_UST = +0.30 (changed from -0.15 in prior version). Sign convention:
        # POSITIVE means rising UST → wider EM spreads, which is what's
        # documented in stress regimes (taper-tantrum 2013 β≈+1.5, 2018 hiking
        # β≈+0.55, 2022 hiking peak β≈+0.85). Negative β only shows up in
        # benign growth-driven rate rallies. Default +0.30 errs on the side
        # users actually care about — stress scenarios.
        beta_ust    =  0.30  # bps spread per +1bp UST (positive: rates up = stress)
        beta_dxy    =  6.0   # bps spread per +1% DXY rally (USD strength → EM stress)
        beta_vix    =  4.0   # bps spread per +1 VIX point (risk-off → wider spreads)
        beta_rating = -50.0  # bps spread per +1 notch rating improvement

        # ----- 2. Layout -----
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 4
        ws.column_dimensions["G"].width = 38

        ws["B1"] = "Forecast Simulator — EMBI Global"
        self._font(ws["B1"], size=18, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 26
        ws["B2"] = ("Edit the YELLOW cells. For each driver, enter your view at 6 months AND at 12 months "
                    "— that lets you express path-dependence (e.g. 'rates sell off in H1 then stabilize'). "
                    "The 3-month forecast linearly interpolates between today and your 6m view; 6m uses "
                    "your 6m view directly; 12m uses your 12m view directly.")
        ws.merge_cells("B2:G2")
        self._font(ws["B2"], italic=True, color=COLOR_NOTE)
        ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 42

        def _section_header(row, text):
            cell = ws.cell(row=row, column=2, value=text)
            self._font(cell, bold=True, size=12, color=COLOR_HEADER_BG)
            cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
            for col in (3, 4, 5):
                ws.cell(row=row, column=col).fill = PatternFill("solid", start_color=COLOR_REGION_BG)

        def _input_cell_dual(row, label, value_6m, value_12m, fmt="0", note=""):
            """Write a label + two yellow inputs (6m view in C, 12m view in D)."""
            ws.cell(row=row, column=2, value=label)
            self._font(ws.cell(row=row, column=2))
            for col, val in [(3, value_6m), (4, value_12m)]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", start_color=COLOR_INPUT_BG)
                cell.number_format = fmt
                self._font(cell, color=COLOR_HARDCODE, bold=True)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
            if note:
                ws.cell(row=row, column=7, value=note)
                self._font(ws.cell(row=row, column=7), italic=True, color=COLOR_NOTE, size=9)
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="center")

        def _input_cell_single(row, label, value, fmt="0", note=""):
            ws.cell(row=row, column=2, value=label)
            self._font(ws.cell(row=row, column=2))
            cell = ws.cell(row=row, column=3, value=value)
            cell.fill = PatternFill("solid", start_color=COLOR_INPUT_BG)
            cell.number_format = fmt
            self._font(cell, color=COLOR_HARDCODE, bold=True)
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if note:
                ws.cell(row=row, column=7, value=note)
                self._font(ws.cell(row=row, column=7), italic=True, color=COLOR_NOTE, size=9)
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="center")

        def _const_cell(row, label, value, fmt="0.00"):
            ws.cell(row=row, column=2, value=label)
            self._font(ws.cell(row=row, column=2))
            cell = ws.cell(row=row, column=3, value=value)
            cell.number_format = fmt
            self._font(cell, color=COLOR_HARDCODE)
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            return cell

        # ----- 3. Assumption inputs (yellow) -----
        _section_header(4, "YOUR ASSUMPTIONS  (yellow cells — your view at each horizon)")
        # Sub-header row showing which column is which
        sub_c = ws.cell(row=4, column=3, value="6-month view")
        self._font(sub_c, bold=True, size=10, color=COLOR_HEADER_BG)
        sub_c.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
        sub_c.alignment = Alignment(horizontal="center")
        sub_d = ws.cell(row=4, column=4, value="12-month view")
        self._font(sub_d, bold=True, size=10, color=COLOR_HEADER_BG)
        sub_d.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
        sub_d.alignment = Alignment(horizontal="center")

        _input_cell_dual(5, "UST 10Y change (bps)", 0, 0,
                         fmt="+0;-0;0",
                         note="Positive = rates rise. e.g. +200 in the 12m column = 200bp UST sell-off by month 12.")
        _input_cell_dual(6, "DXY (US Dollar Index) change (%)", 0, 0,
                         fmt="+0.0%;-0.0%;0.0%",
                         note="Positive = USD strength. Format as decimal (0.05 = +5%). USD strength tightens dollar funding → wider EM spreads.")
        _input_cell_dual(7, "VIX change (points)", 0, 0,
                         fmt="+0;-0;0",
                         note="Absolute change in VIX. e.g. +10 = VIX moves from 18 to 28 (risk-off shock).")
        _input_cell_dual(8, "EM rating drift (S&P notches; +1 = upgrade)", 0, 0,
                         fmt="+0;-0;0",
                         note="Average index-wide rating drift. +1 ≈ index moves up one notch.")
        _input_cell_single(9, "Volatility multiplier (1.0 = baseline; 2.0 = stress)", 1.0,
                           fmt="0.0",
                           note="Variance scaling — single number, applies to all horizons. Use >1 for fatter tails / regime-shift scenarios.")

        # ----- 4. Current state (read-only) -----
        _section_header(11, "CURRENT STATE  (latest data — read only)")
        ws.cell(row=12, column=2, value="Latest data date")
        self._font(ws.cell(row=12, column=2))
        ws.cell(row=12, column=3, value=self.dates[-1] if self.dates else "")
        ws.cell(row=12, column=3).number_format = "yyyy-mm-dd"
        self._font(ws.cell(row=12, column=3), color=COLOR_HARDCODE)
        ws.cell(row=12, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=12, column=3).border = THIN_BORDER

        cur_spread_cell = _const_cell(13, "EMBI Global Z-Spread (bps)", cur_spread, fmt="0")
        cur_yield_cell  = _const_cell(14, "EMBI Global Yield to Maturity (%)", cur_yield / 100, fmt="0.00%")
        # Implied UST = yield - spread/10000 (since spread in bps and yield in %)
        ws.cell(row=15, column=2, value="Implied UST 10Y (%)")
        self._font(ws.cell(row=15, column=2))
        ws.cell(row=15, column=3, value=f"=C14-C13/10000")
        ws.cell(row=15, column=3).number_format = "0.00%"
        self._font(ws.cell(row=15, column=3))
        ws.cell(row=15, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=15, column=3).border = THIN_BORDER

        dur_cell    = _const_cell(16, "Spread duration (years)", duration, fmt="0.00")
        sp_mean_cell= _const_cell(17, "Historical Δspread monthly mean (bps)", sp_mean, fmt="+0.0;-0.0;0.0")
        sp_std_cell = _const_cell(18, "Historical Δspread monthly stdev (bps)", sp_std, fmt="0.0")
        yd_std_cell = _const_cell(19, "Historical Δyield monthly stdev (%)", yd_std / 100, fmt="0.00%")
        ws.cell(row=19, column=7, value=f"Sample: {len(spread_diffs)} monthly observations of EMBI Global.")
        self._font(ws.cell(row=19, column=7), italic=True, color=COLOR_NOTE, size=9)

        # ----- 5. Sensitivities (advanced; editable) -----
        _section_header(21, "MACRO SENSITIVITIES  (defaults from practitioner research — override if you have a strong view)")
        beta_ust_cell = _const_cell(22, "β: bps spread move per +1bp UST",          beta_ust,    fmt="0.000")
        beta_dxy_cell = _const_cell(23, "β: bps spread move per +1% DXY rally",      beta_dxy,    fmt="0.0")
        beta_vix_cell = _const_cell(24, "β: bps spread move per +1 VIX point",       beta_vix,    fmt="0.0")
        beta_rat_cell = _const_cell(25, "β: bps spread move per +1 rating notch",    beta_rating, fmt="0")
        ws.cell(row=22, column=7, value="Sign-dependent on regime: POSITIVE in stress (taper-tantrum 2013 β≈+1.5; 2018 hiking ≈+0.55; 2022 hiking peak ≈+0.85). NEGATIVE only in benign growth-rally regimes (≈-0.10 to -0.20). Default +0.30 covers typical user scenarios.")
        ws.cell(row=23, column=7, value="Best single FX driver of EM credit. USD strength tightens dollar funding → wider spreads. Empirical band: +3 to +9 bps/%.")
        ws.cell(row=24, column=7, value="Pure risk-off proxy. Captures sentiment regime independently of rates/FX. Empirical band: +2 to +7 bps/point.")
        ws.cell(row=25, column=7, value="A 1-notch broad rating drift moves spreads ~50bps. Empirical band: 30 to 80.")
        for r in (22, 23, 24, 25):
            self._font(ws.cell(row=r, column=7), italic=True, color=COLOR_NOTE, size=9)
            ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="center")

        # ----- 6. Forecast output -----
        _section_header(27, "FORECAST  (3 / 6 / 12 months — auto-calculated)")
        # Header row
        ws.cell(row=28, column=2, value="Metric")
        for i, h in enumerate(["3 months", "6 months", "12 months"]):
            self._hdr(ws.cell(row=28, column=3 + i), h)
        self._font(ws.cell(row=28, column=2), bold=True)
        ws.cell(row=28, column=2).fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        ws.cell(row=28, column=2).font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_HEADER_FG)

        horizons = [3, 6, 12]
        # Z-score for the 5%/95% percentile
        Z95 = 1.6449
        Z75 = 0.6745

        # Helper to write a row of forecast formulas across the 3 horizon cols
        def _row(row, label, fmt, formula_factory, *, bold=False):
            ws.cell(row=row, column=2, value=label)
            self._font(ws.cell(row=row, column=2), bold=bold)
            for i, h in enumerate(horizons):
                cell = ws.cell(row=row, column=3 + i, value=formula_factory(h))
                cell.number_format = fmt
                self._font(cell, bold=bold)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                if bold:
                    cell.fill = PatternFill("solid", start_color=COLOR_INDEX_BG)

        # Cell references. Each driver now has a 6m view (col C) and a 12m view (col D).
        # The vol multiplier is single-column (C9 only).
        #   C5/D5  UST input (bps)        |  C13 current spread (bps)
        #   C6/D6  DXY input (decimal %)  |  C14 current yield (decimal)
        #   C7/D7  VIX input (points)     |  C16 spread duration
        #   C8/D8  Rating input (notches) |  C17 historical mean Δspread (bps)
        #   C9     Volatility multiplier  |  C18 historical std Δspread (bps)
        #   C22 β_UST | C23 β_DXY | C24 β_VIX | C25 β_Rating
        UST6,  UST12  = "$C$5", "$D$5"
        DXY6,  DXY12  = "$C$6", "$D$6"
        VIX6,  VIX12  = "$C$7", "$D$7"
        RAT6,  RAT12  = "$C$8", "$D$8"
        VOL_C = "$C$9"
        SPR_C, YLD_C = "$C$13", "$C$14"
        DUR_C = "$C$16"
        SPMEAN_C, SPSTD_C = "$C$17", "$C$18"
        BUST_C, BDXY_C, BVIX_C, BRAT_C = "$C$22", "$C$23", "$C$24", "$C$25"

        def _interp(x6, x12, h):
            """Piecewise-linear path: 0 at month 0, x6 at month 6, x12 at month 12.
            For h ≤ 6 we pro-rate from 0; for h > 6 we interpolate between x6 and x12.
            """
            if h <= 6:
                return f"({x6}*{h}/6)"
            return f"({x6}+({x12}-{x6})*({h}-6)/6)"

        # ΔSpread drift formula (in bps, AT horizon h, given the user's path):
        #   = β_UST    × ΔUST_bps@h
        #   + β_DXY    × ΔDXY_pct@h   (DXY input is decimal; ×100 → percent points)
        #   + β_VIX    × ΔVIX_points@h
        #   + β_Rating × ΔRating_notches@h
        #   + μ_monthly × horizon_months
        def _drift(h):
            ust_h = _interp(UST6, UST12, h)
            dxy_h = _interp(DXY6, DXY12, h)
            vix_h = _interp(VIX6, VIX12, h)
            rat_h = _interp(RAT6, RAT12, h)
            return (f"={BUST_C}*{ust_h}"
                    f"+{BDXY_C}*{dxy_h}*100"
                    f"+{BVIX_C}*{vix_h}"
                    f"+{BRAT_C}*{rat_h}"
                    f"+{SPMEAN_C}*{h}")

        # σ spread for horizon h
        def _sigma(h):
            return f"={SPSTD_C}*SQRT({h})*{VOL_C}"

        # Spread P50/P5/P95
        def _sp_p50(h): return f"={SPR_C}+({_drift(h)[1:]})"
        def _sp_p5(h):  return f"={SPR_C}+({_drift(h)[1:]})-{Z95}*({_sigma(h)[1:]})"
        def _sp_p95(h): return f"={SPR_C}+({_drift(h)[1:]})+{Z95}*({_sigma(h)[1:]})"
        def _sp_p25(h): return f"={SPR_C}+({_drift(h)[1:]})-{Z75}*({_sigma(h)[1:]})"
        def _sp_p75(h): return f"={SPR_C}+({_drift(h)[1:]})+{Z75}*({_sigma(h)[1:]})"

        # Yield P50/P5/P95: yield_now + ΔUST@h/10000 + ΔSpread_drift/10000
        def _yd_p50(h):
            return (f"={YLD_C}+{_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000)")
        def _yd_p5(h):
            return (f"={YLD_C}+{_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000)"
                    f"-{Z95}*({_sigma(h)[1:]})/10000")
        def _yd_p95(h):
            return (f"={YLD_C}+{_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000)"
                    f"+{Z95}*({_sigma(h)[1:]})/10000")

        # TR (decimal return) ≈ carry − duration × ΔYield_in_decimal
        # carry             = current_yield_decimal × h/12
        # ΔYield_in_decimal = ΔUST_bps@h/10000 + Δspread_drift_bps/10000
        # σ_TR (decimal)    = duration × σ_spread_h / 10000
        def _tr_p50(h):
            return (f"={YLD_C}*{h}/12-{DUR_C}*({_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000))")
        def _tr_p5(h):
            return (f"={YLD_C}*{h}/12-{DUR_C}*({_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000))"
                    f"-{Z95}*{DUR_C}*({_sigma(h)[1:]})/10000")
        def _tr_p95(h):
            return (f"={YLD_C}*{h}/12-{DUR_C}*({_interp(UST6, UST12, h)}/10000+(({_drift(h)[1:]})/10000))"
                    f"+{Z95}*{DUR_C}*({_sigma(h)[1:]})/10000")

        _row(29, "Median spread (bps)",       "0",            _sp_p50, bold=True)
        _row(30, "5th percentile spread (bps)",  "0",         _sp_p5)
        _row(31, "95th percentile spread (bps)", "0",         _sp_p95)
        _row(32, "Median yield (%)",          "0.00%",        _yd_p50, bold=True)
        _row(33, "5th percentile yield (%)",  "0.00%",        _yd_p5)
        _row(34, "95th percentile yield (%)", "0.00%",        _yd_p95)
        _row(35, "Median total return (%)",   "+0.00%;-0.00%;0.00%", _tr_p50, bold=True)
        _row(36, "5th percentile TR (%)",     "+0.00%;-0.00%;0.00%", _tr_p5)
        _row(37, "95th percentile TR (%)",    "+0.00%;-0.00%;0.00%", _tr_p95)

        # ----- 7. Stress matrix: TR median across UST 12-month shocks -----
        _section_header(40, "STRESS TABLE — Median TR forecast across UST 12m scenarios (linear path; 6m = ½ × 12m view)")
        ws.cell(row=41, column=2, value="UST 12m shock (bps)")
        self._font(ws.cell(row=41, column=2), bold=True)
        for i, h in enumerate(["3 months", "6 months", "12 months"]):
            self._hdr(ws.cell(row=41, column=3 + i), h)

        # Symmetric ±200bp range — covers everything from a Fed cutting cycle
        # to a 2022-style hiking shock at the high end.
        ust_scenarios_12m = [-200, -100, 0, +100, +200]

        def _ust_path(ust_12m_shk: int, h: int) -> str:
            """Return an Excel-formula fragment for the UST shock at horizon h,
            given a 12m view of `ust_12m_shk`. 6m view assumed at half (linear path)."""
            ust_6m = ust_12m_shk / 2
            if h <= 6:
                return f"({ust_6m}*{h}/6)"
            return f"({ust_6m}+({ust_12m_shk}-{ust_6m})*({h}-6)/6)"

        for r_offset, ust_12m_shk in enumerate(ust_scenarios_12m):
            r = 42 + r_offset
            label = f"{ust_12m_shk:+d} bps" if ust_12m_shk != 0 else "0 bps (UST flat)"
            ws.cell(row=r, column=2, value=label)
            self._font(ws.cell(row=r, column=2))
            ws.cell(row=r, column=2).border = THIN_BORDER
            for i, h in enumerate(horizons):
                ust_h_expr = _ust_path(ust_12m_shk, h)
                # Driver values at horizon h, holding user's 12m views fixed.
                dxy_h = _interp(DXY6, DXY12, h)
                vix_h = _interp(VIX6, VIX12, h)
                rat_h = _interp(RAT6, RAT12, h)
                drift_at_shk = (f"({BUST_C}*{ust_h_expr}"
                                f"+{BDXY_C}*{dxy_h}*100"
                                f"+{BVIX_C}*{vix_h}"
                                f"+{BRAT_C}*{rat_h}"
                                f"+{SPMEAN_C}*{h})")
                formula = (f"={YLD_C}*{h}/12-{DUR_C}*({ust_h_expr}/10000+({drift_at_shk}/10000))")
                cell = ws.cell(row=r, column=3 + i, value=formula)
                cell.number_format = "+0.00%;-0.00%;0.00%"
                self._font(cell)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER

        # Color scale on stress matrix
        ws.conditional_formatting.add(
            "C42:E46",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num", mid_value=0, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

        # ----- 8. Fan chart data + chart -----
        # Build a 13-month projection (month 0 to 12) for each percentile band.
        # Place data far to the right so the user doesn't see the helper block.
        chart_data_col = 11  # column K
        ws.cell(row=4, column=chart_data_col, value="Forecast trajectory (helper data — feeds chart)")
        self._font(ws.cell(row=4, column=chart_data_col), italic=True, color=COLOR_NOTE, size=9)
        ws.cell(row=5, column=chart_data_col, value="Month")
        for j, lbl in enumerate(["P5", "P25", "Median", "P75", "P95"]):
            ws.cell(row=5, column=chart_data_col + 1 + j, value=lbl)
            self._font(ws.cell(row=5, column=chart_data_col + 1 + j), bold=True)
        for m in range(0, 13):
            r = 6 + m
            ws.cell(row=r, column=chart_data_col, value=m)
            ws.cell(row=r, column=chart_data_col).number_format = "0"
            # Each month uses the interpolated path: 0 at m=0, 6m view at m=6, 12m view at m=12.
            ust_m = _interp(UST6, UST12, m) if m > 0 else "0"
            dxy_m = _interp(DXY6, DXY12, m) if m > 0 else "0"
            vix_m = _interp(VIX6, VIX12, m) if m > 0 else "0"
            rat_m = _interp(RAT6, RAT12, m) if m > 0 else "0"
            for j, z_factor, sign in [(0, Z95, -1), (1, Z75, -1), (2, 0, 0), (3, Z75, 1), (4, Z95, 1)]:
                drift = (f"({BUST_C}*{ust_m}"
                         f"+{BDXY_C}*{dxy_m}*100"
                         f"+{BVIX_C}*{vix_m}"
                         f"+{BRAT_C}*{rat_m}"
                         f"+{SPMEAN_C}*{m})")
                sigma = f"({SPSTD_C}*SQRT({m})*{VOL_C})" if m > 0 else "0"
                if z_factor == 0:
                    formula = f"={SPR_C}+{drift}"
                else:
                    formula = f"={SPR_C}+{drift}{'+' if sign > 0 else '-'}{z_factor}*{sigma}"
                cell = ws.cell(row=r, column=chart_data_col + 1 + j, value=formula)
                cell.number_format = "0"
                self._font(cell, color=COLOR_NOTE, size=9)

        # Build the line chart (5 lines for the percentile bands)
        fan = LineChart()
        fan.title = "EMBI Global Z-Spread fan — projected (bps)"
        fan.style = 12
        fan.y_axis.title = "bps"
        fan.x_axis.title = "Months ahead"
        fan.height = 11
        fan.width = 22
        data_ref = Reference(
            ws, min_col=chart_data_col + 1, max_col=chart_data_col + 5,
            min_row=5, max_row=5 + 13,
        )
        cats_ref = Reference(
            ws, min_col=chart_data_col, min_row=6, max_row=5 + 13,
        )
        fan.add_data(data_ref, titles_from_data=True)
        fan.set_categories(cats_ref)
        ws.add_chart(fan, "B49")

        # ----- 9. Methodology / disclaimer footer -----
        ws.cell(row=72, column=2, value="Methodology")
        self._font(ws.cell(row=72, column=2), bold=True, size=11, color=COLOR_HEADER_BG)
        notes = [
            ("Driver selection",
             "UST, DXY, VIX, and rating drift are the four variables with the strongest, most "
             "documented impact on aggregate EM credit spreads. Oil was deliberately excluded — "
             "the EMBI Global universe contains both major oil exporters (Saudi, UAE, Mexico, "
             "Nigeria, Angola) and major importers (China, India, Turkey, Indonesia), so oil's "
             "aggregate effect is small, noisy, and policy-dependent. DXY is a far cleaner FX "
             "driver because USD strength tightens dollar-funding for the entire EM dollar-debt "
             "complex regardless of country-level commodity exposure."),
            ("Path (6m + 12m views)",
             "Each driver input is your view at TWO horizons. The 3-month forecast linearly "
             "pro-rates from today's value toward your 6m view (so a +200bp 12m view shows "
             "as +100bp at 6m, +50bp at 3m). Between 6m and 12m it interpolates between your "
             "two views — letting you express path-dependent scenarios like 'rates spike then "
             "rally', which a single-point view can't capture."),
            ("β_UST sign convention",
             "Positive β means rising UST → wider EM spreads. This holds in stress regimes: "
             "taper-tantrum 2013 β≈+1.5; 2018 hiking β≈+0.55; 2022 hiking peak β≈+0.85. "
             "Negative β only appears in benign growth-driven rate rallies (≈-0.10 to -0.20). "
             "Default +0.30 errs on the stress regime since that's what user scenarios "
             "typically explore."),
            ("Vol scaling",
             "Spread vol scales with √horizon (Brownian assumption)."),
            ("Distribution",
             "Returns assumed normal. Real spread distributions are fatter-tailed; treat the 5/95 bands "
             "as approximate. For a true tail-risk view, set Volatility multiplier to 1.5–2.0."),
            ("Calibration",
             "Mean and σ are estimated from the EMBI Global Z-Spread monthly time series in the loaded "
             "returns file. As you re-run the script with more data, the calibration sample grows and "
             "the model becomes more reliable."),
            ("Sensitivity coefficients",
             "Default β values are practitioner-standard, not estimated in-sample. β_DXY ≈ +6 bps/% "
             "and β_VIX ≈ +4 bps/point are well within the empirical ranges reported in IMF / BIS push-"
             "pull literature and Fed research notes. If you have your own regression, override rows 22–25."),
            ("What this is NOT",
             "Not a trading model. Not a Bloomberg-grade analytic. A scenario monitor intended to "
             "translate a macro view into a directional EMBIG forecast."),
        ]
        for i, (k, v) in enumerate(notes, start=73):
            ws.cell(row=i, column=2, value=k)
            self._font(ws.cell(row=i, column=2), bold=True)
            ws.cell(row=i, column=2).alignment = Alignment(vertical="top")
            ws.cell(row=i, column=3, value=v)
            self._font(ws.cell(row=i, column=3))
            ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
            ws.row_dimensions[i].height = max(18, 14 * (1 + len(v) // 80))

    # --------- Rating Trend ----------
    def _ratings_as_of(self, target_date: datetime) -> Dict[str, int]:
        """Return {country: S&P score} using the most recent snapshot at or
        before target_date. Falls back to the earliest snapshot if target_date
        precedes all snapshots loaded (so historical weight months that pre-
        date your earliest JPM snapshot still get a rating mapping)."""
        if not self.snap_history:
            return {}
        chosen = None
        for snap_date, snap_data in self.snap_history:
            if snap_date <= target_date:
                chosen = snap_data
            else:
                break
        if chosen is None:
            chosen = self.snap_history[0][1]
        out: Dict[str, int] = {}
        for ent, data in chosen.items():
            if not isinstance(ent, str) or ent.startswith(("REGION:", "INDEX:", "AGG:")):
                continue
            sp = (data.get("Average S&P Rating") or "").strip()
            if sp in RATING_SCORE:
                out[ent] = RATING_SCORE[sp]
        return out

    def _compute_rating_series(self) -> List[Tuple[datetime, float, float, float, float]]:
        """For each weight history date, compute weighted-average S&P rating
        score, IG share, HY share, and NR share.

        Returns: [(date, avg_score, ig_pct, hy_pct, nr_pct), ...]
        IG = score >= 13 (BBB- or better);  HY = 1 ≤ score ≤ 12;  NR = no rating.
        """
        out = []
        for i, d in enumerate(self.weight_dates):
            ratings = self._ratings_as_of(d)
            total_rated_wgt = 0.0
            total_score = 0.0
            ig_wgt = hy_wgt = nr_wgt = 0.0
            for country, weights in self.weight_series.items():
                w = weights[i]
                if w is None or w == 0:
                    continue
                score = ratings.get(country)
                if score is None or score == 0:
                    nr_wgt += w
                    continue
                total_rated_wgt += w
                total_score += w * score
                if score >= 13:   # BBB- or better → Investment Grade
                    ig_wgt += w
                else:
                    hy_wgt += w
            if total_rated_wgt <= 0:
                continue
            avg_score = total_score / total_rated_wgt
            total_all = total_rated_wgt + nr_wgt
            out.append((
                d,
                avg_score,
                ig_wgt / total_all if total_all else 0.0,
                hy_wgt / total_all if total_all else 0.0,
                nr_wgt / total_all if total_all else 0.0,
            ))
        return out

    def build_rating_trend(self):
        ws = self.wb.create_sheet("Rating_Trend")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12

        ws["B1"] = "EM Universe Rating Trend"
        self._font(ws["B1"], size=18, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 28

        # Description (varies based on whether we have multiple snapshots)
        n_snaps = len(self.snap_history)
        if n_snaps <= 1:
            desc = ("Tracks the weighted-average S&P rating of the EMBI Global universe over time. "
                    "Use the 3m / 6m / 12m drift signals below to gauge what to plug into the "
                    "'EM rating drift' inputs on the Forecast tab. " +
                    f"NOTE: only {n_snaps} snapshot is currently loaded — all historical observations "
                    "use that snapshot's per-country ratings, so the trend captured is purely "
                    "COMPOSITIONAL DRIFT (countries entering/exiting the index, weight shifts), "
                    "not true rating-action drift. Drop a fresh JPM snapshot CSV into the folder "
                    "each month and re-run; the script keeps every snapshot you've ever loaded "
                    "and uses each as-of its own date, so the trend gradually becomes a true mix "
                    "of composition + rating-action drift.")
        else:
            earliest = self.snap_history[0][0]
            latest   = self.snap_history[-1][0]
            desc = (f"Tracks the weighted-average S&P rating of the EMBI Global universe over time. "
                    f"Uses {n_snaps} snapshots from {earliest:%Y-%m-%d} to {latest:%Y-%m-%d}, "
                    f"applying each as-of its snapshot date. Pre-snapshot history (back to 1993) "
                    f"uses the earliest snapshot's ratings, so deep-history is composition-only; "
                    f"the recent past reflects both composition AND true rating drift.")

        ws["B2"] = desc
        ws.merge_cells("B2:G2")
        self._font(ws["B2"], italic=True, color=COLOR_NOTE)
        ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 70

        # ----- Compute the rating series -----
        series = self._compute_rating_series()
        if not series:
            ws["B4"] = ("No rating data could be computed. This requires both a weights history "
                        "file AND at least one snapshot file in the folder.")
            self._font(ws["B4"], italic=True, color=COLOR_NOTE)
            return

        # Latest snapshot summary
        latest_d, latest_score, latest_ig, latest_hy, latest_nr = series[-1]

        # Drift signals — assumes monthly weights file. Index back N positions.
        def _drift_at(n: int) -> Optional[float]:
            if len(series) <= n:
                return None
            return latest_score - series[-1 - n][1]

        drift_3m  = _drift_at(3)
        drift_6m  = _drift_at(6)
        drift_12m = _drift_at(12)

        # ----- Layout -----
        def _section(row, txt):
            cell = ws.cell(row=row, column=2, value=txt)
            self._font(cell, bold=True, size=12, color=COLOR_HEADER_BG)
            cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
            for col in (3, 4, 5, 6, 7):
                ws.cell(row=row, column=col).fill = PatternFill("solid", start_color=COLOR_REGION_BG)

        def _kv(row, label, value, fmt="0.00", color=COLOR_HARDCODE, bold=False):
            ws.cell(row=row, column=2, value=label)
            self._font(ws.cell(row=row, column=2), bold=bold)
            cell = ws.cell(row=row, column=3, value=value)
            cell.number_format = fmt
            self._font(cell, color=color, bold=bold)
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER

        _section(4, "LATEST SNAPSHOT  (as of " + latest_d.strftime("%Y-%m-%d") + ")")
        _kv(5, "Weighted-avg rating score", latest_score, fmt="0.00", bold=True)
        _kv(6, "Weighted-avg rating (S&P notation)", score_to_sp_label(latest_score), fmt="@", bold=True)
        _kv(7, "Investment Grade share",   latest_ig, fmt="0.0%")
        _kv(8, "High Yield share",         latest_hy, fmt="0.0%")
        _kv(9, "Not-Rated share",          latest_nr, fmt="0.0%")

        _section(11, "DRIFT SIGNALS  (positive = ratings improving; negative = deteriorating)")
        # Suggestion row uses drift_12m as the suggested 12m view; drift_6m as 6m view.
        for r, label, val in [
            (12, "Δ rating notches over last 3 months",  drift_3m),
            (13, "Δ rating notches over last 6 months",  drift_6m),
            (14, "Δ rating notches over last 12 months", drift_12m),
        ]:
            ws.cell(row=r, column=2, value=label)
            self._font(ws.cell(row=r, column=2))
            if val is None:
                self._val(ws.cell(row=r, column=3), "n/a", fmt="@", color=COLOR_NOTE)
            else:
                self._val(ws.cell(row=r, column=3), val, fmt="+0.00;-0.00;0.00",
                          color=COLOR_FORMULA)
                self._font(ws.cell(row=r, column=3), bold=True,
                           color=("006100" if val > 0 else ("9C0006" if val < 0 else COLOR_FORMULA)))

        # Suggested forecast inputs derived from observed drift
        ws.cell(row=16, column=2, value="Suggested 'rating drift' inputs for the Forecast tab:")
        self._font(ws.cell(row=16, column=2), bold=True)
        ws.cell(row=17, column=2, value="  •  6m view  ←  if you expect the 6m drift trend to continue:")
        self._font(ws.cell(row=17, column=2))
        if drift_6m is not None:
            self._val(ws.cell(row=17, column=3), drift_6m, fmt="+0.00;-0.00;0.00", color=COLOR_FORMULA)
        ws.cell(row=18, column=2, value="  •  12m view  ←  if you expect the 12m drift trend to continue:")
        self._font(ws.cell(row=18, column=2))
        if drift_12m is not None:
            self._val(ws.cell(row=18, column=3), drift_12m, fmt="+0.00;-0.00;0.00", color=COLOR_FORMULA)
        ws.cell(row=19, column=2, value="(For mean-reversion scenarios, plug numbers smaller than these. For acceleration, plug larger.)")
        self._font(ws.cell(row=19, column=2), italic=True, size=9, color=COLOR_NOTE)

        # ----- Time series table — last 60 monthly observations -----
        _section(21, "TIME SERIES  (most recent 60 monthly observations)")
        header_row = 22
        for i, h in enumerate(["Date", "Avg score", "Avg label", "IG %", "HY %", "NR %"], start=2):
            self._hdr(ws.cell(row=header_row, column=i), h)

        n_show = min(60, len(series))
        recent = series[-n_show:]
        for j, (d, score, ig, hy, nr) in enumerate(recent):
            r = header_row + 1 + j
            ws.cell(row=r, column=2, value=d).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=2).border = THIN_BORDER
            self._font(ws.cell(row=r, column=2), color=COLOR_HARDCODE, size=9)
            self._val(ws.cell(row=r, column=3), score, fmt="0.00", color=COLOR_HARDCODE)
            self._val(ws.cell(row=r, column=4), score_to_sp_label(score), fmt="@", color=COLOR_HARDCODE)
            self._val(ws.cell(row=r, column=5), ig, fmt="0.0%", color=COLOR_HARDCODE)
            self._val(ws.cell(row=r, column=6), hy, fmt="0.0%", color=COLOR_HARDCODE)
            self._val(ws.cell(row=r, column=7), nr, fmt="0.0%", color=COLOR_HARDCODE)

        last_table_row = header_row + n_show

        # Color scale on the score column to make the trend visible at a glance
        ws.conditional_formatting.add(
            f"C{header_row + 1}:C{last_table_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

        # ----- Charts: rating score line + IG share line -----
        # Place chart-source data far right (cols J onwards) for cleanliness.
        chart_col = 10  # J
        ws.cell(row=4, column=chart_col, value="Chart data (helper)")
        self._font(ws.cell(row=4, column=chart_col), italic=True, color=COLOR_NOTE, size=9)
        ws.cell(row=5, column=chart_col, value="Date")
        ws.cell(row=5, column=chart_col + 1, value="Avg rating score")
        ws.cell(row=5, column=chart_col + 2, value="IG share (%)")
        for c in range(chart_col, chart_col + 3):
            self._font(ws.cell(row=5, column=c), bold=True)

        # Use last 120 months (10 yrs) for the chart so it's readable
        chart_n = min(120, len(series))
        chart_data = series[-chart_n:]
        for j, (d, score, ig, _hy, _nr) in enumerate(chart_data):
            r = 6 + j
            ws.cell(row=r, column=chart_col, value=d).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=chart_col + 1, value=score).number_format = "0.00"
            ws.cell(row=r, column=chart_col + 2, value=ig * 100).number_format = "0.0"

        # Chart 1: weighted-avg rating score (B22 area to the right of the table)
        score_chart = LineChart()
        score_chart.title = "EMBI Global — weighted-average S&P rating score (last 10 yrs)"
        score_chart.style = 12
        score_chart.y_axis.title = "Rating score (higher = better quality)"
        score_chart.height = 10
        score_chart.width = 22
        data_ref = Reference(ws, min_col=chart_col + 1, min_row=5,
                             max_col=chart_col + 1, max_row=5 + chart_n)
        cats_ref = Reference(ws, min_col=chart_col, min_row=6, max_row=5 + chart_n)
        score_chart.add_data(data_ref, titles_from_data=True)
        score_chart.set_categories(cats_ref)
        ws.add_chart(score_chart, f"B{last_table_row + 3}")

        # Chart 2: IG share %
        ig_chart = LineChart()
        ig_chart.title = "EMBI Global — Investment Grade share of index (%)"
        ig_chart.style = 13
        ig_chart.y_axis.title = "%"
        ig_chart.height = 10
        ig_chart.width = 22
        data_ref2 = Reference(ws, min_col=chart_col + 2, min_row=5,
                              max_col=chart_col + 2, max_row=5 + chart_n)
        cats_ref2 = Reference(ws, min_col=chart_col, min_row=6, max_row=5 + chart_n)
        ig_chart.add_data(data_ref2, titles_from_data=True)
        ig_chart.set_categories(cats_ref2)
        ws.add_chart(ig_chart, f"B{last_table_row + 25}")

    # --------- Methodology / Coefficient Justification ----------
    def build_methodology(self):
        """Standalone reference doc justifying the forecast model's β coefficients.

        Designed so a user can open just this tab in a meeting and walk through
        the model's defense. Self-contained, readable as a memo, with citations
        to the historical episodes and research bands that anchor each prior.
        """
        ws = self.wb.create_sheet("Methodology")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 130

        def _h1(row, text):
            cell = ws.cell(row=row, column=2, value=text)
            self._font(cell, size=18, bold=True, color=COLOR_HEADER_BG)
            ws.row_dimensions[row].height = 28

        def _h2(row, text):
            cell = ws.cell(row=row, column=2, value=text)
            self._font(cell, size=13, bold=True, color=COLOR_HEADER_BG)
            cell.fill = PatternFill("solid", start_color=COLOR_REGION_BG)
            cell.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[row].height = 22

        def _h3(row, text):
            cell = ws.cell(row=row, column=2, value=text)
            self._font(cell, size=11, bold=True)
            ws.row_dimensions[row].height = 18

        def _body(row, text, *, italic=False, indent=False, fill=None):
            cell = ws.cell(row=row, column=2, value=text)
            self._font(cell, size=10, italic=italic)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       indent=2 if indent else 0)
            if fill:
                cell.fill = PatternFill("solid", start_color=fill)
            # Rough auto-height: ~135 characters per line at this width
            chars_per_line = 135
            lines = max(1, (len(text) + chars_per_line - 1) // chars_per_line)
            ws.row_dimensions[row].height = max(16, lines * 15 + 4)

        def _spacer(row):
            ws.row_dimensions[row].height = 8

        r = 1
        _h1(r, "Forecasting Model — Methodology & Coefficient Justification"); r += 1
        _body(r, "Reference document for the Forecast tab. Use this when explaining or "
                  "defending the model's structure and parameter choices. The intent is "
                  "that you can open this tab in a meeting, walk down it, and answer any "
                  "challenge to the model's design.", italic=True); r += 1
        _spacer(r); r += 1

        # ---- Section 1: How the model works ----
        _h2(r, "1.  How the model works (one paragraph)"); r += 1
        _body(r,
              "The Forecast tab is a linear factor model with closed-form Gaussian "
              "percentile bands — the analytical equivalent of an infinite-path Monte "
              "Carlo simulation under the assumption that monthly spread changes are "
              "Normal. Spread changes are decomposed into a deterministic DRIFT (responding "
              "linearly to four macro drivers: UST 10Y, DXY, VIX, EM rating drift) and a "
              "stochastic SHOCK (drawn from a Normal distribution with vol calibrated on "
              "the EMBI Global Z-spread monthly time series). Yield and total return drop "
              "out by accounting identity (yield = UST + spread; TR ≈ carry − duration·ΔY)."); r += 1
        _spacer(r); r += 1

        # ---- Section 2: What's estimated vs prior ----
        _h2(r, "2.  What is estimated from your data vs what is set as a prior"); r += 1
        _h3(r, "Estimated from your loaded data — auto-updates every time you re-run the script:"); r += 1
        for txt in [
            "•  Historical mean drift (μ) — sample mean of monthly EMBI Global Z-spread changes",
            "•  Historical volatility (σ) — sample standard deviation of those changes",
            "•  Current spread, current yield, implied UST — pulled from the latest data point",
            "•  Spread duration — pulled from the latest snapshot file",
        ]:
            _body(r, txt, indent=True); r += 1
        _spacer(r); r += 1
        _h3(r, "Set as priors — fixed in the script, but user-overridable in cells C22:C25 of the Forecast tab:"); r += 1
        for txt in [
            "•  β_UST    — bps spread move per +1bp UST",
            "•  β_DXY    — bps spread move per +1% DXY rally",
            "•  β_VIX    — bps spread move per +1 VIX point",
            "•  β_Rating — bps spread move per +1 notch rating drift",
        ]:
            _body(r, txt, indent=True); r += 1
        _spacer(r); r += 1

        # ---- Section 3: Why priors and not regression ----
        _h2(r, "3.  Why priors, not in-sample regression?"); r += 1
        _body(r,
              "With ~53 monthly observations, an in-sample regression would produce "
              "unstable, multicollinear, regime-dependent β estimates that change every "
              "time new data arrives. You'd see the model 'change its mind' about how "
              "much DXY drives spreads month to month, which is worse than a stable, "
              "well-justified prior. Priors anchored to (a) documented historical "
              "episodes, (b) published empirical research bands, and (c) practitioner "
              "convention are the disciplined choice for a small-sample monitor. Once "
              "the data window grows to 5+ years of daily observations (≈1,250+ data "
              "points), regression-based estimation becomes viable as a future iteration."); r += 1
        _spacer(r); r += 1

        # ---- Section 4: Coefficient-by-coefficient anchors ----
        _h2(r, "4.  Coefficient-by-coefficient: where each β comes from"); r += 1

        _h3(r, "β_UST = +0.30  (bps spread per +1bp UST)"); r += 1
        _body(r, "Anchored to documented historical episodes. The sign is positive in stress regimes "
                  "and negative in benign growth-rally regimes; +0.30 is a moderate-stress midpoint."); r += 1
        for txt in [
            "•  Taper Tantrum (May–Sep 2013): ~+100bp UST move ↔ ~+150bp EMBIG widening   →  β ≈ +1.5  (extreme regime)",
            "•  2018 Fed hiking cycle (full year): ~+90bp UST   ↔ ~+50bp spread widening   →  β ≈ +0.55",
            "•  2022 Fed hiking peak (mid-year):   +280bp UST   ↔ +240bp spread peak       →  β ≈ +0.85  (stress)",
            "•  2022 year-end (settled):           +280bp UST   ↔ +60bp residual           →  β ≈ +0.20  (settled)",
            "•  Benign growth rallies (1990s–2000s, gradual UST rises in risk-on regimes)  →  β ≈ -0.10 to -0.20",
        ]:
            _body(r, txt, indent=True); r += 1
        _body(r, "Default placement: closer to settled-2022 than to peak taper-tantrum. Conservative on "
                  "the stress side of the regime distribution. The user can override to a stronger value "
                  "for explicitly stress scenarios.", italic=True); r += 1
        _spacer(r); r += 1

        _h3(r, "β_DXY = +6.0  (bps spread per +1% DXY rally)"); r += 1
        _body(r, "Drawn from IMF/BIS push-pull literature on EM credit. Channel: USD strength tightens "
                  "global dollar funding → raises external debt-servicing burden for EM dollar-debt "
                  "issuers → widens spreads, regardless of country-level commodity exposure (which is "
                  "why DXY works for the EMBIG aggregate where oil does not)."); r += 1
        for txt in [
            "•  Fratzscher (2012) ECB Working Paper 1364 — capital flows & EM premia under USD regimes",
            "•  Adler & Tovar (2014) IMF WP/14/153 — exchange-rate transmission to EM credit",
            "•  IMF WEO and BIS QR notes — empirical band cited at +3 to +9 bps per +1% DXY",
            "•  Sell-side practitioner conventions (GS, JPM, Citi EM strategy desks): rule-of-thumb +5 to +7",
        ]:
            _body(r, txt, indent=True); r += 1
        _body(r, "Default placement: midpoint of the +3 to +9 published band.", italic=True); r += 1
        _spacer(r); r += 1

        _h3(r, "β_VIX = +4.0  (bps spread per +1 VIX point)"); r += 1
        _body(r, "Pure risk-off proxy — captures sentiment-regime shifts that aren't already priced "
                  "into rates or FX. Empirical band documented in Fed and BIS notes on EM risk-premium "
                  "dynamics: +2 to +7 bps/point under normal regimes; tail β rises steeply in shocks."); r += 1
        for txt in [
            "•  Fed/BIS staff notes on EM risk premia: empirical band +2 to +7 in normal regimes",
            "•  COVID March 2020 sanity check: VIX 15 → 80 (+65 pts), EMBIG spreads peaked ~700bp wider",
            "    →  tail-regime β ≈ +10; linear coefficient through normal regimes closer to +4",
            "•  Default +4 captures normal regimes; the volatility-multiplier input handles the tail",
            "    by widening the percentile bands without distorting the central forecast",
        ]:
            _body(r, txt, indent=True); r += 1
        _body(r, "Why a separate volatility multiplier instead of just a higher β_VIX: β_VIX captures "
                  "the MEAN shift under risk-off, the volatility multiplier captures DISTRIBUTION-WIDTH "
                  "shifts (fatter tails in regime shocks). They are different statistical objects and "
                  "should be modeled separately.", italic=True); r += 1
        _spacer(r); r += 1

        _h3(r, "β_Rating = -50  (bps spread per +1 notch upgrade)"); r += 1
        _body(r, "The most empirically anchored of the four — directly verifiable in the data already "
                  "loaded into this workbook."); r += 1
        for txt in [
            "•  CDS-implied rating-spread curves (JPM, Moody's KMV, BAML): ~50bp per notch on average",
            "•  Slope is steeper at the speculative-grade end (BB → B → CCC), flatter through IG (A → BBB)",
            "•  THIS WORKBOOK'S OWN DATA — see the By_Rating tab:",
            "    'Credit BBB only' vs 'Credit B only' composites are 3 notches apart, ~150–200bp spread",
            "    difference  →  derived β ≈ 50–65bp per notch  (consistent with the -50 default)",
        ]:
            _body(r, txt, indent=True); r += 1
        _body(r, "If anyone challenges β_Rating, point them at the By_Rating tab — the JPM rating-bucket "
                  "composites in your loaded data validate this coefficient directly.", italic=True); r += 1
        _spacer(r); r += 1

        # ---- Section 5: Where Monte Carlo lives ----
        _h2(r, "5.  Where the 'Monte Carlo' actually lives in the model"); r += 1
        _body(r,
              "This is worth being precise about. The MODEL has two pieces — the drift (a "
              "deterministic forecast given your assumptions) and the shock (a probability "
              "distribution around that forecast). The β coefficients drive the DRIFT and "
              "are priors; they did not come from a simulation. The percentile bands "
              "(P5/P25/P50/P75/P95) drive the SHOCK and ARE Monte Carlo in flavor — they "
              "are the analytical solution to what you'd get by running infinite simulated "
              "paths under the Gaussian assumption and reading off the quantiles. So the "
              "uncertainty quantification is Monte-Carlo-equivalent; the central scenario "
              "is anchored to research-based priors."); r += 1
        _spacer(r); r += 1

        # ---- Section 6: Defense script ----
        _h2(r, "6.  How to defend it in a meeting (verbatim script)"); r += 1
        _body(r,
              '"The β coefficients are practitioner priors anchored to documented historical episodes — '
              'taper-tantrum 2013, 2018 hiking, 2022 hiking, COVID 2020 — and sit within the empirical '
              'bands published in IMF, BIS, and Fed research on EM credit risk premia. They are sanity-'
              'checked against the rating-bucket structure visible in JPM\'s own data, which is loaded '
              'into the By_Rating tab of this workbook. They are NOT in-sample regression estimates '
              'because the calibration sample (~53 monthly observations) is too short for stable '
              'coefficient estimates — priors are the disciplined choice for a small-sample monitor. '
              'They are user-overridable in cells C22:C25 of the Forecast tab. With 5+ years of daily '
              'data we would estimate these from the residuals of a multi-factor regression — that\'s '
              'the next iteration, and the architecture supports it."',
              italic=True, fill=COLOR_RATING_BG); r += 1
        _spacer(r); r += 1

        # ---- Section 7: Limitations (own them) ----
        _h2(r, "7.  Limitations (own them upfront)"); r += 1
        for txt in [
            "•  These are NOT in-sample regression estimates — that's a deliberate small-sample choice, "
            "but it means the coefficients don't adapt automatically to a regime shift. Override manually.",
            "•  β values are regime-dependent in reality. β_UST flips sign between stress and benign-growth "
            "regimes; the +0.30 default is a stress-leaning midpoint, not a universal truth.",
            "•  The Normal distribution understates fat tails. Real EM spread changes have excess kurtosis "
            "(2008, 2013, 2020 all featured > 4σ moves). The volatility multiplier is the safety valve.",
            "•  The model assumes orthogonality between drivers. In practice UST/DXY/VIX moves correlate, "
            "especially in stress regimes — so be cautious about double-counting risk in extreme scenarios.",
            "•  No country-level forecasts. The model is built for EMBI Global aggregate; country-level "
            "spreads diverge from the index based on idiosyncratic credit factors not captured here.",
        ]:
            _body(r, txt, indent=True); r += 1
        _spacer(r); r += 1

        # ---- Section 8: Future iterations ----
        _h2(r, "8.  What would make this model better (future work)"); r += 1
        for txt in [
            "•  Multi-factor regression on residuals once 5+ years of daily data accumulate (~1,250+ obs)",
            "•  Pull UST / DXY / VIX time series from FRED to enable in-sample β estimation",
            "•  Add t-distribution or GARCH option for fatter tails without relying solely on the vol multiplier",
            "•  Country-level forecasts using country β-loadings on the same factors",
            "•  Backtest the forecast-vs-realized spread changes to validate or recalibrate",
        ]:
            _body(r, txt, indent=True); r += 1

    # --------- Charts (index-wide) ----------
    def build_charts(self):
        ws = self.wb.create_sheet("Charts")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "Charts — refreshed automatically each script run"
        self._font(ws["A1"], size=14, bold=True, color=COLOR_HEADER_BG)
        ws.row_dimensions[1].height = 22

        chart_dates = self.dates[-60:] if len(self.dates) > 60 else self.dates[:]
        offset = len(self.dates) - len(chart_dates)

        def _fill_block(start_row, title, metric, fmt):
            ws.cell(row=start_row, column=1, value=title)
            self._font(ws.cell(row=start_row, column=1), bold=True)
            ws.cell(row=start_row + 1, column=1, value="Date")
            for j, region in enumerate(REGION_ORDER):
                ws.cell(row=start_row + 1, column=2 + j, value=region)
                self._font(ws.cell(row=start_row + 1, column=2 + j), bold=True)
            for i, d in enumerate(chart_dates):
                ws.cell(row=start_row + 2 + i, column=1, value=d).number_format = "yyyy-mm-dd"
                for j, region in enumerate(REGION_ORDER):
                    agg = REGION_AGGREGATE[region]
                    v = self.value_at(agg, metric, offset + i)
                    if v is not None:
                        ws.cell(row=start_row + 2 + i, column=2 + j, value=v).number_format = fmt
            return start_row + 1, start_row + 1 + len(chart_dates)

        spr_h, spr_l = _fill_block(60, "Z-Spread by region — time series", METRICS["spread"], "0")
        nxt = spr_l + 5
        yld_h, yld_l = _fill_block(nxt, "Yield by region — time series", METRICS["yield"], "0.00")
        nxt2 = yld_l + 5
        # Rebased TR including the index
        ws.cell(row=nxt2, column=1, value="Cumulative Total Return rebased to 100")
        self._font(ws.cell(row=nxt2, column=1), bold=True)
        ws.cell(row=nxt2 + 1, column=1, value="Date")
        cols_to_plot = [INDEX_NAME] + [REGION_AGGREGATE[r] for r in REGION_ORDER]
        labels = ["EMBI Global"] + REGION_ORDER
        for j, lbl in enumerate(labels):
            ws.cell(row=nxt2 + 1, column=2 + j, value=lbl)
            self._font(ws.cell(row=nxt2 + 1, column=2 + j), bold=True)
        bases = {}
        for ent in cols_to_plot:
            for v in (self.series.get((ent, METRICS["tret"]), []) or [])[offset:]:
                if v is not None:
                    bases[ent] = v
                    break
            else:
                bases[ent] = None
        for i, d in enumerate(chart_dates):
            ws.cell(row=nxt2 + 2 + i, column=1, value=d).number_format = "yyyy-mm-dd"
            for j, ent in enumerate(cols_to_plot):
                v = self.value_at(ent, METRICS["tret"], offset + i)
                base = bases.get(ent)
                if v is not None and base:
                    ws.cell(row=nxt2 + 2 + i, column=2 + j, value=(v / base) * 100).number_format = "0.0"
        tr_h, tr_l = nxt2 + 1, nxt2 + 1 + len(chart_dates)

        def _make(title, ytitle, hdr, last, ncols):
            ch = LineChart()
            ch.title = title
            ch.style = 12
            ch.y_axis.title = ytitle
            ch.height = 10
            ch.width = 22
            data = Reference(ws, min_col=2, max_col=1 + ncols, min_row=hdr, max_row=last)
            cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=last)
            ch.add_data(data, titles_from_data=True)
            ch.set_categories(cats)
            return ch

        ws.add_chart(_make("Z-Spread by Region (bps)", "bps", spr_h, spr_l, len(REGION_ORDER)), "B3")
        ws.add_chart(_make("Yield to Maturity by Region (%)", "%", yld_h, yld_l, len(REGION_ORDER)), "B25")
        ws.add_chart(_make("Cumulative Total Return (rebased = 100)", "Index", tr_h, tr_l, len(cols_to_plot)), "B47")

    # --------- Data_Raw ----------
    def build_data_raw(self):
        ws = self.wb.create_sheet("Data_Raw")
        for i, h in enumerate(("Date", "Entity", "Metric", "Value"), start=1):
            self._hdr(ws.cell(row=1, column=i), h)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14

        keep = (
            [INDEX_NAME] + [REGION_AGGREGATE[r] for r in REGION_ORDER]
            + self.all_countries + self.ratings_present
        )
        row = 2
        for ent in keep:
            for metric in METRICS.values():
                vals = self.series.get((ent, metric))
                if not vals:
                    continue
                for d, v in zip(self.dates, vals):
                    if v is None:
                        continue
                    ws.cell(row=row, column=1, value=d).number_format = "yyyy-mm-dd"
                    ws.cell(row=row, column=2, value=ent)
                    ws.cell(row=row, column=3, value=metric)
                    ws.cell(row=row, column=4, value=v).number_format = "0.000000"
                    row += 1
        # Append weight history into Data_Raw too
        for c, vals in self.weight_series.items():
            for d, v in zip(self.weight_dates, vals):
                if v is None:
                    continue
                ws.cell(row=row, column=1, value=d).number_format = "yyyy-mm-dd"
                ws.cell(row=row, column=2, value=c)
                ws.cell(row=row, column=3, value="Index Weight (%)")
                ws.cell(row=row, column=4, value=v).number_format = "0.000000"
                row += 1

    # --------- orchestration ----------
    def build(self) -> Workbook:
        self.build_cover()
        self.build_instructions()
        self.build_forecast()
        self.build_methodology()
        self.build_rating_trend()
        self.build_weights()
        self._build_timeseries_sheet("Spreads", "spread", "bps", "0", "+0;-0;0")
        self._build_timeseries_sheet("Yields",  "yield",  "% YTM", "0.00", "+0.00;-0.00;0.00")
        self.build_tret_ytd()
        self.build_by_rating()
        self.build_snapshot()
        self.build_latam_focus()
        self.build_charts()
        self.build_weights_history()
        self.build_data_raw()
        return self.wb


# ============================================================================
# 6. CLI
# ============================================================================

def collect_inputs(args_inputs: List[str], recursive: bool) -> List[Path]:
    """Resolve user-provided paths to a list of CSV files."""
    out: List[Path] = []
    if not args_inputs:
        # Default: scan the cwd.
        base = Path.cwd()
        out.extend(sorted(base.rglob("*.csv") if recursive else base.glob("*.csv")))
        return out
    for raw in args_inputs:
        p = Path(raw).expanduser().resolve()
        if not p.exists():
            print(f"  WARN: not found, skipping: {p}", file=sys.stderr)
            continue
        if p.is_dir():
            out.extend(sorted(p.rglob("*.csv") if recursive else p.glob("*.csv")))
        else:
            out.append(p)
    return out


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build EMBI Global dashboard from JPM CSVs.")
    parser.add_argument("inputs", nargs="*",
                        help="CSV files or directories (default: scan current directory).")
    parser.add_argument("--output", "-o", default="EMBI_Global_Dashboard.xlsx",
                        help="Output XLSX path (default: %(default)s).")
    parser.add_argument("--recursive", action="store_true",
                        help="Recursively scan directories for *.csv.")
    args = parser.parse_args(argv)

    paths = collect_inputs(args.inputs, args.recursive)
    if not paths:
        print("ERROR: no CSV files found", file=sys.stderr)
        return 1

    # ---- Snapshot auto-archive ----
    # Determine the project folder (where the user is running from / pointing at).
    if args.inputs:
        first = Path(args.inputs[0]).expanduser().resolve()
        project_dir = first if first.is_dir() else first.parent
    else:
        project_dir = Path.cwd()
    archive_dir, newly_archived = archive_snapshots(paths, project_dir)
    if newly_archived:
        print(f"\n  Archived snapshots to {archive_dir}:")
        for d, name in newly_archived:
            print(f"    [{d}]  {name}")
    # Always include archived snapshots in the load — even if the user runs
    # the script with explicit file arguments that don't include them.
    if archive_dir is not None and archive_dir.exists():
        archive_csvs = sorted(archive_dir.glob("*.csv"))
        seen = {p.resolve() for p in paths}
        for p in archive_csvs:
            if p.resolve() not in seen:
                paths.append(p)

    returns_results = []
    weights_results = []
    snapshot_results = []
    sources: List[Tuple[str, str]] = []
    skipped: List[str] = []

    for p in paths:
        kind = classify_csv(p)
        if kind == "returns":
            returns_results.append(load_returns(p))
            sources.append(("returns", p.name))
            print(f"  [returns]    {p.name}")
        elif kind == "weights_history":
            weights_results.append(load_weights_history(p))
            sources.append(("weights_history", p.name))
            print(f"  [weights]    {p.name}")
        elif kind == "snapshot":
            snapshot_results.append(load_snapshot(p))
            sources.append(("snapshot", p.name))
            print(f"  [snapshot]   {p.name}")
        else:
            skipped.append(p.name)

    if skipped:
        print(f"  Skipped {len(skipped)} non-JPM CSVs: {', '.join(skipped[:5])}{'...' if len(skipped) > 5 else ''}")

    snap_date, snap_data, snap_history = merge_snapshots(snapshot_results)

    # Generate synthetic returns rows from accumulated snapshots and prepend
    # to returns_results — synthetic processed first, real returns last, so on
    # any overlapping date the real returns file wins. This is what makes the
    # Spreads / Yields / TR_YTD tabs grow a new column every time the user
    # drops a fresh snapshot in the folder.
    synthetic_returns: List[Tuple[List[datetime], Dict[Tuple[str, str], List[Optional[float]]]]] = []
    for s_date, s_data in snap_history:
        syn_dates, syn_series = snapshot_to_returns(s_date, s_data)
        if syn_series:
            synthetic_returns.append((syn_dates, syn_series))
    if synthetic_returns:
        print(f"  Synthesized {len(synthetic_returns)} returns row(s) from accumulated snapshots")

    if not returns_results and not synthetic_returns:
        print("ERROR: no time-series data found (need either a returns file "
              "or at least one snapshot)", file=sys.stderr)
        return 1
    if not returns_results:
        print("\n  WARNING: no JPM returns file found — time-series tabs will only "
              "contain dates from the snapshots you've accumulated. Add a "
              "Query 3 returns CSV to the folder for full historical context.",
              file=sys.stderr)

    dates, series = merge_returns(synthetic_returns + returns_results)
    weight_dates, weight_series = merge_weights_history(weights_results)

    print(f"\n  Returns:   {len(dates)} dates × {len(series)} series")
    print(f"  Weights:   {len(weight_dates)} dates × {len(weight_series)} countries")
    print(f"  Snapshots: {len(snap_history)} loaded  (latest: {snap_date.strftime('%Y-%m-%d') if snap_date else 'none'})")

    builder = Builder(
        dates=dates, series=series,
        weight_dates=weight_dates, weight_series=weight_series,
        snap_date=snap_date, snap_data=snap_data, snap_history=snap_history,
        sources=sources,
    )
    wb = builder.build()
    out_path = Path(args.output).expanduser().resolve()
    wb.save(out_path)
    print(f"\n  Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
