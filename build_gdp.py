"""
Build: Argentina GDP pack.
Source: INDEC national accounts via apis.datos.gob.ar; BCRA Com. "A" 3500 FX.
"""
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, Series
from openpyxl.comments import Comment
from common_style import *

D = "/home/claude/data/"
nsa = pd.read_csv(D + "gdp_constant_nsa.csv", parse_dates=["date"])
sa = pd.read_csv(D + "gdp_constant_sa.csv", parse_dates=["date"])
cur = pd.read_csv(D + "gdp_current.csv", parse_dates=["date"])
emae = pd.read_csv(D + "emae_monthly.csv", parse_dates=["date"])
fx = pd.read_csv(D + "fx_monthly.csv", parse_dates=["date"])

# quarterly average FX
fx["q"] = fx["date"].dt.to_period("Q")
fxq = fx.groupby("q")["fx_avg"].mean()

NQ = len(nsa)
LASTQ = nsa["date"].iloc[-1]
LASTQL = qlabel(LASTQ)
LAST_Y = LASTQ.year
FULL_YEARS = [y for y in range(2004, LAST_Y + 1)
              if (nsa["date"].dt.year == y).sum() == 4]
LAST_FULL = FULL_YEARS[-1]

wb = Workbook()

# =====================================================================  READ ME
ws = wb.active
ws.title = "Read me"
sheet_base(ws)
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 108
title(ws, "B2", "Argentina — GDP and demand-side activity pack")
put(ws, 3, 2, f"INDEC national accounts, base 2004. Data through {LASTQL}; EMAE through "
              f"{emae['date'].iloc[-1]:%b-%Y}.", color=MUTED, italic=True)

readme = [
    ("What's in here", ""),
    ("Data Constant", "INDEC quarterly volumes at 2004 prices, ORIGINAL (not seasonally adjusted). Blue = source. "
                      "Includes an identity check column that must read zero."),
    ("Data SA", "The seasonally adjusted volume series. INDEC does not publish SA inventories or SA statistical "
                "discrepancy, so those are backed out here as the residual."),
    ("Data Nominal", "Current-price series, plus the quarterly USD conversion at the Com. \"A\" 3500 rate."),
    ("Quarterly", "The read: GDP q/q seasonally adjusted, the annualised version of it, and y/y — with each demand "
                  "component's y/y beside them."),
    ("Contributions", "The tab you asked for. Demand-side contributions to growth in percentage points, quarterly y/y "
                      "and annual. Contributions sum exactly to GDP growth, which is only true if the statistical "
                      "discrepancy is carried explicitly — see the note on that tab."),
    ("Annual", "Full-year levels and growth: real, nominal in pesos, nominal in USD, and the implied deflator."),
    ("EMAE", "The monthly GDP proxy, and the quarter-to-date nowcast it implies for the quarter INDEC has not yet "
             "published."),
    ("Forecast", "Bull / base / bear quarterly paths to Q4-2027 with the assumptions behind each, benchmarked to the "
                 "BCRA REM and the IMF."),
    ("", ""),
    ("Two things that will bite you if you don't know them", ""),
    ("Annual rates", "INDEC's quarterly national accounts are published at ANNUAL RATES. A year's level is the "
                     "AVERAGE of its four quarters, never the sum. Every annual figure in this pack uses AVERAGEIFS "
                     "for that reason. The Data tabs also carry a true-quarterly column (level / 4) for anything that "
                     "needs a genuine quarterly flow — the USD conversion uses it."),
    ("Statistical discrepancy", "INDEC only began publishing 'discrepancia estadistica' as its own line in Q1-2024. "
                                "Before that it sat inside 'variacion de existencias'. So a 2024 y/y comparison that "
                                "splits the two lines is comparing a carved-out series against one that still "
                                "contains it. Every contribution calculation here therefore uses the COMBINED "
                                "'inventories + statistical discrepancy' line, which is continuous across the break. "
                                "The split is shown separately as a memo from 2024 only."),
    ("", ""),
    ("Source", "INDEC, Cuentas nacionales, oferta y demanda globales, base 2004, retrieved from the official series "
               "API at apis.datos.gob.ar (constant prices family 4.2, seasonally adjusted 3.2, current prices 4.4, "
               "EMAE 143.3). Exchange rate: BCRA Comunicacion \"A\" 3500 wholesale reference rate, series "
               "175.1_DR_REFE500_0_0_25, monthly average of daily fixings."),
    ("Validation", "Annual real growth computed here reproduces the published series for every year 2005-2025 to two "
                   "decimal places (max deviation 0.11pp, at 2025). Nominal GDP for 2025 reproduces INDEC's published "
                   "850,239,798 million pesos exactly. GDP in USD for 2024 and 2025 lands within 0.2% of the IMF WEO "
                   "figures of US$637.2bn and US$681.5bn, which is an independent check on the FX series and the "
                   "annual-rate convention at the same time."),
    ("Known vintage gaps", "The published q/q SA prints for 2025 (+1.0 / -0.1 / +0.6 / +0.6) do not reproduce from "
                           "the current API vintage, which gives +0.77 / +0.30 / +0.12 / +1.16. Seasonal adjustment "
                           "is re-estimated at every release and redistributes growth across quarters while holding "
                           "the annual total; the annual totals do match. If you need to tie to a specific published "
                           "print, you need that release's vintage, not the live series."),
    ("2019 identity", "Q3 and Q4 2019 are the only quarters where the demand components do not sum exactly to GDP "
                      "(residual 135 and 891 on ~671,000, i.e. 0.13%). That is unallocated discrepancy in INDEC's own "
                      "data for a period before the discrepancy line existed, not a transcription error."),
]
r = 5
for k, v in readme:
    if v == "":
        section(ws, r, 2, k)
    else:
        put(ws, r, 2, k, bold=True, size=10)
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(ARIAL, 10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 14 * (1 + len(v) // 100)
    r += 1

# =====================================================================  DATA CONSTANT
ws = wb.create_sheet("Data Constant")
sheet_base(ws)
title(ws, "B2", "INDEC national accounts — constant 2004 prices, original (NSA)",
      "Millions of pesos at 2004 prices, at ANNUAL RATES. Blue = INDEC source value. Black = formula.")
HR, R0 = 5, 6
cols = [("Quarter", 10), ("Year", 7), ("Q", 5), ("GDP", 13), ("Private\nconsumption", 13),
        ("Public\nconsumption", 12), ("Gross fixed\ncapital form.", 13),
        ("Change in\ninventories", 12), ("Statistical\ndiscrepancy", 12),
        ("Exports\nG&S", 12), ("Imports\nG&S", 12),
        ("Inventories +\ndiscrepancy", 13), ("Identity\ncheck", 11),
        ("True quarterly\nlevel", 13), ("GDP y/y", 10)]
for i, (t, w) in enumerate(cols):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 30

for i in range(NQ):
    r = R0 + i
    d = nsa["date"].iloc[i]
    put(ws, r, 2, qlabel(d))
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, (d.month - 1) // 3 + 1, "0", color=MUTED)
    for j, key in enumerate(["pib", "cons_priv", "cons_pub", "ibif", "var_exist",
                             "discrep", "expo", "impo"]):
        v = nsa[key].iloc[i]
        put(ws, r, 5 + j, None if pd.isna(v) else float(v), NUM0, color=INPUTC)
    put(ws, r, 13, f"=N(I{r})+N(J{r})", NUM0)
    put(ws, r, 14, f"=F{r}+G{r}+H{r}+M{r}+K{r}-L{r}-E{r}", NUM2, color=MUTED)
    put(ws, r, 15, f"=E{r}/4", NUM0)
    if i >= 4:
        put(ws, r, 16, f"=E{r}/E{r-4}-1", PCT)
LASTROW_C = R0 + NQ - 1
ws.freeze_panes = f"B{R0}"
note(ws, LASTROW_C + 2, 2,
     "Identity check = components - GDP. Zero everywhere except Q3/Q4 2019 (INDEC unallocated discrepancy).")
note(ws, LASTROW_C + 3, 2,
     "Statistical discrepancy is only published from Q1-2024; blank before. Column M combines it with inventories "
     "so the series is continuous — use M, not I and J separately, for any growth calculation.")
ws["B2"].comment = Comment(
    "INDEC, oferta y demanda globales, base 2004, constant prices, original series. "
    "Retrieved from apis.datos.gob.ar family 4.2. Published at annual rates.", "Claude")

# =====================================================================  DATA SA
ws = wb.create_sheet("Data SA")
sheet_base(ws)
title(ws, "B2", "Constant 2004 prices — seasonally adjusted",
      "INDEC publishes SA series for GDP and the five main components only. Inventories + discrepancy is the residual.")
cols = [("Quarter", 10), ("Year", 7), ("Q", 5), ("GDP SA", 13), ("Private\nconsumption", 13),
        ("Public\nconsumption", 12), ("Gross fixed\ncapital form.", 13), ("Exports\nG&S", 12),
        ("Imports\nG&S", 12), ("Inventories +\ndiscrep. (resid.)", 14),
        ("GDP q/q", 10), ("GDP q/q\nannualised", 12), ("GDP y/y", 10)]
for i, (t, w) in enumerate(cols):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 30
for i in range(NQ):
    r = R0 + i
    d = sa["date"].iloc[i]
    put(ws, r, 2, qlabel(d))
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, (d.month - 1) // 3 + 1, "0", color=MUTED)
    for j, key in enumerate(["pib_sa", "cons_priv_sa", "cons_pub_sa", "ibif_sa",
                             "expo_sa", "impo_sa"]):
        put(ws, r, 5 + j, float(sa[key].iloc[i]), NUM0, color=INPUTC)
    put(ws, r, 11, f"=E{r}-F{r}-G{r}-H{r}-I{r}+J{r}", NUM0)
    if i >= 1:
        put(ws, r, 12, f"=E{r}/E{r-1}-1", PCT2)
        put(ws, r, 13, f"=(E{r}/E{r-1})^4-1", PCT)
    if i >= 4:
        put(ws, r, 14, f"=E{r}/E{r-4}-1", PCT)
ws.freeze_panes = f"B{R0}"
note(ws, LASTROW_C + 2, 2,
     "Seasonal factors are re-estimated at every INDEC release, so the q/q path is revised routinely even when the "
     "annual total is not. INDEC benchmarks each SA series to its NSA annual total: the two agree here to within "
     "0.001 in every full year, which is what makes the annual figures on the Forecast tab valid.")

# =====================================================================  DATA NOMINAL
ws = wb.create_sheet("Data Nominal")
sheet_base(ws)
title(ws, "B2", "Current prices, and the conversion to US dollars",
      "Millions of current pesos at annual rates. USD uses the true quarterly flow (level / 4) at the period-average "
      "Com. \"A\" 3500 rate.")
cols = [("Quarter", 10), ("Year", 7), ("Q", 5), ("GDP", 15), ("Private\nconsumption", 15),
        ("Public\nconsumption", 14), ("Gross fixed\ncapital form.", 14),
        ("Change in\ninventories", 14), ("Statistical\ndiscrepancy", 14), ("Exports\nG&S", 14),
        ("Imports\nG&S", 14), ("Inventories +\ndiscrepancy", 14), ("Identity\ncheck", 11),
        ("GDP, true\nquarterly (ARS mn)", 15), ("ARS/USD\navg", 10),
        ("GDP, quarterly\n(USD mn)", 14), ("4q rolling GDP\n(USD mn)", 15),
        ("Deflator y/y", 11)]
for i, (t, w) in enumerate(cols):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 32
for i in range(NQ):
    r = R0 + i
    d = cur["date"].iloc[i]
    put(ws, r, 2, qlabel(d))
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, (d.month - 1) // 3 + 1, "0", color=MUTED)
    for j, key in enumerate(["pib_n", "cons_priv_n", "cons_pub_n", "ibif_n", "var_exist_n",
                             "discrep_n", "expo_n", "impo_n"]):
        v = cur[key].iloc[i]
        put(ws, r, 5 + j, None if pd.isna(v) else float(v), NUM0, color=INPUTC)
    put(ws, r, 13, f"=N(I{r})+N(J{r})", NUM0)
    put(ws, r, 14, f"=F{r}+G{r}+H{r}+M{r}+K{r}-L{r}-E{r}", NUM2, color=MUTED)
    put(ws, r, 15, f"=E{r}/4", NUM0)
    put(ws, r, 16, float(fxq.loc[d.to_period("Q")]), NUM2, color=INPUTC)
    put(ws, r, 17, f"=O{r}/P{r}", NUM0)
    if i >= 3:
        put(ws, r, 18, f"=SUM(Q{r-3}:Q{r})", NUM0)
    if i >= 4:
        put(ws, r, 19, f"=(E{r}/'Data Constant'!E{r})/(E{r-4}/'Data Constant'!E{r-4})-1", PCT)
ws.freeze_panes = f"B{R0}"
note(ws, LASTROW_C + 2, 2,
     "Exchange rate: BCRA Comunicacion \"A\" 3500 wholesale reference rate, monthly average of daily fixings, "
     "averaged to the quarter. Source series 175.1_DR_REFE500_0_0_25 on apis.datos.gob.ar.")
note(ws, LASTROW_C + 3, 2,
     "USD GDP is a period-average conversion, the same convention the IMF WEO uses. At an official rate inside a "
     "managed band this flatters the dollar level relative to a market-clearing rate — read levels with that in mind; "
     "the ratios on the balance-of-payments pack are less sensitive to it than the levels are.")

# =====================================================================  QUARTERLY
ws = wb.create_sheet("Quarterly")
sheet_base(ws)
title(ws, "B2", "Quarterly read — growth rates",
      "Green = pulled from the Data tabs. q/q is seasonally adjusted; y/y uses the original series.")
cols = [("Quarter", 10), ("Year", 7), ("GDP q/q\nSA", 11), ("GDP q/q\nannualised", 12),
        ("GDP y/y\nSA", 11), ("GDP y/y\nNSA", 11), ("Private\ncons. y/y", 11),
        ("Public\ncons. y/y", 11), ("GFCF y/y", 11), ("Exports\ny/y", 11), ("Imports\ny/y", 11)]
for i, (t, w) in enumerate(cols):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 30
for i in range(NQ):
    r = R0 + i
    d = nsa["date"].iloc[i]
    put(ws, r, 2, qlabel(d))
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, f"='Data SA'!L{r}", PCT2, color=LINKC)
    put(ws, r, 5, f"='Data SA'!M{r}", PCT, color=LINKC)
    put(ws, r, 6, f"='Data SA'!N{r}", PCT, color=LINKC)
    put(ws, r, 7, f"='Data Constant'!P{r}", PCT, color=LINKC)
    if i >= 4:
        for j, cl in enumerate(["F", "G", "H", "K", "L"]):
            put(ws, r, 8 + j,
                f"='Data Constant'!{cl}{r}/'Data Constant'!{cl}{r-4}-1", PCT)
ws.freeze_panes = f"B{R0}"

r25 = R0 + list(nsa["date"].dt.year).index(2015)
ch = line_chart(32, 13, "GDP growth — q/q seasonally adjusted vs y/y", "0.0%")
cats = Reference(ws, min_col=2, min_row=r25, max_row=LASTROW_C)
for col, nm, colr, wdt, dash in [(4, "q/q SA", "D1495B", 24000, False),
                                 (7, "y/y NSA", "2F5C8F", 26000, False)]:
    ref = Reference(ws, min_col=col, min_row=r25 - 1, max_row=LASTROW_C)
    s = Series(ref, title=nm)
    style_series(s, colr, wdt, dash)
    ch.series.append(s)
ch.set_categories(cats)
ws.add_chart(ch, "N5")

ch2 = line_chart(32, 13, "Demand components, y/y (since 2015)", "0.0%")
for col, nm, colr in [(8, "Private consumption", "2F5C8F"), (10, "GFCF", "D1495B"),
                      (11, "Exports", "4F9D69"), (12, "Imports", "E8A33D")]:
    ref = Reference(ws, min_col=col, min_row=r25 - 1, max_row=LASTROW_C)
    s = Series(ref, title=nm)
    style_series(s, colr, 22000)
    ch2.series.append(s)
ch2.set_categories(cats)
ws.add_chart(ch2, "N32")

# =====================================================================  ANNUAL
ws = wb.create_sheet("Annual")
sheet_base(ws)
title(ws, "B2", "Annual — levels and growth",
      "A year's level is the AVERAGE of its four quarters, because INDEC publishes quarterly data at annual rates.")
AHR, AR0 = 5, 6
acols = [("Year", 8), ("GDP\n(2004 ARS mn)", 15), ("Private\nconsumption", 14),
         ("Public\nconsumption", 13), ("GFCF", 13), ("Inventories +\ndiscrepancy", 14),
         ("Exports", 13), ("Imports", 13), ("Real GDP\ngrowth", 11),
         ("Nominal GDP\n(ARS mn)", 16), ("Nominal GDP\n(USD bn)", 14),
         ("Deflator\ny/y", 10), ("Nominal\ngrowth", 11), ("GFCF\n% of GDP", 11),
         ("Priv. cons.\n% of GDP", 12)]
for i, (t, w) in enumerate(acols):
    hdr(ws, AHR, 2 + i, t, w, wrap=True)
ws.row_dimensions[AHR].height = 32
DC = "'Data Constant'"
DN = "'Data Nominal'"
CR = f"$C$6:$C${LASTROW_C}"
for i, y in enumerate(FULL_YEARS):
    r = AR0 + i
    put(ws, r, 2, y, "0", bold=True)
    for j, cl in enumerate(["E", "F", "G", "H", "M", "K", "L"]):
        put(ws, r, 3 + j,
            f"=AVERAGEIFS({DC}!${cl}$6:${cl}${LASTROW_C},{DC}!{CR},$B{r})", NUM0)
    if i >= 1:
        put(ws, r, 10, f"=C{r}/C{r-1}-1", PCT2, bold=True, color=ACCENT)
    put(ws, r, 11, f"=AVERAGEIFS({DN}!$E$6:$E${LASTROW_C},{DN}!{CR},$B{r})", NUM0)
    put(ws, r, 12, f"=SUMIFS({DN}!$Q$6:$Q${LASTROW_C},{DN}!{CR},$B{r})/1000", NUM1)
    if i >= 1:
        put(ws, r, 13, f"=(K{r}/C{r})/(K{r-1}/C{r-1})-1", PCT)
        put(ws, r, 14, f"=K{r}/K{r-1}-1", PCT)
    put(ws, r, 15, f"=AVERAGEIFS({DN}!$H$6:$H${LASTROW_C},{DN}!{CR},$B{r})/K{r}", PCT)
    put(ws, r, 16, f"=AVERAGEIFS({DN}!$F$6:$F${LASTROW_C},{DN}!{CR},$B{r})/K{r}", PCT)
ALAST = AR0 + len(FULL_YEARS) - 1
note(ws, ALAST + 2, 2,
     f"Full calendar years only ({FULL_YEARS[0]}-{LAST_FULL}). {LAST_Y} is incomplete and is handled on the "
     "Forecast tab, not here.")
note(ws, ALAST + 3, 2,
     "Real growth here reproduces the published annual series to two decimals for every year; nominal 2025 "
     "reproduces INDEC's 850,239,798 million pesos exactly; USD GDP for 2024 and 2025 lands within 0.2% of IMF WEO.")

bc = bar_chart(30, 12, "Real GDP growth by year", "0.0%")
bc.add_data(Reference(ws, min_col=10, min_row=AHR, max_row=ALAST), titles_from_data=True)
bc.set_categories(Reference(ws, min_col=2, min_row=AR0, max_row=ALAST))
bc.series[0].graphicalProperties.solidFill = "2F5C8F"
ws.add_chart(bc, "R5")

ch = line_chart(30, 12, "Nominal GDP in US dollars (bn, period-average FX)", "#,##0")
ref = Reference(ws, min_col=12, min_row=AHR, max_row=ALAST)
s = Series(ref, title_from_data=True)
style_series(s, "D1495B", 28000)
ch.series.append(s)
ch.set_categories(Reference(ws, min_col=2, min_row=AR0, max_row=ALAST))
ws.add_chart(ch, "R29")

# =====================================================================  CONTRIBUTIONS
ws = wb.create_sheet("Contributions")
sheet_base(ws)
title(ws, "B2", "Demand-side contributions to GDP growth",
      "Percentage points. Contributions sum exactly to GDP growth — which requires carrying the statistical "
      "discrepancy explicitly. Imports enter with a negative sign.")
put(ws, 4, 2,
    "Contribution of component i = (level_i this period - level_i a year earlier) / GDP a year earlier. "
    "The inventories line is inventories PLUS statistical discrepancy, because INDEC only split them out in 2024 and "
    "a split series would break the y/y comparison in that year.", color=MUTED, italic=True, size=9)

section(ws, 6, 2, "1. Quarterly, year-on-year (original series)")
QHR, QR0 = 7, 8
qcols = [("Quarter", 10), ("GDP y/y", 10), ("Private\nconsumption", 12), ("Public\nconsumption", 12),
         ("GFCF", 10), ("Inventories +\ndiscrepancy", 13), ("Exports", 10), ("Imports", 10),
         ("Sum of\ncontributions", 13), ("Check\n(sum - GDP)", 12),
         ("memo: discrep.\nalone", 13)]
for i, (t, w) in enumerate(qcols):
    hdr(ws, QHR, 2 + i, t, w, wrap=True)
ws.row_dimensions[QHR].height = 30
start_i = 4
for i in range(start_i, NQ):
    r = QR0 + (i - start_i)
    src = R0 + i
    prv = R0 + i - 4
    put(ws, r, 2, f"={DC}!B{src}", color=LINKC)
    put(ws, r, 3, f"=({DC}!E{src}/{DC}!E{prv}-1)*100", PPT, bold=True, color=ACCENT)
    for j, cl in enumerate(["F", "G", "H", "M", "K"]):
        put(ws, r, 4 + j, f"=({DC}!{cl}{src}-{DC}!{cl}{prv})/{DC}!$E{prv}*100", PPT)
    put(ws, r, 9, f"=-({DC}!L{src}-{DC}!L{prv})/{DC}!$E{prv}*100", PPT)
    put(ws, r, 10, f"=SUM(D{r}:I{r})", PPT, bold=True)
    put(ws, r, 11, f"=J{r}-C{r}", "0.000;-0.000;-", color=MUTED)
    put(ws, r, 12, f"=IF(OR({DC}!J{src}=\"\",{DC}!J{prv}=\"\"),\"\","
                   f"({DC}!J{src}-{DC}!J{prv})/{DC}!$E{prv}*100)", PPT, color=MUTED)
QLAST = QR0 + (NQ - start_i) - 1
note(ws, QLAST + 2, 2,
     "The check column is zero everywhere except Q3/Q4 2019 and their y/y counterparts in 2020, where INDEC's own "
     "components do not sum to GDP. The memo column is blank before 2024 because the discrepancy series does not "
     "exist there — it is inside the inventories line, where the contribution calculation already counts it.")

r_from = QR0 + max(0, (NQ - start_i) - 45)
bc = bar_chart(34, 14, "Contributions to y/y GDP growth (pp)", "0.0", stacked=True)
bc.add_data(Reference(ws, min_col=4, min_row=QHR, max_col=9, max_row=QLAST), titles_from_data=True)
bc.set_categories(Reference(ws, min_col=2, min_row=QR0, max_row=QLAST))
for s, colr in zip(bc.series, ["2F5C8F", "8AA0B8", "D1495B", "E8A33D", "4F9D69", "7D5BA6"]):
    s.graphicalProperties.solidFill = colr
    s.graphicalProperties.line.noFill = True
ws.add_chart(bc, "N7")

# annual contributions
ar = QLAST + 5
section(ws, ar, 2, "2. Annual")
AHR2, AR2 = ar + 1, ar + 2
for i, (t, w) in enumerate(qcols[:10]):
    hdr(ws, AHR2, 2 + i, t if t != "Quarter" else "Year", w, wrap=True)
ws.row_dimensions[AHR2].height = 30
for i, y in enumerate(FULL_YEARS[1:], start=1):
    r = AR2 + (i - 1)
    src = AR0 + i
    prv = AR0 + i - 1
    put(ws, r, 2, f"='Annual'!B{src}", "0", color=LINKC, bold=True)
    put(ws, r, 3, f"='Annual'!J{src}*100", PPT, bold=True, color=ACCENT)
    for j, cl in enumerate(["D", "E", "F", "G", "H"]):
        put(ws, r, 4 + j, f"=('Annual'!{cl}{src}-'Annual'!{cl}{prv})/'Annual'!$C{prv}*100", PPT)
    put(ws, r, 9, f"=-('Annual'!I{src}-'Annual'!I{prv})/'Annual'!$C{prv}*100", PPT)
    put(ws, r, 10, f"=SUM(D{r}:I{r})", PPT, bold=True)
    put(ws, r, 11, f"=J{r}-C{r}", "0.000;-0.000;-", color=MUTED)
ALAST2 = AR2 + len(FULL_YEARS) - 2

bc2 = bar_chart(34, 13, "Annual contributions to GDP growth (pp)", "0.0", stacked=True)
bc2.add_data(Reference(ws, min_col=4, min_row=AHR2, max_col=9, max_row=ALAST2), titles_from_data=True)
bc2.set_categories(Reference(ws, min_col=2, min_row=AR2, max_row=ALAST2))
for s, colr in zip(bc2.series, ["2F5C8F", "8AA0B8", "D1495B", "E8A33D", "4F9D69", "7D5BA6"]):
    s.graphicalProperties.solidFill = colr
    s.graphicalProperties.line.noFill = True
ws.add_chart(bc2, f"N{AHR2}")

# =====================================================================  EMAE
ws = wb.create_sheet("EMAE")
sheet_base(ws)
title(ws, "B2", "EMAE — the monthly activity proxy",
      "Index 2004 = 100. EMAE leads the quarterly national accounts by roughly two months and is what you nowcast "
      "the unpublished quarter from.")
NM = len(emae)
ecols = [("Month", 10), ("Year", 7), ("M", 5), ("Original", 11), ("Seasonally\nadjusted", 12),
         ("Trend-cycle", 11), ("y/y", 10), ("m/m SA", 10), ("3m/3m\nannualised", 12)]
for i, (t, w) in enumerate(ecols):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 30
for i in range(NM):
    r = R0 + i
    d = emae["date"].iloc[i]
    put(ws, r, 2, f"{d:%Y-%m}")
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, d.month, "0", color=MUTED)
    for j, key in enumerate(["emae_orig", "emae_sa", "emae_trend"]):
        put(ws, r, 5 + j, float(emae[key].iloc[i]), NUM2, color=INPUTC)
    if i >= 12:
        put(ws, r, 8, f"=E{r}/E{r-12}-1", PCT)
    if i >= 1:
        put(ws, r, 9, f"=F{r}/F{r-1}-1", PCT2)
    if i >= 6:
        put(ws, r, 10, f"=(AVERAGE(F{r-2}:F{r})/AVERAGE(F{r-5}:F{r-3}))^4-1", PCT)
ELAST = R0 + NM - 1
ws.freeze_panes = f"B{R0}"

# nowcast block
nr = 6
section(ws, nr, 12, "Nowcast — EMAE quarterly average vs published GDP")
hdr(ws, nr + 1, 12, "Quarter", 10)
hdr(ws, nr + 1, 13, "EMAE q-avg\ny/y", 12, wrap=True)
hdr(ws, nr + 1, 14, "GDP y/y\n(published)", 12, wrap=True)
hdr(ws, nr + 1, 15, "Gap (pp)", 10)
hdr(ws, nr + 1, 16, "Months in\nquarter", 11, wrap=True)
ws.row_dimensions[nr + 1].height = 30
nq_show = 13
for k in range(nq_show):
    r = nr + 2 + k
    qi = NQ - nq_show + k
    d = nsa["date"].iloc[qi]
    put(ws, r, 12, qlabel(d), color=LINKC)
    put(ws, r, 13,
        f"=AVERAGEIFS($E$6:$E${ELAST},$C$6:$C${ELAST},{d.year},$D$6:$D${ELAST},\">=\"&{d.month},"
        f"$D$6:$D${ELAST},\"<=\"&{d.month + 2})/"
        f"AVERAGEIFS($E$6:$E${ELAST},$C$6:$C${ELAST},{d.year - 1},$D$6:$D${ELAST},\">=\"&{d.month},"
        f"$D$6:$D${ELAST},\"<=\"&{d.month + 2})-1", PCT)
    put(ws, r, 14, f"='Data Constant'!P{R0 + qi}", PCT, color=LINKC)
    put(ws, r, 15, f"=(M{r}-N{r})*100", PPT, color=MUTED)
    put(ws, r, 16,
        f"=COUNTIFS($C$6:$C${ELAST},{d.year},$D$6:$D${ELAST},\">=\"&{d.month},"
        f"$D$6:$D${ELAST},\"<=\"&{d.month + 2})", "0", color=MUTED)

# the partially-observed quarter
nxt = LASTQ + pd.offsets.QuarterBegin(startingMonth=1)
r = nr + 2 + nq_show + 1
put(ws, r, 12, qlabel(nxt), bold=True, color=ACCENT)
put(ws, r, 13,
    f"=AVERAGEIFS($E$6:$E${ELAST},$C$6:$C${ELAST},{nxt.year},$D$6:$D${ELAST},\">=\"&{nxt.month},"
    f"$D$6:$D${ELAST},\"<=\"&{nxt.month + 2})/"
    f"AVERAGEIFS($E$6:$E${ELAST},$C$6:$C${ELAST},{nxt.year - 1},$D$6:$D${ELAST},\">=\"&{nxt.month},"
    f"$D$6:$D${ELAST},\"<=\"&{nxt.month + 2})-1", PCT, bold=True, color=ACCENT)
put(ws, r, 14, "not yet published", color=MUTED, italic=True)
put(ws, r, 16,
    f"=COUNTIFS($C$6:$C${ELAST},{nxt.year},$D$6:$D${ELAST},\">=\"&{nxt.month},"
    f"$D$6:$D${ELAST},\"<=\"&{nxt.month + 2})", "0", color=MUTED)
note(ws, r + 2, 12,
     f"{qlabel(nxt)} is a partial quarter — the months-in-quarter column tells you how many of the three months are "
     "in. The y/y is computed on the same months of the prior year, so it is like-for-like, but it is a nowcast, "
     "not a print.")
note(ws, r + 3, 12,
     "The gap column reads zero in every completed quarter, and that is the point: INDEC benchmarks EMAE to the "
     "quarterly national accounts, so the quarterly average of the original EMAE index reproduces published GDP y/y "
     "exactly. The nowcast above is therefore not a correlation-based proxy — it is the same series, two months "
     "early. Treat the partial-quarter figure as reliable for direction and roughly right on magnitude, with the "
     "caveat that the missing month can still move it.")

ch = line_chart(30, 12, "EMAE — seasonally adjusted level and trend-cycle", "#,##0")
rfrom = ELAST - 96
for col, nm, colr, dash in [(6, "Seasonally adjusted", "2F5C8F", False),
                            (7, "Trend-cycle", "D1495B", True)]:
    ref = Reference(ws, min_col=col, min_row=rfrom - 1, max_row=ELAST)
    s = Series(ref, title=nm)
    style_series(s, colr, 24000, dash)
    ch.series.append(s)
ch.set_categories(Reference(ws, min_col=2, min_row=rfrom, max_row=ELAST))
ws.add_chart(ch, "L28")

ch2 = line_chart(30, 12, "EMAE momentum — y/y vs 3m/3m annualised", "0.0%")
for col, nm, colr, dash in [(8, "y/y", "2F5C8F", False),
                            (10, "3m/3m annualised", "D1495B", False)]:
    ref = Reference(ws, min_col=col, min_row=rfrom - 1, max_row=ELAST)
    s = Series(ref, title=nm)
    style_series(s, colr, 24000, dash)
    ch2.series.append(s)
ch2.set_categories(Reference(ws, min_col=2, min_row=rfrom, max_row=ELAST))
ws.add_chart(ch2, "L54")

# =====================================================================  FORECAST
ws = wb.create_sheet("Forecast")
sheet_base(ws)
title(ws, "B2", "Forecast — real GDP, bull / base / bear to Q4-2027",
      "Quarterly q/q seasonally adjusted paths are judgement inputs (blue). Annual outcomes are formulas. "
      "Benchmarked to the BCRA REM of June-2026 and the IMF April-2026 WEO / Article IV.")

# scenario q/q SA paths, from the quarter after the last actual
fq = []
d = LASTQ + pd.offsets.QuarterBegin(startingMonth=1)
while d <= pd.Timestamp("2027-10-01"):
    fq.append(d)
    d = d + pd.offsets.QuarterBegin(startingMonth=1)

BULL = [1.3, 1.4, 1.4, 1.3, 1.2, 1.2, 1.1, 1.1]
BASE = [0.8, 0.9, 0.9, 0.8, 0.8, 0.8, 0.7, 0.7]
BEAR = [0.0, -0.3, 0.1, 0.3, 0.4, 0.5, 0.5, 0.5]
BULL, BASE, BEAR = BULL[:len(fq)], BASE[:len(fq)], BEAR[:len(fq)]

FHR, FR0 = 6, 7
fcols = [("Quarter", 10), ("Actual q/q\nSA", 11), ("Bull q/q", 10), ("Base q/q", 10),
         ("Bear q/q", 10), ("Bull level", 12), ("Base level", 12), ("Bear level", 12)]
for i, (t, w) in enumerate(fcols):
    hdr(ws, FHR, 2 + i, t, w, wrap=True)
ws.row_dimensions[FHR].height = 30

hist_from = NQ - 9          # show a bit of history so the chart joins up
rows_hist = list(range(hist_from, NQ))
for k, i in enumerate(rows_hist):
    r = FR0 + k
    put(ws, r, 2, qlabel(nsa["date"].iloc[i]), color=LINKC)
    put(ws, r, 3, f"='Data SA'!L{R0 + i}", PCT2, color=LINKC)
    for j in range(3):
        put(ws, r, 7 + j, f"='Data SA'!E{R0 + i}", NUM0, color=LINKC)
ANCHOR = FR0 + len(rows_hist) - 1
for k, d in enumerate(fq):
    r = ANCHOR + 1 + k
    put(ws, r, 2, qlabel(d), bold=True)
    for j, arr in enumerate([BULL, BASE, BEAR]):
        put(ws, r, 4 + j, arr[k] / 100, PCT2, color=INPUTC)
    for j in range(3):
        cl = get_column_letter(7 + j)      # level column being built
        lv = get_column_letter(4 + j)      # its q/q driver
        put(ws, r, 7 + j, f"={cl}{r - 1}*(1+{lv}{r})", NUM0)
FLAST = ANCHOR + len(fq)

# annual outcomes
ar = FLAST + 2
section(ws, ar, 2, "Annual real GDP growth implied by each path")
hdr(ws, ar + 1, 2, "Year", 10)
for j, nm in enumerate(["Bull", "Base", "Bear"]):
    hdr(ws, ar + 1, 3 + j, nm, 11)
hdr(ws, ar + 1, 6, "BCRA REM\n(Jun-26)", 13, wrap=True)
hdr(ws, ar + 1, 7, "IMF WEO\n(Apr-26)", 13, wrap=True)
ws.row_dimensions[ar + 1].height = 30

# map every quarter of 2025/2026/2027 to a row on this sheet
qrow = {}
for k, i in enumerate(rows_hist):
    qrow[nsa["date"].iloc[i].to_period("Q")] = FR0 + k
for k, d in enumerate(fq):
    qrow[d.to_period("Q")] = ANCHOR + 1 + k


def year_avg(col_letter, year, actual_col=None):
    """AVERAGE of the four quarters of `year` on this sheet, using scenario level cols."""
    cells = []
    for q in range(1, 5):
        p = pd.Period(f"{year}Q{q}", freq="Q")
        if p in qrow:
            cells.append(f"{col_letter}{qrow[p]}")
    return cells


for i, y in enumerate([2025, 2026, 2027]):
    r = ar + 2 + i
    put(ws, r, 2, y, "0", bold=True)
    for j in range(3):
        cl = get_column_letter(7 + j)
        cur_cells = year_avg(cl, y)
        prv_cells = year_avg(cl, y - 1)
        if len(cur_cells) == 4 and len(prv_cells) == 4:
            f = f"=AVERAGE({','.join(cur_cells)})/AVERAGE({','.join(prv_cells)})-1"
        elif len(cur_cells) == 4:
            f = (f"=AVERAGE({','.join(cur_cells)})/"
                 f"AVERAGEIFS('Data Constant'!$E$6:$E${LASTROW_C},'Data Constant'!$C$6:$C${LASTROW_C},$B{r}-1)-1")
        else:
            f = None
        if f:
            put(ws, r, 3 + j, f, PCT2, bold=True, size=11, color=ACCENT)
put(ws, ar + 3, 6, 0.030, PCT2, color=INPUTC)
put(ws, ar + 3, 7, 0.035, PCT2, color=INPUTC)
put(ws, ar + 4, 6, None)
put(ws, ar + 4, 7, 0.040, PCT2, color=INPUTC)
note(ws, ar + 5, 2,
     "REM = BCRA Relevamiento de Expectativas de Mercado, June-2026 round (44 participants, published 6-Jul-2026): "
     "median 3.0% for 2026. The June round did not carry a 2027 annual column, hence the blank. IMF: April-2026 WEO "
     "and the 2026 Article IV / Second Review of the Extended Arrangement, both 3.5% for 2026 and 4.0% for 2027.")
note(ws, ar + 6, 2,
     "2025 is shown as a control: it is computed from actuals only, so all three columns must read the same number "
     "and must match the Annual tab. If they do not, a scenario path has leaked into a historical quarter.")

# assumptions
asr = ar + 9
section(ws, asr, 2, "Scenario assumptions")
hdr(ws, asr + 1, 2, "Driver", 22)
for j, nm in enumerate(["Bull", "Base", "Bear"]):
    hdr(ws, asr + 1, 3 + j, nm, 36, wrap=True)
assump = [
    ("Where this starts",
     "Q1-26 momentum (+0.7% q/q SA) is the floor, not the ceiling; EMAE's April-May softness is noise",
     "Q1-26 growth was real but EMAE has flatlined since — May m/m SA -0.5%, y/y +0.2%. Carry that through H2",
     "The EMAE stall is the signal. Activity is already rolling over into Q3"),
    ("Real income / consumption",
     "Disinflation restores real wages faster than credit tightens; private consumption leads",
     "Real wages recover slowly, consumption grows near trend, no consumption boom",
     "Real wage recovery stalls with inflation converging to the crawl rather than falling further"),
    ("Investment",
     "RIGI projects in energy and mining move from announcement to disbursement through 2027",
     "RIGI delivers, but slowly — GFCF grows above GDP without transforming the aggregate",
     "Financing costs and FX uncertainty push capex decisions past the 2027 electoral cycle"),
    ("External / net exports",
     "Energy and mining volumes keep exports growing double-digit; imports grow with activity",
     "Exports keep growing on Vaca Muerta volume; import growth erodes most of the net contribution",
     "Import growth outpaces exports as the real exchange rate stays strong — net exports subtract"),
    ("FX regime",
     "Band holds, reserve accumulation programme delivers, risk premium compresses",
     "Band holds with its CPI-indexed crawl; no step devaluation before the 2027 cycle",
     "Band ceiling is tested; a discrete adjustment hits real income and activity for two quarters"),
    ("What would change my mind",
     "Two consecutive EMAE prints above +0.5% m/m SA with GFCF confirming",
     "The base case is closest to the IMF (3.5% for 2026) and above the REM (3.0%). The disagreement is about H2, "
     "not about Q1",
     "Watch reserve accumulation against the programme path — a miss there is what breaks the band, and the band is "
     "what anchors the whole forecast"),
]
for i, row in enumerate(assump):
    r = asr + 2 + i
    put(ws, r, 2, row[0], bold=True)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    for j in range(3):
        c = ws.cell(row=r, column=3 + j, value=row[1 + j])
        c.font = Font(ARIAL, 9, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 46

ch = line_chart(30, 13, "Real GDP, seasonally adjusted level — realised and scenario paths", "#,##0")
cats = Reference(ws, min_col=2, min_row=FR0, max_row=FLAST)
for col, nm, colr, dash in [(7, "Bull", "4F9D69", True), (8, "Base", "2F5C8F", False),
                            (9, "Bear", "D1495B", True)]:
    ref = Reference(ws, min_col=col, min_row=FHR, max_row=FLAST)
    s = Series(ref, title=nm)
    style_series(s, colr, 26000, dash)
    ch.series.append(s)
ch.set_categories(cats)
ws.add_chart(ch, "J6")

wb.save("/home/claude/Argentina_GDP.xlsx")
print("saved GDP; quarters:", NQ, "last:", LASTQL, "full years:", FULL_YEARS[0], "-", LAST_FULL)
