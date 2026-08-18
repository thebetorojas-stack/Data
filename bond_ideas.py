#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bond_ideas.py — "give me the best <COUNTRY> bonds" screener
============================================================

Reads the SAME weekly text feeds the EM Bond List is built from, filters the
universe down to one country (or issuer), and ranks what is left twice:

    1. TACTICAL      — short-term plays: cheapness vs the country's own curve,
                       1-year carry + roll-down, house view, and the risk/reward
                       between a 100bp EM rally and a 100bp EM sell-off.
    2. HOLD-TO-MATURITY — buy-and-forget: yield locked in, credit quality,
                       maturity fit, pull-to-par, clean senior structure.

Output is one Excel workbook with the two ranked idea lists, the full scored
universe, the curve/peer working, and a methodology tab you can show a client.


HOW TO RUN (the 30-second version)
----------------------------------
Put this file in the SAME folder as gem_report_builder_v3.py and run:

    python bond_ideas.py                        # Colombia sovereigns, USD
    python bond_ideas.py --country Argentina    # same thing for Argentina
    python bond_ideas.py --country Brazil --universe all --top 15
    python bond_ideas.py --issuer Ecopetrol
    python bond_ideas.py --list-countries       # what's in this week's feed
    python bond_ideas.py --list-issuers --country CO   # exact issuer names

Or, if you prefer pressing F5 in Spyder: edit the CONFIG block below
(TARGET = 'Argentina', etc.) and run with no arguments.

Everything in the CONFIG block is meant to be edited. The scoring weights are
there too — if you think roll-down matters more than the house view, change the
number, re-run, and the workbook's Methodology tab documents whatever you used.


WHAT IT NEEDS
-------------
The weekly feed files, in the usual place:

    data/current/CurrentPublishableBondData.txt
    data/current/CurrentPublishableIssuerData.txt
    data/current/PublishableBondDataUpdate.txt
    data/current/PublishableIssuerDataUpdate.txt
    data/current/PublishableColorFlags.txt
    data/current/IssuerTexts.txt
    data/current/IssuerRatings.txt

(Point somewhere else with --data-dir. The previous-week files and the PRIIPS
reference are optional here — they only affect the week-on-week diff and the
Restrictions column, both of which this script treats as nice-to-have.)


A WORD ON WHAT THIS IS AND ISN'T
--------------------------------
Every number comes from the feed or is derived from it with standard bond
maths. The script does NOT invent prices, spreads or forecasts. Where the feed
has no yield (pure floaters), the bond is set aside rather than scored with a
made-up number — you'll find those on the "Not scored" block of the full tab.

The ranking is a systematic screen, not a recommendation. It is designed to get
you from ~1,900 bonds to a shortlist of ten worth an analyst's attention.
"""

# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIG — edit this block, that's the whole point
# ══════════════════════════════════════════════════════════════════════════════

# ---- What to screen ---------------------------------------------------------
TARGET          = 'Colombia'   # country name or ISO code ('CO'). Overridden by --country
TARGET_ISSUER   = ''           # optional: name fragment, e.g. 'Ecopetrol'. Overridden by --issuer
UNIVERSE        = 'sovereign'  # 'sovereign' | 'quasi' | 'all'
CURRENCIES      = ['USD']      # [] = every currency in the feed
TOP_N           = 10           # ideas per list

# ---- Horizon assumptions ----------------------------------------------------
TACTICAL_MIN_YEARS  = 0.75     # below this a bond is really a cash proxy
HOLD_BAND_YEARS     = (5.0, 15.0)   # the sweet spot for a buy-and-hold client
HOLD_MAX_YEARS      = 30.0     # beyond this, maturity-fit score decays to zero

# ---- Relative-value settings ------------------------------------------------
CURVE_MIN_POINTS    = 4        # need at least this many bonds to fit a curve
CURVE_FALLBACK      = True     # if the target alone is too thin, widen to the
                               # whole country (sovereign + quasi) for the fit
PEER_RATING_BAND    = 1        # +/- notches for the cross-EM peer comparison
PEER_MATURITY_BAND  = 2.0      # +/- years
PEER_MIN_COUNT      = 3        # fewer peers than this -> no peer score
ROLLDOWN_HORIZON    = 1.0      # years of roll-down to credit in the carry calc

# ---- Scenario analysis: what if EM spreads move? ----------------------------
SCENARIO_SHIFT_BP   = 100      # size of the parallel move, applied both ways:
                               #   RALLY   = yields -100bp (spreads compress)
                               #   SELL-OFF= yields +100bp (EM spread widening)
# Beta of each grade to that move. 1.0 / 1.0 means everything moves in parallel,
# which is the only assumption-free choice and is the default. If you want the
# sell-off to hit high yield harder than investment grade — usually what happens
# in a real EM risk-off — set e.g. {'IG': 0.8, 'HY': 1.3}. Whatever you put here
# is printed on the Methodology tab, so the client always sees the assumption.
SELLOFF_BETA        = {'IG': 1.0, 'HY': 1.0}
RALLY_BETA          = {'IG': 1.0, 'HY': 1.0}

# ---- Scoring weights (each list sums to 100; change freely) -----------------
TACTICAL_WEIGHTS = {
    'curve_cheap':     22,   # yield above the country's own fitted curve
    'carry_rolldown':  18,   # 1-yr carry + roll-down return
    'house_view':      18,   # UBS recommendation + Top List membership
    'risk_reward':     12,   # 1-yr upside in a rally vs downside in a sell-off
    'upside_convexity':10,   # price gain alone in the rally case
    'peer_pickup':     10,   # yield vs same-rating/same-maturity EM peers
    'liquidity_access':10,   # denomination size, restriction codes
}
# NOTE: weights do NOT have to sum to 100 — the score divides by whatever is
# present. Set any line to 0 to switch that factor off entirely.
HOLD_WEIGHTS = {
    'yield_level':     30,   # the yield you actually lock in
    'credit_quality':  25,   # worst-of S&P/Moody's
    'maturity_fit':    15,   # sits inside HOLD_BAND_YEARS
    'pull_to_par':     10,   # discount bonds pull up to 100 at maturity
    'structure':       10,   # senior, non-callable, fixed coupon, dated
    'house_view':      10,
}

# ---- Penalties (points subtracted from the weighted score) ------------------
PENALTIES = {
    'sell_rated_tactical':   40,
    'sell_rated_hold':       50,
    'underperform':          10,
    'too_short_tactical':    25,
    'floater':               10,
    'callable_hold':         15,
    'subordinated_hold':     20,
    'perpetual_hold':        30,
    'kid_missing':            8,   # restriction code '2'
}

# ---- Output -----------------------------------------------------------------
OUTPUT_DIR      = 'outputs'
OUTPUT_TEMPLATE = 'Best_{target}_Bonds.xlsx'
QUIET           = False        # True = print nothing but the output path


# ══════════════════════════════════════════════════════════════════════════════
# 2. IMPORTS / PLUMBING
# ══════════════════════════════════════════════════════════════════════════════

import argparse
import contextlib
import io
import math
import os
import statistics
import sys
from datetime import date, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

try:
    import gem_report_builder_v3 as G
except ImportError as e:                                   # pragma: no cover
    sys.exit('Could not import gem_report_builder_v3.py — this script must sit '
             'in the same folder as the weekly builder.\n  (%s)' % e)


def say(*a):
    if not QUIET:
        print(*a)


# ══════════════════════════════════════════════════════════════════════════════
# 3. DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

FEED_FILES = {
    'bond_data':      'CurrentPublishableBondData.txt',
    'issuer_data':    'CurrentPublishableIssuerData.txt',
    'bond_update':    'PublishableBondDataUpdate.txt',
    'issuer_update':  'PublishableIssuerDataUpdate.txt',
    'color_flags':    'PublishableColorFlags.txt',
    'issuer_texts':   'IssuerTexts.txt',
    'issuer_ratings': 'IssuerRatings.txt',
}


def find_data_dir(explicit=None):
    """Locate data/current. Same convention as run_weekly.py."""
    candidates = []
    if explicit:
        candidates.append(explicit)
    candidates += [
        os.path.join(HERE, 'data', 'current'),
        os.path.join(HERE, 'data'),
        HERE,
    ]
    for c in candidates:
        if c and os.path.isfile(os.path.join(c, FEED_FILES['bond_data'])):
            return c
    sys.exit(
        'Could not find the weekly feed files.\n'
        'Looked for %s in:\n    %s\n'
        'Point at the right folder with --data-dir "C:\\path\\to\\current".'
        % (FEED_FILES['bond_data'], '\n    '.join(c for c in candidates if c)))


def find_optional(patterns, folders):
    """Newest file whose name contains any of `patterns` (case-insensitive)."""
    hits = []
    for folder in folders:
        if not folder or not os.path.isdir(folder):
            continue
        for f in os.listdir(folder):
            lf = f.lower()
            if f.startswith('~$'):
                continue
            if any(p in lf for p in patterns) and lf.endswith(
                    ('.csv', '.txt', '.xls', '.xlsx')):
                hits.append(os.path.join(folder, f))
    return max(hits, key=os.path.getmtime) if hits else None


def load_data(data_dir):
    """Build the GEMData object — the exact same business logic the weekly
    PDF/Excel outputs use, so nothing here can disagree with the published
    list about what a bond is or whether it belongs on it."""
    paths = {k: os.path.join(data_dir, v) for k, v in FEED_FILES.items()}
    missing = [v for k, v in FEED_FILES.items() if not os.path.isfile(paths[k])]
    if missing:
        sys.exit('Missing feed file(s) in %s:\n    %s' % (data_dir, '\n    '.join(missing)))

    parent = os.path.dirname(data_dir.rstrip(os.sep))
    search = [data_dir, parent, HERE]
    paths['priips_ref'] = find_optional(['priips'], search)
    legal = find_optional(['legal', 'exclusion'], search)
    if legal and any(s in os.path.basename(legal).lower()
                     for s in ('template', 'example', 'sample')):
        legal = None
    paths['legal_exclusions'] = legal
    paths['prev_bond_data'] = None
    paths['prev_bond_update'] = None
    return G.GEMData(paths)


# ══════════════════════════════════════════════════════════════════════════════
# 4. SMALL NUMERIC HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def num(raw):
    """Float or None — never raises, never returns a misleading zero."""
    if raw is None:
        return None
    s = str(raw).strip().replace(',', '')
    if not s or s.lower() in ('n/a', 'na', 'n.a.', '-', 'nr'):
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return None if math.isnan(v) or math.isinf(v) else v


def years_to(d, ref=None):
    if d is None:
        return None
    ref = ref or date.today()
    if isinstance(d, datetime):
        d = d.date()
    return (d - ref).days / 365.25


def pct_rank(value, population):
    """Percentile rank 0-100 of `value` within `population` (ties averaged).
    Used so every scoring component lands on the same 0-100 scale regardless
    of its natural units (bp, years, %, price points)."""
    vals = [v for v in population if v is not None]
    if value is None or not vals:
        return None
    if len(vals) == 1:
        return 50.0
    below = sum(1 for v in vals if v < value)
    equal = sum(1 for v in vals if v == value)
    return 100.0 * (below + 0.5 * equal) / len(vals)


def clip(v, lo=0.0, hi=100.0):
    return max(lo, min(hi, v))


# ─── bond maths ──────────────────────────────────────────────────────────────

def duration_convexity(price, coupon_pct, years, ytm_pct, freq=2):
    """Modified duration (years) and convexity from clean price, coupon,
    time to maturity and yield. Standard textbook discounting at `freq` periods
    per year — no library needed, and it reconciles to any calculator.

    Returns (mod_duration, convexity, price_check) or (None, None, None) when
    the inputs can't support a calculation.
    """
    if ytm_pct is None or years is None or years <= 0:
        return None, None, None
    if coupon_pct is None:
        coupon_pct = 0.0
    y = ytm_pct / 100.0
    if y <= -0.99:
        return None, None, None
    n = max(1, int(round(years * freq)))
    c = coupon_pct / 100.0 / freq          # cash coupon per period, per 1 face
    r = y / freq                           # periodic yield
    pv_sum = 0.0
    t_pv = 0.0
    tt_pv = 0.0
    for i in range(1, n + 1):
        cf = c + (1.0 if i == n else 0.0)
        df = (1.0 + r) ** (-i)
        pv = cf * df
        pv_sum += pv
        t_pv += i * pv
        tt_pv += i * (i + 1) * pv
    if pv_sum <= 0:
        return None, None, None
    macaulay = (t_pv / pv_sum) / freq
    mod_dur = macaulay / (1.0 + r)
    convexity = (tt_pv / pv_sum) / ((1.0 + r) ** 2) / (freq ** 2)
    return mod_dur, convexity, pv_sum * 100.0


def perp_duration(ytm_pct):
    """A perpetual's modified duration collapses to ~1/y."""
    if not ytm_pct or ytm_pct <= 0:
        return None
    return min(25.0, 100.0 / ytm_pct)


def price_gain_pct(mod_dur, convexity, shift_bp=-100):
    """Approximate % price change for a parallel yield shift."""
    if mod_dur is None:
        return None
    dy = shift_bp / 10000.0
    g = -mod_dur * dy
    if convexity:
        g += 0.5 * convexity * dy * dy
    return g * 100.0


# ─── least squares (pure python, no numpy required) ──────────────────────────

def _solve(A, b):
    """Gaussian elimination with partial pivoting. Returns None if singular."""
    n = len(A)
    M = [row[:] + [b[i]] for i, row in enumerate(A)]
    for col in range(n):
        piv = max(range(col, n), key=lambda r: abs(M[r][col]))
        if abs(M[piv][col]) < 1e-12:
            return None
        M[col], M[piv] = M[piv], M[col]
        for r in range(col + 1, n):
            f = M[r][col] / M[col][col]
            for k in range(col, n + 1):
                M[r][k] -= f * M[col][k]
    x = [0.0] * n
    for r in range(n - 1, -1, -1):
        s = M[r][n] - sum(M[r][k] * x[k] for k in range(r + 1, n))
        x[r] = s / M[r][r]
    return x


def _basis(t, k):
    """Curve shape functions. Two forms so a thin curve doesn't over-fit:
       k=2 -> y = a + b*ln(1+t)          (needs >= 3 bonds)
       k=3 -> y = a + b*ln(1+t) + c*t    (needs >= 5 bonds)"""
    if k == 2:
        return [1.0, math.log(1.0 + max(t, 0.0))]
    return [1.0, math.log(1.0 + max(t, 0.0)), t]


def fit_curve(points, min_points=CURVE_MIN_POINTS):
    """Least-squares yield curve through (years, yield%) pairs.

    Fits, then drops any bond more than 2.5 standard deviations from the fit
    and re-fits — one distressed or mis-priced line shouldn't drag the curve
    everything else is measured against.

    Returns dict(coef, k, n, rmse_bp, dropped) or None.
    """
    pts = [(t, y) for t, y in points if t is not None and y is not None and t > 0]
    if len(pts) < max(3, min_points):
        return None

    def _fit(sample, k):
        A = [[0.0] * k for _ in range(k)]
        b = [0.0] * k
        for t, y in sample:
            x = _basis(t, k)
            for i in range(k):
                b[i] += x[i] * y
                for j in range(k):
                    A[i][j] += x[i] * x[j]
        return _solve(A, b)

    k = 3 if len(pts) >= 5 else 2
    coef = _fit(pts, k)
    if coef is None:
        return None
    resid = [y - sum(c * x for c, x in zip(coef, _basis(t, k))) for t, y in pts]
    dropped = []
    if len(pts) >= max(6, min_points + 2):
        sd = statistics.pstdev(resid) or 0.0
        if sd > 0:
            keep = [(p, r) for p, r in zip(pts, resid) if abs(r) <= 2.5 * sd]
            dropped = [p for p, r in zip(pts, resid) if abs(r) > 2.5 * sd]
            if len(keep) >= max(3, min_points):
                pts2 = [p for p, _ in keep]
                k = 3 if len(pts2) >= 5 else 2
                coef2 = _fit(pts2, k)
                if coef2:
                    coef = coef2
                    resid = [y - sum(c * x for c, x in zip(coef, _basis(t, k)))
                             for t, y in pts2]
                    pts = pts2
    rmse = math.sqrt(sum(r * r for r in resid) / len(resid)) * 100.0  # in bp
    return {'coef': coef, 'k': k, 'n': len(pts), 'rmse_bp': rmse,
            'dropped': dropped}


def curve_yield(fit, t):
    if not fit or t is None or t <= 0:
        return None
    return sum(c * x for c, x in zip(fit['coef'], _basis(t, fit['k'])))


# ══════════════════════════════════════════════════════════════════════════════
# 5. TARGET RESOLUTION (country / issuer)
# ══════════════════════════════════════════════════════════════════════════════

_NAME_TO_CODE = {}
for _code, _name in G.COUNTRY_NAMES.items():
    _NAME_TO_CODE[_name.lower()] = _code
# a few things people actually type
_NAME_TO_CODE.update({
    'uae': 'AE', 'united arab emirates': 'AE', 'korea': 'KR',
    'south korea': 'KR', 'russia': 'RU', 'turkey': 'TR', 'turkiye': 'TR',
    'saudi': 'SA', 'saudi arabia': 'SA', 'china': 'CN', 'hong kong': 'HK',
    'dominican republic': 'DO', 'ivory coast': 'CI', "cote d'ivoire": 'CI',
})


def resolve_country(text):
    """'Colombia' / 'colombia' / 'CO' -> ('CO', 'Colombia'). None if unknown."""
    s = (text or '').strip()
    if not s:
        return None, None
    if len(s) == 2 and s.upper() in G.COUNTRY_NAMES:
        return s.upper(), G.COUNTRY_NAMES[s.upper()]
    code = _NAME_TO_CODE.get(s.lower())
    if code:
        return code, G.COUNTRY_NAMES.get(code, s)
    # last resort: unique prefix match on the country-name list
    hits = [(c, n) for n, c in _NAME_TO_CODE.items() if n.startswith(s.lower())]
    if len(set(c for c, _ in hits)) == 1:
        code = hits[0][0]
        return code, G.COUNTRY_NAMES.get(code, s)
    return None, None


TIER_TO_SP = {}
for _sym, _t in G.RATING_SCALE.items():
    if _sym[0] in 'ABCD' and _sym not in ('Aaa',) and _sym.upper() == _sym:
        TIER_TO_SP.setdefault(_t, _sym)
TIER_TO_SP[0] = 'AAA'


def tier_label(t):
    if t is None:
        return 'n/a'
    lo, hi = int(math.floor(t)), int(math.ceil(t))
    a, b = TIER_TO_SP.get(lo, '?'), TIER_TO_SP.get(hi, '?')
    return a if a == b else '%s/%s' % (a, b)


ITYPE_LABEL = {'SOV': 'Sovereign', 'SUPRA': 'Supranational',
               'FIN': 'Financial', 'CORP': 'Corporate'}


def matches_target(rec, cfg):
    """Does this enriched bond belong to what we were asked to screen?"""
    if cfg['issuer_fragment']:
        return cfg['issuer_fragment'] in rec['issuer_raw'].lower()
    if rec['country'] != cfg['country_code']:
        # A sovereign is occasionally booked under an offshore country code;
        # its published name still carries the country, so check that too.
        if not (rec['itype'] in ('SOV', 'SUPRA')
                and cfg['country_name'].lower() in rec['issuer_raw'].lower()):
            return False
    if cfg['universe'] == 'sovereign' and rec['itype'] not in ('SOV', 'SUPRA'):
        return False
    if cfg['universe'] == 'quasi' and rec['itype'] in ('SOV', 'SUPRA'):
        return False
    return True


# ══════════════════════════════════════════════════════════════════════════════
# 6. ENRICHMENT — one flat, fully numeric record per bond
# ══════════════════════════════════════════════════════════════════════════════

REC_LABEL = {'OP': 'Attractive (OP)', 'MP': 'Fair (MP)', '': 'Fair',
             'UP': 'Expensive (UP)', 'SELL': 'Sell'}
REC_SCORE = {'OP': 100.0, 'MP': 55.0, '': 55.0, 'UP': 15.0, 'SELL': 0.0}


def enrich(b, data):
    """Turn one raw feed row into a scored-ready record. Display fields come
    from GEMData.bond_row so names/ratings/regions match the published list
    exactly; the numeric fields are parsed from the same raw feed values."""
    try:
        row = data.bond_row(b)
    except Exception:
        return None
    isin = row['isin']
    upd = data.bond_updates.get(isin, {})
    gk = row['gk']

    mat_dt = row['maturity_date']
    yrs = years_to(mat_dt)
    is_perp = mat_dt is None

    price = num(b.get('PXASK_ExecDesk'))
    coupon = num(b.get('Coupon'))
    ytm = num(b.get('YLDASK_ExecDesk'))

    cpn_type = (b.get('CpnType') or '').strip().lower()
    fo_type = (b.get('FOType') or '').strip().lower()
    is_floater = cpn_type in ('variable', 'fixed/variable') or 'float' in fo_type
    if is_floater and (ytm is None or abs(ytm) < 1e-6):
        ytm = None                        # the feed has no real yield for these

    if is_perp:
        mod_dur = perp_duration(ytm)
        convex = (2.0 / ((ytm / 100.0) ** 2) if ytm else None)
        if convex:
            convex = min(convex, 900.0)
    else:
        mod_dur, convex, _ = duration_convexity(price, coupon, yrs, ytm)

    eff = data.effective_issuer_rating(gk, b, upd)
    tier = eff.get('worst_tier')

    rec_code = (upd.get('WMR_Bond_Recommendation')
                or b.get('WMR_Bond_Recommendation') or '').strip().upper()

    min_amt = num(b.get('MinAmt'))
    restr = row['restrictions'] or ''

    callable_ = ((b.get('redeemable') or '').strip().upper() == 'Y'
                 or 'call' in fo_type)

    return {
        # identity
        'isin': isin,
        'valor': row['valor'],
        'gk': gk,
        'issuer': row['issuer'],
        'issuer_raw': row['issuer_raw'],
        'country': row['country'],
        'country_display': row['country_display'],
        'itype': row['itype'],
        'itype_label': ITYPE_LABEL.get(row['itype'], row['itype'] or 'n/a'),
        'region': row['region'],
        'ccy': row['ccy'],
        # terms
        'coupon': coupon,
        'maturity_dt': mat_dt,
        'maturity': row['maturity'],
        'years': yrs,
        'is_perp': is_perp,
        'price': price,
        'ytm': ytm,
        'mod_dur': mod_dur,
        'convexity': convex,
        'is_floater': is_floater,
        'is_callable': bool(callable_),
        'is_sub': G.is_subordinated_bond(b),
        'cpn_type': cpn_type or 'fixed',
        # credit
        'sp': eff.get('sp_token') or 'n/a',
        'mdy': eff.get('mdy_token') or 'n/a',
        'ratings': row['ratings'],
        'tier': tier,
        'grade': 'IG' if row['grade'].startswith('Investment') else 'HY',
        # house view
        'rec': rec_code,
        'rec_label': REC_LABEL.get(rec_code, rec_code or 'Fair'),
        'top_list': (b.get('Product_Use') or '').strip() == '7',
        'comment': row['comment'],
        # practicalities
        'min_amt': min_amt,
        'min_denom': row['min_denom'],
        'restrictions': restr,
        'green': row['green'],
    }


# ══════════════════════════════════════════════════════════════════════════════
# 7. RELATIVE VALUE — own curve, roll-down, cross-EM peers
# ══════════════════════════════════════════════════════════════════════════════

def attach_curve(cands, curve_pool, cfg):
    """Fit a yield curve to `curve_pool` and measure every candidate against it.

    curve_cheap_bp : how much MORE yield the bond offers than the curve says a
                     bond of that maturity should — positive = cheap.
    rolldown_bp    : how far the bond slides DOWN the curve over the next year
                     if the curve itself doesn't move.
    carry_roll_pct : 1-year total return estimate = coupon carry + roll-down
                     price gain. The classic tactical metric.
    """
    pool = [(r['years'], r['ytm']) for r in curve_pool
            if r['years'] and r['ytm'] is not None and not r['is_perp']]
    fit = fit_curve(pool, cfg['curve_min_points'])
    for r in cands:
        r['curve_fitted'] = None
        r['curve_cheap_bp'] = None
        r['rolldown_bp'] = None
        r['carry_roll_pct'] = None
        if not fit or r['ytm'] is None or not r['years'] or r['years'] <= 0:
            continue
        fy = curve_yield(fit, r['years'])
        r['curve_fitted'] = fy
        if fy is not None:
            r['curve_cheap_bp'] = (r['ytm'] - fy) * 100.0
        t2 = r['years'] - cfg['rolldown_horizon']
        if fy is not None and t2 > 0.25:
            fy2 = curve_yield(fit, t2)
            if fy2 is not None:
                r['rolldown_bp'] = (fy - fy2) * 100.0
                carry = r['ytm'] * (cfg['rolldown_horizon'])
                roll_gain = (r['mod_dur'] or 0.0) * (r['rolldown_bp'] / 100.0)
                r['carry_roll_pct'] = carry + roll_gain
        elif fy is not None:
            # inside a year of maturity there is no roll-down, only carry
            r['rolldown_bp'] = 0.0
            r['carry_roll_pct'] = r['ytm'] * cfg['rolldown_horizon']
    return fit


def attach_peers(cands, universe, cfg):
    """For each candidate, find comparable bonds ACROSS THE WHOLE EM UNIVERSE
    (same currency, rating within N notches, maturity within M years, different
    issuer) and report the yield pick-up vs the peer median.

    This is the sanity check on the country curve: a bond can look rich against
    its own sovereign and still be the best-paid BB+ 10-year in EM."""
    pool = [u for u in universe
            if u['ytm'] is not None and u['years'] and u['tier'] is not None]
    for r in cands:
        r['peer_n'] = 0
        r['peer_median_ytm'] = None
        r['peer_pickup_bp'] = None
        r['peer_examples'] = ''
        if r['ytm'] is None or r['tier'] is None or not r['years']:
            continue
        peers = [u for u in pool
                 if u['ccy'] == r['ccy']
                 and u['gk'] != r['gk']
                 and abs(u['tier'] - r['tier']) <= cfg['peer_rating_band']
                 and abs(u['years'] - r['years']) <= cfg['peer_maturity_band']]
        r['peer_n'] = len(peers)
        if len(peers) < cfg['peer_min_count']:
            continue
        med = statistics.median(u['ytm'] for u in peers)
        r['peer_median_ytm'] = med
        r['peer_pickup_bp'] = (r['ytm'] - med) * 100.0
        near = sorted(peers, key=lambda u: abs(u['years'] - r['years']))[:3]
        r['peer_examples'] = '; '.join(
            '%s %s %.1fy %.2f%%' % (u['issuer_raw'][:28], u['ccy'],
                                    u['years'], u['ytm']) for u in near)
    return


def country_context(universe, cfg):
    """Fit a curve per country (same currency filter as the screen) and read
    each one off at 3 / 5 / 10 years. Answers the question a client always asks
    next: "fine, but is Colombia paying me enough versus the rest of EM?"""
    ccys = cfg['currencies']
    by_country = {}
    for r in universe:
        if ccys and r['ccy'] not in ccys:
            continue
        if not r['country'] or r['ytm'] is None or not r['years'] or r['is_perp']:
            continue
        by_country.setdefault((r['country'], r['country_display']), []).append(r)
    out = []
    for (code, name), rows in by_country.items():
        sov = [x for x in rows if x['itype'] in ('SOV', 'SUPRA')]
        pool = sov if len(sov) >= cfg['curve_min_points'] else rows
        fit = fit_curve([(x['years'], x['ytm']) for x in pool],
                        cfg['curve_min_points'])
        tiers = [x['tier'] for x in pool if x['tier'] is not None]
        out.append({
            'code': code, 'name': name,
            'n': len(pool), 'basis': 'sovereign' if pool is sov else 'all issuers',
            'y3': curve_yield(fit, 3.0), 'y5': curve_yield(fit, 5.0),
            'y10': curve_yield(fit, 10.0),
            'median_tier': tier_label(statistics.median(tiers) if tiers else None),
        })
    out.sort(key=lambda d: (d['y10'] is None, -(d['y10'] or 0)))
    return out


RR_NO_LOSS = 'no loss'        # sentinel shown when a 1-yr sell-off still pays
RR_RANK_SENTINEL = 20.0       # ratio used for RANKING those bonds (well above
                              # anything a real bond produces, so they sort top)


def attach_scenarios(cands, cfg):
    """Two symmetric scenarios, and the risk/reward that falls out of them.

        RALLY    yields -{shift}bp  (EM spreads compress)
        SELL-OFF yields +{shift}bp  (EM spread widening)

    For each we report the pure PRICE move, and — where the curve gave us a
    carry + roll-down number — the 1-YEAR TOTAL RETURN, which is the honest
    version: in a sell-off the coupon you collect over the year offsets part of
    the mark-to-market hit, and a fat-carry short bond can end the year up even
    though its price fell.

    risk_reward = 1-yr upside in the rally / size of the 1-yr downside in the
    sell-off. Above 1.0 means the trade pays you more when you're right than it
    costs you when you're wrong. Where carry more than covers a {shift}bp
    widening the ratio is undefined and reported as '{noloss}'.

    breakeven_bp = how far yields can widen over the next year before the
    position is flat — carry + roll-down divided by duration. The single most
    useful number for a tactical EM trade.
    """
    shift = cfg['scenario_shift_bp']
    for r in cands:
        grade = r['grade']
        up_bp = shift * cfg['selloff_beta'].get(grade, 1.0)
        dn_bp = -shift * cfg['rally_beta'].get(grade, 1.0)
        r['scenario_rally_bp'] = dn_bp
        r['scenario_selloff_bp'] = up_bp
        r['rally_gain_pct'] = price_gain_pct(r['mod_dur'], r['convexity'], dn_bp)
        r['selloff_loss_pct'] = price_gain_pct(r['mod_dur'], r['convexity'], up_bp)

        carry = r.get('carry_roll_pct')
        if carry is not None and r['rally_gain_pct'] is not None:
            r['ret_rally_1y'] = carry + r['rally_gain_pct']
            r['ret_selloff_1y'] = carry + r['selloff_loss_pct']
            up, dn = r['ret_rally_1y'], r['ret_selloff_1y']
            r['rr_basis'] = '1y total return (incl. carry)'
        else:
            r['ret_rally_1y'] = None
            r['ret_selloff_1y'] = None
            up, dn = r['rally_gain_pct'], r['selloff_loss_pct']
            r['rr_basis'] = 'price only (no carry available)'

        if up is None or dn is None:
            r['risk_reward'] = None
            r['risk_reward_rank'] = None
        elif dn >= 0:
            # A year of carry more than absorbs the whole widening: there is no
            # downside to divide by. Report the cushion instead, and rank these
            # bonds above every finite ratio, ordered by how big the cushion is.
            r['risk_reward'] = '%s (%+.1f%%)' % (RR_NO_LOSS, dn)
            r['risk_reward_rank'] = RR_RANK_SENTINEL + dn
        else:
            r['risk_reward'] = up / abs(dn)
            r['risk_reward_rank'] = r['risk_reward']

        if carry is not None and r['mod_dur']:
            r['breakeven_bp'] = carry / r['mod_dur'] * 100.0
        else:
            r['breakeven_bp'] = None


# ══════════════════════════════════════════════════════════════════════════════
# 8. SCORING
# ══════════════════════════════════════════════════════════════════════════════

def _house_view_score(r):
    s = REC_SCORE.get(r['rec'], 55.0)
    if r['top_list'] and r['rec'] != 'SELL':
        s = min(100.0, s + 15.0)
    return s


def _liquidity_score(r):
    a = r['min_amt']
    if a is None:
        s = 55.0
    elif a <= 1000:
        s = 100.0
    elif a <= 10000:
        s = 88.0
    elif a <= 100000:
        s = 65.0
    elif a <= 200000:
        s = 45.0
    else:
        s = 30.0
    if '2' in r['restrictions']:
        s -= 15.0
    if '1' in r['restrictions']:
        s -= 8.0
    return clip(s)


def _credit_score(r):
    """AAA=100 down to CCC-/D=0, linear on the shared rating tier."""
    if r['tier'] is None:
        return 25.0                       # unrated is not the same as bad, but
                                          # it isn't a hold-to-maturity comfort
    return clip(100.0 - r['tier'] * 5.2)


def _maturity_fit_score(r, cfg):
    lo, hi = cfg['hold_band']
    t = r['years']
    if r['is_perp'] or t is None:
        return 0.0                        # you can't hold a perp to maturity
    if lo <= t <= hi:
        return 100.0
    if t < lo:
        return clip(100.0 * (t / lo) ** 1.5)
    span = max(cfg['hold_max_years'] - hi, 1.0)
    return clip(100.0 * (1.0 - (t - hi) / span) ** 1.2)


def _structure_score(r):
    s = 100.0
    if r['is_sub']:
        s -= 30.0
    if r['is_callable']:
        s -= 20.0
    if r['is_floater']:
        s -= 20.0
    if r['is_perp']:
        s -= 35.0
    if r['cpn_type'] not in ('fixed', ''):
        s -= 5.0
    return clip(s)


def score_all(cands, cfg):
    """Two weighted scores per bond, with every component kept on the record so
    the workbook can show exactly where a score came from."""
    scored = [r for r in cands if r['ytm'] is not None]
    if not scored:
        return scored, []

    pop = {
        'curve_cheap':      [r.get('curve_cheap_bp') for r in scored],
        'carry_rolldown':   [r.get('carry_roll_pct') for r in scored],
        'peer_pickup':      [r.get('peer_pickup_bp') for r in scored],
        'upside_convexity': [r.get('rally_gain_pct') for r in scored],
        'risk_reward':      [r.get('risk_reward_rank') for r in scored],
        'yield_level':      [r['ytm'] for r in scored],
        'pull_to_par':      [(100.0 - r['price']) if r['price'] else None
                             for r in scored],
    }

    for r in scored:
        c = {}
        # market-relative components -> percentile within the screened set
        c['curve_cheap'] = pct_rank(r.get('curve_cheap_bp'), pop['curve_cheap'])
        c['carry_rolldown'] = pct_rank(r.get('carry_roll_pct'), pop['carry_rolldown'])
        c['peer_pickup'] = pct_rank(r.get('peer_pickup_bp'), pop['peer_pickup'])
        c['upside_convexity'] = pct_rank(r.get('rally_gain_pct'),
                                         pop['upside_convexity'])
        c['risk_reward'] = pct_rank(r.get('risk_reward_rank'),
                                    pop['risk_reward'])
        c['yield_level'] = pct_rank(r['ytm'], pop['yield_level'])
        c['pull_to_par'] = pct_rank((100.0 - r['price']) if r['price'] else None,
                                    pop['pull_to_par'])
        # absolute components
        c['house_view'] = _house_view_score(r)
        c['liquidity_access'] = _liquidity_score(r)
        c['credit_quality'] = _credit_score(r)
        c['maturity_fit'] = _maturity_fit_score(r, cfg)
        c['structure'] = _structure_score(r)
        r['components'] = c

        # ---- tactical ------------------------------------------------------
        t_score, t_wsum, t_parts = 0.0, 0.0, []
        for key, w in cfg['tactical_weights'].items():
            v = c.get(key)
            if v is None:
                continue                  # missing input -> weight redistributed
            t_score += w * v
            t_wsum += w
            t_parts.append((key, w * v / 100.0))
        t_score = (t_score / t_wsum) if t_wsum else 0.0
        t_pen, t_why = [], []
        if r['rec'] == 'SELL':
            t_pen.append(('house view is Sell', cfg['penalties']['sell_rated_tactical']))
        elif r['rec'] == 'UP':
            t_pen.append(('house view is Expensive', cfg['penalties']['underperform']))
        if r['years'] is not None and r['years'] < cfg['tactical_min_years']:
            t_pen.append(('under %.2gy to maturity' % cfg['tactical_min_years'],
                          cfg['penalties']['too_short_tactical']))
        if r['is_floater']:
            t_pen.append(('floating coupon', cfg['penalties']['floater']))
        if '2' in r['restrictions']:
            t_pen.append(('no PRIIPs KID', cfg['penalties']['kid_missing']))
        r['tactical_score'] = clip(t_score - sum(p for _, p in t_pen))
        r['tactical_parts'] = sorted(t_parts, key=lambda kv: -kv[1])
        r['tactical_penalties'] = t_pen
        r['tactical_coverage'] = t_wsum

        # ---- hold to maturity ---------------------------------------------
        h_score, h_wsum, h_parts = 0.0, 0.0, []
        for key, w in cfg['hold_weights'].items():
            v = c.get(key)
            if v is None:
                continue
            h_score += w * v
            h_wsum += w
            h_parts.append((key, w * v / 100.0))
        h_score = (h_score / h_wsum) if h_wsum else 0.0
        h_pen = []
        if r['rec'] == 'SELL':
            h_pen.append(('house view is Sell', cfg['penalties']['sell_rated_hold']))
        elif r['rec'] == 'UP':
            h_pen.append(('house view is Expensive', cfg['penalties']['underperform']))
        if r['is_callable']:
            h_pen.append(('callable — yield may not be locked in',
                          cfg['penalties']['callable_hold']))
        if r['is_sub']:
            h_pen.append(('subordinated', cfg['penalties']['subordinated_hold']))
        if r['is_perp']:
            h_pen.append(('perpetual — no maturity date',
                          cfg['penalties']['perpetual_hold']))
        if r['is_floater']:
            h_pen.append(('floating coupon — income not fixed',
                          cfg['penalties']['floater']))
        if '2' in r['restrictions']:
            h_pen.append(('no PRIIPs KID', cfg['penalties']['kid_missing']))
        r['hold_score'] = clip(h_score - sum(p for _, p in h_pen))
        r['hold_parts'] = sorted(h_parts, key=lambda kv: -kv[1])
        r['hold_penalties'] = h_pen
        r['hold_coverage'] = h_wsum

        # ---- plain-English income figure for the HTM client ---------------
        if r['ytm'] is not None and r['price']:
            r['income_per_100k'] = 100000.0 * (r['coupon'] or 0.0) / 100.0
            r['cost_per_100k_nominal'] = 1000.0 * r['price']
        else:
            r['income_per_100k'] = None
            r['cost_per_100k_nominal'] = None

    unscored = [r for r in cands if r['ytm'] is None]
    return scored, unscored


# ══════════════════════════════════════════════════════════════════════════════
# 9. RATIONALE — why this bond, in words you can paste into an email
# ══════════════════════════════════════════════════════════════════════════════

PART_PHRASE = {
    'curve_cheap':      'cheap to its own curve',
    'carry_rolldown':   'best carry + roll-down',
    'house_view':       'house view',
    'upside_convexity': 'most price upside in a rally',
    'risk_reward':      'best risk/reward',
    'peer_pickup':      'pays up vs EM peers',
    'liquidity_access': 'easy to buy',
    'yield_level':      'high yield locked in',
    'credit_quality':   'strong credit',
    'maturity_fit':     'maturity fits the horizon',
    'pull_to_par':      'discount price pulls to par',
    'structure':        'clean senior structure',
}


def tactical_rationale(r):
    bits = []
    if r.get('curve_cheap_bp') is not None:
        if r['curve_cheap_bp'] >= 5:
            bits.append('trades %.0fbp cheap to %s\u2019s own fitted curve'
                        % (r['curve_cheap_bp'], r['country_display']))
        elif r['curve_cheap_bp'] <= -5:
            bits.append('%.0fbp rich to the fitted curve'
                        % abs(r['curve_cheap_bp']))
        else:
            bits.append('sits on the fitted curve')
    if r.get('carry_roll_pct') is not None:
        bits.append('~%.1f%% 1-yr carry + roll-down' % r['carry_roll_pct'])
    if r.get('rally_gain_pct') is not None and r.get('mod_dur'):
        shift = abs(r.get('scenario_selloff_bp') or 100)
        bits.append('%.1fy duration \u2192 %+.1f%% if EM rallies %.0fbp vs '
                    '%+.1f%% if it sells off %.0fbp'
                    % (r['mod_dur'], r['rally_gain_pct'], shift,
                       r['selloff_loss_pct'], shift))
    if r.get('ret_selloff_1y') is not None:
        bits.append('over a year, carry turns that into %+.1f%% / %+.1f%%'
                    % (r['ret_rally_1y'], r['ret_selloff_1y']))
    rr = r.get('risk_reward')
    if isinstance(rr, float):
        bits.append('risk/reward %.1fx' % rr)
    elif isinstance(rr, str) and rr.startswith(RR_NO_LOSS):
        bits.append('carry covers the whole widening \u2014 no 1-yr loss even in '
                    'the sell-off')
    if r.get('breakeven_bp'):
        bits.append('breakeven widening %.0fbp' % r['breakeven_bp'])
    if r.get('peer_pickup_bp') is not None and abs(r['peer_pickup_bp']) >= 10:
        bits.append('%+.0fbp vs %d same-rating EM peers'
                    % (r['peer_pickup_bp'], r['peer_n']))
    if r['top_list']:
        bits.append('on the Top List')
    elif r['rec'] == 'OP':
        bits.append('rated Attractive')
    elif r['rec'] in ('UP', 'SELL'):
        bits.append('house view is %s' % r['rec_label'])
    if r['tactical_penalties']:
        bits.append('watch: ' + ', '.join(p for p, _ in r['tactical_penalties']))
    return _sentence('; '.join(bits)) + '.'


def hold_rationale(r):
    bits = []
    if r['ytm'] is not None:
        bits.append('locks in %.2f%% to %s' % (
            r['ytm'], 'perpetuity' if r['is_perp'] else r['maturity']))
    if r['tier'] is not None:
        bits.append('%s (%s)' % (r['ratings'], r['grade']))
    if r['price']:
        if r['price'] < 99:
            bits.append('bought at %.1f, pulls to 100 at maturity' % r['price'])
        elif r['price'] > 101:
            bits.append('premium price %.1f — coupon is above market' % r['price'])
    if r.get('income_per_100k'):
        bits.append('~USD %s coupon income a year per 100k nominal'
                    % ('{:,.0f}'.format(r['income_per_100k'])))
    if not r['is_callable'] and not r['is_sub'] and not r['is_perp']:
        bits.append('senior, non-callable, fixed coupon')
    if r['min_amt']:
        bits.append('min %s' % '{:,.0f}'.format(r['min_amt']))
    if r['hold_penalties']:
        bits.append('watch: ' + ', '.join(p for p, _ in r['hold_penalties']))
    return _sentence('; '.join(bits)) + '.'


def _sentence(s):
    """Upper-case the first letter only — .capitalize() would flatten ISINs,
    ratings and country names to lower case."""
    s = s.strip()
    return (s[:1].upper() + s[1:]) if s else s


def top_drivers(parts, n=3):
    return ', '.join(PART_PHRASE.get(k, k) for k, _ in parts[:n])


# ══════════════════════════════════════════════════════════════════════════════
# 10. EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL = '305496'
BAND_FILL = 'DDEBF7'


def _style_header(ws, ncols, row=1):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=HDR_FILL)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


def _autosize(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


IDEA_COLS = [
    ('#',            6,  None),
    ('Score',        8,  '0.0'),
    ('ISIN',        15,  None),
    ('Issuer',      34,  None),
    ('Type',        12,  None),
    ('Ccy',          6,  None),
    ('Coupon %',     9,  '0.000'),
    ('Maturity',    11,  None),
    ('Yrs',          6,  '0.0'),
    ('Offer px',     9,  '0.00'),
    ('YTM %',        8,  '0.00'),
    ('Mod dur',      8,  '0.0'),
    ('Ratings',     14,  None),
    ('IG/HY',        7,  None),
    ('House view',  16,  None),
    ('Top List',     9,  None),
    ('vs own curve (bp)', 12, '+0;-0'),
    ('vs EM peers (bp)', 12, '+0;-0'),
    ('1y carry+roll %',  12, '0.00'),
    ('RALLY -100bp\nprice %',      12, '+0.00;-0.00'),
    ('SELL-OFF +100bp\nprice %',   12, '+0.00;-0.00'),
    ('RALLY -100bp\n1y total %',   12, '+0.00;-0.00'),
    ('SELL-OFF +100bp\n1y total %',12, '+0.00;-0.00'),
    ('Risk/reward\n(up : down)',   12, '0.00"x"'),
    ('Breakeven\nwidening (bp)',   12, '0'),
    ('Min denom',   14,  None),
    ('Restrictions', 11, None),
    ('Key drivers', 34,  None),
    ('Why this bond', 78, None),
]


def _idea_row(i, r, score_key, rationale_fn, parts_key):
    return [
        i,
        r[score_key],
        r['isin'],
        r['issuer'],
        r['itype_label'],
        r['ccy'],
        r['coupon'],
        'Perpetual' if r['is_perp'] else r['maturity'],
        r['years'],
        r['price'],
        r['ytm'],
        r['mod_dur'],
        r['ratings'],
        r['grade'],
        r['rec_label'],
        'Yes' if r['top_list'] else '',
        r.get('curve_cheap_bp'),
        r.get('peer_pickup_bp'),
        r.get('carry_roll_pct'),
        r.get('rally_gain_pct'),
        r.get('selloff_loss_pct'),
        r.get('ret_rally_1y'),
        r.get('ret_selloff_1y'),
        r.get('risk_reward'),
        r.get('breakeven_bp'),
        r['min_denom'],
        r['restrictions'] or '',
        top_drivers(r[parts_key]),
        rationale_fn(r),
    ]


def _idea_cols(cfg):
    """Column spec with the scenario size taken from CONFIG, so changing
    SCENARIO_SHIFT_BP relabels the workbook automatically."""
    bp = cfg['scenario_shift_bp']
    out = []
    for name, w, fmt in IDEA_COLS:
        out.append((name.replace('100bp', '%dbp' % bp), w, fmt))
    return out


def write_idea_sheet(wb, title, blurb, rows, score_key, rationale_fn, parts_key,
                     coverage_key=None, total_weight=100, cfg=None):
    from openpyxl.styles import Font, Alignment, PatternFill
    ws = wb.create_sheet(title)
    ws['A1'] = blurb
    ws['A1'].font = Font(bold=True, size=11, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(_idea_cols(cfg) if cfg else IDEA_COLS))
    ws.append([])
    cols = _idea_cols(cfg) if cfg else IDEA_COLS
    ws.append([c[0] for c in cols])
    _style_header(ws, len(cols), row=3)
    ws.row_dimensions[3].height = 44
    for i, r in enumerate(rows, 1):
        ws.append(_idea_row(i, r, score_key, rationale_fn, parts_key))
    for ridx in range(4, 4 + len(rows)):
        for cidx, (_, _, fmt) in enumerate(cols, 1):
            cell = ws.cell(row=ridx, column=cidx)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(vertical='top',
                                       wrap_text=(cidx == len(cols)))
        if ridx % 2 == 0:
            for cidx in range(1, len(cols) + 1):
                ws.cell(row=ridx, column=cidx).fill = PatternFill(
                    'solid', fgColor='F2F7FC')
        ws.row_dimensions[ridx].height = 42
    foot = ws.max_row + 2
    if coverage_key and rows:
        cov = sum(r.get(coverage_key, 0) for r in rows) / len(rows)
        ws.cell(row=foot, column=1, value=(
            'Scoring weight actually available for these bonds: %.0f of %d points '
            '(%.0f%%). Where an input is missing from the feed, that component is '
            'dropped and its weight is spread across the rest — no bond is marked '
            'down for a data gap.' % (cov, total_weight, 100.0 * cov / total_weight)
        )).font = Font(italic=True, size=9, color='595959')
        foot += 1
    ws.cell(row=foot, column=1, value=(
        'Prices and yields are the indicative offer levels in this week\'s feed, '
        'not executable quotes — check with the desk before dealing. This is a '
        'systematic screen, not a recommendation, and it takes no view on '
        'suitability, concentration or the client\'s existing holdings.'
    )).font = Font(italic=True, size=9, color='9C0006')
    _autosize(ws, [c[1] for c in cols])
    ws.freeze_panes = 'D4'
    # colour-scale the score column
    try:
        from openpyxl.formatting.rule import ColorScaleRule
        ws.conditional_formatting.add(
            'B4:B%d' % (3 + max(len(rows), 1)),
            ColorScaleRule(start_type='min', start_color='FFC7CE',
                           mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                           end_type='max', end_color='C6EFCE'))
    except Exception:
        pass
    return ws


FULL_COLS = [
    ('ISIN', 15, None), ('Valor', 11, None), ('Issuer', 34, None),
    ('Type', 12, None), ('Country', 14, None), ('Region', 16, None),
    ('Ccy', 6, None), ('Coupon %', 9, '0.000'), ('Maturity', 11, None),
    ('Yrs', 6, '0.0'), ('Offer px', 9, '0.00'), ('YTM %', 8, '0.00'),
    ('Mod dur', 8, '0.0'), ('Convexity', 10, '0.0'),
    ('S&P', 8, None), ("Moody's", 9, None), ('Tier', 6, None),
    ('IG/HY', 7, None), ('House view', 16, None), ('Top List', 9, None),
    ('Fitted yld %', 11, '0.00'), ('vs curve (bp)', 12, '+0;-0'),
    ('Roll-down (bp)', 12, '+0;-0'), ('1y carry+roll %', 13, '0.00'),
    ('Peer median YTM %', 14, '0.00'), ('vs peers (bp)', 12, '+0;-0'),
    ('# peers', 8, None),
    ('RALLY -100bp price %', 14, '+0.00;-0.00'),
    ('SELL-OFF +100bp price %', 15, '+0.00;-0.00'),
    ('RALLY -100bp 1y total %', 15, '+0.00;-0.00'),
    ('SELL-OFF +100bp 1y total %', 16, '+0.00;-0.00'),
    ('Risk/reward (up : down)', 14, '0.00"x"'),
    ('Breakeven widening (bp)', 14, '0'),
    ('Sub?', 7, None), ('Callable?', 9, None), ('Floater?', 8, None),
    ('Min denom', 14, None), ('Restrictions', 11, None), ('Green', 12, None),
    ('TACTICAL score', 13, '0.0'), ('HOLD score', 12, '0.0'),
    ('Analyst comment', 60, None),
]


def write_full_sheet(wb, cands, unscored, cfg=None):
    from openpyxl.styles import Alignment, Font
    bp = cfg['scenario_shift_bp'] if cfg else 100
    cols = [(n.replace('100bp', '%dbp' % bp), w, f) for n, w, f in FULL_COLS]
    ws = wb.create_sheet('Full scored universe')
    ws.append([c[0] for c in cols])
    _style_header(ws, len(cols))

    def row_of(r):
        return [
            r['isin'], r['valor'], r['issuer'], r['itype_label'],
            r['country_display'], r['region'], r['ccy'], r['coupon'],
            'Perpetual' if r['is_perp'] else r['maturity'], r['years'],
            r['price'], r['ytm'], r['mod_dur'], r['convexity'],
            r['sp'], r['mdy'], r['tier'], r['grade'],
            r['rec_label'], 'Yes' if r['top_list'] else '',
            r.get('curve_fitted'), r.get('curve_cheap_bp'), r.get('rolldown_bp'),
            r.get('carry_roll_pct'),
            r.get('peer_median_ytm'), r.get('peer_pickup_bp'), r.get('peer_n'),
            r.get('rally_gain_pct'), r.get('selloff_loss_pct'),
            r.get('ret_rally_1y'), r.get('ret_selloff_1y'),
            r.get('risk_reward'), r.get('breakeven_bp'),
            'Yes' if r['is_sub'] else '', 'Yes' if r['is_callable'] else '',
            'Yes' if r['is_floater'] else '',
            r['min_denom'], r['restrictions'] or '', r['green'],
            r.get('tactical_score'), r.get('hold_score'), r['comment'],
        ]

    for r in sorted(cands, key=lambda x: -(x.get('tactical_score') or 0)):
        ws.append(row_of(r))
    if unscored:
        ws.append([])
        ws.append(['NOT SCORED — the feed carries no usable yield for these: '
                   'pure floating-rate notes (no meaningful yield-to-maturity), '
                   'or a gap in the data. Listed for completeness — none of '
                   'them is being rejected on merit.'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True,
                                                      color='9C0006')
        for r in unscored:
            ws.append(row_of(r))
    fmts = {i: f for i, (_, _, f) in enumerate(cols, 1) if f}
    for ridx in range(2, ws.max_row + 1):
        for cidx, fmt in fmts.items():
            ws.cell(row=ridx, column=cidx).number_format = fmt
        ws.cell(row=ridx, column=len(cols)).alignment = Alignment(
            wrap_text=True, vertical='top')
    _autosize(ws, [c[1] for c in cols])
    ws.freeze_panes = 'C2'
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(cols)), ws.max_row)
    return ws


def write_curve_sheet(wb, fit, cands, cfg, curve_label, context=None):
    from openpyxl.styles import Font
    ws = wb.create_sheet('Curve & peers')
    ws['A1'] = 'Relative-value working'
    ws['A1'].font = Font(bold=True, size=12, color='1F3864')
    r = 3
    if fit:
        shape = ('yield = a + b*ln(1+t)' if fit['k'] == 2
                 else 'yield = a + b*ln(1+t) + c*t')
        lines = [
            ('Curve fitted to', curve_label),
            ('Bonds used in the fit', fit['n']),
            ('Functional form', shape),
            ('a (level)', round(fit['coef'][0], 4)),
            ('b (ln slope)', round(fit['coef'][1], 4)),
            ('c (linear slope)', round(fit['coef'][2], 4) if fit['k'] == 3 else 'n/a'),
            ('Fit error (RMSE, bp)', round(fit['rmse_bp'], 1)),
            ('Outliers dropped before re-fit', len(fit['dropped'])),
        ]
    else:
        lines = [('Curve fit', 'NOT POSSIBLE — fewer than %d bonds with a '
                               'usable yield. Curve-based scores are skipped '
                               'and their weight is redistributed to the '
                               'remaining components.' % cfg['curve_min_points'])]
    for k, v in lines:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Per-bond: actual vs fitted, and peer set'
            ).font = Font(bold=True, size=11)
    r += 1
    hdr = ['ISIN', 'Issuer', 'Yrs', 'Actual YTM %', 'Fitted YTM %',
           'Cheap/(rich) bp', 'Roll-down bp', 'Risk/reward measured on',
           'Peer median YTM %', 'vs peers bp', '# peers',
           'Closest EM peers (same ccy, +/-%d notch, '
           '+/-%.0fy)' % (cfg['peer_rating_band'], cfg['peer_maturity_band'])]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, len(hdr), row=r)
    hdr_row = r
    for b in sorted(cands, key=lambda x: (x['years'] or 0)):
        r += 1
        vals = [b['isin'], b['issuer'], b['years'], b['ytm'],
                b.get('curve_fitted'), b.get('curve_cheap_bp'),
                b.get('rolldown_bp'), b.get('rr_basis'), b.get('peer_median_ytm'),
                b.get('peer_pickup_bp'), b.get('peer_n'), b.get('peer_examples')]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
    for rr in range(hdr_row + 1, r + 1):
        for cc, fmt in {3: '0.00', 4: '0.00', 5: '0.00', 6: '+0;-0',
                        7: '+0;-0', 9: '0.00', 10: '+0;-0'}.items():
            ws.cell(row=rr, column=cc).number_format = fmt
    if context:
        r += 3
        ws.cell(row=r, column=1,
                value='Country context — every EM country in this week\'s feed, '
                      'its own fitted curve read off at 3 / 5 / 10 years'
                ).font = Font(bold=True, size=11)
        r += 1
        chdr = ['Country', 'Bonds in fit', 'Curve fitted on', 'Median rating',
                '3y yield %', '5y yield %', '10y yield %', '']
        for c, h in enumerate(chdr, 1):
            ws.cell(row=r, column=c, value=h)
        _style_header(ws, len(chdr), row=r)
        for d in context:
            r += 1
            mark = ('\u25c0  screened here'
                    if d['code'] == cfg['country_code'] else '')
            vals = [d['name'], d['n'], d['basis'], d['median_tier'],
                    d['y3'], d['y5'], d['y10'], mark]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if c in (5, 6, 7):
                    cell.number_format = '0.00'
                if mark:
                    cell.font = Font(bold=True)
    _autosize(ws, [22, 32, 9, 14, 14, 14, 14, 26, 16, 12, 8, 70])
    return ws


METHOD_TEXT = """\
WHAT THIS WORKBOOK IS
    A systematic screen of one country's bonds out of the weekly EM Bond List
    universe, ranked two ways: for a tactical investor with a short horizon,
    and for a client who intends to hold to maturity. It is a shortlisting
    tool. The final call stays with the analyst and the client's suitability
    profile.

WHERE THE DATA COMES FROM
    The same weekly feed files the published EM Bond List is built from, read
    through the same GEMData class. Eligibility, issuer names, ratings
    resolution, IG/HY placement and the Restrictions column therefore match the
    published PDF and Excels exactly. No prices, spreads or forecasts are
    invented anywhere in this file.

HOW EACH SCORE IS BUILT
    Every component is put on a common 0-100 scale and then weighted. Market-
    relative components (cheapness, carry, peer pick-up, yield, price) use the
    bond's PERCENTILE within the screened set, so the ranking answers "best of
    this country's bonds", not "best in the world". Absolute components (credit
    quality, maturity fit, structure, house view, dealability) use fixed
    mappings that don't move with the sample.

    Where an input is missing for a bond, that component is dropped and its
    weight is spread across the components that do have data — so a bond is
    never punished for a gap in the feed. The "coverage" figure below each list
    shows how much of the weight was actually available.

TACTICAL SCORE — components and weights
{tactical_weights}

    curve_cheap       Yield minus the yield the country's own fitted curve says
                      a bond of that maturity should pay. Positive = cheap.
    carry_rolldown    Estimated 1-year return from coupon carry plus the price
                      gain as the bond rolls down the fitted curve, holding the
                      curve still. The standard tactical carry trade metric.
    house_view        UBS recommendation (Attractive / Fair / Expensive / Sell)
                      plus a bonus for Top List membership.
    risk_reward       What you make if EM rallies divided by what you lose if it
                      sells off — see the scenario section below. Above 1.0 means
                      the trade pays you more when you're right than it costs you
                      when you're wrong.
    upside_convexity  Modelled % price gain in the rally case alone, using modified
                      duration and convexity computed from the bond's own terms.
    peer_pickup       Yield versus the median of comparable EM bonds in the same
                      currency, within {peer_rating_band} rating notch(es) and
                      {peer_maturity_band:.0f} years of maturity, excluding the
                      same issuer. This is the cross-universe sanity check.
    liquidity_access  Minimum denomination and MiFID/PRIIPs restriction codes —
                      how easy the bond actually is to put in a client account.

HOLD-TO-MATURITY SCORE — components and weights
{hold_weights}

    yield_level       The yield you actually lock in.
    credit_quality    Worst-of S&P / Moody's on the shared tier scale. Weighted
                      heavily: over a long hold, default risk dominates.
    maturity_fit      Full marks inside the {hold_lo:.0f}-{hold_hi:.0f} year band,
                      decaying outside it. Perpetuals score zero here.
    pull_to_par       Discount prices pull up to 100 at redemption. A premium
                      price means part of the coupon is a return of capital.
    structure         Senior, dated, fixed-coupon, non-callable = full marks.
    house_view        Same input as tactical, smaller weight — a hold-to-maturity
                      buyer cares less about near-term richness.

THE TWO SCENARIOS AND THE RISK/REWARD COLUMN
    Every bond is run through a symmetric pair of parallel moves, sized
    {shift:.0f}bp:

        RALLY     yields -{shift:.0f}bp   EM spreads compress
        SELL-OFF  yields +{shift:.0f}bp   EM spread widening

    Grade betas applied to the move — sell-off IG {sell_ig:.2f} / HY {sell_hy:.2f},
    rally IG {rally_ig:.2f} / HY {rally_hy:.2f}. At 1.00 everything moves in
    parallel, which is the assumption-free default; raise the HY beta if you want
    the sell-off to hit speculative-grade harder, as it usually does in a genuine
    EM risk-off.

    Two versions of each scenario are reported:

      PRICE %       the pure mark-to-market move, from modified duration and
                    convexity. Convexity is why the rally gain is a little larger
                    than the sell-off loss on the same bond.
      1y TOTAL %    the same move plus a year of carry and roll-down. This is the
                    honest version: in a sell-off the coupon you collect over the
                    year offsets part of the hit, so a fat-carry short bond can
                    finish the year positive even though its price fell.

    RISK/REWARD is the 1-year upside in the rally divided by the size of the
    1-year downside in the sell-off. Where no carry number is available (no curve
    could be fitted) it falls back to the price-only ratio, which is flagged per
    bond on the Curve & peers tab. Where a year of carry more than absorbs the
    whole widening, there is no downside to divide by: the cell reads
    "{noloss} (+x.x%)", showing the cushion left over after the move, and those
    bonds sort above every finite ratio, ordered by the size of that cushion.

    BREAKEVEN WIDENING is the same idea stated in basis points: how far yields
    can back up over the next year before the position is flat. Carry plus
    roll-down, divided by duration. For a tactical EM trade it is usually the
    single most useful number on the sheet.

    What the scenarios are NOT: a forecast, and not a credit event. A parallel
    {shift:.0f}bp move assumes the bond stays money-good. A default or
    restructuring is a different question, which is what the credit-quality
    component of the hold-to-maturity score is there for.

PENALTIES (subtracted after weighting)
{penalties}

MATHS NOTES
    Modified duration and convexity are computed by discounting the bond's own
    cash flows semi-annually at the feed's offer yield — no approximations, no
    external library. Perpetuals use the 1/y limit, capped at 25 years.
    The curve is a least-squares fit of the form a + b*ln(1+t) (+ c*t when there
    are at least 5 bonds), refitted once after dropping any bond more than 2.5
    standard deviations away, so one distressed line can't distort the reference
    everything else is measured against.

LIMITATIONS — read before sending anything to a client
    * Yields and prices are the indicative offer levels in the weekly feed, not
      executable quotes. Check the desk before dealing.
    * The screen has no view on FX, on the sovereign's fiscal path, or on
      anything published after the feed date below.
    * Roll-down assumes the curve does not move. It won't hold.
    * Peer comparison uses the shared rating tier only; it does not adjust for
      sector, ESG or documentation differences.
    * Suitability, concentration limits and the client's existing holdings are
      not considered anywhere in this file.
"""


def write_method_sheet(wb, cfg, meta):
    from openpyxl.styles import Font, Alignment
    ws = wb.create_sheet('Methodology')

    def fmt_weights(d):
        return '\n'.join('        %-18s %3d%%' % (k, v)
                         for k, v in sorted(d.items(), key=lambda kv: -kv[1]))

    text = METHOD_TEXT.format(
        tactical_weights=fmt_weights(cfg['tactical_weights']),
        hold_weights=fmt_weights(cfg['hold_weights']),
        penalties='\n'.join('        %-24s -%d pts' % (k.replace('_', ' '), v)
                            for k, v in sorted(cfg['penalties'].items())),
        peer_rating_band=cfg['peer_rating_band'],
        peer_maturity_band=cfg['peer_maturity_band'],
        hold_lo=cfg['hold_band'][0], hold_hi=cfg['hold_band'][1],
        shift=cfg['scenario_shift_bp'], noloss=RR_NO_LOSS,
        sell_ig=cfg['selloff_beta'].get('IG', 1.0),
        sell_hy=cfg['selloff_beta'].get('HY', 1.0),
        rally_ig=cfg['rally_beta'].get('IG', 1.0),
        rally_hy=cfg['rally_beta'].get('HY', 1.0))

    ws['A1'] = 'Methodology — %s' % meta['title']
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    r = 3
    for k, v in meta['run'].items():
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    r += 1
    for line in text.split('\n'):
        c = ws.cell(row=r, column=1, value=line)
        if line and not line.startswith(' '):
            c.font = Font(bold=True, size=10, color='1F3864')
        else:
            c.font = Font(name='Consolas', size=9)
        r += 1
    _autosize(ws, [34, 60])
    ws.column_dimensions['A'].width = 100
    return ws


def build_workbook(path, cfg, meta, tactical, hold, cands, unscored, fit,
                   curve_label):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    t_blurb = ('TACTICAL IDEAS — %s.  Ranked for a short-horizon investor: '
               'cheap to the curve, best carry and roll-down, most price upside '
               'if spreads rally. Sorted best first.' % meta['scope'])
    h_blurb = ('HOLD TO MATURITY — %s.  Ranked for a buy-and-hold investor: '
               'yield locked in, credit quality, maturity fit, clean senior '
               'structure. Sorted best first.' % meta['scope'])

    write_idea_sheet(wb, 'Tactical ideas', t_blurb, tactical,
                     'tactical_score', tactical_rationale, 'tactical_parts',
                     'tactical_coverage', sum(cfg['tactical_weights'].values()),
                     cfg=cfg)
    write_idea_sheet(wb, 'Hold to maturity', h_blurb, hold,
                     'hold_score', hold_rationale, 'hold_parts',
                     'hold_coverage', sum(cfg['hold_weights'].values()), cfg=cfg)
    write_full_sheet(wb, cands, unscored, cfg)
    write_curve_sheet(wb, fit, [c for c in cands if c['ytm'] is not None],
                      cfg, curve_label, context=meta.get('context'))
    write_method_sheet(wb, cfg, meta)
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# 11. MAIN
# ══════════════════════════════════════════════════════════════════════════════

def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description='Rank the best bonds of one country (or issuer) out of the '
                    'weekly EM Bond List universe.')
    p.add_argument('--country', default=None,
                   help="Country name or ISO code, e.g. Colombia / CO / Argentina")
    p.add_argument('--issuer', default=None,
                   help="Screen one issuer instead, e.g. 'Ecopetrol'")
    p.add_argument('--universe', default=None,
                   choices=['sovereign', 'quasi', 'all'],
                   help='sovereign (default) | quasi (everything else) | all')
    p.add_argument('--ccy', default=None,
                   help="Comma-separated currencies, or 'all'. Default USD")
    p.add_argument('--top', type=int, default=None, help='Ideas per list')
    p.add_argument('--data-dir', default=None,
                   help='Folder holding the weekly .txt feed files')
    p.add_argument('--out', default=None, help='Output .xlsx path')
    p.add_argument('--list-countries', action='store_true',
                   help='Print what is in this week\'s feed and exit')
    p.add_argument('--list-issuers', action='store_true',
                   help='Print the issuer names in the feed (optionally '
                        'narrowed by --country) and exit')
    p.add_argument('--quiet', action='store_true')
    return p.parse_args(argv)


def build_config(args):
    country_in = args.country or TARGET
    issuer_in = args.issuer if args.issuer is not None else TARGET_ISSUER
    code, name = resolve_country(country_in)
    if not code and not issuer_in:
        sys.exit("Don't recognise the country %r. Use an ISO code (CO, AR, BR) "
                 "or run --list-countries to see what's in the feed." % country_in)
    if args.ccy is None:
        ccys = [c.upper() for c in CURRENCIES]
    elif args.ccy.lower() in ('all', '*', ''):
        ccys = []
    else:
        ccys = [c.strip().upper() for c in args.ccy.split(',') if c.strip()]
    return {
        'country_code': code,
        'country_name': name or country_in,
        'issuer_fragment': (issuer_in or '').strip().lower(),
        'universe': args.universe or UNIVERSE,
        'currencies': ccys,
        'top_n': args.top or TOP_N,
        'tactical_min_years': TACTICAL_MIN_YEARS,
        'hold_band': HOLD_BAND_YEARS,
        'hold_max_years': HOLD_MAX_YEARS,
        'curve_min_points': CURVE_MIN_POINTS,
        'curve_fallback': CURVE_FALLBACK,
        'peer_rating_band': PEER_RATING_BAND,
        'peer_maturity_band': PEER_MATURITY_BAND,
        'peer_min_count': PEER_MIN_COUNT,
        'rolldown_horizon': ROLLDOWN_HORIZON,
        'scenario_shift_bp': SCENARIO_SHIFT_BP,
        'selloff_beta': dict(SELLOFF_BETA),
        'rally_beta': dict(RALLY_BETA),
        'tactical_weights': dict(TACTICAL_WEIGHTS),
        'hold_weights': dict(HOLD_WEIGHTS),
        'penalties': dict(PENALTIES),
    }


def list_countries(universe):
    from collections import Counter
    cnt = Counter()
    for r in universe:
        if r['country']:
            cnt[(r['country'], r['country_display'])] += 1
    print('\nCountries in this week\'s eligible universe '
          '(code, name, bonds):\n')
    for (code, nm), n in sorted(cnt.items(), key=lambda kv: -kv[1]):
        sov = sum(1 for r in universe
                  if r['country'] == code and r['itype'] in ('SOV', 'SUPRA'))
        print('   %-4s %-28s %4d bonds  (%d sovereign)' % (code, nm, n, sov))
    print('\nRun e.g.:  python bond_ideas.py --country CO\n')


def list_issuers(universe, code=None, name=None):
    from collections import Counter
    cnt = Counter()
    for r in universe:
        if code and r['country'] != code:
            continue
        cnt[(r['issuer_raw'], r['itype_label'], r['country_display'])] += 1
    where = (' in %s' % name) if code else ''
    print('\nIssuers in this week\'s eligible universe%s '
          '(name, type, country, bonds):\n' % where)
    for (nm, ityp, cc), n in sorted(cnt.items()):
        print('   %-40s %-14s %-18s %3d' % (nm[:40], ityp, cc, n))
    print('\n--issuer matches any part of the name above, case-insensitive.'
          '\ne.g.  python bond_ideas.py --issuer "Ecopetrol"\n')


def main(argv=None):
    global QUIET
    args = parse_args(argv)
    QUIET = QUIET or args.quiet

    data_dir = find_data_dir(args.data_dir)
    say('[bond-ideas] reading the weekly feed from %s' % data_dir)
    if QUIET:                       # the loader prints its own progress lines
        with contextlib.redirect_stdout(io.StringIO()):
            data = load_data(data_dir)
    else:
        data = load_data(data_dir)

    say('[bond-ideas] enriching %d eligible bonds…' % len(data.em_bonds))
    universe = [e for e in (enrich(b, data) for b in data.em_bonds) if e]

    if args.list_countries:
        list_countries(universe)
        return 0

    cfg = build_config(args)
    if args.list_issuers:
        list_issuers(universe,
                     cfg['country_code'] if args.country else None,
                     cfg['country_name'])
        return 0
    scope_bits = []
    if cfg['issuer_fragment']:
        scope_bits.append("issuer matching '%s'" % cfg['issuer_fragment'])
    else:
        scope_bits.append(cfg['country_name'])
        scope_bits.append({'sovereign': 'sovereign only',
                           'quasi': 'quasi-sovereign & corporate',
                           'all': 'all issuers'}[cfg['universe']])
    scope_bits.append(', '.join(cfg['currencies']) if cfg['currencies']
                      else 'all currencies')
    scope = ' · '.join(scope_bits)
    say('[bond-ideas] screening: %s' % scope)

    cands = [r for r in universe if matches_target(r, cfg)]
    if cfg['currencies']:
        cands = [r for r in cands if r['ccy'] in cfg['currencies']]
    if not cands:
        hint = ''
        if cfg['issuer_fragment']:
            frag = cfg['issuer_fragment']
            words = [w for w in frag.split() if len(w) > 3] or [frag]
            near = sorted({r['issuer_raw'] for r in universe
                           if any(w in r['issuer_raw'].lower() for w in words)})
            if not near:
                hint = ('\n  No published issuer name contains %r. Matching is a '
                        'plain substring of the name as it appears on the list '
                        '(so \'Petroleos\', not \'Pemex\').\n'
                        '  Run:  python %s --list-issuers   (add --country XX to '
                        'narrow it)' % (frag, os.path.basename(__file__)))
            else:
                hint = ('\n  Did you mean one of these?\n    %s'
                        % '\n    '.join(near[:25]))
        sys.exit(
            'No bonds matched.\n'
            '  Screened for: %s%s\n'
            '  Try --universe all, --ccy all, or run --list-countries to see '
            'what this week\'s feed actually contains.' % (scope, hint))
    say('[bond-ideas] %d bond(s) matched' % len(cands))

    # When screening a named issuer, the country used for the fallback curve and
    # for the context table should be the issuer's own, not whatever TARGET says.
    if cfg['issuer_fragment']:
        codes = [r['country'] for r in cands if r['country']]
        if codes:
            cfg['country_code'] = max(set(codes), key=codes.count)
            cfg['country_name'] = G.COUNTRY_NAMES.get(cfg['country_code'],
                                                      cfg['country_code'])

    # ---- curve pool: the target set, widened to the whole country if thin ---
    curve_pool = cands
    curve_label = scope
    usable = sum(1 for r in cands
                 if r['ytm'] is not None and r['years'] and not r['is_perp'])
    if usable < cfg['curve_min_points'] and cfg['curve_fallback'] and cfg['country_code']:
        wider = [r for r in universe if r['country'] == cfg['country_code']
                 and (not cfg['currencies'] or r['ccy'] in cfg['currencies'])]
        if len(wider) > usable:
            curve_pool = wider
            curve_label = ('%s — all issuers (widened: too few bonds in the '
                           'screened set to fit a curve)' % cfg['country_name'])
            say('[bond-ideas] curve widened to all %s issuers (%d bonds)'
                % (cfg['country_name'], len(wider)))

    fit = attach_curve(cands, curve_pool, cfg)
    attach_peers(cands, universe, cfg)
    attach_scenarios(cands, cfg)
    scored, unscored = score_all(cands, cfg)
    if not scored:
        sys.exit('Matched %d bond(s) but none of them carries a usable yield in '
                 'the feed, so nothing can be ranked.' % len(cands))

    context = country_context(universe, cfg)
    tactical = sorted(scored, key=lambda r: -r['tactical_score'])[:cfg['top_n']]
    hold = sorted(scored, key=lambda r: -r['hold_score'])[:cfg['top_n']]

    label = (cfg['issuer_fragment'] or cfg['country_name'] or 'Selection')
    label = ''.join(ch for ch in label.title().replace(' ', '_')
                    if ch.isalnum() or ch == '_')
    out = args.out
    if not out:
        out_dir = os.path.join(HERE, OUTPUT_DIR)
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, OUTPUT_TEMPLATE.format(target=label))

    meta = {
        'title': scope,
        'scope': scope,
        'run': {
            'Screened': scope,
            'Bonds matched': len(cands),
            'Bonds ranked': len(scored),
            'Not ranked (no usable yield)': len(unscored),
            'Feed folder': data_dir,
            'Eligible EM universe this week': len(universe),
            'Peer comparison pool': 'whole eligible EM universe',
            'Scenario shift': '+/- %dbp parallel' % cfg['scenario_shift_bp'],
            'Sell-off beta (IG / HY)': '%.2f / %.2f' % (
                cfg['selloff_beta'].get('IG', 1.0),
                cfg['selloff_beta'].get('HY', 1.0)),
            'Rally beta (IG / HY)': '%.2f / %.2f' % (
                cfg['rally_beta'].get('IG', 1.0),
                cfg['rally_beta'].get('HY', 1.0)),
            'Generated': datetime.now().strftime('%d.%m.%Y %H:%M'),
            'Script': os.path.basename(__file__),
        },
        'context': context,
    }
    build_workbook(out, cfg, meta, tactical, hold, cands, unscored, fit,
                   curve_label)

    say('')
    say('[bond-ideas] TACTICAL top %d:' % len(tactical))
    for i, r in enumerate(tactical, 1):
        say('   %2d. %-14s %-30s %-11s %5s%%  score %.1f'
            % (i, r['isin'], r['issuer_raw'][:30],
               'Perp' if r['is_perp'] else r['maturity'],
               '%.2f' % r['ytm'], r['tactical_score']))
    say('[bond-ideas] HOLD-TO-MATURITY top %d:' % len(hold))
    for i, r in enumerate(hold, 1):
        say('   %2d. %-14s %-30s %-11s %5s%%  score %.1f'
            % (i, r['isin'], r['issuer_raw'][:30],
               'Perp' if r['is_perp'] else r['maturity'],
               '%.2f' % r['ytm'], r['hold_score']))
    say('')
    print('Workbook written to:\n   %s' % out)
    return 0


if __name__ == '__main__':
    sys.exit(main())
