"""Tagging — weekly (Friday) ISIN tagging + week-on-week comparison workbook.

Produces ONE Excel file (outputs/EMBL_Tagging.xlsx) with two tabs:

  Tab 1 — "Offshore ISINs"
      Every ISIN published this week, with its issuer name — the SAME universe
      rendered in the PDF: the reference (currency/region) lists plus the Sell
      Recommendations page. Built from data.reference_list_bonds() +
      data.sell_list_bonds(), so it matches the PDF exactly (currency
      exclusions and all), not merely the GEM-eligibility flag.

  Tab 2 — "Weekly comparison"
      The week-on-week diff: bonds Added, Deleted, Upgraded (and to what) and
      Downgraded (and to what) this week, each with ISIN, issuer, the
      previous/new recommendation view, and a column flagging whether the
      change is relevant for the US Onshore list.

DEFINITIONS (kept consistent with the rest of the pipeline)
-----------------------------------------------------------
  • Offshore / Onshore eligibility ...... reused verbatim from
        gem_excel_builder (is_offshore_eligible / is_onshore_eligible), so any
        future rule change there flows through here automatically.
  • Upgrade / Downgrade ................. change in the WMR recommendation
        view (attr. / fair / exp.), the SAME definition used by the onshore
        "Changes this week" sheet. "To what" = the New view column.
  • Onshore relevant .................... mirrors gem_excel_builder's
        _onshore_change: onshore-eligible THIS week (adds/upgrades/downgrades)
        or LAST week (deletions).

HOW TO RUN
----------
Open in Spyder and press F5 (same pattern as run_weekly.py), or run
`python tagging.py` from the project folder. It reads the same data/current
and data/previous inputs as the weekly pipeline. It does NOT rebuild the PDF
or the published Excels — it only reads the data and writes the tagging file.
"""

import datetime
import glob
import os
import sys

# Run from this file's folder so the relative data/ paths resolve.
HERE = os.path.dirname(os.path.abspath(__file__))
os.chdir(HERE)
if HERE not in sys.path:
    sys.path.insert(0, HERE)

DATA = 'data'
CURR = os.path.join(DATA, 'current')
PREV = os.path.join(DATA, 'previous')
OUT = 'outputs'
os.makedirs(OUT, exist_ok=True)

OUTPUT_PATH = os.path.join(OUT, 'EMBL_Tagging.xlsx')


# ── Reference-file discovery (same conventions as run_weekly.py) ────────────
def _find_priips_reference():
    patterns = []
    for folder in (DATA, CURR, HERE):
        for ext in ('xls', 'xlsx', 'csv'):
            patterns.append(os.path.join(folder, '*[Pp][Rr][Ii][Ii][Pp][Ss]*.' + ext))
    candidates = []
    for pat in patterns:
        candidates.extend(glob.glob(pat))
    candidates = [c for c in set(candidates)
                  if not os.path.basename(c).startswith('~$')]
    return max(candidates, key=os.path.getmtime) if candidates else None


def _find_legal_exclusions():
    patterns = []
    for folder in (DATA, CURR, HERE):
        for ext in ('txt', 'csv'):
            for stem in ('*[Ll]egal*', '*[Ee]xclusion*'):
                patterns.append(os.path.join(folder, stem + '.' + ext))
    candidates = []
    for pat in patterns:
        candidates.extend(glob.glob(pat))
    skip = ('~$', 'template', 'example', 'sample')
    candidates = [c for c in set(candidates)
                  if not any(s in os.path.basename(c).lower() for s in skip)]
    return max(candidates, key=os.path.getmtime) if candidates else None


def build_data():
    """Construct GEMData from the standard weekly inputs (no PDF/Excel build)."""
    import gem_report_builder_v3 as grb

    paths = {
        'bond_data':       os.path.join(CURR, 'CurrentPublishableBondData.txt'),
        'issuer_data':     os.path.join(CURR, 'CurrentPublishableIssuerData.txt'),
        'bond_update':     os.path.join(CURR, 'PublishableBondDataUpdate.txt'),
        'issuer_update':   os.path.join(CURR, 'PublishableIssuerDataUpdate.txt'),
        'color_flags':     os.path.join(CURR, 'PublishableColorFlags.txt'),
        'issuer_texts':    os.path.join(CURR, 'IssuerTexts.txt'),
        'issuer_ratings':  os.path.join(CURR, 'IssuerRatings.txt'),
        'prev_bond_data':  os.path.join(PREV, 'CurrentPublishableBondData.txt'),
        'prev_bond_update':os.path.join(PREV, 'PublishableBondDataUpdate.txt'),
        'priips_ref':      _find_priips_reference(),
        'legal_exclusions':_find_legal_exclusions(),
    }
    return grb.GEMData(paths)


# ── Tab content builders ────────────────────────────────────────────────────
def offshore_isin_rows(data):
    """[(isin, issuer)] for every bond PUBLISHED this week, sorted by issuer.

    This is the exact PDF universe: the reference (currency/region) lists
    (reference_list_bonds — excludes Sell bonds and EXCLUDED_CURRENCIES) plus
    the Sell Recommendations page (sell_list_bonds). Deduplicated by ISIN so a
    bond appearing in more than one section is listed once.
    """
    seen = set()
    rows = []
    for b in list(data.reference_list_bonds()) + list(data.sell_list_bonds()):
        isin = (b.get('Isin') or '').strip()
        if not isin or isin in seen:
            continue
        seen.add(isin)
        gk = (b.get('GK_Nummer') or '').strip()
        issuer = data.issuer_display_name(gk, fallback=b.get('IssuerName', '')).strip()
        rows.append((isin, issuer))
    rows.sort(key=lambda r: (r[1].lower(), r[0]))
    return rows


def _is_onshore_relevant(change_row, data):
    """Mirror gem_excel_builder._onshore_change: onshore-eligible this week
    (adds / upgrades / downgrades) or last week (deletions)."""
    from gem_excel_builder import is_onshore_eligible

    isin = (change_row.get('isin') or '').strip()
    if not isin:
        return False
    bond = data.bond_by_isin.get(isin)
    if bond is not None and is_onshore_eligible(bond, data):
        return True
    prev_bond = getattr(data, 'prev_bonds', {}).get(isin)
    if not prev_bond:
        return False
    prev_upd = getattr(data, 'prev_bond_updates', {}).get(isin, {})
    return is_onshore_eligible(prev_bond, data, upd=prev_upd)


def comparison_rows(data):
    """Flat list of comparison rows across the four change categories.

    Each row: (category, isin, issuer, prev_view, new_view, onshore_relevant).
    Categories: Added, Deleted, Upgraded, Downgraded.
    """
    changes = data.recommendation_changes()
    order = [
        ('Added', 'additions'),
        ('Deleted', 'deletions'),
        ('Upgraded', 'upgrades'),
        ('Downgraded', 'downgrades'),
    ]
    out = []
    for label, key in order:
        for r in changes.get(key, []):
            out.append((
                label,
                r.get('isin', ''),
                r.get('issuer', ''),
                r.get('view_prior') or '-',
                r.get('view_new') or '-',
                'Yes' if _is_onshore_relevant(r, data) else 'No',
            ))
    return out


# ── Excel writer ────────────────────────────────────────────────────────────
def write_workbook(offshore, comparison, output_path=OUTPUT_PATH):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill('solid', fgColor='1F3864')   # UBS-ish dark blue
    HDR_FONT = Font(bold=True, color='FFFFFF')
    SEC_FILL = PatternFill('solid', fgColor='D9E1F2')    # light blue band
    SEC_FONT = Font(bold=True)
    CENTER = Alignment(horizontal='center')

    wb = Workbook()

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(ncols)}{ws.max_row}'

    def _autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 1 — Offshore ISINs ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Offshore ISINs'
    ws1.append(['ISIN', 'Issuer'])
    for isin, issuer in offshore:
        ws1.append([isin, issuer])
    _style_header(ws1, 2)
    _autosize(ws1, [18, 60])

    # ── Tab 2 — Weekly comparison ───────────────────────────────────────
    ws2 = wb.create_sheet('Weekly comparison')
    headers = ['Change', 'ISIN', 'Issuer', 'Previous view',
               'New view', 'Onshore relevant']
    ws2.append(headers)
    for row in comparison:
        ws2.append(list(row))
    if ws2.max_row == 1:
        ws2.append(['No changes this week', '', '', '', '', ''])
    _style_header(ws2, len(headers))
    for c in (2, 4, 5, 6):   # center ISIN / views / flag
        for r in range(2, ws2.max_row + 1):
            ws2.cell(row=r, column=c).alignment = CENTER
    _autosize(ws2, [13, 18, 50, 14, 12, 16])

    wb.save(output_path)
    return output_path


def main():
    print('Building GEMData from data/current and data/previous ...')
    data = build_data()

    offshore = offshore_isin_rows(data)
    comparison = comparison_rows(data)

    path = write_workbook(offshore, comparison)

    n_add = sum(1 for r in comparison if r[0] == 'Added')
    n_del = sum(1 for r in comparison if r[0] == 'Deleted')
    n_up = sum(1 for r in comparison if r[0] == 'Upgraded')
    n_dn = sum(1 for r in comparison if r[0] == 'Downgraded')
    n_onshore = sum(1 for r in comparison if r[5] == 'Yes')

    print(f'\n✓ Wrote {path}')
    print(f'  Tab 1 "Offshore ISINs":      {len(offshore):,} bonds')
    print(f'  Tab 2 "Weekly comparison":   {len(comparison):,} changes '
          f'({n_add} added, {n_del} deleted, {n_up} upgraded, {n_dn} downgraded; '
          f'{n_onshore} onshore-relevant)')
    if not getattr(data, 'prev_bond_updates', None):
        print('  NOTE: no previous-week data found, so the comparison tab is '
              'empty. Check data/previous.')
    return path


if __name__ == '__main__':
    main()
