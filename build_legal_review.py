"""Build the Legal review workbook for the EM Bond List.

WHAT THIS PRODUCES  ->  EMBL_Legal_Review.xlsx
-----------------------------------------------
One sheet with these stacked sections, matching the format Legal uses:

  1. "Bonds we weren't allowed to add in the previous update ..."
        From legal_blocked.csv  (a list YOU maintain — the removed/added-back
        dates are Legal/compliance decisions, not in the bond feed).
  2. Additions          \
  3. Upgrades            >  computed from the week-on-week diff
  4. Downgrades         /   (data.recommendation_changes()), highlighted as
                            "this week" changes.
  5. Global Restriction Lists
        From global_restrictions.csv (a hard-coded list YOU maintain — these
        are US-government / sanctions calls with no flag in the data).

Columns per row: ISIN | Issuer (CTL_List_Name) | Coupon | Maturity | US Onshore?
(The old "Duplicate?" column is intentionally dropped.)

HOW TO RUN
----------
    python build_legal_review.py                 # uses data/current + data/previous
    python build_legal_review.py --date 2026-06-23

Run it after this week's snapshot is in data/current.

MAINTAINED INPUT FILES (edit these as needed, no code changes):
  • legal_blocked.csv       columns: isin,issuer,coupon,maturity,us_onshore,
                                     date_removed,added_back
  • global_restrictions.csv columns: isin,issuer
"""

import os
import sys
import csv
import argparse
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

import gem_report_builder_v3 as gem
from gem_excel_builder import is_onshore_eligible

DATA = os.path.join(HERE, 'data')
CURR = os.path.join(DATA, 'current')
PREV = os.path.join(DATA, 'previous')

LEGAL_BLOCKED_CSV = os.path.join(HERE, 'legal_blocked.csv')
GLOBAL_RESTR_CSV  = os.path.join(HERE, 'global_restrictions.csv')
SOFT_RESTR_CSV    = os.path.join(HERE, 'soft_restrictions.csv')
OUT_XLSX          = os.path.join(HERE, 'EMBL_Legal_Review.xlsx')


# ----------------------------------------------------------------- data hookup
def build_data():
    """Construct the production GEMData (same inputs as run_weekly.py)."""
    paths = {
        'bond_data':        os.path.join(CURR, 'CurrentPublishableBondData.txt'),
        'issuer_data':      os.path.join(CURR, 'CurrentPublishableIssuerData.txt'),
        'bond_update':      os.path.join(CURR, 'PublishableBondDataUpdate.txt'),
        'issuer_update':    os.path.join(CURR, 'PublishableIssuerDataUpdate.txt'),
        'color_flags':      os.path.join(CURR, 'PublishableColorFlags.txt'),
        'issuer_texts':     os.path.join(CURR, 'IssuerTexts.txt'),
        'issuer_ratings':   os.path.join(CURR, 'IssuerRatings.txt'),
        'prev_bond_data':   os.path.join(PREV, 'CurrentPublishableBondData.txt'),
        'prev_bond_update': os.path.join(PREV, 'PublishableBondDataUpdate.txt'),
        'priips_ref':       None,
        'legal_exclusions': None,
    }
    return gem.GEMData(paths)


# ------------------------------------------------------------------- row helpers
def _bond(data, isin):
    return (getattr(data, 'bond_by_isin', {}) or {}).get(isin)


def onshore_flag(data, isin, default=''):
    """'Yes' / 'No' from the production onshore-eligibility rule, or `default`
    when the bond isn't in this week's feed."""
    b = _bond(data, isin)
    if b is None:
        return default
    try:
        return 'Yes' if is_onshore_eligible(b, data) else 'No'
    except Exception:
        return default


def issuer_short(data, isin, fallback=''):
    """CTL_List_Name for the ISIN (the column Legal calls 'Issuer')."""
    b = _bond(data, isin) or {}
    gk = (b.get('GK_Nummer') or '').strip()
    return data.issuer_display_name(gk, fallback=fallback or b.get('IssuerName', ''))


def _read_csv(path):
    if not os.path.exists(path):
        print(f'[legal] note: {os.path.basename(path)} not found — section left empty.')
        return []
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))


# ------------------------------------------------------------------- sections
def section_blocked(data):
    """Section 1 — maintained list of legally-blocked bonds."""
    rows = []
    for r in _read_csv(LEGAL_BLOCKED_CSV):
        isin = (r.get('isin') or '').strip()
        if not isin:
            continue
        rows.append({
            'isin':     isin,
            'issuer':   (r.get('issuer') or '').strip() or issuer_short(data, isin),
            'coupon':   (r.get('coupon') or '').strip(),
            'maturity': (r.get('maturity') or '').strip(),
            'onshore':  (r.get('us_onshore') or '').strip() or onshore_flag(data, isin),
            'date_removed': (r.get('date_removed') or '').strip(),
            'added_back':   (r.get('added_back') or '').strip(),
        })
    return rows


def _changes_rows(data, items):
    """Map recommendation_changes() rows -> legal columns."""
    out = []
    for it in items:
        isin = (it.get('isin') or '').strip()
        out.append({
            'isin':     isin,
            'issuer':   issuer_short(data, isin, fallback=it.get('issuer', '')),
            'coupon':   it.get('coupon', ''),
            'maturity': it.get('maturity', ''),
            'onshore':  onshore_flag(data, isin),
        })
    return out


def section_global_restrictions(data):
    rows = []
    for r in _read_csv(GLOBAL_RESTR_CSV):
        isin = (r.get('isin') or '').strip()
        if not isin:
            continue
        b = _bond(data, isin) or {}
        rows.append({
            'isin':     isin,
            'issuer':   (r.get('issuer') or '').strip() or issuer_short(data, isin),
            'coupon':   gem.format_percent(b.get('Coupon')) if b else '',
            'maturity': (gem.format_date(b.get('Maturity')) if b else '') or '',
            'onshore':  onshore_flag(data, isin),
        })
    return rows


def section_soft_restrictions(data):
    """Bottom section — maintained list of 'soft restriction' bonds.
    Columns differ: ISIN | Issuer | Maturity | Currency | US Onshore?
    (no coupon). Maturity/Currency come from the file, falling back to the
    feed when the bond is present this week."""
    rows = []
    for r in _read_csv(SOFT_RESTR_CSV):
        isin = (r.get('isin') or '').strip()
        if not isin:
            continue
        b = _bond(data, isin) or {}
        rows.append({
            'isin':     isin,
            'issuer':   (r.get('issuer') or '').strip() or issuer_short(data, isin),
            'maturity': (r.get('maturity') or '').strip()
                        or (gem.format_date(b.get('Maturity')) if b else '') or '',
            'currency': (r.get('currency') or '').strip()
                        or (b.get('CCY', '').strip().upper() if b else ''),
            'onshore':  (r.get('us_onshore') or '').strip() or onshore_flag(data, isin),
        })
    return rows


# ------------------------------------------------------------------- rendering
def write_xlsx(data, changes, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Legal review'

    GREEN  = PatternFill('solid', fgColor='6B7B6B')   # section bar
    HILITE = PatternFill('solid', fgColor='FFF2CC')   # this-week changes
    COLHDR = PatternFill('solid', fgColor='D9D9D9')
    white_bold = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(bottom=thin, right=thin)

    COLS_BASE  = ['ISIN', 'Issuer', 'Coupon', 'Maturity', 'US Onshore?']
    COLS_BLOCK = COLS_BASE + ['Date removed', 'Added back']
    NCOLS = len(COLS_BLOCK)

    row = 1

    def section_bar(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = white_bold
        for c in range(1, NCOLS + 1):
            ws.cell(row=row, column=c).fill = GREEN
        row += 1

    def col_header(cols):
        nonlocal row
        for i, h in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = bold
            cell.fill = COLHDR
        row += 1

    def data_rows(rows, cols_keys, highlight=False):
        nonlocal row
        if not rows:
            ws.cell(row=row, column=1, value='No changes this week')
            row += 1
            return
        for r in rows:
            for i, k in enumerate(cols_keys, start=1):
                cell = ws.cell(row=row, column=i, value=r.get(k, ''))
                cell.border = border
                if highlight:
                    cell.fill = HILITE
            row += 1

    def blank():
        nonlocal row
        row += 1

    base_keys  = ['isin', 'issuer', 'coupon', 'maturity', 'onshore']
    block_keys = base_keys + ['date_removed', 'added_back']

    # 1) Blocked
    section_bar("Bonds we weren't allowed to add in the previous update and "
                "would like to see if we can add them now")
    col_header(COLS_BLOCK)
    data_rows(section_blocked(data), block_keys)
    blank()

    # 2-4) This-week changes (highlighted)
    for title, key in (('Additions', 'additions'),
                       ('Upgrades', 'upgrades'),
                       ('Downgrades', 'downgrades')):
        section_bar(title)
        col_header(COLS_BASE)
        data_rows(_changes_rows(data, changes.get(key, [])), base_keys,
                  highlight=True)
        blank()

    # 5) Global restriction lists
    section_bar('Global Restriction Lists (no actions, previously approved, '
                'flagged by system)')
    col_header(COLS_BASE)
    data_rows(section_global_restrictions(data), base_keys)
    blank()

    # 6) Soft restrictions (own column layout: Maturity + Currency, no coupon)
    COLS_SOFT  = ['ISIN', 'Issuer', 'Maturity', 'Currency', 'US Onshore?']
    soft_keys  = ['isin', 'issuer', 'maturity', 'currency', 'onshore']
    section_bar('Soft restrictions (no actions, previously approved, '
                'flagged by system)')
    col_header(COLS_SOFT)
    data_rows(section_soft_restrictions(data), soft_keys)

    widths = [16, 34, 12, 12, 12, 13, 12]
    for i, wch in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wch
    ws.freeze_panes = 'A1'
    wb.save(path)


# ------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--date', default=datetime.date.today().isoformat(),
                    help='As-of date (currently informational).')
    args = ap.parse_args()

    data = build_data()
    changes = data.recommendation_changes()
    write_xlsx(data, changes, OUT_XLSX)

    n = sum(len(changes.get(k, [])) for k in ('additions', 'upgrades', 'downgrades'))
    print(f'[legal] {n} week-on-week changes rendered')
    print(f'[legal] wrote {OUT_XLSX}')


if __name__ == '__main__':
    main()
