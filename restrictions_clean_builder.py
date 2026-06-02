"""Clean, shareable Restrictions workbook (EMBL_Restrictions_CLEAN.xlsx).

Produces a two-tab workbook each weekly run:
  • "Read me - India team"   — what file India must keep current, where it
                                goes (the data/ folder), the required columns,
                                the rule, and ownership.
  • "Restrictions (Offshore)"— one clean row per Offshore bond: ISIN, Valor,
                                Issuer, final Restriction, and the source
                                MiFID/PRIIPS flags it came from. New issues not
                                yet in the extract are highlighted and marked.

It is driven by the freshly-generated GEM_List_Offshore.xlsx (so the bond set
and Restriction values are EXACTLY what was published this week — additions and
removals included) plus the loaded PRIIPS reference on the GEMData object (for
the source flags). Call build_restrictions_clean_xlsx() after the Excels are
built. See run_weekly.py.
"""

import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
HDR = "D9001D"; DARK = "333333"; GREY = "595959"
AMBER = "FFF2CC"; LIGHT = "F7F7F7"
_thin = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Offshore BondList layout (rows 1-7 are banner/title, row 8 = header, data from 9)
_HDR_ROW = 8
_COL_ISIN = 1          # B (0-based 1)
_COL_VALOR = 2
_COL_ISSUER = 3
_COL_RESTR = 22        # 'Restrictions**'


def _read_offshore(offshore_path):
    """Yield (isin, valor, issuer) for each bond on the published Offshore list.

    Only the bond *identity* is taken from the file — the Restriction value is
    recomputed from the authoritative flags below, so the clean workbook is
    correct even if read against a stale Offshore file."""
    wb = openpyxl.load_workbook(offshore_path, read_only=True, data_only=True)
    ws = wb["BondList"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[_HDR_ROW:]:
        isin = (str(r[_COL_ISIN]).strip() if r[_COL_ISIN] else "")
        valor = (str(r[_COL_VALOR]).strip() if r[_COL_VALOR] else "")
        if not isin and not valor:
            continue
        issuer = (str(r[_COL_ISSUER]).strip() if r[_COL_ISSUER] else "")
        yield isin, valor, issuer


def _lookup_flags(data, isin, valor):
    """Return (mifid, priips_relevant, priips_kid, found) from the loaded reference."""
    ref = getattr(data, "priips_ref", {}) or {}
    refv = getattr(data, "priips_ref_valor", {}) or {}
    rec = ref.get(isin)
    if rec is None and valor:
        v = valor
        try:
            v = str(int(float(valor)))
        except (ValueError, TypeError):
            pass
        rec = refv.get(v)
    if rec is None:
        return "", "", "", False
    mc, pr, kid = rec
    return (mc or ""), (pr or ""), (kid or ""), True


def _instructions_tab(s):
    s.sheet_view.showGridLines = False

    def put(r, c, val, font=None, fill=None, align="left", wrap=False, merge=None, h=None):
        cell = s.cell(row=r, column=c, value=val)
        cell.font = font or Font(name=FONT, size=10.5, color=DARK)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        if merge:
            s.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge)
        if h:
            s.row_dimensions[r].height = h
        return cell

    put(1, 1, "EM Bond List - PRIIPS / MiFID Classification File",
        Font(name=FONT, size=15, bold=True, color=HDR), merge=6)
    put(2, 1, "What the India team needs to provide, where it goes, and how it is used",
        Font(name=FONT, size=10.5, italic=True, color=GREY), merge=6)

    r = 4
    put(r, 1, "In one line", Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    put(r, 1, "Keep ONE up-to-date PRIIPS classification extract in the pipeline's \"data\" folder. "
              "The weekly run reads it automatically and stamps the Restrictions column on every bond. "
              "No manual editing of the bond list, no coding of values - just keep this file current.",
        wrap=True, merge=6, h=44); r += 2

    put(r, 1, "Where the file goes", Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    put(r, 1, "Place the file directly in the  data\\  folder - NOT inside data\\current or data\\previous.",
        wrap=True, merge=6); r += 1
    put(r, 1, "The runner picks up the newest file whose name contains \"PRIIPS\". You do NOT need to rename it; "
              "the long auto-generated export name is fine. If several are present, the most recent wins.",
        wrap=True, merge=6, h=30); r += 1
    put(r, 1, "Preferred format: .xlsx or .csv. (.xls also works if the Python environment has the 'xlrd' package.)",
        Font(name=FONT, size=10, italic=True, color=GREY), wrap=True, merge=6); r += 2

    put(r, 1, "Required columns (exact header names)", Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    put(r, 1, "The file must contain at least these five columns. Extra columns are ignored.", wrap=True, merge=6); r += 1
    colhdr = r
    for i, h in enumerate(["Column header", "What it holds", "Allowed values"]):
        c = s.cell(row=r, column=1 + i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR)
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = _BORDER
    r += 1
    coldefs = [
        ("U_ISIN", "Instrument ISIN - the join key", "e.g. XS2816006303"),
        ("VALOREN_NO", "Valoren number - fallback join key", "e.g. 135000666"),
        ("MIFIDII_COMPLEXITY", "MiFID II complexity classification",
         "\"Complex Instrument (M10C01)\" / \"Non Complex Instrument (M10C02)\" / blank"),
        ("PRIIPS_RELEVANT", "Whether the instrument is PRIIPS-relevant", "\"PRIIPS_RELEVANT\" or blank"),
        ("PRIIPS_KID", "Whether a PRIIPS KID exists", "\"Y\" or blank"),
    ]
    for name, desc, vals in coldefs:
        s.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, bold=True, color="C00000")
        s.cell(row=r, column=2, value=desc).font = Font(name=FONT, size=10, color=DARK)
        s.cell(row=r, column=3, value=vals).font = Font(name=FONT, size=9.5, color=GREY)
        for cc in range(1, 4):
            cell = s.cell(row=r, column=cc); cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT if (r - colhdr) % 2 else "FFFFFF")
        r += 1
    r += 1

    put(r, 1, "How the Restriction is decided (for reference - the code does this automatically)",
        Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    rules = [
        ("MIFIDII_COMPLEXITY = \"Complex Instrument (M10C01)\"", "->  shows  1"),
        ("MIFIDII_COMPLEXITY = \"Non Complex Instrument (M10C02)\"", "->  shows  (blank)"),
        ("MIFIDII_COMPLEXITY missing / unrecognised", "->  shows  n/a"),
        ("PRIIPS_RELEVANT  AND  PRIIPS_KID != \"Y\"", "->  adds  , 2"),
        ("PRIIPS_RELEVANT  AND  PRIIPS_KID = \"Y\"", "->  adds nothing"),
    ]
    for cond, res in rules:
        s.cell(row=r, column=1, value=cond).font = Font(name=FONT, size=10, color=DARK)
        s.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        s.cell(row=r, column=3, value=res).font = Font(name=FONT, size=10, bold=True, color="C00000")
        r += 1
    r += 1
    put(r, 1, "Meaning:  1 = MiFID complex   -   2 = PRIIPS-relevant, no KID   -   1, 2 = both   -   "
              "blank = neither   -   n/a = MiFID classification unavailable",
        Font(name=FONT, size=9.5, italic=True, color=GREY), wrap=True, merge=6, h=26); r += 2

    put(r, 1, "What happens if the file is missing or a bond is not in it",
        Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    put(r, 1, "- If the file is missing entirely, the run still completes but falls back to the old (less accurate) "
              "estimate and prints a warning - so please keep it current.", wrap=True, merge=6, h=30); r += 1
    put(r, 1, "- If a specific bond (usually a brand-new issue) is not yet in the file, it shows \"n/a\" and is "
              "listed in outputs\\priips_unmatched_report.csv so it can be added next time.",
        wrap=True, merge=6, h=30); r += 2

    put(r, 1, "Ownership", Font(name=FONT, size=12, bold=True, color=HDR)); r += 1
    put(r, 1, "The India team owns keeping this file accurate and current in the data\\ folder. Everything after "
              "that - reading it, applying the rule, and publishing the PDF and Excel - is fully automatic.",
        wrap=True, merge=6, h=30)

    for col, w in zip("ABCDEF", [40, 40, 34, 3, 3, 3]):
        s.column_dimensions[col].width = w


def build_restrictions_clean_xlsx(data, offshore_path, output_path):
    """Write the clean two-tab restrictions workbook. Returns the bond count."""
    # Use the generator's own rule so this file matches the published lists
    # exactly. (gem_report_builder_v3 is already imported by the weekly run.)
    from gem_report_builder_v3 import _restriction_from_flags, PRIIPS_UNMATCHED_RESTRICTION

    bonds = list(_read_offshore(offshore_path))

    wb = openpyxl.Workbook()
    _instructions_tab(wb.active)
    wb.active.title = "Read me - India team"

    d = wb.create_sheet("Restrictions (Offshore)")
    d.sheet_view.showGridLines = False
    today = datetime.date.today().isoformat()
    t = d.cell(row=1, column=1, value=f"EM Bond List - Restrictions (Offshore)  -  generated {today}")
    t.font = Font(name=FONT, size=12, bold=True, color=HDR); d.merge_cells("A1:H1")
    n = d.cell(row=2, column=1, value="Source of truth: PRIIPS classification extract joined by ISIN. "
                                      "'New issue' rows are not yet in the extract and show n/a.")
    n.font = Font(name=FONT, size=9, italic=True, color=GREY); d.merge_cells("A2:H2")

    cols = ["ISIN", "Valor", "Issuer", "Restriction", "MiFID Complexity (source)",
            "PRIIPS Relevant", "PRIIPS KID", "Status"]
    hr = 4
    for i, h in enumerate(cols):
        c = d.cell(row=hr, column=1 + i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR); c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d.freeze_panes = "A5"

    out_rows = []
    for isin, valor, issuer in bonds:
        mc, pr, kid, found = _lookup_flags(data, isin, valor)
        if found:
            restr = _restriction_from_flags(mc, pr, kid)
            status = "OK"
        else:
            restr = PRIIPS_UNMATCHED_RESTRICTION
            status = "New issue - not yet in extract"
        out_rows.append((isin, valor, issuer, restr, mc, pr, kid, status))

    rr = hr + 1
    for isin, valor, issuer, restr, mc, pr, kid, status in sorted(
            out_rows, key=lambda x: (x[7] != "OK", x[2].lower())):
        vals = [isin, valor, issuer, (restr or "(blank)"), mc, pr, kid, status]
        for i, v in enumerate(vals):
            cell = d.cell(row=rr, column=1 + i, value=v)
            cell.font = Font(name=FONT, size=9, color=DARK); cell.border = _BORDER
            cell.alignment = Alignment(horizontal=("left" if i in (2, 4, 5) else "center"), vertical="center")
        if status != "OK":
            for i in range(8):
                d.cell(row=rr, column=1 + i).fill = PatternFill("solid", fgColor=AMBER)
        d.cell(row=rr, column=4).font = Font(name=FONT, size=9, bold=True, color="C00000")
        rr += 1

    last = rr - 1
    if last >= hr + 1:
        d.auto_filter.ref = f"A{hr}:H{last}"
    for col, w in zip("ABCDEFGH", [15, 11, 32, 12, 26, 16, 11, 26]):
        d.column_dimensions[col].width = w

    wb.save(output_path)
    return len(out_rows)
