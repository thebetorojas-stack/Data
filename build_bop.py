"""
Build: Argentina balance of payments pack.
Source: INDEC balanza de pagos MBP6 (BPM6) via apis.datos.gob.ar, quarterly, USD mn.
USD GDP for the ratio tab comes from the companion GDP pack (INDEC current prices / Com. "A" 3500).
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import Reference, Series
from openpyxl.comments import Comment
from common_style import *

D = "/home/claude/data/"
bop = pd.read_csv(D + "bop_quarterly.csv", parse_dates=["date"])
cur = pd.read_csv(D + "gdp_current.csv", parse_dates=["date"])
fx = pd.read_csv(D + "fx_monthly.csv", parse_dates=["date"])
fx["q"] = fx["date"].dt.to_period("Q")
fxq = fx.groupby("q")["fx_avg"].mean()

# quarterly nominal GDP in USD mn (level is at annual rate -> /4), aligned to BoP dates
cur["q"] = cur["date"].dt.to_period("Q")
usd_gdp = {}
for _, row in cur.iterrows():
    p = row["q"]
    if p in fxq.index:
        usd_gdp[p] = (row["pib_n"] / 4.0) / fxq.loc[p]

SER = ["ca_total", "gs_total", "goods_bal", "goods_x", "goods_m", "serv_bal", "serv_cr",
       "serv_dr", "prim_inc", "prim_comp", "prim_invinc", "prim_fdi_inc", "prim_port_inc",
       "prim_other_inc", "sec_inc", "ka_total", "fa_total", "fdi_total", "fdi_a_equity",
       "fdi_a_debt", "fdi_l_equity", "fdi_l_debt", "port_total", "port_a_equity",
       "port_a_debt", "port_l_equity", "port_l_debt", "deriv", "oi_total", "oi_assets",
       "oi_liab_col", "reserves", "eo"]
COL = {k: 5 + i for i, k in enumerate(SER)}          # E .. AK
C_OILIAB, C_IDENT, C_FABUILD, C_RESINC, C_GDPQ, C_GDP4 = 38, 39, 40, 41, 42, 43

NQ = len(bop)
LASTQ = bop["date"].iloc[-1]
LASTQL = qlabel(LASTQ)
LAST_Y = LASTQ.year
FULL_YEARS = [y for y in range(2006, LAST_Y + 1)
              if (bop["date"].dt.year == y).sum() == 4]
LAST_FULL = FULL_YEARS[-1]
R0, HR = 6, 5
LASTROW = R0 + NQ - 1
DB = "'Data BoP'"
YR = f"{DB}!$C${R0}:$C${LASTROW}"

wb = Workbook()

# =====================================================================  READ ME
ws = wb.active
ws.title = "Read me"
sheet_base(ws)
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 106
title(ws, "B2", "Argentina — balance of payments and external funding")
put(ws, 3, 2, f"INDEC balanza de pagos, BPM6 basis, millions of US dollars. Data through {LASTQL}.",
    color=MUTED, italic=True)

readme = [
    ("What's in here", ""),
    ("Data BoP", "All 33 published BPM6 lines, quarterly from 2006. Blue = INDEC source. The identity and build-up "
                 "check columns must read zero — they are how you know nothing has been mistyped or mis-signed."),
    ("Current Account", "Every current-account line in three frames: quarterly, four-quarter rolling, and annual. "
                        "The rolling frame is the one to read — quarterly Argentine BoP is violently seasonal "
                        "(the harvest lands in Q2) and the annual frame is too slow to trade on."),
    ("Financial Account", "The same three frames for the financial account, split into direct, portfolio, "
                          "derivatives, other investment and reserve assets, each further split into what residents "
                          "did with assets and what non-residents did with liabilities."),
    ("Funding", "How the current account is actually funded. Reads down from the current account to the change in "
                "reserves, showing every source of financing in between. This is the tab that answers what builds "
                "and what dents reserves."),
    ("% of GDP", "The four-quarter rolling frame divided by four-quarter rolling USD GDP. Scale-free, so 2006 and "
                 "2026 are comparable."),
    ("Reserves", "The reserve-asset line on its own: quarterly, rolling, annual, and cumulated since 2006."),
    ("Forecast", "Bull / base / bear current-account paths for 2026 and 2027, built up from the components, with "
                 "the assumptions behind each and the IMF projection as a benchmark."),
    ("", ""),
    ("Sign conventions — read this before using any number", ""),
    ("Financial account", "BPM6 net lending(+) / net borrowing(-). The financial account equals net acquisition of "
                          "financial assets MINUS net incurrence of liabilities, and reserve assets sit inside it. "
                          "So a NEGATIVE financial account is a net capital INFLOW. This trips people up constantly; "
                          "the Funding tab restates everything in inflow-positive terms so you never have to hold it "
                          "in your head."),
    ("Reserve assets", "POSITIVE = reserves BUILT. Verified against the known years: 2019 -21,375 and 2023 -21,675 "
                       "(the two big loss years), 2024 +6,093, 2025 +7,221."),
    ("The identity", "Current account + capital account - financial account + errors and omissions = 0. It holds to "
                     "machine precision in 77 of 81 quarters; the four 2016 quarters carry a residual of up to 2.1 "
                     "that is internal to INDEC's own vintage."),
    ("A trap in the source data", "The other-investment liabilities series is stored with its sign FLIPPED from "
                                  "Q1-2017 onward, but in natural sign before that. Column AL of the Data tab "
                                  "corrects for it with an explicit date test, and everything downstream uses the "
                                  "corrected column. If you pull this series yourself and assume one sign "
                                  "convention, you will silently corrupt 2006-2016."),
    ("", ""),
    ("Two gaps in what INDEC publishes", ""),
    ("Investment income", "Investment income does not decompose into direct + portfolio + other. A fourth component "
                          "— income earned ON reserve assets — is in the total but not published separately. It "
                          "averages about +88mn a quarter and is always a credit. Column AN backs it out as the "
                          "residual so the decomposition closes."),
    ("Sector detail", "INDEC's API does not carry other investment split by sector (central bank, government, banks, "
                      "other) or by instrument (currency and deposits, loans, trade credit). That detail exists only "
                      "in the informe tecnico spreadsheets. If you need to see which sector borrowed, that is the "
                      "source, not this."),
    ("", ""),
    ("Source", "INDEC, Balanza de pagos, posicion de inversion internacional y deuda externa, MBP6, quarterly, "
               "retrieved from the official series API at apis.datos.gob.ar, dataset 160 (quarterly family 160.2). "
               "USD GDP for the ratio tab is INDEC current-price GDP converted at the BCRA Comunicacion \"A\" 3500 "
               "wholesale rate, period average — the same series used in the companion GDP pack."),
    ("Validation", "Annual current account reproduces INDEC's published figures for 2023 (-22,355), 2024 (+5,891) "
                   "and 2025 (-7,788), and every quarterly print from Q1-2025 to Q1-2026, to within rounding. "
                   "Reserve assets reproduce 2019, 2023, 2024 and 2025 exactly. One benchmark did NOT reproduce: "
                   "2021 computes to 6,624.8 against a published 6,645 — INDEC has since revised 2021 down by 20.2, "
                   "so the live series is right and the old print is stale."),
    ("Provisional data", "INDEC marks everything from 2022 onward as provisional and has revised it more than once. "
                         "Do not treat the last four years as settled."),
]
r = 5
for k, v in readme:
    if v == "":
        section(ws, r, 2, k)
    else:
        put(ws, r, 2, k, bold=True, size=10)
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(ARIAL, 10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 14 * (1 + len(v) // 98)
    r += 1

# =====================================================================  DATA
ws = wb.create_sheet("Data BoP")
sheet_base(ws)
title(ws, "B2", "INDEC balance of payments, BPM6 — quarterly source data",
      "Millions of US dollars. Blue = INDEC source value. Black = formula. Check columns must read zero.")
heads = [
    ("Quarter", 10), ("Year", 7), ("Q", 5),
    ("CURRENT\nACCOUNT", 13), ("Goods &\nservices", 12), ("Goods\nbalance", 12),
    ("Goods\nexports FOB", 13), ("Goods\nimports FOB", 13), ("Services\nbalance", 12),
    ("Services\ncredits", 12), ("Services\ndebits", 12), ("Primary\nincome", 12),
    ("of which:\ncompensation", 13), ("Investment\nincome", 12), ("Income on\ndirect inv.", 12),
    ("Income on\nportfolio", 12), ("Income on\nother inv.", 12), ("Secondary\nincome", 12),
    ("CAPITAL\nACCOUNT", 12), ("FINANCIAL\nACCOUNT", 13),
    ("Direct inv.\ntotal", 12), ("DI assets\nequity", 12), ("DI assets\ndebt", 12),
    ("DI liab.\nequity", 12), ("DI liab.\ndebt", 12),
    ("Portfolio\ntotal", 12), ("PI assets\nequity", 12), ("PI assets\ndebt", 12),
    ("PI liab.\nequity", 12), ("PI liab.\ndebt", 12), ("Financial\nderivatives", 12),
    ("Other inv.\ntotal", 12), ("OI assets", 12), ("OI liab.\n(as published)", 14),
    ("RESERVE\nASSETS", 12), ("Errors &\nomissions", 12),
    ("OI liab., sign\ncorrected", 14), ("Identity\ncheck", 11), ("Fin. acct\nbuild check", 12),
    ("Income on\nreserves (resid.)", 14), ("Nominal GDP\n(USD mn, qtr)", 14),
    ("4q rolling GDP\n(USD mn)", 14)]
for i, (t, w) in enumerate(heads):
    hdr(ws, HR, 2 + i, t, w, wrap=True)
ws.row_dimensions[HR].height = 34

for i in range(NQ):
    r = R0 + i
    d = bop["date"].iloc[i]
    put(ws, r, 2, qlabel(d))
    put(ws, r, 3, d.year, "0", color=MUTED)
    put(ws, r, 4, (d.month - 1) // 3 + 1, "0", color=MUTED)
    for k in SER:
        v = bop[k].iloc[i]
        put(ws, r, COL[k], None if pd.isna(v) else float(v), NUM0, color=INPUTC)
    ail = gcl(COL["oi_liab_col"])
    put(ws, r, C_OILIAB, f"=IF($C{r}<2017,{ail}{r},-{ail}{r})", NUM0)
    put(ws, r, C_IDENT,
        f"={gcl(COL['ca_total'])}{r}+{gcl(COL['ka_total'])}{r}-{gcl(COL['fa_total'])}{r}"
        f"+{gcl(COL['eo'])}{r}", NUM2, color=MUTED)
    put(ws, r, C_FABUILD,
        f"={gcl(COL['fdi_total'])}{r}+{gcl(COL['port_total'])}{r}+{gcl(COL['deriv'])}{r}"
        f"+{gcl(COL['oi_total'])}{r}+{gcl(COL['reserves'])}{r}-{gcl(COL['fa_total'])}{r}",
        NUM2, color=MUTED)
    put(ws, r, C_RESINC,
        f"={gcl(COL['prim_invinc'])}{r}-{gcl(COL['prim_fdi_inc'])}{r}"
        f"-{gcl(COL['prim_port_inc'])}{r}-{gcl(COL['prim_other_inc'])}{r}", NUM0)
    p = d.to_period("Q")
    put(ws, r, C_GDPQ, float(usd_gdp[p]) if p in usd_gdp else None, NUM0, color=INPUTC)
    if i >= 3:
        put(ws, r, C_GDP4, f"=SUM({gcl(C_GDPQ)}{r-3}:{gcl(C_GDPQ)}{r})", NUM0)
ws.freeze_panes = f"E{R0}"
note(ws, LASTROW + 2, 2,
     "Identity check = current account + capital account - financial account + errors and omissions. Zero except "
     "the four 2016 quarters (residual up to 2.1), which is internal to INDEC's vintage.")
note(ws, LASTROW + 3, 2,
     "Other-investment liabilities are published in natural sign to Q4-2016 and sign-flipped from Q1-2017. Column AL "
     "corrects for that; use AL, never AI, in any calculation. In AL a POSITIVE value is a net incurrence of "
     "liabilities, i.e. an inflow.")
note(ws, LASTROW + 4, 2,
     "Income on reserve assets is not published separately; it is the residual that makes investment income close. "
     "Always a credit, averaging about +88mn a quarter.")
note(ws, LASTROW + 5, 2,
     "Nominal GDP in USD: INDEC current-price GDP (published at annual rates, so divided by four for a true "
     "quarterly flow) converted at the period-average Com. \"A\" 3500 wholesale rate. Same construction as the "
     "companion GDP pack.")
ws["B2"].comment = Comment(
    "INDEC, balanza de pagos MBP6, quarterly, millions of USD, from apis.datos.gob.ar dataset 160.2. "
    "Provisional from 2022 onward.", "Claude")


# =====================================================================  helper
def frame_sheet(name, title_text, subtitle, items, chart_specs=None, pct_of_gdp=False):
    """items: list of (label, formula_template, is_total). Template uses {r} and {DB}."""
    w = wb.create_sheet(name)
    sheet_base(w)
    title(w, "B2", title_text, subtitle)
    nrows = len(items)

    blocks = [("Quarterly", 0), ("Four-quarter rolling", 1)]
    col0 = 2
    for bi, (bname, kind) in enumerate(blocks):
        base_col = col0 + bi * (nrows + 2)
        section(w, 4, base_col, bname)
        hdr(w, 5, base_col, "Quarter", 10)
        for j, (lab, _, _) in enumerate(items):
            hdr(w, 5, base_col + 1 + j, lab, 13, wrap=True)
        w.row_dimensions[5].height = 40
        for i in range(NQ):
            r = R0 + i
            put(w, r, base_col, f"={DB}!B{r}", color=LINKC)
            for j, (lab, tmpl, is_tot) in enumerate(items):
                if kind == 0:
                    f = "=" + tmpl.format(r=r)
                else:
                    if i < 3:
                        continue
                    parts = [f"({tmpl.format(r=rr)})" for rr in range(r - 3, r + 1)]
                    f = "=" + "+".join(parts)
                put(w, r, base_col + 1 + j, f, NUM0,
                    bold=is_tot, color=ACCENT if is_tot else INK)
    # annual block
    abase = col0
    ar = LASTROW + 3
    section(w, ar, abase, "Annual")
    hdr(w, ar + 1, abase, "Year", 10)
    for j, (lab, _, _) in enumerate(items):
        hdr(w, ar + 1, abase + 1 + j, lab, 13, wrap=True)
    w.row_dimensions[ar + 1].height = 40
    for i, y in enumerate(FULL_YEARS):
        r = ar + 2 + i
        put(w, r, abase, y, "0", bold=True)
        for j, (lab, tmpl, is_tot) in enumerate(items):
            f = "=" + tmpl.format(r="{r}")
            # build a SUMIFS-equivalent by summing the four quarters via SUMIFS on each column
            f = "=" + sumifs_of(tmpl, r)
            put(w, r, abase + 1 + j, f, NUM0, bold=is_tot,
                color=ACCENT if is_tot else INK)
    return w, ar + 2 + len(FULL_YEARS) - 1


def sumifs_of(tmpl, arow):
    """Turn a per-row formula over Data BoP columns into an annual SUMIFS expression."""
    import re
    out = tmpl.format(r="@@")
    def rep(m):
        col = m.group(1)
        return (f"SUMIFS({DB}!${col}${R0}:${col}${LASTROW},{YR},$B{arow})")
    return re.sub(r"'Data BoP'!\$?([A-Z]{1,2})@@", rep, out)


DBC = lambda k: f"{DB}!{gcl(COL[k])}{{r}}"
DBX = lambda c: f"{DB}!{gcl(c)}{{r}}"

# =====================================================================  CURRENT ACCOUNT
ca_items = [
    ("CURRENT ACCOUNT", DBC("ca_total"), True),
    ("Goods balance", DBC("goods_bal"), False),
    ("  Exports FOB", DBC("goods_x"), False),
    ("  Imports FOB", DBC("goods_m"), False),
    ("Services balance", DBC("serv_bal"), False),
    ("  Services credits", DBC("serv_cr"), False),
    ("  Services debits", DBC("serv_dr"), False),
    ("Primary income", DBC("prim_inc"), False),
    ("  Compensation", DBC("prim_comp"), False),
    ("  Investment income", DBC("prim_invinc"), False),
    ("    on direct inv.", DBC("prim_fdi_inc"), False),
    ("    on portfolio", DBC("prim_port_inc"), False),
    ("    on other inv.", DBC("prim_other_inc"), False),
    ("    on reserves", DBX(C_RESINC), False),
    ("Secondary income", DBC("sec_inc"), False),
    ("Capital account", DBC("ka_total"), False),
]
wsca, ca_alast = frame_sheet(
    "Current Account",
    "Current account — quarterly, four-quarter rolling and annual",
    "Millions of US dollars. Positive = credit / surplus. The rolling frame is the one to read: Argentine quarterly "
    "BoP is dominated by the Q2 harvest.",
    ca_items)

r15 = R0 + list(bop["date"].dt.year).index(2013)
roll0 = 2 + (len(ca_items) + 2)
ch = line_chart(34, 14, "Current account and its main components — four-quarter rolling (USD mn)", "#,##0")
cats = Reference(wsca, min_col=roll0, min_row=r15, max_row=LASTROW)
for off, nm, colr, wdt in [(1, "Current account", "1A1A1A", 32000),
                           (2, "Goods balance", "2F5C8F", 24000),
                           (5, "Services balance", "E8A33D", 22000),
                           (8, "Primary income", "D1495B", 22000),
                           (15, "Secondary income", "8AA0B8", 18000)]:
    ref = Reference(wsca, min_col=roll0 + off, min_row=5, max_row=LASTROW)
    s = Series(ref, title=nm)
    style_series(s, colr, wdt)
    ch.series.append(s)
ch.set_categories(cats)
wsca.add_chart(ch, f"B{ca_alast + 4}")

# =====================================================================  FINANCIAL ACCOUNT
fa_items = [
    ("FINANCIAL ACCOUNT", DBC("fa_total"), True),
    ("Direct investment, net", DBC("fdi_total"), False),
    ("  DI assets (equity)", DBC("fdi_a_equity"), False),
    ("  DI assets (debt)", DBC("fdi_a_debt"), False),
    ("  DI liabilities (equity)", DBC("fdi_l_equity"), False),
    ("  DI liabilities (debt)", DBC("fdi_l_debt"), False),
    ("Portfolio investment, net", DBC("port_total"), False),
    ("  PI assets (equity)", DBC("port_a_equity"), False),
    ("  PI assets (debt)", DBC("port_a_debt"), False),
    ("  PI liabilities (equity)", DBC("port_l_equity"), False),
    ("  PI liabilities (debt)", DBC("port_l_debt"), False),
    ("Financial derivatives", DBC("deriv"), False),
    ("Other investment, net", DBC("oi_total"), False),
    ("  OI assets", DBC("oi_assets"), False),
    ("  OI liabilities (corrected)", DBX(C_OILIAB), False),
    ("RESERVE ASSETS", DBC("reserves"), True),
    ("Errors & omissions", DBC("eo"), False),
]
wsfa, fa_alast = frame_sheet(
    "Financial Account",
    "Financial account — quarterly, four-quarter rolling and annual",
    "Millions of US dollars, BPM6 net lending(+) / net borrowing(-). A NEGATIVE total is a net capital INFLOW. "
    "Liabilities are shown in natural sign: positive = non-residents put money in.",
    fa_items)

roll0f = 2 + (len(fa_items) + 2)
ch = line_chart(34, 14, "Financial account by instrument — four-quarter rolling (USD mn)", "#,##0")
cats = Reference(wsfa, min_col=roll0f, min_row=r15, max_row=LASTROW)
for off, nm, colr, wdt in [(1, "Financial account", "1A1A1A", 32000),
                           (2, "Direct investment", "2F5C8F", 22000),
                           (7, "Portfolio investment", "D1495B", 22000),
                           (13, "Other investment", "E8A33D", 22000),
                           (16, "Reserve assets", "4F9D69", 26000)]:
    ref = Reference(wsfa, min_col=roll0f + off, min_row=5, max_row=LASTROW)
    s = Series(ref, title=nm)
    style_series(s, colr, wdt)
    ch.series.append(s)
ch.set_categories(cats)
wsfa.add_chart(ch, f"B{fa_alast + 4}")

# =====================================================================  FUNDING
ws = wb.create_sheet("Funding")
sheet_base(ws)
title(ws, "B2", "How the current account is funded — and what that leaves for reserves",
      "Four-quarter rolling, millions of US dollars. Everything is restated INFLOW-POSITIVE: a positive number adds "
      "to Argentina's dollars, a negative number takes them away.")
put(ws, 4, 2,
    "Read it down. The current account plus the capital account is the external balance. Non-residents putting money "
    "in adds to it; residents taking money out subtracts. Errors and omissions is unrecorded flow — in Argentina it "
    "is large and usually a leak. What is left is the change in reserve assets, which is the bottom line and ties "
    "exactly to INDEC's published reserve-asset figure.", color=MUTED, italic=True, size=9)

fund_items = [
    ("Current account", DBC("ca_total"), 0),
    ("Capital account", DBC("ka_total"), 0),
    ("Total: external balance", f"{DBC('ca_total')}+{DBC('ka_total')}", 1),
    ("FDI into Argentina", f"{DBC('fdi_l_equity')}+{DBC('fdi_l_debt')}", 0),
    ("Portfolio into Argentina", f"{DBC('port_l_equity')}+{DBC('port_l_debt')}", 0),
    ("Other inv. borrowing", DBX(C_OILIAB), 0),
    ("Total: gross inflows (liabilities)",
     f"{DBC('fdi_l_equity')}+{DBC('fdi_l_debt')}+{DBC('port_l_equity')}+{DBC('port_l_debt')}"
     f"+{DBX(C_OILIAB)}", 1),
    ("Residents' FDI abroad", f"-({DBC('fdi_a_equity')}+{DBC('fdi_a_debt')})", 0),
    ("Residents' portfolio abroad", f"-({DBC('port_a_equity')}+{DBC('port_a_debt')})", 0),
    ("Residents' other assets", f"-{DBC('oi_assets')}", 0),
    ("Total: gross outflows (assets)",
     f"-({DBC('fdi_a_equity')}+{DBC('fdi_a_debt')}+{DBC('port_a_equity')}+{DBC('port_a_debt')}"
     f"+{DBC('oi_assets')})", 1),
    ("Financial derivatives", f"-{DBC('deriv')}", 0),
    ("Total: net capital inflow",
     f"-({DBC('fa_total')}-{DBC('reserves')})", 1),
    ("Errors & omissions", DBC("eo"), 0),
    ("CHANGE IN RESERVES", DBC("reserves"), 2),
]
FHR, FR0 = 6, 7
hdr(ws, FHR, 2, "Quarter", 10)
for j, (lab, _, _) in enumerate(fund_items):
    hdr(ws, FHR, 3 + j, lab, 14, wrap=True)
ws.row_dimensions[FHR].height = 44
for i in range(NQ):
    r = FR0 + i
    src = R0 + i
    put(ws, r, 2, f"={DB}!B{src}", color=LINKC)
    if i < 3:
        continue
    for j, (lab, tmpl, lvl) in enumerate(fund_items):
        parts = [f"({tmpl.format(r=rr)})" for rr in range(src - 3, src + 1)]
        put(ws, r, 3 + j, "=" + "+".join(parts), NUM0,
            bold=(lvl > 0), color=ACCENT if lvl == 1 else ("D1495B" if lvl == 2 else INK))
FUND_LAST = FR0 + NQ - 1
ws.freeze_panes = f"C{FR0}"
note(ws, FUND_LAST + 2, 2,
     "Check: external balance + net capital inflow + errors and omissions = change in reserves. This is the BPM6 "
     "identity rearranged, so it holds by construction — if it ever does not, a source cell has been edited.")
note(ws, FUND_LAST + 3, 2,
     "'Other investment borrowing' is where IMF disbursements land. The April-2025 EFF disbursement is the "
     "single largest positive entry in that column in the whole history.")
note(ws, FUND_LAST + 4, 2,
     "Errors and omissions in Argentina is not noise to be ignored. It is persistently negative and large, and is "
     "conventionally read as unrecorded capital flight. Treat it as a funding line, which is why it sits here rather "
     "than in a footnote.")

ch = bar_chart(36, 15, "What builds and what dents reserves — four-quarter rolling (USD mn)", "#,##0", stacked=True)
r13 = FR0 + list(bop["date"].dt.year).index(2013)
for off, nm, colr in [(2, "External balance", "2F5C8F"), (7, "Gross inflows", "4F9D69"),
                      (11, "Gross outflows", "D1495B"), (12, "Derivatives", "BFC9D4"),
                      (14, "Errors & omissions", "E8A33D")]:
    ref = Reference(ws, min_col=3 + off, min_row=FHR, max_row=FUND_LAST)
    s = Series(ref, title_from_data=True)
    s.graphicalProperties.solidFill = colr
    s.graphicalProperties.line.noFill = True
    ch.series.append(s)
ch.set_categories(Reference(ws, min_col=2, min_row=FR0, max_row=FUND_LAST))
ws.add_chart(ch, f"B{FUND_LAST + 6}")

lc = line_chart(36, 13, "Change in reserve assets, four-quarter rolling (USD mn)", "#,##0")
ref = Reference(ws, min_col=3 + 14, min_row=FHR, max_row=FUND_LAST)
s = Series(ref, title_from_data=True)
style_series(s, "D1495B", 30000)
lc.series.append(s)
lc.set_categories(Reference(ws, min_col=2, min_row=FR0, max_row=FUND_LAST))
ws.add_chart(lc, f"B{FUND_LAST + 38}")

# =====================================================================  % OF GDP
ws = wb.create_sheet("% of GDP")
sheet_base(ws)
title(ws, "B2", "Four-quarter rolling, as a share of GDP",
      "Rolling four-quarter flows divided by rolling four-quarter nominal GDP in US dollars. Scale-free, so any two "
      "eras are directly comparable.")
gdp_items = [
    ("Current account", DBC("ca_total")),
    ("Goods balance", DBC("goods_bal")),
    ("Services balance", DBC("serv_bal")),
    ("Primary income", DBC("prim_inc")),
    ("Secondary income", DBC("sec_inc")),
    ("Financial account", DBC("fa_total")),
    ("FDI liabilities", f"{DBC('fdi_l_equity')}+{DBC('fdi_l_debt')}"),
    ("Portfolio liabilities", f"{DBC('port_l_equity')}+{DBC('port_l_debt')}"),
    ("Other inv. liabilities", DBX(C_OILIAB)),
    ("Resident asset accumulation",
     f"{DBC('fdi_a_equity')}+{DBC('fdi_a_debt')}+{DBC('port_a_equity')}+{DBC('port_a_debt')}"
     f"+{DBC('oi_assets')}"),
    ("Errors & omissions", DBC("eo")),
    ("Change in reserves", DBC("reserves")),
]
GHR, GR0 = 6, 7
hdr(ws, GHR, 2, "Quarter", 10)
for j, (lab, _) in enumerate(gdp_items):
    hdr(ws, GHR, 3 + j, lab, 14, wrap=True)
hdr(ws, GHR, 3 + len(gdp_items), "memo: 4q GDP\n(USD mn)", 14, wrap=True)
ws.row_dimensions[GHR].height = 44
for i in range(NQ):
    r = GR0 + i
    src = R0 + i
    put(ws, r, 2, f"={DB}!B{src}", color=LINKC)
    if i < 3:
        continue
    den = f"{DB}!${gcl(C_GDP4)}{src}"
    for j, (lab, tmpl) in enumerate(gdp_items):
        parts = [f"({tmpl.format(r=rr)})" for rr in range(src - 3, src + 1)]
        put(ws, r, 3 + j, f"=IF({den}=0,\"\",({'+'.join(parts)})/{den})", PCT)
    put(ws, r, 3 + len(gdp_items), f"={den}", NUM0, color=LINKC)
GLAST = GR0 + NQ - 1
ws.freeze_panes = f"C{GR0}"
note(ws, GLAST + 2, 2,
     "Denominator is nominal GDP converted at the official Com. \"A\" 3500 rate. Under a managed band that rate is "
     "not a market-clearing one, so the USD GDP level — and therefore every ratio here — moves with the regime as "
     "well as with the economy. In periods with a wide parallel gap (2012-2015, 2020-2023) these ratios understate "
     "the true burden.")

ch = line_chart(34, 14, "Current account and reserve change, % of GDP (4q rolling)", "0.0%")
r13 = GR0 + list(bop["date"].dt.year).index(2013)
for off, nm, colr, wdt in [(0, "Current account", "1A1A1A", 30000),
                           (1, "Goods balance", "2F5C8F", 22000),
                           (3, "Primary income", "D1495B", 22000),
                           (11, "Change in reserves", "4F9D69", 24000)]:
    ref = Reference(ws, min_col=3 + off, min_row=GHR, max_row=GLAST)
    s = Series(ref, title_from_data=True)
    style_series(s, colr, wdt)
    ch.series.append(s)
ch.set_categories(Reference(ws, min_col=2, min_row=GR0, max_row=GLAST))
ws.add_chart(ch, f"B{GLAST + 5}")

# =====================================================================  RESERVES
ws = wb.create_sheet("Reserves")
sheet_base(ws)
title(ws, "B2", "Reserve assets — the BPM6 transaction line",
      "This is the change in reserves attributable to transactions, which is what the balance of payments explains. "
      "It is not the change in the headline stock: valuation, gold price and cross-rate moves are excluded.")
RHR, RR0 = 6, 7
hdr(ws, RHR, 2, "Quarter", 10)
for j, lab in enumerate(["Quarterly", "4q rolling", "Cumulative\nsince 2006", "% of GDP\n(4q rolling)"]):
    hdr(ws, RHR, 3 + j, lab, 14, wrap=True)
ws.row_dimensions[RHR].height = 34
rescol = gcl(COL["reserves"])
for i in range(NQ):
    r = RR0 + i
    src = R0 + i
    put(ws, r, 2, f"={DB}!B{src}", color=LINKC)
    put(ws, r, 3, f"={DB}!{rescol}{src}", NUM0)
    if i >= 3:
        put(ws, r, 4, f"=SUM({DB}!{rescol}{src-3}:{DB}!{rescol}{src})", NUM0, bold=True, color=ACCENT)
        put(ws, r, 6, f"=IF({DB}!${gcl(C_GDP4)}{src}=0,\"\",D{r}/{DB}!${gcl(C_GDP4)}{src})", PCT)
    put(ws, r, 5, f"=SUM({DB}!${rescol}${R0}:{DB}!${rescol}{src})", NUM0)
RLAST = RR0 + NQ - 1

ar = RLAST + 3
section(ws, ar, 2, "Annual")
hdr(ws, ar + 1, 2, "Year", 10)
for j, lab in enumerate(["Reserve assets\n(USD mn)", "% of GDP", "Current account\n(USD mn)",
                        "Net capital inflow\n(USD mn)", "Errors & omissions\n(USD mn)"]):
    hdr(ws, ar + 1, 3 + j, lab, 16, wrap=True)
ws.row_dimensions[ar + 1].height = 34
for i, y in enumerate(FULL_YEARS):
    r = ar + 2 + i
    put(ws, r, 2, y, "0", bold=True)
    put(ws, r, 3, f"=SUMIFS({DB}!${rescol}${R0}:${rescol}${LASTROW},{YR},$B{r})", NUM0,
        bold=True, color=ACCENT)
    put(ws, r, 4, f"=C{r}/SUMIFS({DB}!${gcl(C_GDPQ)}${R0}:${gcl(C_GDPQ)}${LASTROW},{YR},$B{r})", PCT)
    put(ws, r, 5, f"=SUMIFS({DB}!${gcl(COL['ca_total'])}${R0}:${gcl(COL['ca_total'])}${LASTROW},{YR},$B{r})", NUM0)
    put(ws, r, 6,
        f"=-(SUMIFS({DB}!${gcl(COL['fa_total'])}${R0}:${gcl(COL['fa_total'])}${LASTROW},{YR},$B{r})-C{r})", NUM0)
    put(ws, r, 7, f"=SUMIFS({DB}!${gcl(COL['eo'])}${R0}:${gcl(COL['eo'])}${LASTROW},{YR},$B{r})", NUM0)
RALAST = ar + 2 + len(FULL_YEARS) - 1
note(ws, RALAST + 2, 2,
     "Reserve assets here are transaction flows on the BPM6 definition. The BCRA's published gross reserve stock "
     "moves for other reasons too — valuation, the gold price, cross rates, and swap lines that are on the balance "
     "sheet but not freely usable. Reconciling this line to the headline stock needs BCRA's balance cambiario, which "
     "is not in this pack.")
note(ws, RALAST + 3, 2,
     "Validated against INDEC's published annual figures: 2019 -21,375; 2023 -21,675; 2024 +6,093; 2025 +7,221.")

bc = bar_chart(30, 12, "Reserve assets by year (USD mn, BPM6 transactions)", "#,##0")
bc.add_data(Reference(ws, min_col=3, min_row=ar + 1, max_row=RALAST), titles_from_data=True)
bc.set_categories(Reference(ws, min_col=2, min_row=ar + 2, max_row=RALAST))
bc.series[0].graphicalProperties.solidFill = "2F5C8F"
ws.add_chart(bc, f"I{ar}")

# =====================================================================  FORECAST
ws = wb.create_sheet("Forecast")
sheet_base(ws)
title(ws, "B2", "Current account forecast — bull / base / bear, 2026 and 2027",
      "Built up from the components in millions of US dollars. Blue = judgement input. Black = formula. "
      "'Bull' means a stronger external position, i.e. a better current account.")
put(ws, 4, 2,
    "2025 and the first quarter of 2026 are actuals pulled from the Data tab. The scenario columns are full-year "
    "figures, so the 2026 columns already contain the realised Q1 — check the memo row at the bottom, which shows "
    "what the remaining three quarters have to deliver for each scenario to land.",
    color=MUTED, italic=True, size=9)

FHR2, FR2 = 7, 8
hdr(ws, FHR2, 2, "US$ mn", 24)
hdr(ws, FHR2, 3, f"{LAST_FULL}\nactual", 12, wrap=True)
hdr(ws, FHR2, 4, "2026 Q1\nactual", 12, wrap=True)
for j, nm in enumerate(["2026 Bull", "2026 Base", "2026 Bear", "2027 Bull", "2027 Base", "2027 Bear"]):
    hdr(ws, FHR2, 5 + j, nm, 12, wrap=True)
ws.row_dimensions[FHR2].height = 34

# (label, data-tab column key or None, bull26, base26, bear26, bull27, base27, bear27)
lines = [
    ("Goods exports FOB", "goods_x", 103000, 98000, 94500, 118000, 109000, 101000),
    ("Goods imports FOB", "goods_m", 76000, 79000, 81000, 84000, 89000, 89000),
    ("Services credits", "serv_cr", 21500, 21000, 20500, 23500, 22500, 21500),
    ("Services debits", "serv_dr", 32000, 34500, 35500, 33500, 36500, 38000),
    ("Primary income", "prim_inc", -15500, -17000, -18500, -16500, -18500, -20000),
    ("Secondary income", "sec_inc", 2900, 2800, 2700, 3000, 2900, 2800),
]
for i, (lab, key, *vals) in enumerate(lines):
    r = FR2 + i
    put(ws, r, 2, lab, bold=False)
    put(ws, r, 3,
        f"=SUMIFS({DB}!${gcl(COL[key])}${R0}:${gcl(COL[key])}${LASTROW},{YR},{LAST_FULL})",
        NUM0, color=LINKC)
    put(ws, r, 4, f"={DB}!{gcl(COL[key])}{LASTROW}", NUM0, color=LINKC)
    for j, v in enumerate(vals):
        put(ws, r, 5 + j, v, NUM0, color=INPUTC)
GB = FR2 + len(lines)
put(ws, GB, 2, "Goods balance", bold=True)
put(ws, GB + 1, 2, "Services balance", bold=True)
put(ws, GB + 2, 2, "CURRENT ACCOUNT", bold=True, size=11)
put(ws, GB + 3, 2, "  % of GDP", bold=True)
put(ws, GB + 4, 2, "Nominal GDP (USD bn)")
for c in range(3, 11):
    L = gcl(c)
    put(ws, GB, c, f"={L}{FR2}-{L}{FR2+1}", NUM0, bold=True)
    put(ws, GB + 1, c, f"={L}{FR2+2}-{L}{FR2+3}", NUM0, bold=True)
    put(ws, GB + 2, c, f"={L}{GB}+{L}{GB+1}+{L}{FR2+4}+{L}{FR2+5}", NUM0,
        bold=True, size=11, color=ACCENT)
    put(ws, GB + 3, c, f"=IF({L}{GB+4}=0,\"\",{L}{GB+2}/({L}{GB+4}*1000))", PCT)
put(ws, GB + 4, 3,
    f"=SUMIFS({DB}!${gcl(C_GDPQ)}${R0}:${gcl(C_GDPQ)}${LASTROW},{YR},{LAST_FULL})/1000",
    NUM1, color=LINKC)
put(ws, GB + 4, 4, f"={DB}!{gcl(C_GDPQ)}{LASTROW}/1000", NUM1, color=LINKC)
for j, v in enumerate([730, 710, 665, 780, 750, 690]):
    put(ws, GB + 4, 5 + j, v, NUM1, color=INPUTC)

# memo: implied remainder of 2026
mr = GB + 6
section(ws, mr, 2, "Memo")
put(ws, mr + 1, 2, "Implied Q2-Q4 2026 current account")
for j in range(3):
    L = gcl(5 + j)
    put(ws, mr + 1, 5 + j, f"={L}{GB+2}-$D${GB+2}", NUM0, color=MUTED)
put(ws, mr + 2, 2, "IMF projection (Apr-2026 WEO / Art. IV)")
put(ws, mr + 2, 6, -5435, NUM0, color=INPUTC)
put(ws, mr + 2, 9, -4070, NUM0, color=INPUTC)
note(ws, mr + 3, 2,
     "IMF World Economic Outlook of April 2026 and the 2026 Article IV / Second Review of the Extended Arrangement, "
     "which both carry a current account of -US$5.4bn for 2026 (-0.8% of GDP) and -US$4.1bn for 2027 (-0.6%). "
     "The BCRA REM does not survey the current account, so there is no market consensus line to show; the REM's "
     "June-2026 trade balance median of +US$23.6bn on a customs basis is the nearest thing, and it is not directly "
     "comparable with the BoP goods balance.")

asr = mr + 6
section(ws, asr, 2, "Scenario assumptions")
hdr(ws, asr + 1, 2, "Driver", 24)
for j, nm in enumerate(["Bull (stronger external)", "Base", "Bear (weaker external)"]):
    hdr(ws, asr + 1, 3 + j, nm, 40, wrap=True)
assump = [
    ("Energy and mining exports",
     "Vaca Muerta liquids and the pipeline build push net energy exports past US$12bn by 2027; Rio Tinto / lithium "
     "volumes add on top",
     "Net energy exports build steadily toward US$9-10bn by 2027 — the volume is real but infrastructure paces it",
     "Pipeline and export-terminal slippage caps energy at current levels; a soft global oil price does the rest"),
    ("Agriculture",
     "Two good harvests and no export-tax reversal; farmer selling stays current",
     "Normal harvest, no drought, retenciones broadly where they are",
     "A La Nina drought year — the 2018 and 2023 template, worth US$15-20bn of exports when it lands"),
    ("Imports",
     "Import growth lags activity as the real exchange rate normalises",
     "Imports grow roughly with domestic demand; no new restrictions and none needed",
     "A strong real exchange rate plus recovering activity pulls imports up faster than exports"),
    ("Services",
     "The tourism deficit narrows as the real exchange rate corrects; knowledge-based exports keep growing",
     "Tourism stays a large and persistent deficit — Argentines travelling out is the single biggest services line "
     "and it responds to a strong peso",
     "An overvalued peso keeps outbound tourism at record levels; the services deficit widens further"),
    ("Primary income",
     "Falling country risk cuts the interest bill; profit remittances stay modest",
     "Interest on the restructured stock plus IMF charges keeps primary income deeply negative; remittances resume "
     "gradually as FX access normalises",
     "Renewed profit and dividend remittance after years of pent-up stock, on top of the interest bill"),
    ("What would change my mind",
     "Monthly trade prints running above US$2.5bn with imports flat — that is the bull case showing up early",
     "The base sits close to the IMF on 2026 but is more cautious on 2027, because the import response to a "
     "recovering economy at a strong real exchange rate is the thing that historically closes these windows",
     "Watch the services line, not the goods line. Goods surpluses in Argentina have repeatedly been eaten by "
     "tourism and profit remittances rather than by imports"),
]
for i, row in enumerate(assump):
    r = asr + 2 + i
    put(ws, r, 2, row[0], bold=True)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    for j in range(3):
        c = ws.cell(row=r, column=3 + j, value=row[1 + j])
        c.font = Font(ARIAL, 9, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 52

wb.save("/home/claude/Argentina_BoP.xlsx")
print("saved BoP; quarters:", NQ, "last:", LASTQL, "years:", FULL_YEARS[0], "-", LAST_FULL)
