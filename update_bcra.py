"""
update_bcra.py
==============
Descarga el archivo oficial 'series.xlsm' del BCRA y actualiza tu tracker
con datos reales de compras/ventas del BCRA en el MULC.

Uso:
    python3 update_bcra.py                    # ejecución normal
    python3 update_bcra.py --inspect          # solo muestra estructura, no actualiza
    python3 update_bcra.py --since 2023-01-01 # solo desde una fecha
    python3 update_bcra.py --no-download      # usa series.xlsm ya descargado

Requiere: openpyxl  (pip install openpyxl)
"""

import argparse
import os
import sys
import urllib.request
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("Falta openpyxl. Instalalo con: pip install openpyxl")
    sys.exit(1)


# =================== CONFIG ===================
SOURCE_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/series.xlsm"
DOWNLOADED_FILE = "series.xlsm"          # se guarda en el mismo directorio
TRACKER_FILE = "BCRA_compras_anuales_10y.xlsx"   # tu tracker

# Hoja y columnas — ajustar si el BCRA cambia el formato
RESERVAS_SHEET_HINT = "Reservas"          # buscamos hoja que contenga este texto
DATE_COL = "A"                             # primera columna = fecha
COMPRAS_COL = "H"                          # columna H = compras/ventas USD (según BCRA)
HEADER_ROW = 1                             # ajustar si los encabezados están en otra fila
# ===============================================


def log(msg):
    print(f"[bcra-update] {msg}")


def download_series(url=SOURCE_URL, dest=DOWNLOADED_FILE):
    """Descarga series.xlsm del BCRA (con un User-Agent amigable)."""
    log(f"Descargando {url}")
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Excel update bot — uso personal)",
    })
    with urllib.request.urlopen(req, timeout=120) as resp:
        size = 0
        with open(dest, "wb") as f:
            while chunk := resp.read(64 * 1024):
                f.write(chunk)
                size += len(chunk)
    log(f"Descarga OK ({size/1024:.0f} KB) → {dest}")


def find_reservas_sheet(wb):
    """Devuelve el nombre de la hoja de reservas (insensible a mayúsculas)."""
    for name in wb.sheetnames:
        if RESERVAS_SHEET_HINT.lower() in name.lower():
            return name
    raise ValueError(
        f"No encontré una hoja que contenga '{RESERVAS_SHEET_HINT}'. "
        f"Hojas disponibles: {wb.sheetnames}"
    )


def inspect(path=DOWNLOADED_FILE):
    """Imprime estructura del archivo para que verifiques antes de procesar."""
    log(f"Inspeccionando {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    log(f"Hojas: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        log(f"--- {name} ({ws.max_row} filas × {ws.max_column} cols) ---")
        # Primeras 3 filas
        for r in range(1, min(4, ws.max_row + 1)):
            row = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
            print(f"   fila {r}: {row}")
    wb.close()


def extract_reservas(path=DOWNLOADED_FILE, since=None):
    """Extrae fecha + compras (col H) + algunas columnas de contexto."""
    log(f"Leyendo {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = find_reservas_sheet(wb)
    ws = wb[sheet_name]
    log(f"Hoja seleccionada: '{sheet_name}'")

    date_idx = column_index_from_string(DATE_COL)
    compras_idx = column_index_from_string(COMPRAS_COL)

    # Capturamos también encabezados de todas las columnas con datos
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=c).value
        if val is not None:
            headers[c] = val
    log(f"Encabezados detectados ({len(headers)} columnas):")
    for c, h in headers.items():
        marker = " ← compras" if c == compras_idx else ""
        print(f"   {get_column_letter(c)}: {h}{marker}")

    data = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not r or len(r) < compras_idx:
            continue
        fecha = r[date_idx - 1]
        compras = r[compras_idx - 1]
        if fecha is None:
            continue
        if isinstance(fecha, datetime):
            fecha = fecha.date()
        if not isinstance(fecha, date):
            continue  # filas de texto / total / etc.
        if since and fecha < since:
            continue
        if not isinstance(compras, (int, float)):
            compras = None  # mantener fila pero marcar nulo
        # Capturar TODA la fila para que tengas las otras variables disponibles
        full_row = [fecha] + [r[c - 1] if c <= len(r) else None for c in range(2, ws.max_column + 1)]
        data.append(full_row)
    wb.close()
    log(f"Filas válidas extraídas: {len(data)}")
    return headers, data


def update_tracker(headers, data, tracker=TRACKER_FILE):
    """Pega los datos en una hoja nueva del tracker y agrega gráficos."""
    if not os.path.exists(tracker):
        log(f"No encontré {tracker} en el directorio actual. Skipping update.")
        return
    log(f"Actualizando {tracker}")
    wb = openpyxl.load_workbook(tracker)

    sheet_name = "Datos Reales BCRA"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Título
    ws["A1"] = f"Datos oficiales BCRA — series.xlsm — actualizado {datetime.now():%Y-%m-%d %H:%M}"
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(headers) + 1, 10))

    # Encabezados (fila 3)
    HEADER_BG = "1F4E79"
    THIN = Side(style="thin", color="BFBFBF")
    BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.cell(row=3, column=1, value="Fecha")
    cols_to_write = sorted(headers.keys())
    for j, c in enumerate(cols_to_write, start=2):
        ws.cell(row=3, column=j, value=str(headers[c])[:60])
    for c in range(1, len(cols_to_write) + 2):
        cell = ws.cell(row=3, column=c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BOX

    # Data
    for i, row in enumerate(data, start=4):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        for j, c in enumerate(cols_to_write, start=2):
            val = row[c - 1] if c - 1 < len(row) else None
            cell = ws.cell(row=i, column=j, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00;(#,##0.00);-"

    last_row = 3 + len(data)
    # Localizar el índice de la columna H en nuestra grilla resultante
    compras_pos = cols_to_write.index(column_index_from_string(COMPRAS_COL)) + 2

    # Gráfico: compras diarias últimos 250 días hábiles
    if len(data) > 10:
        start_chart = max(4, last_row - 250)
        ch = BarChart()
        ch.type = "col"
        ch.style = 11
        ch.title = "Compras/ventas diarias del BCRA — fuente: series.xlsm BCRA"
        ch.y_axis.title = "USD millones"
        data_ref = Reference(ws, min_col=compras_pos, min_row=3,
                             max_col=compras_pos, max_row=last_row)
        cat_ref = Reference(ws, min_col=1, min_row=start_chart, max_row=last_row)
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cat_ref)
        ch.height = 10; ch.width = 22
        ws.add_chart(ch, "B" + str(last_row + 3))

    # Auto-ajuste de columnas
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(cols_to_write) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = "B4"
    wb.save(tracker)
    log(f"Tracker actualizado: hoja '{sheet_name}' con {len(data)} filas.")


def main():
    parser = argparse.ArgumentParser(description="Actualizar tracker con datos reales del BCRA.")
    parser.add_argument("--no-download", action="store_true",
                        help="No descargar; usar series.xlsm local existente.")
    parser.add_argument("--inspect", action="store_true",
                        help="Solo inspeccionar estructura, no actualizar tracker.")
    parser.add_argument("--since", type=str, default=None,
                        help="Filtrar desde fecha (YYYY-MM-DD).")
    parser.add_argument("--source", type=str, default=SOURCE_URL,
                        help="URL alternativa.")
    parser.add_argument("--tracker", type=str, default=TRACKER_FILE,
                        help="Ruta del tracker a actualizar.")
    args = parser.parse_args()

    if not args.no_download:
        download_series(args.source)

    if args.inspect:
        inspect()
        return

    since = None
    if args.since:
        since = datetime.strptime(args.since, "%Y-%m-%d").date()

    headers, data = extract_reservas(since=since)
    update_tracker(headers, data, tracker=args.tracker)
    log("Listo. Abrí tu Excel para ver la hoja 'Datos Reales BCRA'.")


if __name__ == "__main__":
    main()
