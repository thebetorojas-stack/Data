"""Underlying data for the Venezuela TAA deck. One tab per chart, plus README and derived calcs."""
import sys
sys.path.insert(0, "/mnt/project")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
INK = "1A1A1A"
MUTED = "6B6B6B"
INPUTC = "0000FF"          # hardcoded source input
RED = "EC0016"             # UBS red
HDR_FILL = PatternFill("solid", fgColor="F2F2F2")
TITLE_FILL = PatternFill("solid", fgColor="EC0016")
NUM0 = "#,##0;(#,##0);-"
NUM1 = "#,##0.0;(#,##0.0);-"
NUM2 = "#,##0.00;(#,##0.00);-"
PCT = "0%"

wb = Workbook()


def sheet_base(ws, first_col_width=2):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED
    ws.column_dimensions["A"].width = first_col_width


def title(ws, cell, text, sub=None):
    ws[cell] = text
    ws[cell].font = Font(ARIAL, 14, bold=True, color="FFFFFF")
    ws[cell].fill = TITLE_FILL
    ws[cell].alignment = Alignment(vertical="center")
    ws.row_dimensions[ws[cell].row].height = 26
    if sub:
        r = ws[cell].row + 1
        c = ws[cell].column_letter
        ws[f"{c}{r}"] = sub
        ws[f"{c}{r}"].font = Font(ARIAL, 9, italic=True, color=MUTED)


def hdr(ws, row, col, text, width=None, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(ARIAL, 9, bold=True, color=INK)
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = Border(bottom=Side("thin", color=RED))
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def put(ws, r, c, value, fmt=None, color=INK, bold=False, size=9, italic=False, wrap=False):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(ARIAL, size, bold=bold, color=color, italic=italic)
    if fmt:
        cell.number_format = fmt
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def note(ws, r, c, text):
    n = ws.cell(row=r, column=c, value=text)
    n.font = Font(ARIAL, 8, italic=True, color=MUTED)
    return n


def section(ws, r, c, text):
    s = ws.cell(row=r, column=c, value=text)
    s.font = Font(ARIAL, 11, bold=True, color=RED)
    return s


# ---------------------------------------------------------------- README ---
ws = wb.active
ws.title = "README"
sheet_base(ws)
title(ws, "B2", "Venezuela TAA deck — underlying chart data",
      "Companion to venezuela_taa_ubs.pptx. Prepared 1 September 2026.")

put(ws, 4, 2, "Tab", bold=True, size=10)
put(ws, 4, 3, "Feeds", bold=True, size=10)
put(ws, 4, 4, "Contents", bold=True, size=10)
rows = [
    ("S1_Output_Exports", "Slide 1, left chart", "Monthly crude output (OPEC secondary sources) and oil exports (Reuters tanker data), Jan–Jul 2026, plus the implied inventory drawdown."),
    ("S1_Rigs", "Slide 1, right chart", "Baker Hughes active rig count and the rig requirement benchmarks."),
    ("S2_Reserves", "Slide 2, left chart", "Reserve funnel from the 303bn book to liftable barrels, and the dollar arithmetic behind the three stat callouts."),
    ("S3_Targets", "Slide 3, main chart", "Announced production targets since 2005 vs. delivered EIA annual crude output, 2005–2030."),
    ("S3_Scorecard", "Slide 3, table", "Target vs. delivered at each deadline."),
    ("Sources", "All slides", "Full source list with dates and the specific figure each supports."),
]
r = 5
for t, feeds, contents in rows:
    put(ws, r, 2, t, size=9, bold=True)
    put(ws, r, 3, feeds, size=9, color=MUTED)
    put(ws, r, 4, contents, size=9, wrap=True)
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
section(ws, r, 2, "Conventions")
r += 1
for line in [
    "Blue figures are hardcoded inputs taken from a published source. Black figures are formulas that recalculate.",
    "Every input row carries its source in the right-hand column. Do not overwrite a blue cell without updating that source.",
    "The PowerPoint charts are native and carry their own embedded worksheets: right-click a chart > Edit Data to change it in place.",
    "This workbook is the audit trail and the place to rebuild a series from scratch; the two are not linked, so update both.",
]:
    put(ws, r, 2, line, size=9)
    r += 1

r += 1
section(ws, r, 2, "Before publication")
r += 1
for line in [
    "OPEC revises the monthly secondary-source series each month — refresh Table 5-x from the latest MOMR before the call.",
    "April 2026 exports (1,230 kb/d) are derived from Reuters' '+14% vs March', not a directly published level.",
    "Upgrader nameplate capacity (~700 kb/d, four units) is from a 2019 CFR piece and should be refreshed if a newer disclosure exists.",
]:
    put(ws, r, 2, line, size=9)
    r += 1

ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 78

# ------------------------------------------------- S1 output and exports ---
ws = wb.create_sheet("S1_Output_Exports")
sheet_base(ws)
title(ws, "B2", "Slide 1 — crude output vs. oil exports",
      "Thousand b/d. Output on OPEC secondary sources; exports are crude + fuel on tanker tracking.")

cols = [("Month", 12), ("Crude output,\nOPEC secondary", 16), ("Exports,\ncrude + fuel", 14),
        ("PDVSA direct\ncommunication", 16), ("Exports less\noutput", 14),
        ("Cumulative drawdown,\nmn bbl", 20), ("Source / note", 62)]
for i, (t, w) in enumerate(cols):
    hdr(ws, 5, 2 + i, t, width=w, wrap=True)
ws.row_dimensions[5].height = 30

data = [
    ("Jan-26",  924,  800, None, "Output: MOMR. Exports: Reuters, 2 Feb 2026 (up from 498 in Dec-25)."),
    ("Feb-26",  903,  737, None, "Exports fell 6.5% m/m on the loss of the Chinese market (Reuters, 4 Mar 2026)."),
    ("Mar-26",  985, 1090, 1095, "60 vessels departed carrying 1.09mn b/d (Reuters)."),
    ("Apr-26", 1036, 1230, 1136, "Exports derived from Reuters' +14% m/m vs March; MOMR later revised output from 1,031 to 1,036."),
    ("May-26", 1072, 1240, 1179, "Export peak for the year (Reuters)."),
    ("Jun-26", 1104, 1200, 1187, "Second month of export decline as storage ran down."),
    ("Jul-26", 1117, 1160, 1200, "Exports to the US 786 kb/d, highest since early 2019 (Reuters, 3 Aug 2026)."),
]
first = 6
r = first
for m, prod, exp, direct, srctxt in data:
    put(ws, r, 2, m, bold=True)
    put(ws, r, 3, prod, NUM0, color=INPUTC)
    put(ws, r, 4, exp, NUM0, color=INPUTC)
    if direct is not None:
        put(ws, r, 5, direct, NUM0, color=INPUTC)
    put(ws, r, 6, f"=D{r}-C{r}", NUM0)
    days = 31 if m in ("Jan-26", "Mar-26", "May-26", "Jul-26") else (28 if m == "Feb-26" else 30)
    prev = "" if r == first else f"+G{r-1}"
    put(ws, r, 7, f"=MAX(0,F{r})*{days}/1000{prev}", NUM1)
    put(ws, r, 8, srctxt, size=8, color=MUTED, wrap=True)
    ws.row_dimensions[r].height = 22
    r += 1

last = r - 1
r += 1
put(ws, r, 2, "2H-2025 average", bold=True)
put(ws, r, 3, 941, NUM0, color=INPUTC)
put(ws, r, 8, "Mean of 3Q25 (946) and 4Q25 (936), MOMR Jan-2026 Table 5-7. Plotted as the dashed reference line.", size=8, color=MUTED, wrap=True)
avg_row = r
r += 1
put(ws, r, 2, "4Q-2025 average", bold=True)
put(ws, r, 3, 936, NUM0, color=INPUTC)
put(ws, r, 8, "Base for the +181 kb/d headline on slide 1.", size=8, color=MUTED)
q4_row = r
r += 2

section(ws, r, 2, "Derived")
r += 1
put(ws, r, 2, "Output change, Jul-26 vs 4Q-25 (kb/d)", size=9)
put(ws, r, 3, f"=C{last}-C{q4_row}", NUM0, bold=True)
r += 1
put(ws, r, 2, "Cumulative drawdown, Mar–Jul (mn bbl)", size=9)
put(ws, r, 3, f"=G{last}-G{first+1}", NUM1, bold=True)
put(ws, r, 8, "Sum of positive (exports less output) from March. A floor: before netting domestic consumption.", size=8, color=MUTED, wrap=True)
r += 1
put(ws, r, 2, "Blockade inventory build, Dec-25/Jan-26 (mn bbl)", size=9)
put(ws, r, 3, 40, NUM0, color=INPUTC, bold=True)
put(ws, r, 8, "More than 40mn bbl of crude and fuel in onshore tanks and vessels (Reuters, 2 Feb 2026).", size=8, color=MUTED, wrap=True)
r += 2
note(ws, r, 2, "Blue = hardcoded from source. Black = formula. Column E is PDVSA's own figure, shown for the reconciliation gap only; it is not plotted.")

# ------------------------------------------------------------- S1 rigs ---
ws = wb.create_sheet("S1_Rigs")
sheet_base(ws)
title(ws, "B2", "Slide 1 — active drilling rigs", "Baker Hughes International Rig Count and rig requirement benchmarks.")
for i, (t, w) in enumerate([("Item", 34), ("Rigs", 10), ("Source / note", 74)]):
    hdr(ws, 5, 2 + i, t, width=w)
rigs = [
    ("Active, Jul-2026", 2, "Baker Hughes, end-July 2026. Also 2 at end-March, both working for Chevron projects (Reuters, 27 Apr 2026)."),
    ("SLB idle rigs, reactivation plan", 15, "SLB preparing to reactivate up to 15 idle drilling and workover rigs; up to 4 by end-2026 if contracts are signed (Reuters, 19 Aug 2026)."),
    ("Needed by 2028 (oil ministry)", 93, "Ministry of Hydrocarbons presentation by minister Paula Henao, mostly for the Orinoco Belt (Reuters, 27 Apr 2026)."),
    ("Peak, May-1997", 119, "All-time high in the Baker Hughes series for Venezuela, which starts in 1995."),
]
r = 6
for item, v, srctxt in rigs:
    put(ws, r, 2, item)
    put(ws, r, 3, v, NUM0, color=INPUTC, bold=True)
    put(ws, r, 4, srctxt, size=8, color=MUTED, wrap=True)
    ws.row_dimensions[r].height = 24
    r += 1
r += 1
put(ws, r, 2, "Active as % of 2028 requirement", size=9)
put(ws, r, 3, "=C6/C9", PCT, bold=True)
r += 1
put(ws, r, 2, "Record low", size=9)
put(ws, r, 3, 0, NUM0, color=INPUTC)
put(ws, r, 4, "Zero rigs in June 2020 (Baker Hughes).", size=8, color=MUTED)

# --------------------------------------------------------- S2 reserves ---
ws = wb.create_sheet("S2_Reserves")
sheet_base(ws)
title(ws, "B2", "Slide 2 — reserve funnel and deal arithmetic",
      "Billion barrels unless stated. Shows how 65bn barrels converts into liftable volume and dollars.")

for i, (t, w) in enumerate([("Step", 42), ("bn bbl", 10), ("Source / basis", 74)]):
    hdr(ws, 5, 2 + i, t, width=w)
funnel = [
    ("Venezuela proven reserves (OPEC / EIA)", 303, None, "~303bn bbl, roughly 17% of the world total (US EIA, cited by AP). Rests on the Magna Reserva certification of Orinoco extra-heavy oil at ~20% recovery."),
    ("Covered by the 17-field deal", 65, None, "Trump Truth Social and Rodriguez communique, 28 Aug 2026. A US official told Newsweek 63bn; Reuters reports 63.7bn recoverable at a ~20% recovery factor; Axios reported ~90bn pre-announcement. Not reconciled."),
    ("Liftable in 25 yrs at the 1.5 mb/d target", None, "=1.5*365*25/1000", "Rodriguez: 25-year term, production could exceed 1.5 mb/d. Straight-line arithmetic."),
    ("Implied by Caracas's $209bn tax claim", None, "PLACEHOLDER", "$209bn of taxes at the $19/bbl implied take flagged by economist Francisco Rodriguez."),
]
r = 6
funnel_rows = {}
for step, v, formula, srctxt in funnel:
    put(ws, r, 2, step)
    if formula:
        put(ws, r, 3, formula, NUM1, bold=True)
    else:
        put(ws, r, 3, v, NUM1, color=INPUTC, bold=True)
    put(ws, r, 4, srctxt, size=8, color=MUTED, wrap=True)
    ws.row_dimensions[r].height = 30
    funnel_rows[step] = r
    r += 1

r += 1
section(ws, r, 2, "Assumptions")
r += 1
assum_start = r
put(ws, r, 2, "Merey price, Jul-26 average ($/bbl)")
put(ws, r, 3, 67.36, NUM2, color=INPUTC)
put(ws, r, 4, "OPEC MOMR August 2026, ORB component table.", size=8, color=MUTED)
merey_row = r
r += 1
put(ws, r, 2, "Production target (mn b/d)")
put(ws, r, 3, 1.5, NUM1, color=INPUTC)
put(ws, r, 4, "Rodriguez, televised address 29 Aug 2026.", size=8, color=MUTED)
r += 1
put(ws, r, 2, "Term (years)")
put(ws, r, 3, 25, NUM0, color=INPUTC)
put(ws, r, 4, "Rodriguez. The White House fact sheet of 31 Aug says 100-year rights; LOH 2026 art. 35 caps mixed companies at 25+15.", size=8, color=MUTED, wrap=True)
r += 1
put(ws, r, 2, "Tax claim to the Venezuelan State ($bn)")
put(ws, r, 3, 209, NUM0, color=INPUTC)
put(ws, r, 4, "Rodriguez communique, 28 Aug 2026.", size=8, color=MUTED)
tax_row = r
r += 1
put(ws, r, 2, "Implied state take ($/bbl)")
put(ws, r, 3, 19, NUM0, color=INPUTC)
put(ws, r, 4, "Francisco Rodriguez, from the $209bn figure. He notes it assumes $65 oil.", size=8, color=MUTED, wrap=True)
r += 1
put(ws, r, 2, "Stated investment ($bn)")
put(ws, r, 3, 100, NUM0, color=INPUTC)
put(ws, r, 4, "Stated objective. No sponsor, lender, ECA or fund identified as of 31 Aug 2026.", size=8, color=MUTED, wrap=True)

# now that the assumption rows exist, point the funnel's last bar at them
ws.cell(row=funnel_rows["Implied by Caracas's $209bn tax claim"], column=3,
        value=f"=C{tax_row}/C{tax_row + 1}").number_format = NUM1

r += 2
section(ws, r, 2, "Slide 2 callouts")
r += 1
put(ws, r, 2, "In-ground headline ($trn)")
put(ws, r, 3, f"=C{funnel_rows['Covered by the 17-field deal']}*C{merey_row}/1000", NUM1, bold=True)
put(ws, r, 4, "65bn bbl x Merey. A stock, not a cash flow.", size=8, color=MUTED)
r += 1
put(ws, r, 2, "25-year gross revenue at target ($trn)")
put(ws, r, 3, f"=C{funnel_rows['Liftable in 25 yrs at the 1.5 mb/d target']}*C{merey_row}/1000", NUM1, bold=True)
put(ws, r, 4, "Before diluent, royalty (up to 30%), the integrated hydrocarbons tax (up to 15%) and $100bn of capex.", size=8, color=MUTED, wrap=True)
r += 1
put(ws, r, 2, "Gross revenue per year ($bn)")
put(ws, r, 3, f"=C{funnel_rows['Liftable in 25 yrs at the 1.5 mb/d target']}*C{merey_row}/25", NUM0, bold=True)
r += 1
put(ws, r, 2, "Years to lift 65bn bbl at 1.5 mn b/d")
put(ws, r, 3, f"=C{funnel_rows['Covered by the 17-field deal']}/C{funnel_rows['Liftable in 25 yrs at the 1.5 mb/d target']}*25", NUM0, bold=True)
put(ws, r, 4, "Why Caracas says 25 years and Washington says 100.", size=8, color=MUTED)
r += 1
put(ws, r, 2, "Deal reserves as % of the national book")
put(ws, r, 3, f"=C{funnel_rows['Covered by the 17-field deal']}/C{funnel_rows['Venezuela proven reserves (OPEC / EIA)']}", PCT, bold=True)
r += 1
put(ws, r, 2, "Tax claim per year ($bn)")
put(ws, r, 3, f"=C{tax_row}/25", NUM1, bold=True)
put(ws, r, 4, "Against $18.4bn the State collected in 2025 at ~941 kb/d (F. Rodriguez).", size=8, color=MUTED, wrap=True)

# ---------------------------------------------------------- S3 targets ---
ws = wb.create_sheet("S3_Targets")
sheet_base(ws)
title(ws, "B2", "Slide 3 — announced targets vs. delivered output",
      "Million b/d. Actual is EIA annual crude production; each target column is a straight line from the announcement year to the target year.")

headers = ["Year", "Actual crude output\n(EIA)", "Plan Siembra\nPetrolera (2005)",
           "Plan de la Patria\n(2013)", "Maduro pledge\n(Mar-2022)", "PDVSA +18%\n(Jan-2026)",
           "US-Venezuela deal\n(Aug-2026)"]
widths = [8, 16, 14, 14, 14, 14, 16]
for i, (t, w) in enumerate(zip(headers, widths)):
    hdr(ws, 5, 2 + i, t, width=w, wrap=True)
ws.row_dimensions[5].height = 32

actual = {2005: 2.84, 2006: 2.79, 2007: 2.82, 2008: 2.51, 2009: 2.52, 2010: 2.41,
          2011: 2.50, 2012: 2.50, 2013: 2.50, 2014: 2.50, 2015: 2.49, 2016: 2.25,
          2017: 2.00, 2018: 1.49, 2019: 0.88, 2020: 0.53, 2021: 0.59, 2022: 0.72,
          2023: 0.77, 2024: 0.86, 2025: 0.97}
segs = {
    "psp":  {2005: 2.84, 2012: 5.84},
    "pat":  {2013: 2.50, 2019: 6.00},
    "mad":  {2021: 0.59, 2022: 2.00},
    "pdv":  {2025: 0.97, 2026: 1.15},
    "deal": {2025: 0.97, 2030: 1.50},
}
r = 6
year_first = r
for y in range(2005, 2031):
    put(ws, r, 2, str(y), bold=True)
    if y in actual:
        put(ws, r, 3, actual[y], NUM2, color=INPUTC)
    for j, k in enumerate(["psp", "pat", "mad", "pdv", "deal"]):
        if y in segs[k]:
            put(ws, r, 4 + j, segs[k][y], NUM2, color=INPUTC)
    r += 1
year_last = r - 1

r += 1
note(ws, r, 2, "Target columns hold only the endpoints; the chart joins them with a straight line. The 2026 deal has no stated date and is drawn to 2030 for illustration only.")
r += 2

section(ws, r, 2, "Headline arithmetic")
r += 1
put(ws, r, 2, "Sum of dated targets (mn b/d)")
put(ws, r, 3, "=D13+E20+F23", NUM2, bold=True)
put(ws, r, 8, "5.84 (2012) + 6.00 (2019) + 2.00 (2022). The undated 1.5 target is excluded.", size=8, color=MUTED)
r += 1
put(ws, r, 2, "Output change, 2005 to 2025 (mn b/d)")
put(ws, r, 3, f"=C{year_first + 20}-C{year_first}", NUM2, bold=True)
r += 1
put(ws, r, 2, "Peak-to-trough decline, 2005 to 2020 (%)")
put(ws, r, 3, f"=C{year_first + 15}/C{year_first}-1", PCT, bold=True)

# -------------------------------------------------------- S3 scorecard ---
ws = wb.create_sheet("S3_Scorecard")
sheet_base(ws)
title(ws, "B2", "Slide 3 — scorecard", "Each dated production target against delivered output in the target year.")
for i, (t, w) in enumerate([("Plan", 26), ("Announced", 12), ("Target\n(mn b/d)", 11),
                            ("Target\nyear", 9), ("Delivered\n(mn b/d)", 11),
                            ("% of\ntarget", 9), ("Source", 66)]):
    hdr(ws, 5, 2 + i, t, width=w, wrap=True)
ws.row_dimensions[5].height = 30

score = [
    ("Plan Siembra Petrolera", "Aug 2005", 5.84, 2012, 2.50,
     "Target: PDVSA 2005-2012 strategy, 5.84 mn b/d (Monaldi, Baker Institute). Some contemporaneous reporting cites 5.4 or 5.8."),
    ("Plan de la Patria / PDVSA", "2013", 6.00, 2019, 0.88,
     "Target: 6 mn b/d by 2019, of which 4 mn b/d from the Orinoco Belt (PDVSA, Homeland Plan oil goals)."),
    ("Maduro pledge", "Mar 2022", 2.00, 2022, 0.72,
     "Target: 2 mn b/d by end-2022, VTV address (EFE via Infobae, 13 Apr 2022)."),
]
r = 6
for plan, ann, tgt, yr, delivered, srctxt in score:
    put(ws, r, 2, plan, bold=True)
    put(ws, r, 3, ann, color=MUTED)
    put(ws, r, 4, tgt, NUM2, color=INPUTC)
    put(ws, r, 5, str(yr), color=INPUTC)
    put(ws, r, 6, delivered, NUM2, color=INPUTC)
    put(ws, r, 7, f"=F{r}/D{r}", PCT, bold=True, color=RED)
    put(ws, r, 8, srctxt, size=8, color=MUTED, wrap=True)
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
put(ws, r, 2, "Targets met", bold=True)
put(ws, r, 4, "=COUNTIF(G6:G8,\">=1\")", NUM0, bold=True)
put(ws, r, 5, "of", size=9, color=MUTED)
put(ws, r, 6, "=COUNT(G6:G8)", NUM0, bold=True)
r += 2
put(ws, r, 2, "Undated target", bold=True)
put(ws, r, 4, 1.5, NUM2, color=INPUTC)
put(ws, r, 8, "Rodriguez, 28-29 Aug 2026: production could exceed 1.5 mn b/d under the 25-year agreement. No date given, so it is not scored.", size=8, color=MUTED, wrap=True)
r += 1
put(ws, r, 2, "Nearest checkable test", bold=True)
put(ws, r, 4, 0.18, "0.0%", color=INPUTC)
put(ws, r, 8, "PDVSA CEO Hector Obregon, 24 Jan 2026: grow output by at least 18% in 2026 from around 1 mn b/d (AFP).", size=8, color=MUTED, wrap=True)

# ----------------------------------------------------------- Sources ---
ws = wb.create_sheet("Sources")
sheet_base(ws)
title(ws, "B2", "Sources", "Every figure in the deck, with the publication and date it came from.")
for i, (t, w) in enumerate([("Slide", 10), ("Figure", 46), ("Source", 60), ("Date", 14)]):
    hdr(ws, 5, 2 + i, t, width=w)
srcs = [
    ("1", "Monthly crude output, secondary sources", "OPEC Monthly Oil Market Report, World Oil Supply tables", "Monthly"),
    ("1", "Monthly oil exports, crude + fuel", "Reuters (Parraga/Guanipa), tanker tracking and PDVSA documents", "Monthly"),
    ("1", "Blockade inventory build, >40mn bbl", "Reuters", "2 Feb 2026"),
    ("1", "Active rig count", "Baker Hughes International Rig Count", "Jul 2026"),
    ("1", "93 rigs needed by 2028", "Reuters, on a Ministry of Hydrocarbons presentation", "27 Apr 2026"),
    ("1", "SLB, 15 idle rigs, up to 4 by end-2026", "Reuters, from an industry conference in Houston", "19 Aug 2026"),
    ("1", "Chevron 280 kb/d, +50% by 2028", "Chevron 2Q26 earnings call, CFO Eimear Bonner", "31 Jul 2026"),
    ("2", "303bn bbl proven, ~17% of world", "US EIA, cited by AP", "Aug 2026"),
    ("2", "65bn bbl, 17 fields, >$100bn, >$209bn taxes", "Truth Social post; Rodriguez communique", "28 Aug 2026"),
    ("2", "63.7bn recoverable at ~20% recovery factor", "Reuters, via Oil & Gas 360", "Aug 2026"),
    ("2", "Merey $67.36/b, July average", "OPEC MOMR, ORB component table", "12 Aug 2026"),
    ("2", "$209bn implies $19/bbl", "Francisco Rodriguez, cited by OilPrice", "31 Aug 2026"),
    ("2", "Four Orinoco upgraders, ~700 kb/d nameplate", "Council on Foreign Relations", "2019"),
    ("2", "81 kb/d of heavy naphtha imported in July", "Reuters", "3 Aug 2026"),
    ("2", "Tankers waiting up to 30 days at Jose", "Andy Lipow, Lipow Oil Associates, via CNBC", "31 Aug 2026"),
    ("3", "Annual crude output 2005-2025", "US EIA International Energy Statistics", "Jul 2026 vintage"),
    ("3", "Plan Siembra Petrolera target 5.84 mn b/d by 2012", "F. Monaldi, Baker Institute, 'The Collapse of the Venezuelan Oil Industry'", "2020"),
    ("3", "6 mn b/d by 2019, 4 mn b/d from the Orinoco Belt", "PDVSA, Homeland Plan oil goals", "2013"),
    ("3", "2 mn b/d by end-2022", "Maduro, VTV address, reported by EFE via Infobae", "Mar/Apr 2022"),
    ("3", "+18% output growth in 2026", "PDVSA CEO Hector Obregon, reported by AFP", "24 Jan 2026"),
    ("3", ">1.5 mn b/d under the agreement, 25-year term", "Rodriguez, statement and televised address", "28-29 Aug 2026"),
    ("3", "4 mn b/d needs ~a decade and ~$100bn", "Analysts, to AP", "Jan 2026"),
    ("4", "OSC 35% stake, 20% at-cost offtake, 100-year rights", "White House fact sheet, via AP", "31 Aug 2026"),
    ("4", "OSC 'does not take equity stakes'", "Pentagon spokesman Sean Parnell", "30 Aug 2026"),
    ("4", "35% passive stake via penny warrants", "Wall Street Journal", "29-30 Aug 2026"),
    ("4", "Constitution arts. 12, 150, 187.9, 303", "Constitucion de la Republica Bolivariana de Venezuela", "1999"),
    ("4", "LOH arts. 35, 40, 68", "Ley de Reforma de la LOH, Gaceta Oficial 6.978 Extraordinario", "29 Jan 2026"),
    ("4", "Exxon: 'un-investable'", "Darren Woods, reported by AP", "2026"),
    ("4", "Conoco $10-12bn award as a precondition", "Ryan Lance, reported by the Houston Chronicle", "2026"),
    ("4", "$98.3bn of bondholder claims incl. past-due interest", "Barclays, via CNBC", "Jan 2026"),
]
r = 6
for sl, fig, source, date in srcs:
    put(ws, r, 2, sl)
    put(ws, r, 3, fig, size=9, wrap=True)
    put(ws, r, 4, source, size=9, color=MUTED, wrap=True)
    put(ws, r, 5, date, size=9, color=MUTED)
    r += 1
r += 1
note(ws, r, 2, "teleSUR is excluded as a source. The private counterparty is attributed only as a private domestic operator, per the compliance decision on the 29 August note.")

wb.save("/home/claude/vz/venezuela_taa_chart_data.xlsx")
print("saved")
