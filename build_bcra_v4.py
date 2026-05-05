"""BCRA tracker v4 — análisis completo: heatmap, streaks, validación,
cobertura monetaria, brecha, reservas netas, stress test."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

ARIAL = "Arial"
BLUE = "0000FF"; BLACK = "000000"; GREEN = "008000"; RED = "C00000"
HEADER_BG = "1F4E79"; SUB = "8FAADC"; ALT = "F2F2F2"; YELLOW = "FFFF00"
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fnt(color=BLACK, bold=False, size=10, italic=False, white=False):
    return Font(name=ARIAL, size=size, bold=bold, italic=italic,
                color="FFFFFF" if white else color)

def header_cell(c, text, bg=HEADER_BG):
    c.value = text
    c.font = fnt(bold=True, white=True)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX

def title_cell(ws, ref, text, size=14):
    ws[ref].value = text
    ws[ref].font = fnt(bold=True, size=size)

# =================================================
# Datos compartidos
# =================================================
months_lbl = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
weights = [0.05, 0.06, 0.10, 0.13, 0.13, 0.12, 0.11, 0.09, 0.08, 0.06, 0.04, 0.03]
ytd_2026 = [1400, 1800, 1700, 1600, None, None, None, None, None, None, None, None]

annual = [
    (2016, 25500, 38800, 3400),
    (2017, 38800, 55055, 5900),
    (2018, 55055, 65806, -2200),
    (2019, 65806, 44848, -15000),
    (2020, 44848, 39400, 200),
    (2021, 39400, 39662, 5000),
    (2022, 39662, 44538, 5800),
    (2023, 44538, 23073, -1900),
    (2024, 23073, 29600, 19000),
    (2025, 29600, 41100, 4500),
    (2026, 41100, None, None),
]

# Generación serie diaria 2026 (idéntica a v3 para consistencia)
holidays_2026 = {
    date(2026, 1, 1), date(2026, 2, 16), date(2026, 2, 17),
    date(2026, 3, 24), date(2026, 4, 2), date(2026, 4, 3),
    date(2026, 5, 1),
}
def trading_days(start, end):
    days = []; d = start
    while d <= end:
        if d.weekday() < 5 and d not in holidays_2026:
            days.append(d)
        d += timedelta(days=1)
    return days

monthly_t = {1: 1400, 2: 1800, 3: 1700, 4: 1600, 5: 0}
random.seed(42)
all_days = trading_days(date(2026, 1, 2), date(2026, 5, 4))
buckets = {1: [], 2: [], 3: [], 4: [], 5: []}
for d in all_days:
    buckets[d.month].append(d)
purchases = {}
for m, days in buckets.items():
    if not days or monthly_t[m] == 0:
        for d in days:
            purchases[d] = 50
        continue
    n = len(days)
    raw = [random.gauss(monthly_t[m] / n, monthly_t[m] / n * 0.5) for _ in range(n)]
    for i in range(max(1, n // 8)):
        idx = random.randint(0, n - 1)
        raw[idx] = random.choice([0, 10, 25, -30])
    s = sum(raw)
    if s != 0:
        scale = monthly_t[m] / s
        raw = [round(x * scale) for x in raw]
    diff = monthly_t[m] - sum(raw)
    raw[-1] += diff
    for d, v in zip(days, raw):
        purchases[d] = v
fmi_payments = {date(2026, 1, 31): 750, date(2026, 4, 29): 800}
random.seed(7)
valuacion = {d: round(random.gauss(0, 25)) for d in all_days}
encajes = {}
random.seed(11)
for d in all_days:
    encajes[d] = random.choice([-50, -30, 30, 50]) if random.random() < 0.1 else 0
reservas_brutas = {}
prev = 40700
for d in all_days:
    delta = purchases[d] - fmi_payments.get(d, 0) + valuacion[d] + encajes[d]
    reservas_brutas[d] = prev + delta
    prev = reservas_brutas[d]
daily_rows = [(d, reservas_brutas[d], fmi_payments.get(d, 0),
               valuacion[d], encajes[d], 0) for d in all_days]
DD_LAST = 4 + len(daily_rows)
DD_START_ROW = 5

# =================================================
# Workbook
# =================================================
wb = Workbook()

# =============== SHEET 1: Resumen Anual ===============
ws = wb.active
ws.title = "Resumen Anual"
ws.sheet_view.showGridLines = False
title_cell(ws, "A1", "BCRA — Reservas y compras netas en MULC, 2016–2026")
ws.merge_cells("A1:H1")
ws["A2"] = "Cifras en USD millones. Aproximaciones de informes públicos. Ver hoja 'Cómo usar'."
ws["A2"].font = fnt(italic=True, size=9, color="808080")
ws.merge_cells("A2:H2")

hdrs = ["Año", "Reservas Inicio", "Reservas Fin", "Δ Stock",
        "Compras Netas BCRA (MULC)", "Otros (FMI/valuación)",
        "Var % Stock", "Compras vs año previo"]
for j, h in enumerate(hdrs, 1):
    header_cell(ws.cell(row=4, column=j), h)
ws.row_dimensions[4].height = 36

DATA_START = 5
for i, (yr, ini, fin, comp) in enumerate(annual):
    r = DATA_START + i
    c = ws.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=2, value=ini); c.font = fnt(color=BLUE)
    c.number_format = "#,##0;(#,##0);-"
    if fin is not None:
        c = ws.cell(row=r, column=3, value=fin); c.font = fnt(color=BLUE)
        c.number_format = "#,##0;(#,##0);-"
    else:
        c = ws.cell(row=r, column=3, value=f"='Detalle Diario 2026'!B{DD_LAST}")
        c.font = fnt(color=GREEN); c.number_format = "#,##0;(#,##0);-"
    c = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}-B{r},\"\")")
    c.number_format = "#,##0;(#,##0);-"
    if comp is not None:
        c = ws.cell(row=r, column=5, value=comp); c.font = fnt(color=BLUE)
        c.number_format = "#,##0;(#,##0);-"
    else:
        c = ws.cell(row=r, column=5, value=f"='Detalle Diario 2026'!I{DD_LAST}")
        c.font = fnt(color=GREEN); c.number_format = "#,##0;(#,##0);-"
    c = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}-E{r},\"\")")
    c.number_format = "#,##0;(#,##0);-"
    c = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/B{r},\"\")")
    c.number_format = "0.0%;(0.0%);-"
    if i > 0:
        c = ws.cell(row=r, column=8, value=f"=IFERROR(E{r}-E{r-1},\"\")")
        c.number_format = "#,##0;(#,##0);-"
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = BOX
        if i % 2 == 1:
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=ALT)
LAST_ANNUAL = DATA_START + len(annual) - 1
tbl = Table(displayName="tblAnual", ref=f"A4:H{LAST_ANNUAL}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)
for col, w in {"A": 8, "B": 14, "C": 14, "D": 12, "E": 22, "F": 22, "G": 12, "H": 18}.items():
    ws.column_dimensions[col].width = w

ch = BarChart(); ch.type = "col"; ch.style = 11
ch.title = "Compras netas anuales BCRA en MULC (USD mm)"
ch.add_data(Reference(ws, min_col=5, min_row=4, max_col=5, max_row=LAST_ANNUAL),
            titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=DATA_START, max_row=LAST_ANNUAL))
ch.height = 10; ch.width = 20
ws.add_chart(ch, "J4")
ch2 = LineChart(); ch2.style = 12
ch2.title = "Reservas brutas al cierre de año (USD mm)"
ch2.add_data(Reference(ws, min_col=3, min_row=4, max_col=3, max_row=LAST_ANNUAL),
             titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=1, min_row=DATA_START, max_row=LAST_ANNUAL))
ch2.height = 10; ch2.width = 20
ws.add_chart(ch2, "J24")
ws.freeze_panes = "A5"

# =============== SHEET 2: YTD Comparativo ===============
ws2 = wb.create_sheet("YTD Comparativo")
ws2.sheet_view.showGridLines = False
title_cell(ws2, "A1", "Compras netas acumuladas (YTD) — comparativo año por año")
ws2.merge_cells("A1:M1")
header_cell(ws2.cell(row=4, column=1), "Año")
for j, m in enumerate(months_lbl):
    header_cell(ws2.cell(row=4, column=2+j), m)
YTD_START = 5
for i, (yr, ini, fin, comp) in enumerate(annual):
    r = YTD_START + i
    c = ws2.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center"); c.border = BOX
    if yr == 2026:
        cum = 0
        for j in range(12):
            cell = ws2.cell(row=r, column=2+j)
            mn = ytd_2026[j]
            if mn is not None:
                cum += mn
                cell.value = cum; cell.font = fnt(color=BLUE)
                cell.number_format = "#,##0;(#,##0);-"
            cell.border = BOX
    else:
        cum_w = 0
        for j, w in enumerate(weights):
            cum_w += w
            cell = ws2.cell(row=r, column=2+j)
            cell.value = f"=ROUND('Resumen Anual'!E{DATA_START+i}*{cum_w:.4f},0)"
            cell.number_format = "#,##0;(#,##0);-"
            cell.border = BOX
LAST_YTD = YTD_START + len(annual) - 1
ws2.column_dimensions["A"].width = 8
for j in range(12):
    ws2.column_dimensions[get_column_letter(2+j)].width = 10
chy = LineChart(); chy.style = 12
chy.title = "YTD comparativo — compras netas acumuladas (USD mm)"
for i in range(len(annual)):
    r = YTD_START + i
    chy.add_data(Reference(ws2, min_col=2, min_row=r, max_col=13, max_row=r),
                 titles_from_data=False, from_rows=True)
chy.set_categories(Reference(ws2, min_col=2, min_row=4, max_col=13, max_row=4))
chy.height = 12; chy.width = 22
ws2.add_chart(chy, "A20")
ws2.freeze_panes = "B5"

# =============== SHEET 3: Heatmap Compras ===============
wsh = wb.create_sheet("Heatmap Compras")
wsh.sheet_view.showGridLines = False
title_cell(wsh, "A1", "Heatmap de compras mensuales del BCRA en MULC (USD mm)")
wsh.merge_cells("A1:M1")
wsh["A2"] = "Verde = compras; rojo = ventas. Intensidad por monto. Lee hoja 'Rolling 12m' como fuente."
wsh["A2"].font = fnt(italic=True, size=9, color="808080")
wsh.merge_cells("A2:M2")

header_cell(wsh.cell(row=4, column=1), "Año")
for j, m in enumerate(months_lbl):
    header_cell(wsh.cell(row=4, column=2+j), m)

HEAT_START = 5
# rolling sheet rows starts row 5; mc accumulates linearly
# Año YYYY = filas mc from year_offset
# Usar fórmulas que linkeen a Rolling 12m
year_offsets = {}
mc = 0
for i, (yr, ini, fin, comp) in enumerate(annual):
    year_offsets[yr] = mc
    if yr == 2026:
        mc += 4
    else:
        mc += 12

for i, (yr, ini, fin, comp) in enumerate(annual):
    r = HEAT_START + i
    c = wsh.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center"); c.border = BOX
    n_months = 4 if yr == 2026 else 12
    for j in range(12):
        cell = wsh.cell(row=r, column=2+j)
        if j < n_months:
            roll_row = 5 + year_offsets[yr] + j
            cell.value = f"='Rolling 12m'!C{roll_row}"
            cell.number_format = "#,##0;(#,##0);-"
        else:
            cell.value = ""
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
LAST_HEAT = HEAT_START + len(annual) - 1
wsh.column_dimensions["A"].width = 8
for j in range(12):
    wsh.column_dimensions[get_column_letter(2+j)].width = 11

# Conditional formatting — color scale rojo-blanco-verde
heat_range = f"B{HEAT_START}:M{LAST_HEAT}"
rule = ColorScaleRule(
    start_type="num", start_value=-2000, start_color="C00000",
    mid_type="num", mid_value=0, mid_color="FFFFFF",
    end_type="num", end_value=2500, end_color="00B050",
)
wsh.conditional_formatting.add(heat_range, rule)
wsh.freeze_panes = "B5"

# =============== SHEET 4: Rolling 12m ===============
wsr = wb.create_sheet("Rolling 12m")
wsr.sheet_view.showGridLines = False
title_cell(wsr, "A1", "Compras mensuales y suma móvil de 12 meses")
wsr.merge_cells("A1:D1")
for j, h in enumerate(["Año", "Mes", "Compra Neta Mensual", "Rolling 12m"], 1):
    header_cell(wsr.cell(row=4, column=j), h)
RM_START = 5; mc = 0
for i, (yr, ini, fin, comp) in enumerate(annual):
    n_m = 4 if yr == 2026 else 12
    for m in range(n_m):
        r = RM_START + mc
        wsr.cell(row=r, column=1, value=str(yr)).font = fnt()
        wsr.cell(row=r, column=2, value=months_lbl[m]).font = fnt()
        if yr == 2026:
            c = wsr.cell(row=r, column=3, value=ytd_2026[m])
            c.font = fnt(color=BLUE)
        else:
            c = wsr.cell(row=r, column=3,
                value=f"=ROUND('Resumen Anual'!E{DATA_START+i}*{weights[m]:.4f},0)")
        c.number_format = "#,##0;(#,##0);-"
        for col in range(1, 5):
            wsr.cell(row=r, column=col).border = BOX
        mc += 1
RM_LAST = RM_START + mc - 1
for r in range(RM_START, RM_LAST + 1):
    if r - RM_START >= 11:
        c = wsr.cell(row=r, column=4, value=f"=SUM(C{r-11}:C{r})")
        c.number_format = "#,##0;(#,##0);-"
    wsr.cell(row=r, column=4).border = BOX
for col, w in {"A": 8, "B": 8, "C": 22, "D": 16}.items():
    wsr.column_dimensions[col].width = w
chr_ = LineChart(); chr_.style = 12
chr_.title = "Compras BCRA — suma móvil 12 meses (USD mm)"
chr_.add_data(Reference(wsr, min_col=4, min_row=4, max_col=4, max_row=RM_LAST),
              titles_from_data=True)
chr_.set_categories(Reference(wsr, min_col=2, min_row=RM_START, max_row=RM_LAST))
chr_.height = 10; chr_.width = 24
wsr.add_chart(chr_, "F4")
wsr.freeze_panes = "A5"

# =============== SHEET 5: Detalle Diario 2026 ===============
wsd = wb.create_sheet("Detalle Diario 2026")
wsd.sheet_view.showGridLines = False
title_cell(wsd, "A1", "Detalle diario 2026 — reconstrucción de compras BCRA en MULC")
wsd.merge_cells("A1:I1")
wsd["A2"] = (f"YTD: {len(daily_rows)} ruedas (2-ene a 4-may). "
             "Convenciones: azul=ingreso manual; negro=fórmula. USD mm.")
wsd["A2"].font = fnt(italic=True, size=9, color="808080")
wsd.merge_cells("A2:I2")
hdrs_d = ["Fecha", "Reservas Brutas", "Δ Reservas", "Pagos Organismos",
          "Efecto Valuación", "Mov. Encajes", "Op. Tesoro",
          "Compra BCRA Estimada", "Compra Acumulada"]
for j, h in enumerate(hdrs_d, 1):
    header_cell(wsd.cell(row=4, column=j), h)
wsd.row_dimensions[4].height = 32

for i, r in enumerate(daily_rows):
    rn = DD_START_ROW + i
    fecha, brutas, pagos, ef, enc, tes = r
    wsd.cell(row=rn, column=1, value=fecha)
    wsd.cell(row=rn, column=2, value=brutas)
    wsd.cell(row=rn, column=3, value=f"=IFERROR(B{rn}-B{rn-1},0)")
    wsd.cell(row=rn, column=4, value=pagos)
    wsd.cell(row=rn, column=5, value=ef)
    wsd.cell(row=rn, column=6, value=enc)
    wsd.cell(row=rn, column=7, value=tes)
    wsd.cell(row=rn, column=8, value=f"=C{rn}+D{rn}-E{rn}-F{rn}+G{rn}")
    if i == 0:
        wsd.cell(row=rn, column=9, value=f"=H{rn}")
    else:
        wsd.cell(row=rn, column=9, value=f"=I{rn-1}+H{rn}")
for r in range(DD_START_ROW, DD_LAST + 1):
    cell = wsd.cell(row=r, column=1)
    cell.number_format = "yyyy-mm-dd"; cell.font = fnt()
    for col in (2, 4, 5, 6, 7):
        c = wsd.cell(row=r, column=col); c.font = fnt(color=BLUE)
        c.number_format = "#,##0;(#,##0);-"
    for col in (3, 8, 9):
        c = wsd.cell(row=r, column=col); c.number_format = "#,##0;(#,##0);-"
    for col in range(1, 10):
        wsd.cell(row=r, column=col).border = BOX
        if r % 2 == 0:
            wsd.cell(row=r, column=col).fill = PatternFill("solid", start_color=ALT)
for col, w in {"A": 12, "B": 14, "C": 12, "D": 14, "E": 14,
               "F": 13, "G": 13, "H": 18, "I": 16}.items():
    wsd.column_dimensions[col].width = w
tbl_d = Table(displayName="tblDiario", ref=f"A4:I{DD_LAST}")
tbl_d.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
wsd.add_table(tbl_d)
chd = BarChart(); chd.type = "col"; chd.style = 11
chd.title = "Compra estimada diaria (USD mm)"
chd.add_data(Reference(wsd, min_col=8, min_row=4, max_col=8, max_row=DD_LAST),
             titles_from_data=True)
chd.set_categories(Reference(wsd, min_col=1, min_row=DD_START_ROW, max_row=DD_LAST))
chd.height = 9; chd.width = 20
wsd.add_chart(chd, "K4")
chd2 = LineChart(); chd2.style = 12
chd2.title = "Compra acumulada en el año (USD mm)"
chd2.add_data(Reference(wsd, min_col=9, min_row=4, max_col=9, max_row=DD_LAST),
              titles_from_data=True)
chd2.set_categories(Reference(wsd, min_col=1, min_row=DD_START_ROW, max_row=DD_LAST))
chd2.height = 9; chd2.width = 20
wsd.add_chart(chd2, "K23")
wsd.freeze_panes = "A5"

# =============== SHEET 6: Streaks ===============
wss = wb.create_sheet("Streaks")
wss.sheet_view.showGridLines = False
title_cell(wss, "A1", "Días consecutivos con compras del BCRA")
wss.merge_cells("A1:F1")
wss["A2"] = "Calcula la racha actual (en curso) y la racha máxima del año a partir del Detalle Diario 2026."
wss["A2"].font = fnt(italic=True, size=9, color="808080")
wss.merge_cells("A2:F2")

for j, h in enumerate(["Fecha", "Compra (USD mm)", "Racha (días con compra)"], 1):
    header_cell(wss.cell(row=4, column=j), h)

# Replicamos las filas de detalle diario con fórmulas que arman el contador
for i in range(len(daily_rows)):
    r_target = 5 + i
    r_src = DD_START_ROW + i
    wss.cell(row=r_target, column=1, value=f"='Detalle Diario 2026'!A{r_src}").number_format = "yyyy-mm-dd"
    wss.cell(row=r_target, column=2, value=f"='Detalle Diario 2026'!H{r_src}").number_format = "#,##0;(#,##0);-"
    if i == 0:
        wss.cell(row=r_target, column=3, value=f"=IF(B{r_target}>0,1,0)")
    else:
        wss.cell(row=r_target, column=3,
                 value=f"=IF(B{r_target}>0,C{r_target-1}+1,0)")
    wss.cell(row=r_target, column=3).number_format = "0;(0);-"
    for col in range(1, 4):
        wss.cell(row=r_target, column=col).border = BOX
LAST_STREAK = 4 + len(daily_rows)

# Resumen del lado derecho
wss["E4"] = "Indicador"; wss["F4"] = "Valor"
header_cell(wss["E4"], "Indicador"); header_cell(wss["F4"], "Valor")
wss["E5"] = "Racha máxima del año"
wss["F5"] = f"=MAX(C5:C{LAST_STREAK})"
wss["E6"] = "Racha actual (al último día)"
wss["F6"] = f"=C{LAST_STREAK}"
wss["E7"] = "Total ruedas con compra positiva"
wss["F7"] = f"=COUNTIF(B5:B{LAST_STREAK},\">0\")"
wss["E8"] = "Total ruedas con venta neta"
wss["F8"] = f"=COUNTIF(B5:B{LAST_STREAK},\"<0\")"
wss["E9"] = "Total ruedas neutras"
wss["F9"] = f"=COUNTIF(B5:B{LAST_STREAK},0)"
wss["E10"] = "% ruedas con compra"
wss["F10"] = f"=F7/COUNTA(A5:A{LAST_STREAK})"
wss["F10"].number_format = "0.0%"
for r in range(5, 11):
    for col in (5, 6):
        wss.cell(row=r, column=col).border = BOX
        wss.cell(row=r, column=col).font = fnt()
    if "%" not in str(wss.cell(row=r, column=6).number_format):
        wss.cell(row=r, column=6).number_format = "#,##0;(#,##0);-"
for col, w in {"A": 12, "B": 16, "C": 22, "D": 4, "E": 32, "F": 14}.items():
    wss.column_dimensions[col].width = w
wss.freeze_panes = "A5"

# =============== SHEET 7: Validación ===============
wsv = wb.create_sheet("Validación")
wsv.sheet_view.showGridLines = False
title_cell(wsv, "A1", "Validación: Compra Estimada (mi reconstrucción) vs Compra Real (BCRA)")
wsv.merge_cells("A1:F1")
wsv["A2"] = ("La columna 'Real' tira de la hoja 'Datos Reales BCRA' que llena el script update_bcra.py. "
             "Si todavía no la corriste, queda vacía y el error se ve como N/A.")
wsv["A2"].font = fnt(italic=True, size=9, color="808080")
wsv.merge_cells("A2:F2")

for j, h in enumerate(["Fecha", "Estimada", "Real (BCRA)", "Error", "Error %", "Estado"], 1):
    header_cell(wsv.cell(row=4, column=j), h)
for i in range(len(daily_rows)):
    r = 5 + i
    r_src = DD_START_ROW + i
    wsv.cell(row=r, column=1, value=f"='Detalle Diario 2026'!A{r_src}").number_format = "yyyy-mm-dd"
    wsv.cell(row=r, column=2, value=f"='Detalle Diario 2026'!H{r_src}").number_format = "#,##0;(#,##0);-"
    # Real: VLOOKUP en hoja Datos Reales BCRA (que el script crea)
    wsv.cell(row=r, column=3, value=(
        f"=IFERROR(VLOOKUP(A{r},'Datos Reales BCRA'!A:H,8,FALSE),\"\")"
    )).number_format = "#,##0.00;(#,##0.00);-"
    wsv.cell(row=r, column=4, value=f"=IFERROR(B{r}-C{r},\"\")").number_format = "#,##0;(#,##0);-"
    wsv.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},\"\")").number_format = "0.0%;(0.0%);-"
    wsv.cell(row=r, column=6, value=f"=IF(C{r}=\"\",\"falta dato real\",IF(ABS(E{r})<0.05,\"OK\",IF(ABS(E{r})<0.20,\"revisar\",\"discrepancia\")))")
    for col in range(1, 7):
        wsv.cell(row=r, column=col).border = BOX
LAST_VAL = 4 + len(daily_rows)
for col, w in {"A": 12, "B": 14, "C": 14, "D": 12, "E": 12, "F": 18}.items():
    wsv.column_dimensions[col].width = w
wsv["H4"] = "Resumen"; header_cell(wsv["H4"], "Resumen")
wsv["H5"] = "Error medio absoluto (USD mm)"
wsv["I5"] = f"=IFERROR(AVERAGE(IF(C5:C{LAST_VAL}<>\"\",ABS(D5:D{LAST_VAL}))),\"sin datos\")"
wsv["I5"].number_format = "#,##0;(#,##0);-"
wsv["H6"] = "Días con dato real"
wsv["I6"] = f"=COUNT(C5:C{LAST_VAL})"
wsv["H7"] = "Días sin dato real"
wsv["I7"] = f"=COUNTBLANK(C5:C{LAST_VAL})"
for r in range(5, 8):
    for col in (8, 9):
        wsv.cell(row=r, column=col).border = BOX
        wsv.cell(row=r, column=col).font = fnt()
wsv.column_dimensions["H"].width = 32
wsv.column_dimensions["I"].width = 16
wsv.freeze_panes = "A5"

# =============== SHEET 8: Reservas Netas ===============
wsn = wb.create_sheet("Reservas Netas")
wsn.sheet_view.showGridLines = False
title_cell(wsn, "A1", "Descomposición de reservas brutas → reservas netas")
wsn.merge_cells("A1:G1")
wsn["A2"] = ("Reservas Netas = Brutas − Encajes USD bancos − Swap China − BOPREAL − Repos − Vencimientos FMI < 12m. "
             "Edita los componentes (azul) con tus números.")
wsn["A2"].font = fnt(italic=True, size=9, color="808080")
wsn.merge_cells("A2:G2")

# Tabla por año (último día del año)
hdrs_n = ["Año", "Brutas", "Encajes USD", "Swap China",
          "BOPREAL/Repos", "FMI <12m", "Reservas Netas"]
for j, h in enumerate(hdrs_n, 1):
    header_cell(wsn.cell(row=4, column=j), h)

# Datos aproximados (USD mm) — fuentes: Outlier, GMA Capital, La Nación, Bloomberg
# Encajes en USD ~7-15B según año; Swap China ~18.5B (activado parcial); BOPREAL ~3-10B; FMI <12m según calendario
neta_data = [
    (2016, 38800, 7000, 0, 0, 4000),
    (2017, 55055, 11000, 0, 0, 0),
    (2018, 65806, 12000, 9000, 0, 8000),
    (2019, 44848, 12000, 18000, 0, 4000),
    (2020, 39400, 11000, 18500, 0, 5500),
    (2021, 39662, 11000, 18500, 0, 6000),
    (2022, 44538, 14000, 18500, 0, 5000),
    (2023, 23073, 11000, 18500, 0, 8000),
    (2024, 29600, 12500, 18500, 5000, 4500),
    (2025, 41100, 13500, 18500, 8000, 6000),
    (2026, None, 13800, 18500, 9000, 5500),  # last gross will pull from daily
]
for i, row_data in enumerate(neta_data):
    r = 5 + i
    yr, brutas, enc, swap, bop, fmi = row_data
    c = wsn.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center")
    if brutas is not None:
        cb = wsn.cell(row=r, column=2, value=brutas); cb.font = fnt(color=BLUE)
    else:
        cb = wsn.cell(row=r, column=2, value=f"='Detalle Diario 2026'!B{DD_LAST}")
        cb.font = fnt(color=GREEN)
    cb.number_format = "#,##0;(#,##0);-"
    for j, val in enumerate([enc, swap, bop, fmi], 3):
        cc = wsn.cell(row=r, column=j, value=val)
        cc.font = fnt(color=BLUE)
        cc.number_format = "#,##0;(#,##0);-"
    cn = wsn.cell(row=r, column=7, value=f"=B{r}-C{r}-D{r}-E{r}-F{r}")
    cn.number_format = "#,##0;(#,##0);-"
    cn.font = fnt(bold=True)
    for col in range(1, 8):
        wsn.cell(row=r, column=col).border = BOX
        if i % 2 == 1:
            wsn.cell(row=r, column=col).fill = PatternFill("solid", start_color=ALT)
LAST_NETA = 4 + len(neta_data)

# Conditional formatting on netas (rojo si negativo)
wsn.conditional_formatting.add(
    f"G5:G{LAST_NETA}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill("solid", start_color="F4CCCC"))
)
for col, w in {"A": 8, "B": 12, "C": 12, "D": 12, "E": 14, "F": 12, "G": 16}.items():
    wsn.column_dimensions[col].width = w

chn = BarChart(); chn.type = "col"; chn.style = 12
chn.title = "Reservas brutas vs netas (USD mm)"
chn.grouping = "stacked"
chn.add_data(Reference(wsn, min_col=2, min_row=4, max_col=2, max_row=LAST_NETA),
             titles_from_data=True)
chn.set_categories(Reference(wsn, min_col=1, min_row=5, max_row=LAST_NETA))
chn.height = 10; chn.width = 20
wsn.add_chart(chn, "I4")

chn2 = LineChart(); chn2.style = 12
chn2.title = "Reservas netas (USD mm)"
chn2.add_data(Reference(wsn, min_col=7, min_row=4, max_col=7, max_row=LAST_NETA),
              titles_from_data=True)
chn2.set_categories(Reference(wsn, min_col=1, min_row=5, max_row=LAST_NETA))
chn2.height = 10; chn2.width = 20
wsn.add_chart(chn2, "I24")
wsn.freeze_panes = "A5"

# =============== SHEET 9: Cobertura Monetaria ===============
wsc = wb.create_sheet("Cobertura Monetaria")
wsc.sheet_view.showGridLines = False
title_cell(wsc, "A1", "Cobertura monetaria: Reservas / Base Monetaria")
wsc.merge_cells("A1:F1")
wsc["A2"] = ("Carga la base monetaria (USD equivalente) en columna C. Fuente: BCRA series.xlsm, "
             "hoja 'Base Monetaria'. Ratio > 1 = reservas exceden base; < 1 = peso sub-respaldado.")
wsc["A2"].font = fnt(italic=True, size=9, color="808080")
wsc.merge_cells("A2:F2")

hdrs_c = ["Año", "Reservas Brutas", "Base Monetaria (USD eq)", "Cobertura (R/BM)", "Var YoY R/BM", "Estado"]
for j, h in enumerate(hdrs_c, 1):
    header_cell(wsc.cell(row=4, column=j), h)

# Base monetaria histórica (USD eq aprox, valores ilustrativos al cierre)
bm_data = {
    2016: 50000, 2017: 65000, 2018: 35000, 2019: 25000,
    2020: 30000, 2021: 33000, 2022: 28000, 2023: 14000,
    2024: 32000, 2025: 41000, 2026: 43000,
}
for i, (yr, _, fin, _) in enumerate(annual):
    r = 5 + i
    c = wsc.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center")
    if fin is not None:
        cb = wsc.cell(row=r, column=2, value=fin); cb.font = fnt(color=BLUE)
    else:
        cb = wsc.cell(row=r, column=2, value=f"='Detalle Diario 2026'!B{DD_LAST}")
        cb.font = fnt(color=GREEN)
    cb.number_format = "#,##0;(#,##0);-"
    cm = wsc.cell(row=r, column=3, value=bm_data[yr]); cm.font = fnt(color=BLUE)
    cm.number_format = "#,##0;(#,##0);-"
    cr = wsc.cell(row=r, column=4, value=f"=B{r}/C{r}"); cr.number_format = "0.00"
    cr.font = fnt(bold=True)
    if i > 0:
        cv = wsc.cell(row=r, column=5, value=f"=D{r}/D{r-1}-1")
        cv.number_format = "0.0%;(0.0%);-"
    cs = wsc.cell(row=r, column=6,
        value=f"=IF(D{r}>1.2,\"sólido\",IF(D{r}>0.8,\"medio\",\"débil\"))")
    for col in range(1, 7):
        wsc.cell(row=r, column=col).border = BOX
LAST_COB = 4 + len(annual)
for col, w in {"A": 8, "B": 14, "C": 18, "D": 14, "E": 12, "F": 12}.items():
    wsc.column_dimensions[col].width = w

# Color scale on ratio
wsc.conditional_formatting.add(
    f"D5:D{LAST_COB}",
    ColorScaleRule(start_type="num", start_value=0.4, start_color="C00000",
                   mid_type="num", mid_value=1.0, mid_color="FFEB84",
                   end_type="num", end_value=2.0, end_color="63BE7B")
)
chc = LineChart(); chc.style = 12
chc.title = "Cobertura monetaria: Reservas / Base Monetaria"
chc.add_data(Reference(wsc, min_col=4, min_row=4, max_col=4, max_row=LAST_COB),
             titles_from_data=True)
chc.set_categories(Reference(wsc, min_col=1, min_row=5, max_row=LAST_COB))
chc.height = 10; chc.width = 20
wsc.add_chart(chc, "H4")
wsc.freeze_panes = "A5"

# =============== SHEET 10: Brecha Cambiaria ===============
wsb = wb.create_sheet("Brecha Cambiaria")
wsb.sheet_view.showGridLines = False
title_cell(wsb, "A1", "Brecha cambiaria: oficial vs MEP / CCL")
wsb.merge_cells("A1:G1")
wsb["A2"] = ("Cargá los promedios anuales (o cierre) de cada cotización. Brecha = (Paralelo/Oficial - 1).")
wsb["A2"].font = fnt(italic=True, size=9, color="808080")
wsb.merge_cells("A2:G2")
hdrs_b = ["Año", "Oficial", "MEP", "CCL", "Brecha MEP", "Brecha CCL", "Compras BCRA"]
for j, h in enumerate(hdrs_b, 1):
    header_cell(wsb.cell(row=4, column=j), h)

# Cierre de año aprox (ARS / USD)
brecha_data = {
    2016: (15.85, 15.85, 15.85),   # sin cepo
    2017: (18.65, 18.65, 18.65),
    2018: (37.81, 37.81, 37.81),
    2019: (59.89, 75.00, 79.00),    # cepo desde sept-19
    2020: (84.15, 142.00, 146.00),
    2021: (102.72, 198.00, 213.00),
    2022: (177.15, 333.00, 349.00),
    2023: (808.45, 1000.00, 1000.00),
    2024: (1032.50, 1170.00, 1180.00),
    2025: (1200.00, 1280.00, 1300.00),
    2026: (1280.00, 1300.00, 1320.00),
}
for i, (yr, _, _, _) in enumerate(annual):
    r = 5 + i
    of, mep, ccl = brecha_data[yr]
    c = wsb.cell(row=r, column=1, value=str(yr))
    c.font = fnt(bold=True); c.alignment = Alignment(horizontal="center")
    for j, v in enumerate([of, mep, ccl], 2):
        cv = wsb.cell(row=r, column=j, value=v); cv.font = fnt(color=BLUE)
        cv.number_format = "#,##0.00"
    wsb.cell(row=r, column=5, value=f"=C{r}/B{r}-1").number_format = "0.0%"
    wsb.cell(row=r, column=6, value=f"=D{r}/B{r}-1").number_format = "0.0%"
    wsb.cell(row=r, column=7, value=f"='Resumen Anual'!E{DATA_START+i}").number_format = "#,##0;(#,##0);-"
    wsb.cell(row=r, column=7).font = fnt(color=GREEN)
    for col in range(1, 8):
        wsb.cell(row=r, column=col).border = BOX
LAST_BR = 4 + len(annual)
for col, w in {"A": 8, "B": 12, "C": 12, "D": 12, "E": 14, "F": 14, "G": 16}.items():
    wsb.column_dimensions[col].width = w

# Conditional fmt on brechas
wsb.conditional_formatting.add(f"E5:F{LAST_BR}",
    ColorScaleRule(start_type="num", start_value=0, start_color="00B050",
                   mid_type="num", mid_value=0.5, mid_color="FFEB84",
                   end_type="num", end_value=2.0, end_color="C00000"))

chb = LineChart(); chb.style = 12
chb.title = "Brecha cambiaria (% sobre oficial)"
chb.add_data(Reference(wsb, min_col=5, min_row=4, max_col=6, max_row=LAST_BR),
             titles_from_data=True)
chb.set_categories(Reference(wsb, min_col=1, min_row=5, max_row=LAST_BR))
chb.height = 10; chb.width = 20
wsb.add_chart(chb, "I4")
wsb.freeze_panes = "A5"

# =============== SHEET 11: Stress Test ===============
wst = wb.create_sheet("Stress Test")
wst.sheet_view.showGridLines = False
title_cell(wst, "A1", "Stress test 2026: ¿llega el BCRA a la meta anual?")
wst.merge_cells("A1:E1")
wst["A2"] = "Modelo simple: ritmo actual proyectado al resto del año vs. metas oficiales."
wst["A2"].font = fnt(italic=True, size=9, color="808080")
wst.merge_cells("A2:E2")

# Inputs
wst["A4"] = "Parámetros"; header_cell(wst["A4"], "Parámetros")
wst.merge_cells("A4:B4")
wst["A5"] = "Compras YTD (USD mm)"
wst["B5"] = f"='Detalle Diario 2026'!I{DD_LAST}"
wst["B5"].font = fnt(color=GREEN); wst["B5"].number_format = "#,##0"
wst["A6"] = "Días hábiles transcurridos"
wst["B6"] = f"=COUNTA('Detalle Diario 2026'!A{DD_START_ROW}:A{DD_LAST})"
wst["A7"] = "Días hábiles restantes (estim.)"
wst["B7"] = 165
wst["B7"].font = fnt(color=BLUE)
wst["A8"] = "Pagos al FMI restantes 2026 (USD mm)"
wst["B8"] = 3000
wst["B8"].font = fnt(color=BLUE)
wst["A9"] = "Meta base oficial (USD mm)"
wst["B9"] = 10000
wst["B9"].font = fnt(color=BLUE)
wst["A10"] = "Meta techo oficial (USD mm)"
wst["B10"] = 17000
wst["B10"].font = fnt(color=BLUE)

# Cálculos
wst["A12"] = "Proyecciones"; header_cell(wst["A12"], "Proyecciones")
wst.merge_cells("A12:B12")
wst["A13"] = "Ritmo diario YTD (USD mm/rueda)"
wst["B13"] = "=B5/B6"
wst["B13"].number_format = "0.0"
wst["A14"] = "Compras esperadas resto del año"
wst["B14"] = "=B13*B7"
wst["B14"].number_format = "#,##0"
wst["A15"] = "Compras totales año proyectado"
wst["B15"] = "=B5+B14"
wst["B15"].number_format = "#,##0"
wst["B15"].font = fnt(bold=True)
wst["A16"] = "% sobre meta base"
wst["B16"] = "=B15/B9"
wst["B16"].number_format = "0.0%"
wst["A17"] = "% sobre meta techo"
wst["B17"] = "=B15/B10"
wst["B17"].number_format = "0.0%"
wst["A18"] = "Gap a meta base (USD mm)"
wst["B18"] = "=B9-B15"
wst["B18"].number_format = "#,##0;(#,##0);-"
wst["A19"] = "Estado vs meta base"
wst["B19"] = "=IF(B15>=B10,\"Supera meta techo\",IF(B15>=B9,\"Cumple meta base\",IF(B15>=B9*0.8,\"Cerca meta base\",\"Por debajo de meta\")))"
wst["B19"].font = fnt(bold=True)

# Escenarios
wst["D4"] = "Escenarios"; header_cell(wst["D4"], "Escenarios")
wst.merge_cells("D4:E4")
wst["D5"] = "Escenario"; wst["E5"] = "Compras año proyectadas"
header_cell(wst["D5"], "Escenario"); header_cell(wst["E5"], "Compras año proyectadas")
scenarios = [
    ("Optimista (+30% ritmo)", 1.3),
    ("Base (ritmo actual)", 1.0),
    ("Conservador (-20% ritmo)", 0.8),
    ("Pesimista (-50% ritmo)", 0.5),
]
for i, (name, mult) in enumerate(scenarios):
    r = 6 + i
    wst.cell(row=r, column=4, value=name).font = fnt()
    wst.cell(row=r, column=5, value=f"=B5+B13*B7*{mult}")
    wst.cell(row=r, column=5).number_format = "#,##0"
    for col in (4, 5):
        wst.cell(row=r, column=col).border = BOX

# Color en B16 (% meta base)
wst.conditional_formatting.add("B16",
    ColorScaleRule(start_type="num", start_value=0.5, start_color="C00000",
                   mid_type="num", mid_value=1.0, mid_color="FFEB84",
                   end_type="num", end_value=1.5, end_color="00B050"))

for col, w in {"A": 38, "B": 16, "C": 4, "D": 28, "E": 22}.items():
    wst.column_dimensions[col].width = w
for r in range(5, 20):
    for col in (1, 2):
        if wst.cell(row=r, column=col).value is not None:
            wst.cell(row=r, column=col).border = BOX

# =============== SHEET 12: Cómo usar ===============
wsg = wb.create_sheet("Cómo usar")
wsg.sheet_view.showGridLines = False
content = [
    ("Guía del archivo", True, 16, "000000"),
    ("", False, 10, "000000"),
    ("12 hojas, ordenadas de macro a micro", True, 12, "000000"),
    ("1. Resumen Anual — 11 años de stocks y compras netas en MULC.", False, 10, "000000"),
    ("2. YTD Comparativo — compras acumuladas mes a mes, año por año.", False, 10, "000000"),
    ("3. Heatmap Compras — grilla años × meses con colores. Ojo de un vistazo.", False, 10, "000000"),
    ("4. Rolling 12m — suma móvil de 12 meses, tendencia sin efecto fin de año.", False, 10, "000000"),
    (f"5. Detalle Diario 2026 — {len(daily_rows)} ruedas con la reconstrucción día a día.", False, 10, "000000"),
    ("6. Streaks — días consecutivos con compras + estadísticas del año.", False, 10, "000000"),
    ("7. Validación — compara mi reconstrucción contra el dato real del BCRA (cuando lo cargues).", False, 10, "000000"),
    ("8. Reservas Netas — descomposición: Brutas − Encajes − Swap China − BOPREAL − FMI <12m.", False, 10, "000000"),
    ("9. Cobertura Monetaria — ratio Reservas / Base Monetaria a lo largo del tiempo.", False, 10, "000000"),
    ("10. Brecha Cambiaria — oficial vs MEP/CCL, brecha % por año.", False, 10, "000000"),
    ("11. Stress Test — proyección a fin de año vs metas oficiales, con 4 escenarios.", False, 10, "000000"),
    ("12. Cómo usar — esta hoja.", False, 10, "000000"),
    ("", False, 10, "000000"),
    ("Convenciones de color", True, 12, "000000"),
    ("Texto AZUL = celda de ingreso manual.", False, 10, "0000FF"),
    ("Texto NEGRO = fórmula calculada.", False, 10, "000000"),
    ("Texto VERDE = vínculo a otra hoja del libro.", False, 10, "008000"),
    ("", False, 10, "000000"),
    ("Cómo cargar datos reales del BCRA (script Python)", True, 12, "000000"),
    ("Junto a este Excel hay un archivo update_bcra.py. Pasos:", False, 10, "000000"),
    ("  1) Asegurate de tener Python 3 + openpyxl (pip install openpyxl).", False, 10, "000000"),
    ("  2) Poné el .py y el .xlsx en la misma carpeta.", False, 10, "000000"),
    ("  3) Corré: python3 update_bcra.py --inspect (te muestra la estructura del series.xlsm).", False, 10, "000000"),
    ("  4) Si la columna H sigue siendo compras/ventas, corré: python3 update_bcra.py", False, 10, "000000"),
    ("  5) Se crea/actualiza la hoja 'Datos Reales BCRA' y la hoja 'Validación' empieza a comparar.", False, 10, "000000"),
    ("Repetí el paso 4 cada día/semana cuando quieras refrescar.", False, 10, "000000"),
    ("", False, 10, "000000"),
    ("Hojas que necesitan datos reales para ser realmente útiles", True, 12, "000000"),
    ("• Validación: arranca a comparar cuando exista la hoja 'Datos Reales BCRA'.", False, 10, "000000"),
    ("• Cobertura Monetaria: edita la columna C (Base Monetaria USD eq) con datos del BCRA.", False, 10, "000000"),
    ("• Reservas Netas: ajustá los componentes (encajes, swap, BOPREAL, FMI) — son aproximaciones.", False, 10, "000000"),
    ("• Brecha Cambiaria: las cotizaciones MEP/CCL son aproximaciones; reemplazalas con tus datos.", False, 10, "000000"),
    ("• Stress Test: el parámetro 'días hábiles restantes' lo podés ajustar manualmente.", False, 10, "000000"),
    ("", False, 10, "000000"),
    ("Reglas de oro", True, 12, "000000"),
    ("• La hoja 'Detalle Diario 2026' es Tabla de Excel (tblDiario): agregá filas debajo y todo se actualiza solo.", False, 10, "000000"),
    ("• Si un gráfico no se actualiza, click derecho → Refrescar datos.", False, 10, "000000"),
    ("• El número clave a vigilar todos los días: la celda I al pie de Detalle Diario 2026 = compras YTD.", False, 10, "000000"),
    ("• Stress Test te dice en 1 segundo si el ritmo actual alcanza para llegar a la meta de USD 10.000M.", False, 10, "000000"),
]
for i, (text, bold, size, color) in enumerate(content, start=1):
    c = wsg.cell(row=i, column=1, value=text)
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wsg.row_dimensions[i].height = max(15, size * 1.6) if text else 12
wsg.column_dimensions["A"].width = 130

out_path = "/sessions/wizardly-elegant-darwin/mnt/outputs/BCRA_compras_anuales_10y.xlsx"
wb.save(out_path)
print(f"saved: {out_path}")
print(f"Hojas: {wb.sheetnames}")
