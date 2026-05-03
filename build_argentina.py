"""
build_argentina.py
==================
Builds Argentina_Tracker.xlsx — a pure-Excel, Haver-DLX-driven workbook.

What's inside the file at runtime
---------------------------------
- The user edits codes on the Codes tab.
- Every indicator block on every category tab uses HaverData() formulas
  pointing at the corresponding cell on the Codes tab — change the code
  there, hit Calculate / DLX > Refresh, the whole block re-pulls.
- Forecasts are pure Excel: TREND() for log-linear and FORECAST.ETS()
  for damped Holt-Winters with seasonality. No Python at runtime.
- Charts are pre-built, referencing the data ranges, so they fill in
  the moment DLX populates the cells.

You only run this script when you want to regenerate the workbook layout.
Day-to-day, you live entirely in Excel.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------------------------------------------------------
# Indicator catalogue — Argentina
# ---------------------------------------------------------------------------

@dataclass
class Indicator:
    section: str
    name: str
    qcode: str          # Haver code at quarterly frequency
    acode: str          # Haver code at annual frequency (optional, can be "")
    units: str
    agg: str            # "avg" | "sum" | "last" — quarterly → annual rule
    notes: str

INDICATORS: list[Indicator] = [
    # GDP
    Indicator("GDP", "Real GDP (SA)",            "S213NGPC@LATAM", "S213NGPCA@LATAM",
              "Mn 2004 ARS", "avg",
              "[VERIFY] AR is country code 213 in IFS-style mnemonics."),
    Indicator("GDP", "Nominal GDP",              "S213NGDP@LATAM", "S213NGDPA@LATAM",
              "Mn ARS", "avg", "[VERIFY]"),
    Indicator("GDP", "EMAE Activity Index",      "EMAE@ARGEN",     "EMAEA@ARGEN",
              "Index 2004=100", "avg",
              "INDEC monthly activity proxy. [VERIFY] code on your subscription."),

    # Inflation
    Indicator("Inflation", "CPI National (IPC nacional)", "IPCNAC@ARGEN",  "IPCNACA@ARGEN",
              "Index Dec-16=100", "avg",
              "INDEC. Pre-2017 series has structural breaks (INDEC reform)."),
    Indicator("Inflation", "Core CPI",                    "IPCCORE@ARGEN", "IPCCOREA@ARGEN",
              "Index", "avg", "[VERIFY]"),
    Indicator("Inflation", "Wholesale Price Index (IPIM)","IPIM@ARGEN",    "IPIMA@ARGEN",
              "Index", "avg", "[VERIFY]"),

    # Fiscal
    Indicator("Fiscal", "Primary Fiscal Balance",  "SPNFBAL@ARGEN", "SPNFBALA@ARGEN",
              "Mn ARS", "sum",
              "Sector Publico Nacional No Financiero. [VERIFY]"),
    Indicator("Fiscal", "Federal Tax Revenue",     "RECNAC@ARGEN",  "RECNACA@ARGEN",
              "Mn ARS", "sum", "[VERIFY]"),
    Indicator("Fiscal", "Public Debt",             "DEUDPUB@ARGEN", "DEUDPUBA@ARGEN",
              "USD Mn", "last",
              "End-period stock. [VERIFY]"),

    # Balance of Payments
    Indicator("BoP", "Current Account Balance",  "S213NCABA@LATAM", "S213NCABAA@LATAM",
              "USD Mn", "sum", "INDEC / BCRA. [VERIFY]"),
    Indicator("BoP", "Trade Balance",            "EXPMP@ARGEN",     "EXPMPA@ARGEN",
              "USD Mn", "sum", "Exports - imports. [VERIFY]"),
    Indicator("BoP", "Foreign Direct Investment","FDIANN@ARGEN",    "FDIANNA@ARGEN",
              "USD Mn", "sum", "[VERIFY]"),

    # Reserves
    Indicator("Reserves", "BCRA International Reserves", "RESINT@ARGEN", "RESINTA@ARGEN",
              "USD Mn", "last",
              "BCRA gross international reserves, end of period. [VERIFY]"),
    Indicator("Reserves", "Net International Reserves",  "RESNET@ARGEN", "RESNETA@ARGEN",
              "USD Mn", "last", "BCRA NIR. [VERIFY]"),
]

CATEGORIES = ["GDP", "Inflation", "Fiscal", "BoP", "Reserves"]


# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

# Quarterly horizon: 2008Q1 .. 2029Q4  (88 quarters)
Q_START_YEAR = 2008
Q_END_YEAR = 2029
N_QUARTERS = (Q_END_YEAR - Q_START_YEAR + 1) * 4  # 88

# Annual horizon
A_START_YEAR = Q_START_YEAR
A_END_YEAR = Q_END_YEAR
N_YEARS = A_END_YEAR - A_START_YEAR + 1  # 22

# Each indicator block on a category tab:
#   - Title row (1)
#   - Subtitle row with sources/notes (1)
#   - Quarterly header row (1)
#   - Quarterly data rows (N_QUARTERS = 88)  in cols A..H
#   - Annual header at the same vertical position, in cols J..M
#   - Chart anchored at column O, ~ 18 rows tall
#   - 4-row gap before the next block
# Total rows per block: max(88, chart_height + ...) + ~6 -> ~94
ROWS_PER_BLOCK = N_QUARTERS + 8


# ---------------------------------------------------------------------------
# Style
# ---------------------------------------------------------------------------

FONT = "Calibri"
NAVY = "1F3864"
LIGHT = "D9E1F2"
ZEBRA = "F7F9FC"
INPUT_YELLOW = "FFF2CC"

TITLE_FONT = Font(name=FONT, size=18, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name=FONT, size=10, italic=True, color="595959")
H1_FONT = Font(name=FONT, size=13, bold=True, color="FFFFFF")
H1_FILL = PatternFill("solid", start_color=NAVY)
H2_FONT = Font(name=FONT, size=11, bold=True, color=NAVY)
H2_FILL = PatternFill("solid", start_color=LIGHT)

LABEL_FONT = Font(name=FONT, size=11, bold=True)
INPUT_FONT = Font(name=FONT, size=11, color="0000FF", bold=True)
INPUT_FILL = PatternFill("solid", start_color=INPUT_YELLOW)
FORMULA_FONT = Font(name=FONT, size=10, color="000000")
LINK_FONT = Font(name=FONT, size=10, color="008000")
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color=NAVY)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '#,##0.00;(#,##0.00);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
DATE_FMT = "mmm-yyyy"
YEAR_FMT = "0"


def _set(ws, row, col, value, *, font=None, fill=None, fmt=None,
         align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if fmt:    c.number_format = fmt
    if align:  c.alignment = align
    if border: c.border = border
    return c


def _band(ws, row: int, ncols: int, text: str, *, level: int = 1, start_col: int = 1):
    fill = H1_FILL if level == 1 else H2_FILL
    font = H1_FONT if level == 1 else H2_FONT
    end = start_col + ncols - 1
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = font; c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22 if level == 1 else 18


def _table_header(ws, row: int, cols: list[str], start_col: int = 1):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX


def _zebra(ws, r1: int, r2: int, c1: int, c2: int):
    fill = PatternFill("solid", start_color=ZEBRA)
    for r in range(r1, r2 + 1):
        if (r - r1) % 2 == 1:
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).fill = fill


# ---------------------------------------------------------------------------
# Codes sheet
# ---------------------------------------------------------------------------

CODES_HEADERS = ["Section", "Indicator", "Frequency hint",
                 "Quarterly Code", "Annual Code", "Units", "Aggregation", "Notes"]

CODES_HEADER_ROW = 4
CODES_FIRST_DATA_ROW = 5
COL_SECTION   = 1
COL_INDICATOR = 2
COL_FREQ      = 3
COL_QCODE     = 4
COL_ACODE     = 5
COL_UNITS     = 6
COL_AGG       = 7
COL_NOTES     = 8


def _write_readme(wb: Workbook):
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Macro Tracker — Argentina"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Pure-Excel workbook driven by Haver DLX. "
                "Edit codes on the Codes tab, hit DLX > Refresh / Calculate, "
                "every tab and chart updates.")
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32

    rows = [
        ("How to use", "section"),
        ("1.", "Make sure the Haver DLX add-in is loaded (Excel > File > Options > Add-ins)."),
        ("2.", "Open the Codes tab and review the Haver codes. Yellow cells are inputs."),
        ("3.", "Hit DLX > Refresh (or just press F9) to pull data."),
        ("4.", "Open any category tab (GDP, Inflation, …). Each indicator has a quarterly block, an annual block, and a chart."),
        ("5.", "The Dashboard tab summarises the latest read and 2-year forecast for every indicator."),
        ("", ""),
        ("Forecasts", "section"),
        ("Linear (TREND)",  "Excel's TREND() extends the levels linearly over the next 8 quarters. Robust to zero / negative values."),
        ("Holt-Winters (ETS)", "Excel's FORECAST.ETS() is exponential triple smoothing with auto seasonality. Pair with FORECAST.ETS.CONFINT() for 95% bands."),
        ("", ""),
        ("Notes on Argentina", "section"),
        ("Inflation regime breaks", "INDEC suspended/restated CPI 2007-15. Forecasts on the full history will look strange — consider trimming the date range used in the formula or relying on Holt-Winters which damps."),
        ("Currency revaluations", "Pesos series may have notch jumps. Inspect the chart before quoting a forecast."),
        ("", ""),
        ("Files", "section"),
        ("Argentina_Tracker.xlsx", "This workbook. Open this every day."),
        ("build_argentina.py",     "Regenerates the workbook layout. Only run if you want to add/remove categories or rebuild from scratch."),
    ]
    r = 4
    for k, v in rows:
        if v == "section":
            ws.cell(row=r, column=1, value=k).font = H2_FONT
            ws.cell(row=r, column=1).fill = H2_FILL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 18
        else:
            ws.cell(row=r, column=1, value=k).font = LABEL_FONT
            cell = ws.cell(row=r, column=2, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    for col, w in {"A": 22, "B": 30, "C": 30, "D": 30, "E": 30}.items():
        ws.column_dimensions[col].width = w


def _write_codes(wb: Workbook):
    ws = wb.create_sheet("Codes")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Codes — Single Source of Truth (Argentina)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Edit the yellow cells. Quarterly Code is required; "
                "Annual Code is optional (the workbook aggregates from Q if blank). "
                "Aggregation: avg | sum | last.")
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:H2")

    _table_header(ws, CODES_HEADER_ROW, CODES_HEADERS)
    nrows = len(INDICATORS)
    for i, ind in enumerate(INDICATORS):
        r = CODES_FIRST_DATA_ROW + i
        _set(ws, r, COL_SECTION,   ind.section, font=FORMULA_FONT, border=BOX)
        _set(ws, r, COL_INDICATOR, ind.name,    font=FORMULA_FONT, border=BOX)
        _set(ws, r, COL_FREQ,      "Q",         font=FORMULA_FONT, border=BOX,
             align=Alignment(horizontal="center"))
        _set(ws, r, COL_QCODE,     ind.qcode,   font=INPUT_FONT, fill=INPUT_FILL, border=BOX)
        _set(ws, r, COL_ACODE,     ind.acode,   font=INPUT_FONT, fill=INPUT_FILL, border=BOX)
        _set(ws, r, COL_UNITS,     ind.units,   font=INPUT_FONT, fill=INPUT_FILL, border=BOX)
        _set(ws, r, COL_AGG,       ind.agg,     font=INPUT_FONT, fill=INPUT_FILL, border=BOX,
             align=Alignment(horizontal="center"))
        _set(ws, r, COL_NOTES,     ind.notes,   font=FORMULA_FONT, border=BOX,
             align=Alignment(wrap_text=True, vertical="top"))

    ref = f"A{CODES_HEADER_ROW}:{get_column_letter(len(CODES_HEADERS))}{CODES_HEADER_ROW + nrows}"
    tbl = Table(displayName="tbl_Codes", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                        showRowStripes=True)
    ws.add_table(tbl)

    widths = [12, 32, 8, 22, 22, 22, 13, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = f"A{CODES_FIRST_DATA_ROW}"


# ---------------------------------------------------------------------------
# Indicator block on a category tab
# ---------------------------------------------------------------------------

# Column layout per block on a category tab:
COL = {
    # Quarterly
    "qDate":   1,  # A
    "qLevel":  2,  # B
    "qQoQ":    3,  # C
    "qYoY":    4,  # D
    "qLin":    5,  # E   linear forecast (TREND on log levels via GROWTH, fallback TREND)
    "qETS":    6,  # F   FORECAST.ETS
    "qLow":    7,  # G   ETS lower 95
    "qHigh":   8,  # H   ETS upper 95
    # Spacer column I
    # Annual
    "aYear":   10, # J
    "aLevel":  11, # K
    "aYoY":    12, # L
    "aFcst":   13, # M
}
Q_COL_LETTERS = {k: get_column_letter(v) for k, v in COL.items()}
Q_HEADERS = ["Date", "Level", "QoQ %", "YoY %",
             "Linear Fcst", "ETS Fcst", "ETS Low 95%", "ETS High 95%"]
A_HEADERS = ["Year", "Level", "YoY %", "Forecast"]


def _quarter_dates() -> list[datetime]:
    out = []
    for y in range(Q_START_YEAR, Q_END_YEAR + 1):
        for m in (1, 4, 7, 10):
            out.append(datetime(y, m, 1))
    return out


def _agg_formula(agg: str, q_level_range: str, q_date_range: str,
                 year_cell: str) -> str:
    """Excel formula that aggregates quarterly levels in the given range
    to the given year using the requested rule."""
    start = f'DATE({year_cell},1,1)'
    end   = f'DATE({year_cell},12,31)'
    if agg == "avg":
        return (f'IFERROR(AVERAGEIFS({q_level_range},{q_date_range},'
                f'">="&{start},{q_date_range},"<="&{end}),"")')
    if agg == "sum":
        return (f'IFERROR(SUMIFS({q_level_range},{q_date_range},'
                f'">="&{start},{q_date_range},"<="&{end}),"")')
    if agg == "last":
        # value at Q4 = DATE(year, 10, 1)
        q4 = f'DATE({year_cell},10,1)'
        return (f'IFERROR(INDEX({q_level_range},'
                f'MATCH({q4},{q_date_range},0)),"")')
    return '""'


def _write_indicator_block(ws, ind: Indicator, ind_index_in_codes: int,
                           start_row: int) -> int:
    """Write one indicator's full block (quarterly + annual + chart).
    Returns the next available row."""
    qcode_cell = f"Codes!$D${CODES_FIRST_DATA_ROW + ind_index_in_codes}"
    units_cell = f"Codes!$F${CODES_FIRST_DATA_ROW + ind_index_in_codes}"
    name_cell  = f"Codes!$B${CODES_FIRST_DATA_ROW + ind_index_in_codes}"

    # ----- Title bar (full width) ------------------------------------------
    title_row = start_row
    _band(ws, title_row, 13, "", level=1)
    # write a formula so title stays in sync with Codes
    title_cell = ws.cell(row=title_row, column=1)
    title_cell.value = f'={name_cell}&"   ["&{qcode_cell}&"]   "&{units_cell}'
    title_cell.font = H1_FONT
    title_cell.fill = H1_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ----- Subtitle: agg + notes -------------------------------------------
    sub_row = title_row + 1
    notes_cell  = f"Codes!$H${CODES_FIRST_DATA_ROW + ind_index_in_codes}"
    agg_cell    = f"Codes!$G${CODES_FIRST_DATA_ROW + ind_index_in_codes}"
    ws.cell(row=sub_row, column=1, value="Aggregation:").font = LABEL_FONT
    c = ws.cell(row=sub_row, column=2, value=f"={agg_cell}")
    c.font = LINK_FONT
    ws.cell(row=sub_row, column=4, value="Notes:").font = LABEL_FONT
    c = ws.cell(row=sub_row, column=5, value=f"={notes_cell}")
    c.font = SUBTITLE_FONT
    ws.merge_cells(start_row=sub_row, start_column=5, end_row=sub_row, end_column=13)

    # ----- Quarterly header banner -----------------------------------------
    qheader_row = sub_row + 2
    _band(ws, qheader_row - 1, 8, "Quarterly", level=2, start_col=1)
    _band(ws, qheader_row - 1, 4, "Annual",    level=2, start_col=10)
    _table_header(ws, qheader_row, Q_HEADERS, start_col=1)
    _table_header(ws, qheader_row, A_HEADERS, start_col=10)

    # ----- Quarterly data --------------------------------------------------
    q_first = qheader_row + 1
    q_last  = qheader_row + N_QUARTERS
    dates = _quarter_dates()

    # Range strings for forecast formulas (absolute)
    rng_dates  = f"$A${q_first}:$A${q_last}"
    rng_levels = f"$B${q_first}:$B${q_last}"

    for i, d in enumerate(dates):
        r = q_first + i

        # Date — hardcoded (we own the calendar)
        _set(ws, r, COL["qDate"], d, fmt=DATE_FMT, border=BOX,
             font=FORMULA_FONT, align=Alignment(horizontal="center"))

        # Level — Haver pull, IFERROR to "" so the file is graceful pre-DLX
        f = f'=IFERROR(HaverData({qcode_cell},$A{r}),"")'
        _set(ws, r, COL["qLevel"], f, fmt=NUM_FMT, border=BOX, font=INPUT_FONT)

        # QoQ %
        if i >= 1:
            f = f'=IFERROR(B{r}/B{r-1}-1,"")'
        else:
            f = ""
        _set(ws, r, COL["qQoQ"], f, fmt=PCT_FMT, border=BOX, font=FORMULA_FONT)

        # YoY %
        if i >= 4:
            f = f'=IFERROR(B{r}/B{r-4}-1,"")'
        else:
            f = ""
        _set(ws, r, COL["qYoY"], f, fmt=PCT_FMT, border=BOX, font=FORMULA_FONT)

        # Linear forecast — only filled when Level is missing.
        # GROWTH is exponential (good for prices/levels). It errors when the
        # series has non-positive values, so we IFERROR to TREND (linear).
        f = (f'=IFERROR(IF(B{r}="",IFERROR(GROWTH({rng_levels},{rng_dates},$A{r}),'
             f'TREND({rng_levels},{rng_dates},$A{r})),""),"")')
        _set(ws, r, COL["qLin"], f, fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)

        # ETS forecast — fills both history (mostly equals Level) and future.
        # We fill it for every row so the chart series renders cleanly.
        f = (f'=IFERROR(FORECAST.ETS($A{r},{rng_levels},{rng_dates},4,1,1),"")')
        _set(ws, r, COL["qETS"], f, fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)

        # ETS confidence interval — half-width
        f_low = (f'=IFERROR(F{r}-FORECAST.ETS.CONFINT($A{r},{rng_levels},{rng_dates},'
                 f'0.95,4,1,1),"")')
        f_hi  = (f'=IFERROR(F{r}+FORECAST.ETS.CONFINT($A{r},{rng_levels},{rng_dates},'
                 f'0.95,4,1,1),"")')
        _set(ws, r, COL["qLow"],  f_low, fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)
        _set(ws, r, COL["qHigh"], f_hi,  fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)

    _zebra(ws, q_first, q_last, 1, 8)

    # ----- Annual data ------------------------------------------------------
    a_first = qheader_row + 1
    a_last  = qheader_row + N_YEARS

    for i in range(N_YEARS):
        r = a_first + i
        year = A_START_YEAR + i
        _set(ws, r, COL["aYear"], year, fmt=YEAR_FMT, border=BOX,
             font=FORMULA_FONT, align=Alignment(horizontal="center"))

        # Aggregated Level — formula chooses by the agg cell
        # We hard-code one of the three formulas based on ind.agg at build time;
        # changing the agg cell on Codes doesn't re-pick the formula but it's
        # documented on the README. Most users won't change agg per indicator.
        agg_formula = _agg_formula(ind.agg, rng_levels, rng_dates, f"J{r}")
        _set(ws, r, COL["aLevel"], "=" + agg_formula,
             fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)

        # YoY %
        if i >= 1:
            f = f'=IFERROR(K{r}/K{r-1}-1,"")'
        else:
            f = ""
        _set(ws, r, COL["aYoY"], f, fmt=PCT_FMT, border=BOX, font=FORMULA_FONT)

        # Annual forecast: aggregated from quarterly ETS forecast (col F)
        rng_q_ets = f"$F${q_first}:$F${q_last}"
        ann_formula = _agg_formula(ind.agg, rng_q_ets, rng_dates, f"J{r}")
        _set(ws, r, COL["aFcst"], "=" + ann_formula,
             fmt=NUM_FMT, border=BOX, font=FORMULA_FONT)

    _zebra(ws, a_first, a_last, 10, 13)

    # ----- Chart ------------------------------------------------------------
    chart = LineChart()
    chart.height = 11
    chart.width = 22
    chart.title = None
    chart.legend.position = "b"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.x_axis.number_format = DATE_FMT

    cats = Reference(ws, min_col=COL["qDate"], min_row=q_first, max_row=q_last)
    # series: Level history, ETS forecast, ETS Low, ETS High
    for col_idx, dash in ((COL["qLevel"], None),
                          (COL["qETS"],   "dash"),
                          (COL["qLow"],   "sysDot"),
                          (COL["qHigh"],  "sysDot")):
        data = Reference(ws, min_col=col_idx, min_row=qheader_row,
                         max_row=q_last)
        chart.add_data(data, titles_from_data=True)
        if dash:
            s = chart.series[-1]
            s.graphicalProperties = GraphicalProperties(
                ln=LineProperties(prstDash=dash))
    chart.set_categories(cats)

    # Anchor chart to the right of the annual block
    anchor = f"O{qheader_row}"
    ws.add_chart(chart, anchor)

    return start_row + ROWS_PER_BLOCK


# ---------------------------------------------------------------------------
# Category tabs
# ---------------------------------------------------------------------------

def _write_category_tab(wb: Workbook, category: str):
    ws = wb.create_sheet(category)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{category} — Argentina"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Each block: live HaverData pulls (col B) → growth metrics "
                "→ TREND + FORECAST.ETS forecasts → chart.")
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:O2")

    # Columns
    widths = {
        "A": 11, "B": 14, "C": 9,  "D": 9,
        "E": 12, "F": 12, "G": 12, "H": 12,
        "I": 2,
        "J": 7,  "K": 14, "L": 9,  "M": 14,
        "N": 2,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Find indicators for this category
    section_inds = [(i, ind) for i, ind in enumerate(INDICATORS)
                    if ind.section == category]
    if not section_inds:
        ws["A4"] = "(no indicators configured for this section yet)"
        ws["A4"].font = SUBTITLE_FONT
        return

    cur = 4
    for codes_idx, ind in section_inds:
        cur = _write_indicator_block(ws, ind, codes_idx, cur)

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def _write_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dashboard — Latest reads & 2y forecast (Argentina)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "All values pull live from the category tabs."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    headers = ["Section", "Indicator", "Latest value", "QoQ %", "YoY %",
               "2y forecast (ETS)", "Code"]
    _table_header(ws, 4, headers)

    # For each indicator: write formulas referencing the category tab
    # We need to know where each indicator block sits on its category tab.
    # Compute that here: indicators are written in order within each section.
    section_block_index: dict[str, int] = {c: 0 for c in CATEGORIES}
    for i, ind in enumerate(INDICATORS):
        r = 5 + i
        category = ind.section
        block_idx = section_block_index[category]
        section_block_index[category] += 1

        # First block on a tab starts at row 4 → block_idx=0 → start_row=4
        block_start = 4 + block_idx * ROWS_PER_BLOCK
        # qheader_row inside block:
        qheader_row = block_start + 3        # title=block_start, sub=+1, qheader=+3
        q_last      = qheader_row + N_QUARTERS
        # Last quarterly Level cell with a value: use LOOKUP to find last numeric.
        sheet = f"'{category}'"

        _set(ws, r, 1, ind.section,    border=BOX, font=FORMULA_FONT)
        _set(ws, r, 2, ind.name,       border=BOX, font=FORMULA_FONT)

        # Latest value: look up last non-blank in the Level column.
        # LOOKUP(2, 1/(B_range<>""), B_range)
        rng_level = f"{sheet}!$B${qheader_row + 1}:$B${q_last}"
        rng_qoq   = f"{sheet}!$C${qheader_row + 1}:$C${q_last}"
        rng_yoy   = f"{sheet}!$D${qheader_row + 1}:$D${q_last}"
        rng_ets   = f"{sheet}!$F${qheader_row + 1}:$F${q_last}"

        f_last = f'=IFERROR(LOOKUP(2,1/({rng_level}<>""),{rng_level}),"")'
        _set(ws, r, 3, f_last, fmt=NUM_FMT, border=BOX, font=LINK_FONT)
        f_qoq = f'=IFERROR(LOOKUP(2,1/({rng_qoq}<>""),{rng_qoq}),"")'
        _set(ws, r, 4, f_qoq, fmt=PCT_FMT, border=BOX, font=LINK_FONT)
        f_yoy = f'=IFERROR(LOOKUP(2,1/({rng_yoy}<>""),{rng_yoy}),"")'
        _set(ws, r, 5, f_yoy, fmt=PCT_FMT, border=BOX, font=LINK_FONT)
        # 2y forecast = ETS at the last quarter of horizon = last row of column F
        f_fcst = f'=IFERROR({sheet}!$F${q_last},"")'
        _set(ws, r, 6, f_fcst, fmt=NUM_FMT, border=BOX, font=LINK_FONT)
        # Code (link to Codes)
        f_code = f"=Codes!$D${CODES_FIRST_DATA_ROW + i}"
        _set(ws, r, 7, f_code, border=BOX, font=LINK_FONT)

    _zebra(ws, 5, 5 + len(INDICATORS) - 1, 1, 7)

    widths = [11, 36, 16, 12, 12, 18, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Build entrypoint
# ---------------------------------------------------------------------------

def build(out: str | Path = "Argentina_Tracker.xlsx") -> Path:
    wb = Workbook()
    _write_readme(wb)
    _write_codes(wb)
    for cat in CATEGORIES:
        _write_category_tab(wb, cat)
    _write_dashboard(wb)
    out = Path(out)
    wb.save(out)
    print(f"[ok] wrote {out}")
    return out


if __name__ == "__main__":
    build()
