#!/usr/bin/env python3
"""
check_isins.py — "why are these ISINs missing?" -> one Excel with the answers.

HOW TO USE (3 steps):
  1. Open  isins_to_check.txt , paste the ISINs people asked about (one per
     line), save it.
  2. Run:  python check_isins.py
  3. Open  outputs/why_missing.xlsx  -> one row per ISIN with the reason.

It uses the SAME eligibility logic as the real weekly build, so its answers
always match the published PDF / Offshore / Onshore outputs.
"""
import os
import sys
import datetime

import gem_report_builder_v3 as G
from gem_report_builder_v3 import EXCLUDED_CURRENCIES
from gem_excel_builder import is_onshore_eligible, is_offshore_eligible, _has_issuer_name

HERE = os.path.dirname(os.path.abspath(__file__))
CURR = os.path.join(HERE, 'data', 'current')
PREV = os.path.join(HERE, 'data', 'previous')
ISIN_FILE  = os.path.join(HERE, 'isins_to_check.txt')
OUT_DIR    = os.path.join(HERE, 'outputs')
OUT_XLSX   = os.path.join(OUT_DIR, 'why_missing.xlsx')


# Plain-English explanation for each master-list reason code.
EXPLAIN = {
    'legal_excluded':       'Legal asked us to pull it (on the legal exclusions list).',
    'not_gem':              "Not tagged into the GEM universe (TopListCategory is not 'GEM' in the bond-update file). Most common cause.",
    'near_maturity':        'Matures within 180 days, so it drops off (except defaulted Venezuelan sovereigns).',
    'subordinated_no_flag': 'Subordinated bond with no WMRFlag and no usable rating, so it is flagged, not included.',
    'eligible':             'Eligible.',
    'subordinated_ig':      'Eligible (subordinated, investment grade).',
    'subordinated_hy':      'Eligible (subordinated, high yield).',
    'subordinated_unrated': 'Eligible (subordinated, unrated - included as HY).',
}


def load_isins():
    if len(sys.argv) > 1:                     # ISINs passed on the command line win
        return [a.strip().upper() for a in sys.argv[1:] if a.strip()]
    if not os.path.exists(ISIN_FILE):
        sys.exit(f"Can't find {ISIN_FILE}. Create it and paste ISINs, one per line.")
    out = []
    for line in open(ISIN_FILE, encoding='utf-8'):
        line = line.strip()
        if line and not line.startswith('#'):
            out.append(line.upper())
    if not out:
        sys.exit(f"No ISINs found in {ISIN_FILE}. Paste some in (one per line) and re-run.")
    return out


def build_data():
    return G.GEMData({
        'bond_data':      os.path.join(CURR, 'CurrentPublishableBondData.txt'),
        'issuer_data':    os.path.join(CURR, 'CurrentPublishableIssuerData.txt'),
        'bond_update':    os.path.join(CURR, 'PublishableBondDataUpdate.txt'),
        'issuer_update':  os.path.join(CURR, 'PublishableIssuerDataUpdate.txt'),
        'color_flags':    os.path.join(CURR, 'PublishableColorFlags.txt'),
        'issuer_texts':   os.path.join(CURR, 'IssuerTexts.txt'),
        'issuer_ratings': os.path.join(CURR, 'IssuerRatings.txt'),
        'prev_bond_data':   os.path.join(PREV, 'CurrentPublishableBondData.txt'),
        'prev_bond_update': os.path.join(PREV, 'PublishableBondDataUpdate.txt'),
        'priips_ref':       None,
        'legal_exclusions': _find_legal_exclusions(),
    })


def _find_legal_exclusions():
    for folder in (os.path.join(HERE, 'data'), CURR, HERE):
        if os.path.isdir(folder):
            for f in os.listdir(folder):
                lf = f.lower()
                if lf.endswith(('.csv', '.txt')) and ('legal' in lf or 'exclusion' in lf) \
                        and not f.startswith('~$'):
                    return os.path.join(folder, f)
    return None


def build_published_sets(data):
    """Reproduce EXACTLY what each real output contains, by ISIN.
    These call the same functions the PDF/Excel builders use, so membership
    can't disagree with the published files."""
    def isins(bonds):
        return {(b.get('Isin') or '').strip() for b in bonds}
    return {
        # PDF pages
        'pdf_reference': isins(data.reference_list_bonds()),   # main currency/region tables
        'pdf_sell':      isins(data.sell_list_bonds()),        # Sell Recommendations page
        'pdf_top':       isins(data.top_list_bonds()),         # Top List overlay
        # Excels (eligibility + must have an issuer name)
        'offshore': isins(b for b in data.em_bonds
                          if is_offshore_eligible(b, data) and _has_issuer_name(b, data)),
        'onshore':  isins(b for b in data.em_bonds
                          if is_onshore_eligible(b, data) and _has_issuer_name(b, data)),
    }


def _why_not_in_pdf(bond, data):
    """A bond can be eligible yet absent from the PDF's main table. Say why."""
    rec = (bond.get('WMR_Bond_Recommendation') or '').strip()
    ccy = (bond.get('CCY') or '').strip().upper()
    if rec == 'Sell':
        return 'It is Sell-rated, so it sits on the "Sell Recommendations" page, not the main table.'
    if not ccy:
        return 'It has NO currency (CCY blank) in the feed, so the PDF skips it. Data problem - flag to the desk.'
    if ccy in EXCLUDED_CURRENCIES:
        return (f'It is in an EXCLUDED currency ({ccy}). Local-currency bonds '
                f'({", ".join(sorted(EXCLUDED_CURRENCIES))}) are deliberately kept off the list.')
    if not _has_issuer_name(bond, data):
        return 'It has NO issuer name in the feed, so it is dropped. Data problem - flag to the desk.'
    return 'Eligible but not in the main PDF table for another reason - inspect the bond record.'


def assess(isin, data, pub):
    """Return a dict of answers for one ISIN, using REAL published membership."""
    r = {'ISIN': isin, 'In feed?': 'NO', 'In PDF?': 'NO',
         'Offshore Excel?': 'NO', 'Onshore Excel?': 'NO',
         'Reason': '', 'Explanation': '', 'Issuer': ''}

    bond = data.bond_by_isin.get(isin)
    if bond is None:
        r['Reason'] = 'not_in_feed'
        r['Explanation'] = ("Not in this week's bond feed at all. It never reached us - "
                            "check the ISIN is correct, or the desk needs to add it upstream.")
        return r

    r['In feed?'] = 'YES'
    upd = data.bond_updates.get(isin, {})
    gk  = (bond.get('GK_Nummer') or '').strip()
    r['Issuer'] = data.issuer_display_name(gk, fallback=bond.get('IssuerName', ''))

    # First the eligibility gate (drives everything).
    decision = data._classify_for_list(bond, upd)
    if not decision['eligible']:
        r['Reason'] = decision['reason']
        r['Explanation'] = EXPLAIN.get(decision['reason'], decision['reason'])
        return r

    # Eligible — now check ACTUAL membership in each real output.
    in_pdf_main = isin in pub['pdf_reference']
    in_sell     = isin in pub['pdf_sell']
    in_off      = isin in pub['offshore']
    in_on       = isin in pub['onshore']
    r['In PDF?']         = 'YES' if (in_pdf_main or in_sell) else 'NO'
    r['Offshore Excel?'] = 'YES' if in_off else 'NO'
    r['Onshore Excel?']  = 'YES' if in_on else 'NO'

    if in_pdf_main and in_off and in_on:
        r['Reason'] = 'present'
        r['Explanation'] = 'It IS in the report (PDF main table + both Excels). Check the tab/section they are looking at.'
        return r

    # Eligible but missing from at least one output — explain the gaps.
    bits = []
    if not (in_pdf_main or in_sell):
        bits.append('PDF: ' + _why_not_in_pdf(bond, data))
    elif in_sell and not in_pdf_main:
        bits.append('PDF: on the Sell Recommendations page, not the main table.')
    if not in_off:
        if not _has_issuer_name(bond, data):
            bits.append('Offshore Excel: dropped - no issuer name in the feed.')
        else:
            bits.append('Offshore Excel: not eligible (GEM tag or 180-day maturity).')
    if not in_on:
        if not _has_issuer_name(bond, data):
            bits.append('Onshore Excel: dropped - no issuer name in the feed.')
        else:
            bits.append('Onshore Excel: not onshore-eligible (not USD, Reg-S/144A corporate, or excluded-country sovereign).')
    r['Reason'] = 'eligible_but_filtered'
    r['Explanation'] = '  |  '.join(bits) if bits else 'Eligible; see membership columns.'
    return r


def write_xlsx(rows):
    os.makedirs(OUT_DIR, exist_ok=True)
    cols = ['ISIN', 'Issuer', 'In feed?', 'In PDF?',
            'Offshore Excel?', 'Onshore Excel?', 'Reason', 'Explanation']
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Why missing'
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='305496')
        for row in rows:
            ws.append([row.get(c, '') for c in cols])
            # Colour the three membership columns (In PDF? / Offshore / Onshore).
            for col_idx in (4, 5, 6):
                status = ws.cell(row=ws.max_row, column=col_idx).value
                yes = status == 'YES'
                ws.cell(row=ws.max_row, column=col_idx).fill = PatternFill(
                    'solid', fgColor='C6EFCE' if yes else 'FFC7CE')
                ws.cell(row=ws.max_row, column=col_idx).font = Font(
                    color='006100' if yes else '9C0006', bold=True)
        widths = [16, 30, 9, 13, 16, 16, 22, 70]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for r_ in ws.iter_rows(min_row=2):
            r_[-1].alignment = Alignment(wrap_text=True, vertical='top')
        ws.freeze_panes = 'A2'
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except ImportError:
        import csv
        path = OUT_XLSX.replace('.xlsx', '.csv')
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            for row in rows:
                w.writerow({c: row.get(c, '') for c in cols})
        return path


def main():
    isins = load_isins()
    print(f'Checking {len(isins)} ISIN(s)...')
    data = build_data()
    pub = build_published_sets(data)
    rows = [assess(i, data, pub) for i in isins]
    path = write_xlsx(rows)
    print('\nDone. Answers written to:')
    print('   ' + path)
    print('\nQuick summary (ISIN | PDF | Offshore | Onshore | reason):')
    for r in rows:
        print(f"   {r['ISIN']:<14} PDF={r['In PDF?']:<3} Off={r['Offshore Excel?']:<3} "
              f"On={r['Onshore Excel?']:<3} {r['Reason']}")


if __name__ == '__main__':
    main()
