"""
macro_tracker.py
================
Build Macro_Tracker.xlsx: a single workbook driven by one Codes sheet.

For every row you put on the Codes sheet, the matching category tab gets:
    - A quarterly block (Date | Level | QoQ % | YoY % | Forecast)
    - An annual block (Year | Level | YoY % | Forecast)
    - A line chart showing history + 2-year forecast

Charts are wired to Excel Tables, so when refresh.py rewrites the Tables
the charts re-read the current ranges automatically. No manual chart edits.

This file is the *builder*. Day-to-day you only ever run refresh.py.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from haver_metrics import Series, fetch_series, project, _periods_per_year


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

FONT = "Calibri"
NAVY = "1F3864"
LIGHT = "D9E1F2"
ZEBRA = "F2F2F2"
INPUT_YELLOW = "FFF2CC"

TITLE_FONT = Font(name=FONT, size=18, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name=FONT, size=11, italic=True, color="595959")
H1_FONT = Font(name=FONT, size=14, bold=True, color="FFFFFF")
H1_FILL = PatternFill("solid", start_color=NAVY)
H2_FONT = Font(name=FONT, size=12, bold=True, color=NAVY)
H2_FILL = PatternFill("solid", start_color=LIGHT)
LABEL_FONT = Font(name=FONT, size=11, bold=True)
INPUT_FONT = Font(name=FONT, size=11, color="0000FF", bold=True)
INPUT_FILL = PatternFill("solid", start_color=INPUT_YELLOW)
FORMULA_FONT = Font(name=FONT, size=11, color="000000")
LINK_FONT = Font(name=FONT, size=11, color="008000")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '#,##0.00;(#,##0.00);"-"'
PCT_FMT = '0.00"%";(0.00"%");"-"'
DATE_FMT = "yyyy-mm-dd"
YEAR_FMT = "0"


# ---------------------------------------------------------------------------
# Aggregation rules per Section
# ---------------------------------------------------------------------------

# For each Section, how to aggregate higher-frequency data to (Q, A).
# "mean" = average of period; "sum" = sum of period; "last" = last value.
SECTION_AGG = {
    "GDP":       ("mean", "mean"),  # SAAR levels and indices
    "Inflation": ("mean", "mean"),  # price indices
    "Fiscal":    ("sum",  "sum"),   # flows; we override stock indicators below
    "BoP":       ("sum",  "sum"),
    "Reserves":  ("last", "last"),  # stock at end of period
}

# Indicator-name overrides (substring match, case-insensitive).
INDICATOR_AGG_OVERRIDES = [
    ("debt",      ("last", "last")),  # Federal Debt is a stock
    ("deflator",  ("mean", "mean")),
    ("index",     ("mean", "mean")),
]

CATEGORIES = ["GDP", "Inflation", "Fiscal", "BoP", "Reserves"]


def _agg_for(section: str, indicator: str) -> tuple[str, str]:
    base = SECTION_AGG.get(section, ("mean", "mean"))
    name = (indicator or "").lower()
    for needle, override in INDICATOR_AGG_OVERRIDES:
        if needle in name:
            return override
    return base


def _aggregate(s: pd.Series, target: str, how: str) -> pd.Series:
    """Aggregate a time series to 'Q' or 'A' using how in {mean, sum, last}."""
    rule = "QS" if target == "Q" else "YS"
    grouped = s.resample(rule)
    if how == "mean":
        return grouped.mean().dropna()
    if how == "sum":
        return grouped.sum(min_count=1).dropna()
    if how == "last":
        return grouped.last().dropna()
    raise ValueError(how)


# ---------------------------------------------------------------------------
# Codes I/O
# ---------------------------------------------------------------------------

CODES_COLUMNS = ["Section", "Indicator", "Country", "Frequency",
                 "Quarterly Code", "Annual Code", "Units", "Notes"]


def read_codes_csv(path: str | Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str).fillna("")
    for c in CODES_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df[CODES_COLUMNS]


def read_codes_from_workbook(path: str | Path) -> Optional[pd.DataFrame]:
    """Read the Codes sheet from an existing workbook. Returns None if missing."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, data_only=True)
        if "Codes" not in wb.sheetnames:
            return None
        ws = wb["Codes"]
    except Exception:
        return None
    rows = list(ws.values)
    if len(rows) < 2:
        return None
    header = list(rows[0])
    body = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
    df = pd.DataFrame(body, columns=header).fillna("")
    for c in CODES_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df[CODES_COLUMNS].astype(str)


# ---------------------------------------------------------------------------
# Indicator data assembly: history + forecast at Q and A frequency
# ---------------------------------------------------------------------------

@dataclass
class IndicatorBundle:
    section: str
    indicator: str
    code: str
    units: str
    notes: str
    source: str
    q_hist: pd.Series          # quarterly history
    q_fcst: pd.DataFrame       # columns: Linear, HW, HW_Lower, HW_Upper
    a_hist: pd.Series          # annual history
    a_fcst: pd.DataFrame       # annual forecast (mean of quarterly)
    forecast_notes: dict


def _build_bundle(row: pd.Series, csv_dir: Path,
                  forecast_lookback_q: int = 40) -> Optional[IndicatorBundle]:
    """forecast_lookback_q caps how many recent quarters feed the linear
    trend fit, so emerging-market regime breaks (e.g. Argentine inflation
    measurement reform, currency revaluations) don't poison the projection.
    Holt-Winters still uses the full history."""
    code = (row["Quarterly Code"] or row["Annual Code"]).strip()
    if not code:
        return None
    series: Series = fetch_series(code, csv_dir=csv_dir)
    q_how, a_how = _agg_for(row["Section"], row["Indicator"])

    # Aggregate history
    s = series.data
    if series.frequency.upper().startswith("Q"):
        q_hist = s.copy()
    else:
        q_hist = _aggregate(s, "Q", q_how)
    a_hist = _aggregate(q_hist, "A", a_how)

    # Forecast: project quarterly 2 years ahead, derive annual from it.
    # Use the configured lookback window for stability on EM data.
    fit_series = q_hist.tail(forecast_lookback_q) if len(q_hist) > forecast_lookback_q else q_hist
    q_proj = project(_series_with_data(series, fit_series), years=2)
    q_fcst = pd.DataFrame({
        "Linear":   q_proj.linear,
        "HW":       q_proj.hw_mean,
        "HW_Lower": q_proj.hw_lower,
        "HW_Upper": q_proj.hw_upper,
    }, index=q_proj.horizon)

    # Annual forecast: aggregate quarterly forecast using same rule
    a_fcst = pd.DataFrame({
        col: _aggregate(q_fcst[col], "A", a_how) for col in q_fcst.columns
    })

    return IndicatorBundle(
        section=row["Section"],
        indicator=row["Indicator"],
        code=code,
        units=row["Units"],
        notes=row["Notes"],
        source=series.source,
        q_hist=q_hist,
        q_fcst=q_fcst,
        a_hist=a_hist,
        a_fcst=a_fcst,
        forecast_notes=q_proj.method_notes,
    )


def _series_with_data(template: Series, data: pd.Series) -> Series:
    """Return a Series object carrying the aggregated data so project() picks
    the right frequency."""
    return Series(
        code=template.code,
        description=template.description,
        frequency="Q",
        units=template.units,
        source=template.source,
        data=data,
    )


# ---------------------------------------------------------------------------
# Workbook building blocks
# ---------------------------------------------------------------------------

def _set(ws, row, col, value, *, font=None, fill=None, fmt=None,
         align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if fmt:    c.number_format = fmt
    if align:  c.alignment = align
    if border: c.border = border
    return c


def _bandh(ws, row: int, ncols: int, text: str, *, level: int = 1,
           start_col: int = 1) -> None:
    fill = H1_FILL if level == 1 else H2_FILL
    font = H1_FONT if level == 1 else H2_FONT
    end = start_col + ncols - 1
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = font
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22 if level == 1 else 18


def _table_header(ws, row: int, cols: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX


def _zebra(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        if (r - r1) % 2 == 1:
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=ZEBRA)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_readme(wb: Workbook, when: str, country: str = "") -> None:
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Macro Tracker" + (f" — {country}" if country else "")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {when}.  Edit the Codes sheet, then run refresh.py."
    ws["A2"].font = SUBTITLE_FONT

    rows = [
        ("How this works", ""),
        ("1. Codes sheet", "One row per indicator. Section drives which tab the indicator lands on. Edit cells in yellow."),
        ("2. Run refresh.py", "Pulls every series from Haver (or CSV / synthetic fallback) and rebuilds every tab and chart."),
        ("3. Open this file", "Each category tab shows quarterly + annual blocks, a 2-year forecast, and a chart that auto-extends with the underlying Tables."),
        ("", ""),
        ("Color convention", ""),
        ("Yellow fill", "Inputs you can change (codes on the Codes tab)."),
        ("Blue text", "Hardcoded values written by refresh.py (raw data, forecasts)."),
        ("Black text", "Excel formulas (growth rates, aggregations)."),
        ("Green text", "Cross-sheet links."),
        ("", ""),
        ("Forecast methods", ""),
        ("Linear trend", "Log-linear regression on the full quarterly history, extended 8 quarters."),
        ("Holt-Winters", "Additive damped-trend exponential smoothing with seasonal terms when history allows. 95% bands are residual-σ × √h."),
        ("", ""),
        ("Aggregation rules", ""),
        ("GDP / Inflation", "Average across periods (works for SAAR levels and price indices)."),
        ("Fiscal / BoP flows", "Sum across periods."),
        ("Stocks (Debt, Reserves)", "Last value of period."),
    ]
    r = 4
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = LABEL_FONT if k and not v else (TITLE_FONT if k and not v else LABEL_FONT)
        if not v and k:
            ws.cell(row=r, column=1).font = H2_FONT
        ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 95


def _write_codes(wb: Workbook, codes: pd.DataFrame, country: str = "") -> None:
    ws = wb.create_sheet("Codes")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Codes — Single Source of Truth" + (f"  ({country})" if country else "")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Edit codes here, then run refresh.py. Frequency: D/W/M/Q/A. Annual Code is optional (auto-derived from quarterly if blank)."
    ws["A2"].font = SUBTITLE_FONT

    headers = CODES_COLUMNS
    _table_header(ws, 4, headers)
    nrows = len(codes)
    for i, (_, row) in enumerate(codes.iterrows()):
        r = 5 + i
        for j, col in enumerate(headers):
            c = ws.cell(row=r, column=j + 1, value=row[col])
            c.font = INPUT_FONT
            c.fill = INPUT_FILL
            c.border = BOX
            c.alignment = Alignment(vertical="center", wrap_text=(col == "Notes"))

    # Excel Table over the codes
    ref = f"A4:{get_column_letter(len(headers))}{4 + nrows}"
    tbl = Table(displayName="tbl_Codes", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                        showRowStripes=True)
    ws.add_table(tbl)

    widths = [12, 32, 10, 11, 22, 22, 22, 50]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A5"


# ---------- Indicator block on a category tab -------------------------------

def _safe_name(text: str) -> str:
    import re
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


def _write_indicator_block(ws, b: IndicatorBundle, start_row: int) -> int:
    """Returns the next available row after the block."""
    safe = _safe_name(f"{b.section}_{b.indicator}")
    title = f"{b.indicator}    [{b.code}]    {b.units}"
    _bandh(ws, start_row, 14, title, level=1)

    # Meta line
    meta_row = start_row + 1
    ws.cell(row=meta_row, column=1, value="Source").font = LABEL_FONT
    ws.cell(row=meta_row, column=2, value=b.source).font = FORMULA_FONT
    ws.cell(row=meta_row, column=4, value="Forecast (linear)").font = LABEL_FONT
    ws.cell(row=meta_row, column=5, value=b.forecast_notes["linear"]).font = SUBTITLE_FONT
    ws.cell(row=meta_row, column=8, value="Forecast (HW)").font = LABEL_FONT
    ws.cell(row=meta_row, column=9, value=b.forecast_notes["holt_winters"]).font = SUBTITLE_FONT

    # ---------- Quarterly block (cols A..G) ---------------------------------
    q_header_row = start_row + 3
    _bandh(ws, q_header_row - 1, 7, "Quarterly", level=2)
    q_cols = ["Date", "Level", "QoQ %", "YoY %",
              "Linear Fcst", "HW Fcst", "HW Range"]
    _table_header(ws, q_header_row, q_cols)

    q_dates = list(b.q_hist.index) + list(b.q_fcst.index)
    n_hist = len(b.q_hist)
    n_fcst = len(b.q_fcst)
    n_q = n_hist + n_fcst

    for i, d in enumerate(q_dates):
        r = q_header_row + 1 + i
        is_hist = i < n_hist
        # Date
        _set(ws, r, 1, d.to_pydatetime(), fmt=DATE_FMT, border=BOX)
        # Level (history only) - hardcoded value
        if is_hist:
            _set(ws, r, 2, float(b.q_hist.iloc[i]),
                 font=INPUT_FONT, fmt=NUM_FMT, border=BOX)
        else:
            _set(ws, r, 2, None, border=BOX)
        # QoQ % (formula on Level column)
        if i >= 1:
            _set(ws, r, 3,
                 f"=IFERROR((B{r}/B{r-1}-1)*100,\"\")",
                 fmt=PCT_FMT, border=BOX)
        else:
            _set(ws, r, 3, "", border=BOX)
        # YoY % (formula on Level column)
        if i >= 4:
            _set(ws, r, 4,
                 f"=IFERROR((B{r}/B{r-4}-1)*100,\"\")",
                 fmt=PCT_FMT, border=BOX)
        else:
            _set(ws, r, 4, "", border=BOX)
        # Forecasts (future only)
        if not is_hist:
            j = i - n_hist
            _set(ws, r, 5, float(b.q_fcst["Linear"].iloc[j]),
                 font=INPUT_FONT, fmt=NUM_FMT, border=BOX)
            _set(ws, r, 6, float(b.q_fcst["HW"].iloc[j]),
                 font=INPUT_FONT, fmt=NUM_FMT, border=BOX)
            _set(ws, r, 7,
                 f"={float(b.q_fcst['HW_Upper'].iloc[j]):.6f}-{float(b.q_fcst['HW_Lower'].iloc[j]):.6f}",
                 fmt=NUM_FMT, border=BOX)
        else:
            for c in (5, 6, 7):
                _set(ws, r, c, "", border=BOX)

    q_first = q_header_row + 1
    q_last = q_header_row + n_q
    _zebra(ws, q_first, q_last, 1, 7)

    # Excel Table over the quarterly block
    q_ref = f"A{q_header_row}:G{q_last}"
    qt = Table(displayName=f"tbl_Q_{safe}", ref=q_ref)
    qt.tableStyleInfo = TableStyleInfo(name="TableStyleLight15",
                                       showRowStripes=True)
    ws.add_table(qt)

    # ---------- Annual block (cols I..L) -----------------------------------
    a_header_row = q_header_row
    _bandh(ws, a_header_row - 1, 4, "Annual", level=2, start_col=9)
    a_cols = ["Year", "Level", "YoY %", "Annual Fcst (HW)"]
    _table_header(ws, a_header_row, a_cols, start_col=9)

    a_dates = list(b.a_hist.index) + list(b.a_fcst.index)
    n_a_hist = len(b.a_hist)
    n_a_fcst = len(b.a_fcst)
    n_a = n_a_hist + n_a_fcst

    for i, d in enumerate(a_dates):
        r = a_header_row + 1 + i
        is_hist = i < n_a_hist
        _set(ws, r, 9, d.year, fmt=YEAR_FMT, border=BOX)
        if is_hist:
            _set(ws, r, 10, float(b.a_hist.iloc[i]),
                 font=INPUT_FONT, fmt=NUM_FMT, border=BOX)
        else:
            _set(ws, r, 10, None, border=BOX)
        if i >= 1:
            _set(ws, r, 11,
                 f"=IFERROR((J{r}/J{r-1}-1)*100,\"\")",
                 fmt=PCT_FMT, border=BOX)
        else:
            _set(ws, r, 11, "", border=BOX)
        if not is_hist:
            j = i - n_a_hist
            _set(ws, r, 12, float(b.a_fcst["HW"].iloc[j]),
                 font=INPUT_FONT, fmt=NUM_FMT, border=BOX)
        else:
            _set(ws, r, 12, "", border=BOX)

    a_first = a_header_row + 1
    a_last = a_header_row + n_a
    _zebra(ws, a_first, a_last, 9, 12)

    a_ref = f"I{a_header_row}:L{a_last}"
    at = Table(displayName=f"tbl_A_{safe}", ref=a_ref)
    at.tableStyleInfo = TableStyleInfo(name="TableStyleLight15",
                                       showRowStripes=True)
    ws.add_table(at)

    # ---------- Chart (anchored to the right of the annual block) ----------
    chart = LineChart()
    chart.title = f"{b.indicator} — quarterly history + 2y forecast"
    chart.height = 9
    chart.width = 22
    chart.y_axis.title = b.units or "Level"
    chart.x_axis.title = "Date"

    cats = Reference(ws, min_col=1, min_row=q_header_row + 1, max_row=q_last)
    # Level history
    data_level = Reference(ws, min_col=2, min_row=q_header_row,
                           max_row=q_last)
    chart.add_data(data_level, titles_from_data=True)
    # Linear forecast
    data_lin = Reference(ws, min_col=5, min_row=q_header_row, max_row=q_last)
    chart.add_data(data_lin, titles_from_data=True)
    # HW forecast
    data_hw = Reference(ws, min_col=6, min_row=q_header_row, max_row=q_last)
    chart.add_data(data_hw, titles_from_data=True)
    chart.set_categories(cats)

    # style the series: make forecasts dashed
    if len(chart.series) >= 3:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        for idx in (1, 2):
            s = chart.series[idx]
            s.graphicalProperties = GraphicalProperties(
                ln=LineProperties(prstDash="dash"))

    chart_anchor = f"N{q_header_row}"
    ws.add_chart(chart, chart_anchor)

    # Block consumes max(n_q, n_a) rows of data + 4 for headers/spacing + 18 for chart space
    used = max(q_last, a_last, q_header_row + 22)
    return used + 3


def _write_category_tab(wb: Workbook, category: str,
                        bundles: list[IndicatorBundle],
                        country: str = "") -> None:
    ws = wb.create_sheet(category)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{category}" + (f" — {country}" if country else "")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{len(bundles)} indicator(s) — pulled from Codes sheet."
    ws["A2"].font = SUBTITLE_FONT

    # Column widths
    widths = {"A": 12, "B": 14, "C": 11, "D": 11, "E": 13, "F": 13, "G": 13,
              "H": 2,
              "I": 9, "J": 14, "K": 11, "L": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    if not bundles:
        ws["A4"] = "(no codes for this section yet — add a row in Codes)"
        ws["A4"].font = SUBTITLE_FONT
        return

    cur = 4
    for b in bundles:
        cur = _write_indicator_block(ws, b, cur)

    ws.freeze_panes = "A4"


def _write_dashboard(wb: Workbook, all_bundles: list[IndicatorBundle],
                     country: str = "") -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dashboard — Latest Reads & 2y Forecast" + (f"  ({country})" if country else "")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Cross-sheet KPIs (links pull live from each category tab)."
    ws["A2"].font = SUBTITLE_FONT

    headers = ["Section", "Indicator", "Latest date", "Latest value",
               "YoY %", "2y forecast (HW)", "Source"]
    _table_header(ws, 4, headers)

    for i, b in enumerate(all_bundles):
        r = 5 + i
        latest_date = b.q_hist.index[-1]
        latest_val = float(b.q_hist.iloc[-1])
        # YoY from history
        yoy = ""
        if len(b.q_hist) > 4:
            prev = float(b.q_hist.iloc[-5])
            if prev != 0:
                yoy = (latest_val / prev - 1) * 100
        last_fcst = float(b.q_fcst["HW"].iloc[-1])
        _set(ws, r, 1, b.section, border=BOX)
        _set(ws, r, 2, b.indicator, border=BOX)
        _set(ws, r, 3, latest_date.to_pydatetime(), fmt=DATE_FMT, border=BOX)
        _set(ws, r, 4, latest_val, fmt=NUM_FMT, border=BOX)
        _set(ws, r, 5, yoy if yoy != "" else "",
             fmt=PCT_FMT, border=BOX)
        _set(ws, r, 6, last_fcst, fmt=NUM_FMT, border=BOX)
        _set(ws, r, 7, b.source, border=BOX)

    _zebra(ws, 5, 5 + len(all_bundles) - 1, 1, 7)

    widths = [12, 36, 14, 16, 12, 18, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_workbook(codes: pd.DataFrame, out_path: str | Path,
                   csv_dir: str | Path = "data",
                   country: str = "") -> Path:
    out_path = Path(out_path)
    csv_dir = Path(csv_dir)
    csv_dir.mkdir(exist_ok=True)

    # Infer country from data if not provided
    if not country and "Country" in codes.columns:
        countries = [c for c in codes["Country"].dropna().unique() if c]
        if len(countries) == 1:
            country = countries[0]

    when = datetime.now().strftime("%Y-%m-%d %H:%M")
    wb = Workbook()
    _write_readme(wb, when, country)
    _write_codes(wb, codes, country)

    # Build bundles (one fetch per row)
    all_bundles: list[IndicatorBundle] = []
    by_section: dict[str, list[IndicatorBundle]] = {c: [] for c in CATEGORIES}
    for _, row in codes.iterrows():
        if row["Section"] not in CATEGORIES:
            print(f"[skip] {row['Indicator']}: section {row['Section']} not recognised")
            continue
        try:
            bundle = _build_bundle(row, csv_dir)
            if bundle is None:
                continue
            all_bundles.append(bundle)
            by_section[row["Section"]].append(bundle)
            print(f"[ok]  {row['Section']:10s}  {row['Indicator']:38s}  ({bundle.source})")
        except Exception as e:
            print(f"[fail] {row['Indicator']}: {e}")

    for cat in CATEGORIES:
        _write_category_tab(wb, cat, by_section[cat], country)

    _write_dashboard(wb, all_bundles, country)
    wb.save(out_path)
    print(f"[done] wrote {out_path}  ({len(all_bundles)} indicators)")
    return out_path
