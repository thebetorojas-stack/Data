"""Excel report writer.

Produces one workbook per country with tabs:
    Cover | Monthly | Quarterly | Annual | Credit | Markets | Notes

Charts are native openpyxl LineChart / BarChart objects bound to cell ranges,
so they regenerate live when the data table is rewritten on every refresh.
Users can email the .xlsx to clients and the charts render with no Bloomberg
or Haver dependency on the receiving end.

Layout per data tab:
    Row 1     Country / Tab / Last refresh banner
    Row 3     Latest values summary block (key indicator KPIs)
    Row 8+    Data table (date column + one column per series)
    Charts    Anchored on the right side, one per series flagged chart=true
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.trendline import Trendline
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config_loader import CountryConfig, SeriesConfig, Settings
from src.transform.frequencies import apply_transform, resample_to
from src.transform.credit import build_curve, curve_history
from .styles import (body_font, center, fill, format_for, header_font,
                     hex_to_argb, right, subtitle_font, thin_border, title_font)


TABS = ["monthly", "quarterly", "annual", "credit", "markets"]
TAB_LABELS = {
    "monthly": "Monthly",
    "quarterly": "Quarterly",
    "annual": "Annual",
    "credit": "Credit",
    "markets": "Markets",
}
TAB_FREQ = {"monthly": "M", "quarterly": "Q", "annual": "A", "credit": "D", "markets": "D"}


# ────────────────────────────────────────────────────────────────────────── #
def build_country_workbook(
    cfg: CountryConfig,
    cache_frame: pd.DataFrame,
    settings: Settings,
    out_path: str | Path,
) -> Path:
    """Main entry point. cache_frame is the wide DataFrame from CacheStore.load_all()."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    palette = settings.excel.get("palette", {})
    primary = palette.get("primary", "1F4E79")
    accent = palette.get("accent", "C00000")

    _write_cover(wb, cfg, cache_frame, primary)

    for tab in TABS:
        specs = cfg.by_tab(tab)
        if not specs:
            continue
        _write_data_tab(wb, tab, cfg, specs, cache_frame, primary, accent)

    if cfg.by_tab("credit"):
        _write_credit_extras(wb, cfg, cache_frame, primary, accent)

    _write_notes(wb, cfg, primary)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────────── #
def _write_cover(wb, cfg: CountryConfig, cache: pd.DataFrame, primary: str):
    ws = wb.create_sheet("Cover")
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 18

    ws["B2"] = f"{cfg.country} — Macro & Hard-Currency Credit"
    ws["B2"].font = title_font(primary)
    ws["B3"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} • currency: {cfg.currency}"
    ws["B3"].font = subtitle_font()

    # Mini KPI grid: pick a handful of headline indicators if present
    kpi_names = [
        ("CPI YoY", "cpi_headline_yoy", "%"),
        ("Activity (IMACEC/EMAE) YoY", "imacec" if cfg.iso == "CL" else "emae", "%"),
        ("Policy rate", "policy_rate_tpm" if cfg.iso == "CL" else "policy_rate_leliq", "%"),
        ("Reserves USD bn", "international_reserves" if cfg.iso == "CL" else "bcra_reserves_gross", "USDbn"),
        ("EMBI spread", f"embi_{cfg.country.lower()}_spread", "bps"),
        ("USD/local FX", "usdclp" if cfg.iso == "CL" else "usdars_official", "lcy"),
    ]
    row = 5
    ws.cell(row=row, column=2, value="Latest readings").font = header_font(primary)
    ws.cell(row=row, column=2).fill = fill("D9E1F2")
    row += 1
    for i, (lbl, sname, _u) in enumerate(kpi_names):
        r = row + i
        ws.cell(row=r, column=2, value=lbl).font = body_font()
        if sname in cache.columns:
            s = cache[sname].dropna()
            if not s.empty:
                ws.cell(row=r, column=3, value=float(s.iloc[-1])).number_format = "#,##0.00"
                ws.cell(row=r, column=4, value=s.index.max().strftime("%Y-%m-%d"))

    ws["B14"] = "Tabs"
    ws["B14"].font = header_font(primary)
    ws["B14"].fill = fill("D9E1F2")
    for i, t in enumerate(TABS):
        ws.cell(row=15 + i, column=2, value=TAB_LABELS[t])
        ws.cell(row=15 + i, column=3, value=f"see '{TAB_LABELS[t]}' tab")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


# ────────────────────────────────────────────────────────────────────────── #
def _write_data_tab(wb, tab: str, cfg: CountryConfig,
                    specs: List[SeriesConfig], cache: pd.DataFrame,
                    primary: str, accent: str):
    ws = wb.create_sheet(TAB_LABELS[tab])
    ws.sheet_view.showGridLines = False

    # banner
    ws["A1"] = f"{cfg.country} — {TAB_LABELS[tab]}"
    ws["A1"].font = title_font(primary)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Refreshed {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = subtitle_font()

    # build the data block: index column + one column per series, after transform & resample
    target_freq = TAB_FREQ[tab]
    data: Dict[str, pd.Series] = {}
    for s in specs:
        if s.name not in cache.columns:
            continue
        col = cache[s.name].dropna()
        if col.empty:
            continue
        if target_freq != "D":
            col = resample_to(col, target_freq)
        col = apply_transform(col, s.transform, target_freq)
        if col is None:
            continue
        data[s.name] = col

    if not data:
        ws["A4"] = "No data available — check Bloomberg/Haver connectivity and ticker map."
        ws["A4"].font = body_font()
        return

    df = pd.concat(data.values(), axis=1, keys=data.keys()).sort_index()
    # for daily tabs, keep last 5 years for performance and chart readability
    if target_freq == "D":
        cutoff = df.index.max() - pd.DateOffset(years=5)
        df = df[df.index >= cutoff]

    # write header row
    HEAD_ROW = 8
    ws.cell(row=HEAD_ROW, column=1, value="Date").font = header_font()
    ws.cell(row=HEAD_ROW, column=1).fill = fill(primary)
    spec_by_name = {s.name: s for s in specs}
    for j, name in enumerate(df.columns, start=2):
        sp = spec_by_name.get(name)
        cell = ws.cell(row=HEAD_ROW, column=j, value=sp.label if sp else name)
        cell.font = header_font()
        cell.fill = fill(primary)
        cell.alignment = center()

    # write data rows
    for i, (idx, vals) in enumerate(df.iterrows(), start=HEAD_ROW + 1):
        ws.cell(row=i, column=1, value=idx.to_pydatetime() if hasattr(idx, "to_pydatetime") else idx).number_format = (
            "yyyy-mm" if target_freq in ("M", "Q") else "yyyy" if target_freq == "A" else "yyyy-mm-dd")
        for j, name in enumerate(df.columns, start=2):
            v = vals[name]
            c = ws.cell(row=i, column=j, value=None if pd.isna(v) else float(v))
            sp = spec_by_name.get(name)
            if sp:
                c.number_format = format_for(sp.units)

    # KPI row at top of body
    _write_kpi_strip(ws, df, spec_by_name, kpi_row=4, primary=primary)

    # column widths
    ws.column_dimensions["A"].width = 13
    for j in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = ws.cell(row=HEAD_ROW + 1, column=2)

    # charts: one per chart=true series, anchored to the right
    chart_anchor_col = len(df.columns) + 3  # leave a gap
    chart_row = HEAD_ROW
    for j, name in enumerate(df.columns, start=2):
        sp = spec_by_name.get(name)
        if not sp or not sp.chart:
            continue
        _add_line_chart(
            ws=ws,
            title=sp.label,
            data_col_letter=get_column_letter(j),
            n_rows=len(df),
            head_row=HEAD_ROW,
            anchor=f"{get_column_letter(chart_anchor_col)}{chart_row}",
            color=primary,
            number_fmt=format_for(sp.units),
        )
        chart_row += 16  # ~chart height in rows


# ────────────────────────────────────────────────────────────────────────── #
def _write_kpi_strip(ws, df: pd.DataFrame, spec_by_name: Dict[str, SeriesConfig],
                     kpi_row: int, primary: str):
    """Compact summary block above the data table."""
    ws.cell(row=kpi_row, column=1, value="Latest").font = header_font()
    ws.cell(row=kpi_row, column=1).fill = fill(primary)
    last_idx = df.index.max()
    for j, name in enumerate(df.columns, start=2):
        sp = spec_by_name.get(name)
        last = df[name].dropna()
        if last.empty:
            continue
        v = float(last.iloc[-1])
        c = ws.cell(row=kpi_row, column=j, value=v)
        c.number_format = format_for(sp.units if sp else "level")
        c.font = body_font()
        c.alignment = right()
        # date subscript
        d = ws.cell(row=kpi_row + 1, column=j, value=last.index.max().strftime("%Y-%m-%d"))
        d.font = subtitle_font()
        d.alignment = right()


# ────────────────────────────────────────────────────────────────────────── #
def _add_line_chart(ws, title: str, data_col_letter: str, n_rows: int,
                    head_row: int, anchor: str, color: str, number_fmt: str):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 7.5  # cm
    chart.width = 14
    chart.y_axis.number_format = number_fmt
    chart.y_axis.majorGridlines = None
    chart.x_axis.number_format = "yyyy-mm-dd"
    chart.legend = None

    data_ref = Reference(
        ws, min_col=ws[f"{data_col_letter}{head_row}"].column,
        min_row=head_row, max_row=head_row + n_rows
    )
    cats_ref = Reference(ws, min_col=1, min_row=head_row + 1, max_row=head_row + n_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # styling: line color from palette
    if chart.series:
        from openpyxl.drawing.colors import ColorChoice
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        ln = LineProperties(solidFill=hex_to_argb(color)[2:], w=18000)  # 1.5pt
        chart.series[0].graphicalProperties = GraphicalProperties(ln=ln)

    ws.add_chart(chart, anchor)


# ────────────────────────────────────────────────────────────────────────── #
def _write_credit_extras(wb, cfg: CountryConfig, cache: pd.DataFrame,
                         primary: str, accent: str):
    """Adds a USD curve snapshot table and chart on a sub-section of the Credit tab."""
    if "Credit" not in wb.sheetnames:
        return
    ws = wb["Credit"]

    curve = build_curve(cfg, cache)
    if curve.empty:
        return

    # find a clean spot below current content
    start_row = (ws.max_row or 8) + 4
    ws.cell(row=start_row, column=1, value="USD Sovereign Curve — latest").font = header_font(primary)
    ws.cell(row=start_row, column=1).fill = fill(primary)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    headers = ["Maturity (yrs)", "Yield (%)", "Bond"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row + 1, column=j, value=h)
        c.font = header_font()
        c.fill = fill(primary)

    for i, row in curve.iterrows():
        ws.cell(row=start_row + 2 + i, column=1, value=float(row["maturity_years"]))
        ws.cell(row=start_row + 2 + i, column=2, value=float(row["yield_pct"])).number_format = '0.00"%"'
        ws.cell(row=start_row + 2 + i, column=3, value=row["bond"])

    # Curve chart (scatter-style line)
    chart = LineChart()
    chart.title = f"{cfg.country} USD curve (latest)"
    chart.style = 2
    chart.height = 8
    chart.width = 14
    chart.y_axis.number_format = '0.00"%"'
    chart.legend = None
    data_ref = Reference(ws, min_col=2, min_row=start_row + 1,
                         max_row=start_row + 1 + len(curve))
    cat_ref = Reference(ws, min_col=1, min_row=start_row + 2,
                        max_row=start_row + 1 + len(curve))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, f"F{start_row}")


# ────────────────────────────────────────────────────────────────────────── #
def _write_notes(wb, cfg: CountryConfig, primary: str):
    ws = wb.create_sheet("Notes")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90

    ws["B2"] = f"{cfg.country} — methodology and source notes"
    ws["B2"].font = title_font(primary)

    notes = [
        "Frequencies: Monthly = month-end last; Quarterly = quarter-end last; Annual = year-end last.",
        "Transforms: yoy = year-over-year %; mom = month-over-month %; saar = annualized QoQ growth (compound).",
        "Bloomberg fields: PX_LAST for prices/spreads, YLD_YTM_MID for bond yields.",
        "Haver databases: country-specific (chile / argent). Verify DB code matches your subscription.",
    ]
    if cfg.iso == "AR":
        notes += [
            "Argentina CPI: structural breaks acknowledged. Series spliced via official INDEC where available; "
            "pre-2017 may use IPC-CABA / IPC-Congreso.",
            "Argentina FX: official (BCRA A3500), MEP (AL30D/AL30 cash market), CCL (blue chip swap), blue (informal).",
            "Brecha = (CCL / Official - 1) × 100 = FX gap.",
            "Fiscal: Tesoro Nacional perimeter for primary balance; SPN where noted.",
        ]
    if cfg.iso == "CL":
        notes += [
            "IMACEC: Banco Central de Chile monthly economic activity indicator.",
            "CPI core excludes food and energy; both headline and core tracked.",
            "Fiscal series shown as 12-month rolling balance for monthly view.",
        ]

    for i, n in enumerate(notes):
        c = ws.cell(row=4 + i, column=2, value=f"• {n}")
        c.font = body_font()
        c.alignment = c.alignment.copy(wrap_text=True)
