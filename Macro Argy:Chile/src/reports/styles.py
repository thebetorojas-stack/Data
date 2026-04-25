"""Centralized openpyxl style helpers."""

from __future__ import annotations

from typing import Dict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00


def hex_to_argb(h: str) -> str:
    h = h.lstrip("#")
    if len(h) == 6:
        return "FF" + h.upper()
    return h.upper()


def header_font(color: str = "FFFFFF") -> Font:
    return Font(name="Calibri", size=11, bold=True, color=hex_to_argb(color))


def title_font(color: str = "1F4E79") -> Font:
    return Font(name="Calibri", size=16, bold=True, color=hex_to_argb(color))


def subtitle_font() -> Font:
    return Font(name="Calibri", size=10, italic=True, color=hex_to_argb("595959"))


def body_font() -> Font:
    return Font(name="Calibri", size=10)


def fill(color: str) -> PatternFill:
    argb = hex_to_argb(color)
    return PatternFill("solid", fgColor=argb)


def thin_border() -> Border:
    side = Side(style="thin", color=hex_to_argb("BFBFBF"))
    return Border(left=side, right=side, top=side, bottom=side)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# number format lookup keyed by the `units` field in the YAML
NUMBER_FORMATS: Dict[str, str] = {
    "pct": '0.00"%"',
    "pct_pp": '+0.00"pp";-0.00"pp"',
    "bps": '#,##0" bps"',
    "yield": '0.00"%"',
    "fx_local": '#,##0.00',
    "money_usd_bn": '#,##0.0',
    "money_lcy_bn": '#,##0.0',
    "index": '#,##0.0',
    "level": '#,##0.00',
}


def format_for(units: str) -> str:
    return NUMBER_FORMATS.get(units, FORMAT_NUMBER_00)
