"""
haver_metrics.py
================
Pull a Haver Analytics series, compute standard growth metrics, project
24 months ahead with two methods (linear trend + Holt-Winters), and
write a polished Excel workbook.

Typical usage
-------------
    from haver_metrics import build_workbook
    build_workbook("GDPH@USECON", out_path="GDPH_USECON.xlsx")

Or just edit and run update.py.

Data sources (auto-detected, in priority order)
-----------------------------------------------
1. Haver Python package + DLX add-in (Windows only).
       pip install haver
   Requires Haver Analytics DLX installed locally with a valid licence.
2. CSV fallback. If a file named "<code>.csv" exists in ./data/ with
   columns [date, value], it is used. Lets Mac/Linux users iterate by
   exporting series from Haver to CSV once.
3. Synthetic demo series. If neither is available, a synthetic
   GDP-like series is generated so the template still renders end-to-end.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing  # type: ignore
    _HAS_STATSMODELS = True
except Exception:
    _HAS_STATSMODELS = False


# ---------------------------------------------------------------------------
# 1. Data fetching
# ---------------------------------------------------------------------------

@dataclass
class Series:
    code: str
    description: str
    frequency: str          # "M", "Q", "A", "D", "W"
    units: str
    source: str             # "Haver DLX", "CSV", "Synthetic demo"
    data: pd.Series         # indexed by Timestamp


def _infer_frequency(idx: pd.DatetimeIndex) -> str:
    if len(idx) < 3:
        return "M"
    diffs = np.diff(idx.values).astype("timedelta64[D]").astype(int)
    median = int(np.median(diffs))
    if median <= 2:
        return "D"
    if median <= 10:
        return "W"
    if median <= 45:
        return "M"
    if median <= 120:
        return "Q"
    return "A"


def _fetch_via_haver(code: str) -> Optional[Series]:
    """Try the official Haver Python package (requires DLX on Windows)."""
    try:
        import haver  # type: ignore
    except Exception:
        return None
    try:
        df = haver.data([code])
        meta = haver.metadata([code]).iloc[0]
        s = df[code].dropna()
        s.index = pd.to_datetime(s.index)
        return Series(
            code=code,
            description=str(meta.get("descriptor", code)),
            frequency=str(meta.get("frequency", _infer_frequency(s.index))),
            units=str(meta.get("magnitude", "")),
            source="Haver DLX",
            data=s,
        )
    except Exception as e:
        print(f"[haver] fetch failed: {e}")
        return None


def _fetch_via_csv(code: str, csv_dir: Path) -> Optional[Series]:
    safe = re.sub(r"[^A-Za-z0-9_]+", "_", code)
    candidates = [csv_dir / f"{safe}.csv", csv_dir / f"{code}.csv"]
    for path in candidates:
        if path.exists():
            df = pd.read_csv(path)
            date_col = next((c for c in df.columns if c.lower() in ("date", "period", "observation_date")), df.columns[0])
            val_col = next((c for c in df.columns if c.lower() in ("value", "val", "level")), df.columns[1])
            s = pd.Series(df[val_col].values, index=pd.to_datetime(df[date_col])).sort_index().dropna()
            return Series(
                code=code,
                description=f"{code} (from {path.name})",
                frequency=_infer_frequency(s.index),
                units="",
                source=f"CSV: {path.name}",
                data=s,
            )
    return None


def _synthetic_series(code: str) -> Series:
    """Plausible-looking series for demo. Shape and frequency depend on
    keywords in the code so a populated workbook looks like distinct
    indicators rather than 15 copies of the same line."""
    seed = abs(hash(code)) % (2**32)
    rng = np.random.default_rng(seed)
    cu = code.upper()

    # ----- pick a flavour from the code name --------------------------------
    if any(k in cu for k in ("CPI", "PCU", "JCXFE", "JGDP")):
        # Monthly price index, 2% trend, mild seasonality
        idx = pd.date_range("2005-01-01", periods=252, freq="MS")
        t = np.arange(len(idx))
        base = 100 * (1 + 0.025 / 12) ** t
        season = 1 + 0.003 * np.sin(2 * np.pi * (t % 12) / 12)
        noise = rng.normal(0, 0.0015, len(idx))
        v = base * season * (1 + noise)
        freq, units = "M", "Index"
    elif "GDP" in cu and "DEFL" not in cu:
        # Quarterly GDP-like, level in trillions
        idx = pd.date_range("2005-01-01", periods=84, freq="QS")
        t = np.arange(len(idx))
        trend = 18000 * (1 + 0.022 / 4) ** t
        shock = np.where((idx.year.isin([2008, 2009, 2020])), -0.04, 0)
        noise = rng.normal(0, 0.005, len(idx))
        v = trend * (1 + shock + noise)
        freq, units = "Q", "Bil $"
    elif any(k in cu for k in ("FFR", "FFE", "FFD")) and "FFDH" not in cu:
        # Monthly fiscal flow, sign depends on receipts vs deficit
        idx = pd.date_range("2005-01-01", periods=252, freq="MS")
        t = np.arange(len(idx))
        sign = -1.0 if "FFD" in cu else 1.0
        base = 250 * (1 + 0.04 / 12) ** t * sign
        seas = 1 + 0.20 * np.sin(2 * np.pi * (t % 12) / 12)
        v = base * seas + rng.normal(0, 20, len(idx))
        freq, units = "M", "Bil $"
    elif "FFDH" in cu:
        # Quarterly debt stock, monotonic-ish
        idx = pd.date_range("2005-01-01", periods=84, freq="QS")
        t = np.arange(len(idx))
        v = 8000 * (1 + 0.06 / 4) ** t + rng.normal(0, 80, len(t))
        freq, units = "Q", "Bil $"
    elif "BPC" in cu or "BPT" in cu:
        # Quarterly current account / trade balance, mostly negative for US
        idx = pd.date_range("2005-01-01", periods=84, freq="QS")
        t = np.arange(len(idx))
        base = -100 + 5 * np.sin(2 * np.pi * t / 16)
        v = base + rng.normal(0, 15, len(t))
        freq, units = "Q", "Bil $"
    elif "LIRRA" in cu or "LIGOLD" in cu:
        # Monthly reserves, slowly rising
        idx = pd.date_range("2005-01-01", periods=252, freq="MS")
        t = np.arange(len(idx))
        base = 130000 if "LIRRA" in cu else 11000
        v = base * (1 + 0.01 / 12) ** t + rng.normal(0, base * 0.01, len(t))
        freq, units = "M", "Mil $"
    else:
        # Default: quarterly index
        idx = pd.date_range("2005-01-01", periods=84, freq="QS")
        t = np.arange(len(idx))
        v = 100 * (1 + 0.02 / 4) ** t + rng.normal(0, 0.5, len(t))
        freq, units = "Q", "Index"

    s = pd.Series(v, index=idx)
    return Series(
        code=code,
        description=f"{code} (SYNTHETIC DEMO — install haver or drop a CSV in ./data)",
        frequency=freq,
        units=units,
        source="Synthetic demo",
        data=s,
    )


def fetch_series(code: str, csv_dir: str | Path = "data") -> Series:
    csv_dir = Path(csv_dir)
    csv_dir.mkdir(exist_ok=True)
    return (
        _fetch_via_haver(code)
        or _fetch_via_csv(code, csv_dir)
        or _synthetic_series(code)
    )


# ---------------------------------------------------------------------------
# 2. Metrics
# ---------------------------------------------------------------------------

def _periods_per_year(freq: str) -> int:
    return {"D": 252, "W": 52, "M": 12, "Q": 4, "A": 1}.get(freq.upper()[:1], 12)


def compute_metrics(series: Series) -> pd.DataFrame:
    s = series.data
    ppy = _periods_per_year(series.frequency)
    df = pd.DataFrame({"Level": s})
    df["MoM %"] = s.pct_change(1) * 100
    df["QoQ %"] = s.pct_change(max(ppy // 4, 1)) * 100
    df["YoY %"] = s.pct_change(ppy) * 100
    df["12M Roll Avg"] = s.rolling(ppy, min_periods=max(2, ppy // 2)).mean()
    return df


# ---------------------------------------------------------------------------
# 3. Projections (linear trend + Holt-Winters)
# ---------------------------------------------------------------------------

@dataclass
class Projection:
    horizon: pd.DatetimeIndex
    linear: pd.Series
    hw_mean: pd.Series
    hw_lower: pd.Series
    hw_upper: pd.Series
    method_notes: dict


def _holt_winters_numpy(y: np.ndarray, n: int, m: Optional[int],
                        alpha: float = 0.4, beta: float = 0.1,
                        gamma: float = 0.2, phi: float = 0.95
                        ) -> tuple[np.ndarray, float]:
    """Pure-numpy additive Holt-Winters with damped trend. Used as fallback
    when statsmodels isn't installed."""
    y = y.astype(float)
    L = float(y[0])
    T = float(y[1] - y[0]) if len(y) > 1 else 0.0
    if m and len(y) >= 2 * m:
        S = [float(y[i] - np.mean(y[:m])) for i in range(m)]
    else:
        m = None
        S = [0.0]
    fitted = np.zeros_like(y)
    for t in range(len(y)):
        s_t = S[t % m] if m else 0.0
        prev_L, prev_T = L, T
        L = alpha * (y[t] - s_t) + (1 - alpha) * (prev_L + phi * prev_T)
        T = beta * (L - prev_L) + (1 - beta) * phi * prev_T
        if m:
            S[t % m] = gamma * (y[t] - L) + (1 - gamma) * s_t
        fitted[t] = prev_L + phi * prev_T + s_t
    resid_std = float(np.nanstd(y - fitted))
    out = np.empty(n)
    cum = 0.0
    for h in range(1, n + 1):
        cum += phi ** h
        s_t = S[(len(y) + h - 1) % m] if m else 0.0
        out[h - 1] = L + cum * T + s_t
    return out, resid_std


def _holt_winters_forecast(s: pd.Series, n: int, ppy: int,
                           seasonal: Optional[str]
                           ) -> tuple[np.ndarray, float, str]:
    sp = ppy if seasonal else None
    if _HAS_STATSMODELS:
        try:
            model = ExponentialSmoothing(
                s.astype(float),
                trend="add",
                damped_trend=True,
                seasonal=seasonal,
                seasonal_periods=sp,
                initialization_method="estimated",
            ).fit(optimized=True)
            mean = np.asarray(model.forecast(n))
            resid_std = float(np.nanstd(model.resid))
            return mean, resid_std, f"statsmodels Holt-Winters (additive, damped, seasonal={seasonal})"
        except Exception as e:
            print(f"[hw] statsmodels failed ({e}); using numpy fallback.")
    mean, resid_std = _holt_winters_numpy(s.values, n, sp)
    return mean, resid_std, f"numpy Holt-Winters fallback (additive, damped, seasonal={seasonal})"


def _future_index(last: pd.Timestamp, freq: str, n: int) -> pd.DatetimeIndex:
    pd_freq = {"D": "B", "W": "W", "M": "MS", "Q": "QS", "A": "YS"}.get(freq.upper()[:1], "MS")
    return pd.date_range(start=last, periods=n + 1, freq=pd_freq)[1:]


def project(series: Series, years: int = 2) -> Projection:
    s = series.data.dropna()
    ppy = _periods_per_year(series.frequency)
    n = years * ppy
    horizon = _future_index(s.index[-1], series.frequency, n)

    # ---- Linear trend extrapolation ---------------------------------------
    # Use log-linear when the series is strictly positive (gives stable %
    # growth extrapolation); fall back to OLS on levels for series that can
    # be negative (deficits, trade balances, current accounts).
    raw = s.values.astype(float)
    x = np.arange(len(raw))
    fut_x = np.arange(len(raw), len(raw) + n)
    if np.all(raw > 0):
        slope, intercept = np.polyfit(x, np.log(raw), 1)
        linear_vals = np.exp(intercept + slope * fut_x)
        annualised = (np.exp(slope * ppy) - 1) * 100
    else:
        slope, intercept = np.polyfit(x, raw, 1)
        linear_vals = intercept + slope * fut_x
        ref = np.mean(np.abs(raw[-ppy:])) if len(raw) >= ppy else np.mean(np.abs(raw))
        annualised = (slope * ppy / ref) * 100 if ref else 0.0
    linear = pd.Series(linear_vals, index=horizon, name="Linear trend")

    # ---- Holt-Winters (additive damped trend, seasonal if enough history) -
    seasonal = "add" if (ppy > 1 and len(s) >= ppy * 3) else None
    hw_arr, resid_std, hw_notes = _holt_winters_forecast(s, n, ppy, seasonal)
    hw_mean_s = pd.Series(hw_arr, index=horizon, name="Holt-Winters mean")
    hh = np.arange(1, n + 1)
    band = 1.96 * resid_std * np.sqrt(hh)
    hw_lower = pd.Series(hw_mean_s.values - band, index=horizon, name="HW lower 95%")
    hw_upper = pd.Series(hw_mean_s.values + band, index=horizon, name="HW upper 95%")

    return Projection(
        horizon=horizon,
        linear=linear,
        hw_mean=hw_mean_s,
        hw_lower=hw_lower,
        hw_upper=hw_upper,
        method_notes={
            "linear": f"Log-linear regression on full history. Implied annualised growth: {annualised:.2f}%",
            "holt_winters": hw_notes,
        },
    )


# ---------------------------------------------------------------------------
# 4. Excel writer
# ---------------------------------------------------------------------------

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
LABEL_FONT = Font(name=FONT, bold=True, size=11)
BLUE_INPUT = Font(name=FONT, color="0000FF", bold=True)
BLACK_FORMULA = Font(name=FONT, color="000000")
GREEN_LINK = Font(name=FONT, color="008000")
INPUT_FILL = PatternFill("solid", start_color="FFFF00")
ALT_ROW_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '#,##0.00;(#,##0.00);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
PCT_FMT_DISP = '0.00"%";(0.00"%");"-"'   # values already in %


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _zebra(ws, start_row: int, end_row: int, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_summary(wb, series: Series, metrics: pd.DataFrame, proj: Projection) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Haver Series Tracker"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    # --- input block --------------------------------------------------------
    ws["A4"] = "Haver code:"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = series.code
    ws["B4"].font = BLUE_INPUT
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = BORDER

    meta_rows = [
        ("Description", series.description),
        ("Source",       series.source),
        ("Frequency",    series.frequency),
        ("Units",        series.units or "n/a"),
        ("Observations", len(series.data)),
        ("First date",   series.data.index[0].strftime("%Y-%m-%d")),
        ("Last date",    series.data.index[-1].strftime("%Y-%m-%d")),
        ("Last value",   float(series.data.iloc[-1])),
    ]
    for i, (k, v) in enumerate(meta_rows, start=5):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=v)
        if k == "Last value":
            c.number_format = NUM_FMT
            c.font = BLACK_FORMULA

    # --- KPI block: latest growth rates, pulled live from Metrics tab -------
    ws["D4"] = "Latest growth"
    ws["D4"].font = LABEL_FONT
    last_metrics_row = 1 + len(metrics)        # header at 1, last data at 1+N
    kpi_rows = [
        ("MoM %", f"=Metrics!C{last_metrics_row}"),
        ("QoQ %", f"=Metrics!D{last_metrics_row}"),
        ("YoY %", f"=Metrics!E{last_metrics_row}"),
    ]
    for i, (k, formula) in enumerate(kpi_rows, start=5):
        ws.cell(row=i, column=4, value=k).font = LABEL_FONT
        c = ws.cell(row=i, column=5, value=formula)
        c.number_format = PCT_FMT_DISP
        c.font = GREEN_LINK

    # --- Projection summary -------------------------------------------------
    end = 13
    ws.cell(row=end, column=1, value="2-Year Projection (last point)").font = LABEL_FONT
    proj_last_row = 1 + len(proj.horizon)      # in Projections sheet
    ws.cell(row=end + 1, column=1, value="Linear trend").font = LABEL_FONT
    c = ws.cell(row=end + 1, column=2, value=f"=Projections!B{proj_last_row}")
    c.number_format = NUM_FMT; c.font = GREEN_LINK
    ws.cell(row=end + 2, column=1, value="Holt-Winters mean").font = LABEL_FONT
    c = ws.cell(row=end + 2, column=2, value=f"=Projections!C{proj_last_row}")
    c.number_format = NUM_FMT; c.font = GREEN_LINK
    ws.cell(row=end + 3, column=1, value="HW 95% lower").font = LABEL_FONT
    c = ws.cell(row=end + 3, column=2, value=f"=Projections!D{proj_last_row}")
    c.number_format = NUM_FMT; c.font = GREEN_LINK
    ws.cell(row=end + 4, column=1, value="HW 95% upper").font = LABEL_FONT
    c = ws.cell(row=end + 4, column=2, value=f"=Projections!E{proj_last_row}")
    c.number_format = NUM_FMT; c.font = GREEN_LINK

    ws.cell(row=end + 6, column=1, value="Projection notes").font = LABEL_FONT
    ws.cell(row=end + 7, column=1, value=f"Linear: {proj.method_notes['linear']}")
    ws.cell(row=end + 8, column=1, value=f"HW: {proj.method_notes['holt_winters']}")

    _autosize(ws, {"A": 24, "B": 22, "C": 4, "D": 18, "E": 14})


def _write_raw(wb, series: Series) -> None:
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Date"; ws["B1"] = "Value"
    _style_header_row(ws, 1, 2)
    for i, (d, v) in enumerate(series.data.items(), start=2):
        ws.cell(row=i, column=1, value=d.to_pydatetime()).number_format = "yyyy-mm-dd"
        c = ws.cell(row=i, column=2, value=float(v))
        c.number_format = NUM_FMT
        c.font = BLUE_INPUT  # raw inputs
    _zebra(ws, 2, 1 + len(series.data), 2)
    _autosize(ws, {"A": 14, "B": 18})
    ws.freeze_panes = "A2"


def _write_metrics(wb, series: Series, metrics: pd.DataFrame) -> None:
    """Write Metrics tab with FORMULAS that reference Raw Data."""
    ws = wb.create_sheet("Metrics")
    ws.sheet_view.showGridLines = False
    headers = ["Date", "Level", "MoM %", "QoQ %", "YoY %", "12M Roll Avg"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    ppy = _periods_per_year(series.frequency)
    qoq_lag = max(ppy // 4, 1)
    yoy_lag = ppy
    n = len(series.data)

    for i in range(n):
        excel_row = i + 2                       # row in Metrics sheet
        raw_row   = i + 2                       # row in Raw Data sheet
        ws.cell(row=excel_row, column=1, value=f"='Raw Data'!A{raw_row}").number_format = "yyyy-mm-dd"
        c = ws.cell(row=excel_row, column=2, value=f"='Raw Data'!B{raw_row}")
        c.number_format = NUM_FMT; c.font = GREEN_LINK

        # MoM
        if i >= 1:
            f = f"=IFERROR(('Raw Data'!B{raw_row}/'Raw Data'!B{raw_row-1}-1)*100,\"\")"
        else:
            f = ""
        c = ws.cell(row=excel_row, column=3, value=f); c.number_format = PCT_FMT_DISP

        # QoQ
        if i >= qoq_lag:
            f = f"=IFERROR(('Raw Data'!B{raw_row}/'Raw Data'!B{raw_row-qoq_lag}-1)*100,\"\")"
        else:
            f = ""
        c = ws.cell(row=excel_row, column=4, value=f); c.number_format = PCT_FMT_DISP

        # YoY
        if i >= yoy_lag:
            f = f"=IFERROR(('Raw Data'!B{raw_row}/'Raw Data'!B{raw_row-yoy_lag}-1)*100,\"\")"
        else:
            f = ""
        c = ws.cell(row=excel_row, column=5, value=f); c.number_format = PCT_FMT_DISP

        # 12M (or ppy-period) rolling average
        win = ppy
        if i + 1 >= max(2, win // 2):
            start = max(2, raw_row - win + 1)
            f = f"=IFERROR(AVERAGE('Raw Data'!B{start}:B{raw_row}),\"\")"
        else:
            f = ""
        c = ws.cell(row=excel_row, column=6, value=f)
        c.number_format = NUM_FMT; c.font = BLACK_FORMULA

    _zebra(ws, 2, 1 + n, len(headers))
    _autosize(ws, {"A": 14, "B": 14, "C": 12, "D": 12, "E": 12, "F": 14})
    ws.freeze_panes = "A2"


def _write_projections(wb, series: Series, proj: Projection) -> None:
    ws = wb.create_sheet("Projections")
    ws.sheet_view.showGridLines = False
    headers = ["Date", "Linear trend", "Holt-Winters mean", "HW 95% lower", "HW 95% upper"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    n = len(proj.horizon)
    for i in range(n):
        r = i + 2
        ws.cell(row=r, column=1, value=proj.horizon[i].to_pydatetime()).number_format = "yyyy-mm-dd"
        for col, vals in enumerate(
            [proj.linear, proj.hw_mean, proj.hw_lower, proj.hw_upper], start=2
        ):
            c = ws.cell(row=r, column=col, value=float(vals.iloc[i]))
            c.number_format = NUM_FMT
            c.font = BLUE_INPUT  # forecasts come from Python — treat as inputs

    _zebra(ws, 2, 1 + n, len(headers))
    _autosize(ws, {"A": 14, "B": 16, "C": 18, "D": 16, "E": 16})
    ws.freeze_panes = "A2"


def _write_chart(wb, series: Series, proj: Projection) -> None:
    ws = wb.create_sheet("Chart")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{series.code} — history + 2-year projection"
    ws["A1"].font = TITLE_FONT

    # combined data block: history then forecast, with separate columns
    n_hist = len(series.data)
    n_fut = len(proj.horizon)
    ws["A3"] = "Date"
    ws["B3"] = "History"
    ws["C3"] = "Linear trend"
    ws["D3"] = "Holt-Winters"
    _style_header_row(ws, 3, 4)

    for i, (d, v) in enumerate(series.data.items(), start=4):
        ws.cell(row=i, column=1, value=d.to_pydatetime()).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=float(v)).number_format = NUM_FMT
    fut_start = 4 + n_hist
    for j in range(n_fut):
        r = fut_start + j
        ws.cell(row=r, column=1, value=proj.horizon[j].to_pydatetime()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3, value=float(proj.linear.iloc[j])).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=float(proj.hw_mean.iloc[j])).number_format = NUM_FMT

    last_row = fut_start + n_fut - 1
    chart = LineChart()
    chart.title = f"{series.code}: levels + 2y projection"
    chart.y_axis.title = "Level"
    chart.x_axis.title = "Date"
    chart.height = 12
    chart.width = 22

    cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
    for col in (2, 3, 4):
        data = Reference(ws, min_col=col, min_row=3, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F3")
    _autosize(ws, {"A": 14, "B": 14, "C": 14, "D": 14})


def build_workbook(
    code: str,
    out_path: str | Path = None,
    csv_dir: str | Path = "data",
    years: int = 2,
) -> Path:
    series = fetch_series(code, csv_dir=csv_dir)
    metrics = compute_metrics(series)
    proj = project(series, years=years)

    if out_path is None:
        safe = re.sub(r"[^A-Za-z0-9_]+", "_", code)
        out_path = f"{safe}_metrics.xlsx"
    out_path = Path(out_path)

    wb = Workbook()
    _write_summary(wb, series, metrics, proj)
    _write_raw(wb, series)
    _write_metrics(wb, series, metrics)
    _write_projections(wb, series, proj)
    _write_chart(wb, series, proj)
    wb.save(out_path)

    print(f"[ok] {series.source}  |  {len(series.data)} obs  |  wrote {out_path}")
    return out_path


if __name__ == "__main__":
    import sys
    code = sys.argv[1] if len(sys.argv) > 1 else "GDPH@USECON"
    build_workbook(code)
